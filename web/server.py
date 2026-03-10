#!/usr/bin/env python3
import argparse
import csv
import json
import os
import sqlite3
import urllib.parse
import urllib.error
import hashlib
import base64
import re
import subprocess
import tempfile
import shutil
import urllib.request
import threading
import time
import secrets
import smtplib
from io import BytesIO
from copy import copy as shallow_copy
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer, ThreadingHTTPServer
from pathlib import Path
import unicodedata
from email.message import EmailMessage

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    load_workbook = None
    OPENPYXL_AVAILABLE = False


ROOT = Path(__file__).resolve().parent
ASSETS = ROOT.parent / "assets"
UPLOADS = ROOT / "uploads"
DB_DEFAULT = ROOT.parent / "data" / "erp_import2.sqlite"
OCR_DB_DEFAULT = ROOT.parent / "data" / "ocr_jobs.sqlite"
TESSDATA_DIR = "/opt/homebrew/share/tessdata"
POSTAL_CATALOG_PATH = ROOT.parent / "data" / "catalogos" / "postal_catalogo.csv"
ENV_PATH = ROOT.parent / ".env"
SEGUROS_COMPANY_HINTS_PATH = ROOT.parent / "data" / "seguros_company_hints.json"
S3_BOTO3_AVAILABLE = True

def load_env_file():
    if not ENV_PATH.exists():
        return
    try:
        with ENV_PATH.open("r", encoding="utf-8") as handle:
            for line in handle:
                raw = line.strip()
                if not raw or raw.startswith("#") or "=" not in raw:
                    continue
                key, value = raw.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        return

load_env_file()
DB_CONFIGURED = Path(os.environ.get("DB_PATH") or os.environ.get("DATABASE_PATH") or str(DB_DEFAULT))
OCR_DB_CONFIGURED = Path(
    os.environ.get("OCR_DB_PATH")
    or os.environ.get("DATABASE_OCR_PATH")
    or str(OCR_DB_DEFAULT)
)
GESTORIA_EXCEL_TEMPLATE = ROOT.parent / "templates" / "Plantilla conversor asientos facturas.xlsx"
S3_BUCKET = os.environ.get("AWS_S3_BUCKET") or os.environ.get("S3_BUCKET")
S3_REGION = os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION")
OCR_SUBPROCESS_TIMEOUT_SECONDS = max(15, int(os.environ.get("OCR_SUBPROCESS_TIMEOUT_SECONDS", "90")))
OCR_JOB_STALE_MINUTES = max(1, int(os.environ.get("OCR_JOB_STALE_MINUTES", "15")))
OCR_WORKERS = max(1, min(8, int(os.environ.get("OCR_WORKERS", "2"))))
OCR_PDF_MAX_PAGES = max(0, int(os.environ.get("OCR_PDF_MAX_PAGES", "4")))
OCR_PDF_DPI = max(120, int(os.environ.get("OCR_PDF_DPI", "280")))
OCR_OPENAI_VISION_PAGES = max(0, int(os.environ.get("OCR_OPENAI_VISION_PAGES", "2")))
OCR_OPENAI_VISION_DPI = max(120, int(os.environ.get("OCR_OPENAI_VISION_DPI", "220")))
OCR_EXPERT_MODE = os.environ.get("OCR_EXPERT_MODE", "1").strip().lower() not in ("0", "false", "no", "off")
APP_SESSION_TTL_SECONDS = max(900, int(os.environ.get("APP_SESSION_TTL_SECONDS", "43200")))
SESSION_COOKIE_NAME = os.environ.get("APP_SESSION_COOKIE", "crm_session")
AUTH_ALLOW_FIRST_PASSWORD_SET = os.environ.get("AUTH_ALLOW_FIRST_PASSWORD_SET", "1").strip().lower() not in ("0", "false", "no", "off")
AUTH_INVITE_TTL_SECONDS = max(1800, int(os.environ.get("AUTH_INVITE_TTL_SECONDS", "172800")))
AUTH_PUBLIC_GET_ENDPOINTS = {"/api/health", "/api/me", "/api/auth_invite_status"}
AUTH_PUBLIC_POST_ENDPOINTS = {"/api/login", "/api/logout", "/api/auth_set_password"}
AUTH_SESSIONS = {}
AUTH_SESSIONS_LOCK = threading.Lock()


def parse_ocr_psms(raw):
    value = (raw or "").strip()
    if not value:
        return (6, 11)
    psms = []
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        try:
            psm = int(item)
        except ValueError:
            continue
        if 0 <= psm <= 13 and psm not in psms:
            psms.append(psm)
    return tuple(psms) if psms else (6, 11)


OCR_TESSERACT_PSMS = parse_ocr_psms(os.environ.get("OCR_TESSERACT_PSMS", "6,11"))


def parse_ocr_dpis(raw, base_dpi):
    value = (raw or "").strip()
    dpis = []
    if value:
        for item in value.split(","):
            item = item.strip()
            if not item:
                continue
            try:
                dpi = int(item)
            except ValueError:
                continue
            if dpi >= 120 and dpi not in dpis:
                dpis.append(dpi)
    if base_dpi not in dpis:
        dpis.insert(0, base_dpi)
    if not dpis:
        dpis = [base_dpi, 340]
    return tuple(dpis)


OCR_PDF_DPI_VARIANTS = parse_ocr_dpis(os.environ.get("OCR_PDF_DPI_VARIANTS", ""), OCR_PDF_DPI)

COMPANY_ALIAS_PATTERNS = [
    (r"\bZURICH\b", "Zurich"),
    (r"\bMAPFRE\b", "Mapfre"),
    (r"\bAXA\b", "AXA"),
    (r"\bALLIANZ\b", "Allianz"),
    (r"\bGENERALI\b", "Generali"),
    (r"\bREALE\b", "Reale"),
    (r"\bOCASO\b", "Ocaso"),
    (r"\bPELAYO\b", "Pelayo"),
    (r"\bSANTA\s*LUCIA\b|\bSANTALUCIA\b", "Santa Lucia"),
    (r"\bFIATC\b", "Fiatc"),
    (r"\bLINEA\s*DIRECTA\b", "Línea Directa"),
    (r"\bLIBERTY\b", "Liberty"),
    (r"\bMUTUA\s*MADRILE[NÑ]A\b", "Mutua Madrileña"),
    (r"\bGRUPO\s+MUTUA\s+PROPIETARIOS\b|\bMUTUA\s+PROPIETARIOS\b", "Mutua Propietarios"),
    (r"\bCAJA\s*RURAL\b", "Caja Rural"),
    (r"\bCASER\b", "Caser"),
    (r"\bPLUS\s*ULTRA\b", "Plus Ultra"),
    (r"\bFENIX\s*DIRECTO\b", "Fénix Directo"),
    (r"\bDIRECT\s*SEGUROS\b", "Direct Seguros"),
    (r"\bHEL\s*VETIA\b|\bHELVETIA\b", "Helvetia"),
    (r"\bGROUPAMA\b", "Groupama"),
    (r"\bNATIONALE\s*NEDERLANDEN\b", "Nationale Nederlanden"),
    (r"\bDAS\b", "DAS"),
    (r"\bARAG\b", "ARAG"),
    (r"\bPREVISORA\b", "Previsora General"),
    (r"\bPREVISORA\s*GENERAL\b", "Previsora General"),
    (r"\bSANITAS\b", "Sanitas"),
    (r"\bDKV\b", "DKV"),
    (r"\bADESLAS\b", "Adeslas"),
    (r"\bASISA\b", "Asisa"),
    (r"\bCATALANA\s*OCCIDENTE\b", "Catalana Occidente"),
    (r"\bNORTEHISPANA\b", "NorteHispana"),
    (r"\bSEGUROS\s*BILBAO\b|\bBILBAO\s*SEGUROS\b", "Seguros Bilbao"),
]

POSTAL_PROVINCES = {
    "01": "Álava",
    "02": "Albacete",
    "03": "Alicante",
    "04": "Almería",
    "05": "Ávila",
    "06": "Badajoz",
    "07": "Islas Baleares",
    "08": "Barcelona",
    "09": "Burgos",
    "10": "Cáceres",
    "11": "Cádiz",
    "12": "Castellón",
    "13": "Ciudad Real",
    "14": "Córdoba",
    "15": "A Coruña",
    "16": "Cuenca",
    "17": "Girona",
    "18": "Granada",
    "19": "Guadalajara",
    "20": "Guipúzcoa",
    "21": "Huelva",
    "22": "Huesca",
    "23": "Jaén",
    "24": "León",
    "25": "Lleida",
    "26": "La Rioja",
    "27": "Lugo",
    "28": "Madrid",
    "29": "Málaga",
    "30": "Murcia",
    "31": "Navarra",
    "32": "Ourense",
    "33": "Asturias",
    "34": "Palencia",
    "35": "Las Palmas",
    "36": "Pontevedra",
    "37": "Salamanca",
    "38": "Santa Cruz de Tenerife",
    "39": "Cantabria",
    "40": "Segovia",
    "41": "Sevilla",
    "42": "Soria",
    "43": "Tarragona",
    "44": "Teruel",
    "45": "Toledo",
    "46": "Valencia",
    "47": "Valladolid",
    "48": "Vizcaya",
    "49": "Zamora",
    "50": "Zaragoza",
    "51": "Ceuta",
    "52": "Melilla",
}

def normalize_postal_code(value):
    code = re.sub(r"\D", "", str(value or ""))[:5]
    if len(code) == 4:
        code = f"0{code}"
    return code if len(code) == 5 else ""

def normalize_phone(value):
    if not value:
        return ""
    digits = re.sub(r"\D+", "", str(value))
    if not digits:
        return ""
    if digits.startswith("34") and len(digits) > 9:
        digits = digits[-9:]
    if len(digits) == 9:
        return digits
    return ""

def parse_services_param(raw):
    if not raw:
        return []
    parts = re.split(r"[|,;/]+", str(raw))
    services = [p.strip().lower() for p in parts if p and p.strip()]
    aliases = {
        "gestoria": "gestoría",
        "administracion fincas": "administración fincas",
        "administracion de fincas": "administración de fincas",
        "inversion": "inversión",
        "direccion": "dirección",
    }
    expanded = []
    for service in services:
        expanded.append(service)
        if service in aliases:
            expanded.append(aliases[service])
    return list(dict.fromkeys(expanded))

def cliente_has_servicio(conn, cliente_id, servicios):
    if not cliente_id or not servicios:
        return True
    placeholders = ",".join(["?"] * len(servicios))
    row = conn.execute(
        f"""
        SELECT 1
        FROM clientes_empresas
        WHERE cliente_id = ?
          AND LOWER(servicio) IN ({placeholders})
        LIMIT 1
        """,
        [cliente_id, *servicios],
    ).fetchone()
    return bool(row)
    return digits

def normalize_email(value):
    if not value:
        return ""
    return str(value).strip().lower()

def normalize_person_name(value):
    if not value:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text

def normalize_company_name(value):
    if not value:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def uploaded_policy_filter(alias=""):
    prefix = f"{alias}." if alias else ""
    return f"(COALESCE({prefix}poliza_key, '') <> '' OR COALESCE({prefix}poliza_url, '') <> '')"


def in_vigor_policy_filter(alias=""):
    prefix = f"{alias}." if alias else ""
    estado_expr = f"LOWER(TRIM(COALESCE({prefix}estado, '')))"
    estado_poliza_expr = f"LOWER(TRIM(COALESCE({prefix}estado_poliza, '')))"
    # Considera variaciones operativas reales en producción (activo/activa/alta/emitida).
    return (
        f"({estado_expr} IN ("
        "'en vigor', 'en_vigor', 'vigente', 'poliza', 'póliza', 'poliza en vigor', "
        "'activo', 'activa', 'alta', 'emitida', 'recibido'"
        f") OR {estado_poliza_expr} IN ('activa', 'activo', 'en vigor', 'vigente'))"
    )

def normalize_poliza_key(value):
    if not value:
        return ""
    text = re.sub(r"\s+", "", str(value).upper())
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text

def normalize_company_key(value):
    if not value:
        return ""
    text = normalize_company_name(value).upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


LEGAL_RAMOS_CANONICAL = (
    "Vida",
    "Accidentes",
    "Salud",
    "Decesos",
    "Auto",
    "Hogar",
    "Comercio",
    "Comunidad",
    "Responsabilidad civil",
    "Defensa jurídica",
    "Protección de pagos",
    "Viaje",
    "Ahorro",
    "Caza",
)


def canonicalize_ramo(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    key = normalize_lookup_text(raw)
    if not key:
        return ""
    canonical_by_key = {normalize_lookup_text(item): item for item in LEGAL_RAMOS_CANONICAL}
    if key in canonical_by_key:
        return canonical_by_key[key]
    # Basura típica OCR/legal: párrafos completos o fragmentos normativos.
    noisy_markers = (
        "EXPECTATIVA RAZONABLE",
        "DIRECCION GENERAL DE SEGUROS",
        "SE CONSIDERARA ACEPTADA",
        "PRODUCTOS COMERCIALIZADOS",
        "ENTRE LOS QUE ESTA EL SEGURO",
        "FEC EFECTO PERIODO PAGO ENVIO DOC",
    )
    if len(key) > 42 or any(marker in key for marker in noisy_markers):
        return ""
    if key in ("S", "RC", "R C", "R C PYME", "RESPONSABILIDAD CIVIL PROFESIONAL"):
        return "Responsabilidad civil"
    if key in ("IT", "INCAPACIDAD TEMPORAL", "INCAPACIDAD PERMANENTE"):
        return "Accidentes"
    if "RESPONSABILIDAD CIVIL" in key:
        return "Responsabilidad civil"
    if key in (
        "ALQUILER",
        "IMPAGO ALQUILER",
        "PROTECCION ALQUILER",
        "PROTECCION DE ALQUILER",
        "PAGO ALQUILER",
        "HOGAR ALQUILER",
    ):
        return "Protección de pagos"
    if (
        "IMPAGO" in key
        or "PROTECCION PAGO" in key
        or "PROTECCION DE PAGO" in key
        or "PROTECCION DE PAGOS" in key
    ):
        return "Protección de pagos"
    if "DEFENSA JURIDICA" in key or key.startswith("ARAG"):
        return "Defensa jurídica"
    if "SALUD" in key or "ASISTENCIA SANITARIA" in key or "DKV INTEGRAL" in key:
        return "Salud"
    if "DECESOS" in key:
        return "Decesos"
    if "ACCIDENT" in key:
        return "Accidentes"
    if "VIDA" in key:
        return "Vida"
    if "AHORRO" in key:
        return "Ahorro"
    if "VIAJE" in key or "VIAJES" in key or "VIAJEROS" in key:
        return "Viaje"
    if "CAZA" in key:
        return "Caza"
    if "COMUNIDAD" in key:
        return "Comunidad"
    if "COMERCIO" in key or "PYME" in key or key == "LOCAL":
        return "Comercio"
    if "HOGAR" in key:
        return "Hogar"
    if (
        "AUTO" in key
        or "AUTOMOVIL" in key
        or "AUTOMOVILES" in key
        or "VEHICULO" in key
        or "MOTO" in key
        or "MOTOR" in key
    ):
        return "Auto"
    # Si no coincide pero no es ruido OCR, se conserva para no perder edición manual.
    return raw


def infer_tipo_vigencia(ramo, explicit=None):
    value = normalize_lookup_text(explicit or "")
    if value in ("TEMPORAL_NO_RENOVABLE", "TEMPORAL", "UN_USO", "NO_RENOVABLE"):
        return "temporal_no_renovable"
    if value in ("ANUAL_RENOVABLE", "RENOVABLE", "ANUAL"):
        return "anual_renovable"
    ramo_key = normalize_lookup_text(ramo or "")
    if any(token in ramo_key for token in ("VIAJE", "VIAJES", "CAZA EVENTUAL")):
        return "temporal_no_renovable"
    return "anual_renovable"


def log_seguro_event(conn, seguro_row, event_type, now, motivo="", payload=None):
    if not seguro_row:
        return
    payload_json = ""
    if payload:
        try:
            payload_json = json.dumps(payload, ensure_ascii=False)
        except Exception:
            payload_json = str(payload)
    conn.execute(
        """
        INSERT INTO seguros_eventos (
          id, seguro_id, cliente_id, empresa_id, tipo, fecha, motivo, payload_json, created_at
        ) VALUES (
          ?, ?, ?, ?, ?, ?, ?, ?, datetime(?)
        )
        """,
        (
            os.urandom(16).hex(),
            seguro_row.get("id") if isinstance(seguro_row, dict) else seguro_row["id"],
            (seguro_row.get("cliente_id") if isinstance(seguro_row, dict) else seguro_row["cliente_id"]),
            (seguro_row.get("empresa_id") if isinstance(seguro_row, dict) else seguro_row["empresa_id"]),
            event_type,
            (seguro_row.get("fecha_efecto") if isinstance(seguro_row, dict) else seguro_row["fecha_efecto"]) or now[:10],
            motivo or "",
            payload_json,
            now,
        ),
    )


def upsert_seguro_comision_contabilidad(conn, seguro_row, now, movimiento="emision", fecha=None, importe=None):
    if not seguro_row:
        return None
    seguro_id = (seguro_row.get("id") if isinstance(seguro_row, dict) else seguro_row["id"]) or ""
    empresa_id = (seguro_row.get("empresa_id") if isinstance(seguro_row, dict) else seguro_row["empresa_id"]) or ""
    if not seguro_id or not empresa_id:
        return None
    cliente_id = (seguro_row.get("cliente_id") if isinstance(seguro_row, dict) else seguro_row["cliente_id"]) or None
    cliente_ids_json = json.dumps([cliente_id], ensure_ascii=False) if cliente_id else None
    poliza_numero = (seguro_row.get("poliza_numero") if isinstance(seguro_row, dict) else seguro_row["poliza_numero"]) or ""
    comision_src = importe
    if comision_src in (None, ""):
        comision_src = seguro_row.get("comision") if isinstance(seguro_row, dict) else seguro_row["comision"]
    comision = round(parse_money_value(comision_src), 2)
    if abs(comision) < 0.005:
        return None
    fecha_base = fecha
    if not fecha_base:
        fecha_base = seguro_row.get("fecha_efecto") if isinstance(seguro_row, dict) else seguro_row["fecha_efecto"]
    fecha_iso = ""
    parsed_fecha = parse_iso_date(fecha_base)
    if parsed_fecha:
        fecha_iso = parsed_fecha.isoformat()
    if not fecha_iso:
        return None
    is_renovacion = normalize_lookup_text(movimiento) in ("RENOVACION", "RENOVAR", "RENEW")
    gestion = "Comisión renovación" if is_renovacion else "Comisión emisión"
    concepto = "Comisión renovación póliza" if is_renovacion else "Comisión emisión póliza"
    if poliza_numero:
        concepto = f"{concepto} {poliza_numero}"
    existing = conn.execute(
        """
        SELECT id
        FROM gestoria_contabilidad
        WHERE seguro_id = ?
          AND fecha = ?
          AND LOWER(COALESCE(gestion, '')) = LOWER(?)
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (seguro_id, fecha_iso, gestion),
    ).fetchone()
    notas = f"Auto CRM Seguros · {gestion.lower()}."
    if existing:
        conn.execute(
            """
            UPDATE gestoria_contabilidad
            SET empresa_id = ?, cliente_id = ?, cliente_ids_json = ?, poliza_numero = ?, concepto = ?, tipo = ?, importe = ?, notas = ?,
                updated_at = datetime(?)
            WHERE id = ?
            """,
            (
                empresa_id,
                cliente_id,
                cliente_ids_json,
                poliza_numero,
                concepto,
                "Ingreso",
                comision,
                notas,
                now,
                existing["id"],
            ),
        )
        return existing["id"]
    record_id = os.urandom(16).hex()
    conn.execute(
        """
        INSERT INTO gestoria_contabilidad (
          id, empresa_id, cliente_id, cliente_ids_json, seguro_id, poliza_numero, fecha, concepto, gestion, tipo, importe, notas,
          created_at, updated_at
        ) VALUES (
          ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
        )
        """,
        (
            record_id,
            empresa_id,
            cliente_id,
            cliente_ids_json,
            seguro_id,
            poliza_numero,
            fecha_iso,
            concepto,
            gestion,
            "Ingreso",
            comision,
            notas,
            now,
            now,
        ),
    )
    return record_id


def resolve_seguro_contabilidad_link(conn, seguro_id):
    seguro_id = str(seguro_id or "").strip()
    if not seguro_id:
        return "", None
    row = conn.execute(
        "SELECT poliza_numero, cliente_id FROM seguros WHERE id = ? LIMIT 1",
        (seguro_id,),
    ).fetchone()
    if not row:
        return "", None
    poliza_numero = (row["poliza_numero"] or "").strip()
    cliente_id = (row["cliente_id"] or "").strip() or None
    return poliza_numero, cliente_id


def parse_cliente_ids_payload(raw_value):
    if raw_value in (None, ""):
        return []
    items = raw_value
    if isinstance(raw_value, str):
        text = raw_value.strip()
        if not text:
            return []
        if text.startswith("["):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    items = parsed
                else:
                    items = [parsed]
            except Exception:
                items = [part.strip() for part in text.split(",")]
        else:
            items = [part.strip() for part in text.split(",")]
    elif not isinstance(raw_value, (list, tuple, set)):
        items = [raw_value]
    out = []
    for item in items:
        value = str(item or "").strip()
        if value and value not in out:
            out.append(value)
    return out


def seguros_contabilidad_where_clause(alias="gc"):
    p = alias.strip() or "gc"
    return (
        f"(({p}.seguro_id IS NOT NULL AND TRIM({p}.seguro_id) <> '') "
        f"OR UPPER(COALESCE({p}.notas, '')) LIKE 'AUTO CRM SEGUROS%' "
        f"OR UPPER(COALESCE({p}.notas, '')) LIKE '[SEGUROS]%' "
        f"OR UPPER(TRIM(COALESCE({p}.gestion, ''))) IN ('COMISION EMISION', 'COMISION RENOVACION', 'REGULARIZACION', 'EXTORNO'))"
    )


def find_existing_seguro_id(conn, empresa_id, poliza_numero, compania, exclude_id=None):
    poliza_norm = normalize_poliza_key(poliza_numero)
    if not poliza_norm:
        return ""
    compania_norm = normalize_company_key(compania)
    rows = conn.execute(
        "SELECT id, poliza_numero, compania, estado FROM seguros WHERE empresa_id = ?",
        (empresa_id,),
    ).fetchall()
    for row in rows:
        row_id = row["id"]
        if exclude_id and row_id == exclude_id:
            continue
        estado_key = normalize_lookup_text(row["estado"] or "")
        if "MIGRADO LEGADO" in estado_key or "LEGACY" in estado_key:
            continue
        row_poliza = normalize_poliza_key(row["poliza_numero"])
        if not row_poliza or row_poliza != poliza_norm:
            continue
        row_comp = normalize_company_key(row["compania"])
        if compania_norm and row_comp and compania_norm != row_comp:
            continue
        return row_id
    return ""


def normalize_lookup_text(value):
    if not value:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_service_key(value):
    text = normalize_lookup_text(value)
    aliases = {
        "GESTORIA": "gestoria",
        "SEGUROS": "seguros",
        "INMOBILIARIA": "inmobiliaria",
        "FINANCIACIONES": "financiaciones",
        "HIPOTECAS": "financiaciones",
        "ADMINISTRACION FINCAS": "administracion fincas",
        "ADMINISTRACION DE FINCAS": "administracion fincas",
    }
    return aliases.get(text, text.lower().strip())


def is_active_service_state(value, fecha_fin=None):
    state = normalize_lookup_text(value or "")
    if state and state in {"INACTIVO", "BAJA", "CANCELADO", "ANULADO", "FINALIZADO"}:
        return False
    end_date = parse_iso_date(fecha_fin)
    if end_date and end_date < datetime.now(timezone.utc).date():
        return False
    return True


def parse_iso_date(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    head = raw.split("T", 1)[0].split(" ", 1)[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(head, fmt).date()
        except Exception:
            continue
    return None


def add_year_to_date(value):
    base = parse_iso_date(value)
    if not base:
        return ""
    try:
        return base.replace(year=base.year + 1).isoformat()
    except ValueError:
        # 29/02 -> 28/02 en años no bisiestos
        return base.replace(month=2, day=28, year=base.year + 1).isoformat()


def parse_money_value(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if not text or text in ("-", "."):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def has_explicit_renewal_action(seguro_row):
    estado = normalize_lookup_text(seguro_row.get("estado"))
    estado_ren = normalize_lookup_text(seguro_row.get("estado_renovacion"))
    if seguro_row.get("renovacion_fecha") or seguro_row.get("nueva_poliza_ref"):
        return True
    if any(token in estado for token in ("BAJA", "CANCEL", "ANULAD")):
        return True
    if not estado_ren:
        return False
    return "AUTOM" not in estado_ren


def compute_seguro_display(seguro_row):
    fecha_efecto = seguro_row.get("fecha_efecto")
    venc = parse_iso_date(seguro_row.get("fecha_vencimiento")) or parse_iso_date(add_year_to_date(fecha_efecto))
    base_estado = seguro_row.get("estado") or "-"
    if not venc:
        return {"vencimiento_display": "", "estado_display": base_estado}
    today = datetime.now(timezone.utc).date()
    has_action = has_explicit_renewal_action(seguro_row)
    while venc < today and not has_action:
        try:
            venc = venc.replace(year=venc.year + 1)
        except ValueError:
            venc = venc.replace(month=2, day=28, year=venc.year + 1)
    estado = base_estado
    if parse_iso_date(seguro_row.get("fecha_vencimiento")) and parse_iso_date(seguro_row.get("fecha_vencimiento")) < today and not has_action:
        estado = "Renovada automática"
    return {"vencimiento_display": venc.isoformat(), "estado_display": estado}


def load_seguros_company_hints(path=SEGUROS_COMPANY_HINTS_PATH):
    hints = {}
    if not path.exists():
        return hints
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle) or {}
    except Exception:
        return hints
    raw_hints = payload.get("hints") if isinstance(payload, dict) else {}
    if not isinstance(raw_hints, dict):
        return hints
    for token, company in raw_hints.items():
        norm_token = normalize_lookup_text(token)
        if norm_token and company:
            hints[norm_token] = normalize_company_name(company)
    return hints


LEARNED_COMPANY_HINTS = load_seguros_company_hints()


def detect_company_from_text(text):
    cleaned = normalize_lookup_text(text)
    if not cleaned:
        return ""
    for pattern, name in COMPANY_ALIAS_PATTERNS:
        if re.search(pattern, cleaned, re.IGNORECASE):
            return name
    return ""


def detect_company_from_metadata(*values):
    joined = " ".join(str(v or "") for v in values).strip()
    if not joined:
        return ""
    cleaned = normalize_lookup_text(joined)
    if not cleaned:
        return ""
    if LEARNED_COMPANY_HINTS:
        for token, company in sorted(LEARNED_COMPANY_HINTS.items(), key=lambda it: len(it[0]), reverse=True):
            if token and token in cleaned:
                return company
    return detect_company_from_text(cleaned)

def normalize_fin_nif(value):
    if not value:
        return ""
    raw = str(value).strip().upper()
    if not raw:
        return ""
    raw = raw.replace(" ", "").replace("-", "").replace(".", "")
    raw = raw.replace("O", "0").replace("I", "1").replace("L", "1").replace("S", "5")
    raw = re.sub(r"[^A-Z0-9]", "", raw)
    if len(raw) > 9:
        raw = raw[:9]
    return raw

def normalize_nif(value):
    if not value:
        return ""
    raw = str(value).strip().upper()
    if not raw:
        return ""
    raw = raw.replace(" ", "").replace("-", "").replace(".", "")
    raw = raw.replace("O", "0").replace("I", "1").replace("L", "1").replace("S", "5")
    raw = re.sub(r"[^A-Z0-9]", "", raw)
    if len(raw) > 12:
        raw = raw[:12]
    return raw

FIN_REQUIRED_FIELDS = [
    ("cliente1_nombre", "Nombre cliente 1"),
    ("cliente1_dni", "DNI cliente 1"),
    ("cliente1_telefono", "Teléfono cliente 1"),
    ("fecha", "Fecha"),
    ("ingresos_conjuntos", "Ingresos conjuntos"),
]

def fin_missing_fields(row):
    missing = []
    if not row:
        return missing
    for key, label in FIN_REQUIRED_FIELDS:
        value = row.get(key) if isinstance(row, dict) else row[key]
        if value is None or str(value).strip() == "":
            missing.append(label)
    return missing

def fin_sync_missing_action(conn, empresa_id, asesoramiento_id, cliente_id, cliente_nombre, missing, now):
    if missing:
        notas = f"Completar datos asesoramiento ({', '.join(missing)}). Asesoramiento ID: {asesoramiento_id}"
        exists = conn.execute(
            """
            SELECT id, estado FROM acciones
            WHERE servicio = 'Financiaciones'
              AND tipo = 'Completar datos asesoramiento'
              AND notas LIKE ?
            LIMIT 1
            """,
            (f"%{asesoramiento_id}%",),
        ).fetchone()
        if exists:
            if (exists["estado"] or "").lower() != "pendiente":
                conn.execute(
                    "UPDATE acciones SET estado = 'Pendiente', updated_at = datetime(?) WHERE id = ?",
                    (now, exists["id"]),
                )
            return
        today = datetime.now().strftime("%Y-%m-%d")
        conn.execute(
            """
            INSERT INTO acciones (
              id, empresa_id, servicio, cliente_id, inmueble_id, cliente_nombre,
              fecha, hora, tipo, responsable, estado, notas, recordatorio_min, created_at, updated_at
            ) VALUES (
              ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
            )
            """,
            (
                os.urandom(16).hex(),
                empresa_id,
                "Financiaciones",
                cliente_id,
                None,
                cliente_nombre or "",
                today,
                None,
                "Completar datos asesoramiento",
                None,
                "Pendiente",
                notas,
                None,
                now,
                now,
            ),
        )
    else:
        conn.execute(
            """
            UPDATE acciones
            SET estado = 'Hecho', updated_at = datetime(?)
            WHERE servicio = 'Financiaciones'
              AND tipo = 'Completar datos asesoramiento'
              AND notas LIKE ?
            """,
            (now, f"%{asesoramiento_id}%"),
        )

def normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

def load_postal_catalog(conn):
    if not POSTAL_CATALOG_PATH.exists():
        return
    count_row = conn.execute("SELECT COUNT(*) AS total FROM postal_catalogo").fetchone()
    total = 0
    if count_row:
        try:
            total = count_row["total"]
        except (TypeError, KeyError, IndexError):
            total = count_row[0]
    if total:
        return
    encodings = ("utf-8-sig", "latin-1")
    for encoding in encodings:
        try:
            with POSTAL_CATALOG_PATH.open("r", encoding=encoding, newline="") as handle:
                reader = csv.DictReader(handle)
                if not reader.fieldnames:
                    return
                headers = {normalize_header(name): name for name in reader.fieldnames}
                cp_key = next(
                    (headers[k] for k in headers if k in ("codigopostal", "codigo_postal", "postalcode", "cp", "codigo")),
                    None,
                )
                pob_key = next(
                    (headers[k] for k in headers if k in ("poblacion", "municipio", "municipionombre", "localidad", "ciudad", "nombre")),
                    None,
                )
                prov_key = next(
                    (headers[k] for k in headers if k in ("provincia", "provincianombre", "province")),
                    None,
                )
                if not cp_key:
                    return
                batch = []
                seen = set()
                for row in reader:
                    cp_raw = row.get(cp_key, "")
                    cp = normalize_postal_code(cp_raw)
                    if not cp:
                        continue
                    poblacion = (row.get(pob_key, "") if pob_key else "") or ""
                    poblacion = str(poblacion).strip()
                    provincia = (row.get(prov_key, "") if prov_key else "") or ""
                    provincia = str(provincia).strip()
                    if not provincia:
                        provincia = POSTAL_PROVINCES.get(cp[:2], "")
                    key = (cp, poblacion, provincia)
                    if key in seen:
                        continue
                    seen.add(key)
                    batch.append(key)
                    if len(batch) >= 5000:
                        conn.executemany(
                            "INSERT INTO postal_catalogo (codigo_postal, poblacion, provincia) VALUES (?, ?, ?)",
                            batch,
                        )
                        batch = []
                if batch:
                    conn.executemany(
                        "INSERT INTO postal_catalogo (codigo_postal, poblacion, provincia) VALUES (?, ?, ?)",
                        batch,
                    )
            return
        except UnicodeDecodeError:
            continue

def hash_password(password):
    if not password:
        return None
    salt = os.urandom(16).hex()
    digest = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return f"{salt}${digest}"


def verify_password(password, password_hash):
    if not password_hash or not password:
        return False
    raw = str(password_hash)
    if "$" not in raw:
        return False
    salt, stored = raw.split("$", 1)
    if not salt or not stored:
        return False
    digest = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return secrets.compare_digest(digest, stored)


def _cleanup_expired_sessions():
    now = time.time()
    expired = []
    for token, session in AUTH_SESSIONS.items():
        if float(session.get("expires_at") or 0) <= now:
            expired.append(token)
    for token in expired:
        AUTH_SESSIONS.pop(token, None)


def create_auth_session(user_row):
    now = time.time()
    session = {
        "token": secrets.token_urlsafe(32),
        "user_id": str(user_row["id"]),
        "usuario": str(user_row["usuario"] or ""),
        "nombre": str(user_row["nombre"] or ""),
        "apellido": str(user_row["apellido"] or ""),
        "rol": str(user_row["rol"] or ""),
        "email": str(user_row["email"] or ""),
        "servicio": str(user_row["servicio"] or ""),
        "expires_at": now + APP_SESSION_TTL_SECONDS,
        "created_at": now,
    }
    with AUTH_SESSIONS_LOCK:
        _cleanup_expired_sessions()
        AUTH_SESSIONS[session["token"]] = session
    return session


def get_auth_session(token):
    if not token:
        return None
    with AUTH_SESSIONS_LOCK:
        session = AUTH_SESSIONS.get(token)
        if not session:
            return None
        if float(session.get("expires_at") or 0) <= time.time():
            AUTH_SESSIONS.pop(token, None)
            return None
        session["expires_at"] = time.time() + APP_SESSION_TTL_SECONDS
        return dict(session)


def delete_auth_session(token):
    if not token:
        return
    with AUTH_SESSIONS_LOCK:
        AUTH_SESSIONS.pop(token, None)


def send_mail_smtp(subject, to_email, text_body, html_body=None):
    host = (os.environ.get("SMTP_HOST") or "").strip()
    if not host:
        raise RuntimeError("SMTP_HOST no configurado")
    port = int((os.environ.get("SMTP_PORT") or "587").strip())
    username = (os.environ.get("SMTP_USER") or "").strip()
    password = (os.environ.get("SMTP_PASS") or "").strip()
    from_email = (os.environ.get("SMTP_FROM") or username or "").strip()
    if not from_email:
        raise RuntimeError("SMTP_FROM/SMTP_USER no configurado")
    use_ssl = (os.environ.get("SMTP_SSL") or "").strip().lower() in ("1", "true", "yes", "on")
    use_tls = (os.environ.get("SMTP_TLS") or "1").strip().lower() not in ("0", "false", "no", "off")
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email
    msg.set_content(text_body)
    if html_body:
        msg.add_alternative(html_body, subtype="html")
    if use_ssl:
        server = smtplib.SMTP_SSL(host, port, timeout=20)
    else:
        server = smtplib.SMTP(host, port, timeout=20)
    try:
        server.ehlo()
        if use_tls and not use_ssl:
            server.starttls()
            server.ehlo()
        if username:
            server.login(username, password)
        server.send_message(msg)
    finally:
        try:
            server.quit()
        except Exception:
            pass

def detect_ocr_lang():
    if os.path.isdir(TESSDATA_DIR):
        has_spa = os.path.exists(os.path.join(TESSDATA_DIR, "spa.traineddata"))
        has_eng = os.path.exists(os.path.join(TESSDATA_DIR, "eng.traineddata"))
        if has_spa and has_eng:
            return "spa+eng"
        if has_spa:
            return "spa"
        if has_eng:
            return "eng"
    return "spa+eng"


def run_subprocess(args, **kwargs):
    kwargs.setdefault("timeout", OCR_SUBPROCESS_TIMEOUT_SECONDS)
    return subprocess.run(args, **kwargs)


def is_stale_ocr_job(started_at):
    if not started_at:
        return False
    raw = str(started_at).strip()
    if not raw:
        return False
    if raw.endswith("Z"):
        raw = raw[:-1]
    try:
        started_dt = datetime.fromisoformat(raw)
    except Exception:
        return False
    if started_dt.tzinfo is None:
        started_dt = started_dt.replace(tzinfo=timezone.utc)
    return datetime.now(timezone.utc) - started_dt > timedelta(minutes=OCR_JOB_STALE_MINUTES)


def s3_config():
    bucket = os.environ.get("AWS_S3_BUCKET") or os.environ.get("S3_BUCKET") or S3_BUCKET
    region = os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION") or S3_REGION
    return bucket, region

def s3_client():
    global S3_BOTO3_AVAILABLE
    try:
        import boto3
        from botocore.config import Config
    except ImportError:
        S3_BOTO3_AVAILABLE = False
        return None
    bucket, region = s3_config()
    if not bucket or not region:
        return None
    return boto3.client(
        "s3",
        region_name=region,
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "virtual"},
        ),
    )


def s3_safe_key(prefix, filename):
    base = os.path.basename(filename or "archivo.pdf")
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", base)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    rand = os.urandom(4).hex()
    prefix = prefix.strip("/").strip() if prefix else "seguros"
    return f"{prefix}/{stamp}_{rand}_{safe}"

def s3_get_object_bytes(key):
    client = s3_client()
    if not client:
        return None, "S3 no configurado"
    bucket, _region = s3_config()
    if not bucket:
        return None, "S3 sin bucket"
    try:
        obj = client.get_object(Bucket=bucket, Key=key)
        return obj["Body"].read(), ""
    except Exception as exc:
        return None, str(exc)


def decode_seguros_payload(payload):
    data_uri = payload.get("file_base64") or payload.get("data")
    s3_key = (payload.get("s3_key") or "").strip()
    pdf_bytes = None
    if s3_key:
        pdf_bytes, s3_err = s3_get_object_bytes(s3_key)
        if not pdf_bytes:
            raise ValueError(f"S3: {s3_err}")
    else:
        if not data_uri:
            raise ValueError("Archivo requerido")
        if "," in data_uri:
            data_uri = data_uri.split(",", 1)[1]
        try:
            pdf_bytes = base64.b64decode(data_uri)
        except Exception:
            raise ValueError("Base64 invalido")
    return pdf_bytes


def decode_document_payload(payload):
    data_uri = payload.get("file_base64") or payload.get("data")
    s3_key = (payload.get("s3_key") or "").strip()
    raw_bytes = None
    mime = ""
    filename = str(payload.get("filename") or "").strip()
    if s3_key:
        raw_bytes, s3_err = s3_get_object_bytes(s3_key)
        if not raw_bytes:
            raise ValueError(f"S3: {s3_err}")
        lower_key = s3_key.lower()
        if lower_key.endswith(".pdf"):
            mime = "application/pdf"
        elif lower_key.endswith((".jpg", ".jpeg")):
            mime = "image/jpeg"
        elif lower_key.endswith(".png"):
            mime = "image/png"
    else:
        if not data_uri:
            raise ValueError("Archivo requerido")
        if "," in data_uri:
            header, data_uri = data_uri.split(",", 1)
            if header.startswith("data:") and ";base64" in header:
                mime = header.split(":", 1)[1].split(";", 1)[0]
        try:
            raw_bytes = base64.b64decode(data_uri)
        except Exception:
            raise ValueError("Base64 invalido")
    if not raw_bytes:
        raise ValueError("Archivo vacio")
    source_hint = " ".join(
        [
            filename,
            str(payload.get("source_hint") or ""),
            s3_key,
        ]
    ).strip()
    return raw_bytes, mime, source_hint


def parse_decimal_eu(value):
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    raw = raw.replace("€", "").replace("EUR", "").replace(" ", "")
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in ("", ".", "-", "-."):
        return 0.0
    try:
        return float(raw)
    except Exception:
        return 0.0


def extract_invoice_amount(text, labels):
    if not text:
        return 0.0
    for label in labels:
        pattern = rf"{label}\s*[:\-]?\s*([0-9][0-9\.\,\s€]*)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            value = parse_decimal_eu(match.group(1))
            if value > 0:
                return value
    return 0.0


def parse_invoice_text(text):
    raw = str(text or "")
    upper = normalize_lookup_text(raw)
    if not raw.strip():
        return {}
    numero = ""
    numero_match = re.search(
        r"(?:factura|fra\.?|n[úu]m(?:ero)?|n[º°])\s*[:#-]?\s*([A-Z0-9][A-Z0-9\-\/\.]{2,})",
        raw,
        re.IGNORECASE,
    )
    if numero_match:
        numero = numero_match.group(1).strip()
    fecha = ""
    fecha_match = re.search(r"\b(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})\b", raw)
    if fecha_match:
        fecha = fecha_match.group(1).strip()
    fecha_iso = ""
    if fecha:
        date_bits = re.split(r"[\/\-]", fecha)
        if len(date_bits) == 3:
            day, month, year = date_bits
            if len(year) == 2:
                year = f"20{year}"
            try:
                fecha_iso = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
            except Exception:
                fecha_iso = ""
    nif_candidates = re.findall(
        r"\b(?:[ABCDEFGHJNPQRSUVW]\d{7}[0-9A-J]|\d{8}[A-Z])\b",
        upper,
        re.IGNORECASE,
    )
    nif = nif_candidates[0].upper() if nif_candidates else ""
    tercero = ""
    for pattern in (
        r"(?:proveedor|acreedor)\s*[:\-]\s*([^\n\r]{3,80})",
        r"(?:raz[oó]n social)\s*[:\-]\s*([^\n\r]{3,80})",
        r"(?:cliente)\s*[:\-]\s*([^\n\r]{3,80})",
    ):
        m = re.search(pattern, raw, re.IGNORECASE)
        if m:
            tercero = m.group(1).strip(" .;:")
            break
    tipo = "compra"
    if any(token in upper for token in ("FACTURA EMITIDA", "TOTAL A COBRAR", "CLIENTE")):
        tipo = "venta"
    base = extract_invoice_amount(raw, ["base imponible", "subtotal", "base"])
    cuota_iva = extract_invoice_amount(raw, ["cuota iva", "iva", "i\\.v\\.a\\."])
    cuota_irpf = extract_invoice_amount(raw, ["retencion", "irpf"])
    total = extract_invoice_amount(raw, ["total factura", "importe total", "total a pagar", "total"])
    if base <= 0 and total > 0:
        base = max(0.0, total - cuota_iva + cuota_irpf)
    if total <= 0 and base > 0:
        total = max(0.0, base + cuota_iva - cuota_irpf)
    iva_pct = 0.0
    pct_match = re.search(r"\b(4|10|21)(?:[.,]0+)?\s*%\b", raw)
    if pct_match:
        iva_pct = parse_decimal_eu(pct_match.group(1))
    elif base > 0 and cuota_iva > 0:
        iva_pct = round((cuota_iva / base) * 100.0, 2)
    descripcion = numero or "Factura"
    if tercero:
        descripcion = f"{descripcion} · {tercero}"
    return {
        "numero": numero,
        "fecha": fecha_iso,
        "nif": nif,
        "tercero": tercero,
        "tipo": tipo,
        "base_imponible": round(base, 2),
        "cuota_iva": round(cuota_iva, 2),
        "cuota_irpf": round(cuota_irpf, 2),
        "total": round(total, 2),
        "iva_pct": iva_pct,
        "descripcion": descripcion,
        "raw_text": raw.strip(),
    }


def infer_expense_account(concepto):
    text = normalize_lookup_text(concepto or "")
    mapping = [
        ("ALQUILER", "621"),
        ("ARRENDAMIENTO", "621"),
        ("REPARACION", "622"),
        ("CONSERVACION", "622"),
        ("PROFESIONAL", "623"),
        ("HONORARIO", "623"),
        ("TRANSPORTE", "624"),
        ("SEGURO", "625"),
        ("BANC", "626"),
        ("PUBLICIDAD", "627"),
        ("SUMINISTRO", "628"),
        ("LUZ", "628"),
        ("AGUA", "628"),
        ("TELEFON", "628"),
        ("INTERNET", "628"),
        ("TRIBUTO", "631"),
        ("IMPUESTO", "631"),
    ]
    for token, account in mapping:
        if token in text:
            return account
    return "629"


def infer_revenue_account(concepto):
    text = normalize_lookup_text(concepto or "")
    if "ALQUILER" in text:
        return "705"
    return "700"


def ensure_gestoria_tercero(conn, empresa_id, nif, nombre, tipo, now):
    nif_clean = (nif or "").strip().upper()
    name_clean = (nombre or "").strip()
    kind = (tipo or "").strip().lower() or "proveedor"
    account_map = {"proveedor": "400", "acreedor": "410", "cliente": "430"}
    account = account_map.get(kind, "400")
    row = None
    if nif_clean:
        row = conn.execute(
            "SELECT id, cuenta_contable FROM gestoria_terceros WHERE empresa_id = ? AND UPPER(COALESCE(nif,'')) = ? LIMIT 1",
            (empresa_id, nif_clean),
        ).fetchone()
    if not row and name_clean:
        row = conn.execute(
            "SELECT id, cuenta_contable FROM gestoria_terceros WHERE empresa_id = ? AND UPPER(COALESCE(nombre,'')) = UPPER(?) LIMIT 1",
            (empresa_id, name_clean),
        ).fetchone()
    if row:
        conn.execute(
            """
            UPDATE gestoria_terceros
            SET nif = COALESCE(NULLIF(?, ''), nif),
                nombre = COALESCE(NULLIF(?, ''), nombre),
                tipo = ?,
                cuenta_contable = COALESCE(NULLIF(cuenta_contable, ''), ?),
                updated_at = datetime(?)
            WHERE id = ?
            """,
            (nif_clean, name_clean, kind, account, now, row["id"]),
        )
        return row["id"], row["cuenta_contable"] or account
    tercero_id = os.urandom(16).hex()
    conn.execute(
        """
        INSERT INTO gestoria_terceros (
          id, empresa_id, nif, nombre, tipo, cuenta_contable, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, datetime(?), datetime(?))
        """,
        (tercero_id, empresa_id, nif_clean, name_clean, kind, account, now, now),
    )
    return tercero_id, account


def build_invoice_asiento(parsed, counterpart_account):
    tipo = (parsed.get("tipo") or "compra").strip().lower()
    base = float(parsed.get("base_imponible") or 0.0)
    iva = float(parsed.get("cuota_iva") or 0.0)
    irpf = float(parsed.get("cuota_irpf") or 0.0)
    total = float(parsed.get("total") or 0.0)
    concepto = parsed.get("descripcion") or parsed.get("numero") or "Factura"
    lines = []
    if tipo == "venta":
        ingreso = infer_revenue_account(concepto)
        if total > 0:
            lines.append({"cuenta": counterpart_account or "430", "descripcion": "Cliente", "debe": round(total, 2), "haber": 0.0})
        if base > 0:
            lines.append({"cuenta": ingreso, "descripcion": "Ingreso", "debe": 0.0, "haber": round(base, 2)})
        if iva > 0:
            lines.append({"cuenta": "477", "descripcion": "IVA repercutido", "debe": 0.0, "haber": round(iva, 2), "impuesto": "IVA", "porcentaje": parsed.get("iva_pct") or 0.0})
    else:
        gasto = infer_expense_account(concepto)
        if base > 0:
            lines.append({"cuenta": gasto, "descripcion": "Gasto", "debe": round(base, 2), "haber": 0.0})
        if iva > 0:
            lines.append({"cuenta": "472", "descripcion": "IVA soportado", "debe": round(iva, 2), "haber": 0.0, "impuesto": "IVA", "porcentaje": parsed.get("iva_pct") or 0.0})
        if irpf > 0:
            lines.append({"cuenta": "4751", "descripcion": "H.P. acreedora retenciones", "debe": 0.0, "haber": round(irpf, 2), "impuesto": "IRPF", "porcentaje": 0.0})
        payable = total if total > 0 else max(0.0, base + iva - irpf)
        if payable > 0:
            lines.append({"cuenta": counterpart_account or "400", "descripcion": "Proveedor/Acreedor", "debe": 0.0, "haber": round(payable, 2)})
    debe = round(sum(float(item.get("debe") or 0.0) for item in lines), 2)
    haber = round(sum(float(item.get("haber") or 0.0) for item in lines), 2)
    return lines, debe, haber

def process_seguros_ocr(payload, conn):
    pdf_bytes = decode_seguros_payload(payload)
    fast_mode = str(payload.get("fast_mode") or "").strip().lower() in ("1", "true", "yes", "on")
    tmp_path = None
    text = ""
    err_detail = ""
    method = ""
    source_hint = " ".join(
        [
            str(payload.get("filename") or ""),
            str(payload.get("s3_key") or ""),
            str(payload.get("source_hint") or ""),
        ]
    ).strip()
    hinted_company = detect_company_from_metadata(source_hint)
    required_keys = ("tomador", "poliza_numero", "compania", "fecha_efecto")
    ai_used = False
    ai_error = ""
    field_sources = {}
    candidate_fields = []
    def candidate_score(quality):
        if not isinstance(quality, dict):
            return 0
        required_valid = len(quality.get("required_valid") or [])
        required_filled = len(quality.get("required_filled") or [])
        confidence = float(quality.get("confidence") or 0)
        return required_valid * 100 + required_filled * 10 + int(confidence * 100)
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_file:
            tmp_file.write(pdf_bytes)
            tmp_path = tmp_file.name
        text, err_detail, method = extract_pdf_text(tmp_path)
        if not text:
            raise RuntimeError(err_detail or "No se pudo extraer texto")
        fields = parse_poliza_text(text, source_hint=source_hint, hinted_company=hinted_company)
        candidate_fields.append(("pdf_text", normalize_extracted_fields(fields), 1.0))
        best_quality = compute_ocr_quality(fields, required_keys)
        if (not fast_mode) and (
            not any(str(value or "").strip() for value in fields.values()) or (
            not fields.get("poliza_numero")
            or not fields.get("tomador")
            or not fields.get("compania")
            or not fields.get("fecha_efecto")
            )
        ):
            ocr_text, ocr_err = ocr_pdf_all_pages(tmp_path, use_external=external_ocr_available())
            if ocr_text:
                ocr_fields = parse_poliza_text(ocr_text, source_hint=source_hint, hinted_company=hinted_company)
                candidate_fields.append(("ocr_all_pages", normalize_extracted_fields(ocr_fields), 1.05))
                ocr_quality = compute_ocr_quality(ocr_fields, required_keys)
                if candidate_score(ocr_quality) >= candidate_score(best_quality):
                    fields = ocr_fields
                    best_quality = ocr_quality
                    text = ocr_text
                    method = "vision" if external_ocr_available() else "tesseract"
            elif ocr_err and not err_detail:
                err_detail = ocr_err
        doc_text = ""
        missing_required = any(not fields.get(key) for key in required_keys)
        if (missing_required or candidate_score(best_quality) < 250) and docai_available() and not fast_mode:
            doc_text, doc_fields, doc_err = ocr_image_docai(pdf_bytes, "application/pdf")
            if doc_err and not err_detail:
                err_detail = doc_err
            doc_mapped = map_docai_poliza_fields(doc_fields)
            doc_parsed = parse_poliza_text(doc_text, source_hint=source_hint, hinted_company=hinted_company) if doc_text else {}
            candidate_fields.append(("docai_fields", normalize_extracted_fields(doc_mapped), 1.1))
            candidate_fields.append(("docai_parsed", normalize_extracted_fields(doc_parsed), 1.05))
            merged = dict(fields)
            for key, value in doc_mapped.items():
                if value and not merged.get(key):
                    merged[key] = value
            for key, value in doc_parsed.items():
                if value and not merged.get(key):
                    merged[key] = value
            merged_quality = compute_ocr_quality(merged, required_keys)
            if candidate_score(merged_quality) >= candidate_score(best_quality):
                fields = merged
                best_quality = merged_quality
            if doc_text and doc_text.strip():
                method = "docai"
        missing_required = any(not fields.get(key) for key in required_keys)
        need_zones = missing_required or (not fast_mode and candidate_score(best_quality) < 320)
        if need_zones:
            zones_text, zones_err = ocr_poliza_key_regions(tmp_path, use_external=external_ocr_available())
            if zones_text:
                zones_fields = parse_poliza_text(
                    zones_text,
                    source_hint=source_hint,
                    hinted_company=hinted_company,
                )
                candidate_fields.append(("zones", normalize_extracted_fields(zones_fields), 1.08))
                zones_merged = merge_fields(fields, zones_fields)
                zones_quality = compute_ocr_quality(zones_merged, required_keys)
                if candidate_score(zones_quality) >= candidate_score(best_quality):
                    fields = zones_merged
                    best_quality = zones_quality
                    if method:
                        method = f"{method}+zones"
                    else:
                        method = "zones"
            elif zones_err and not err_detail:
                err_detail = zones_err
        if hinted_company and not fields.get("compania"):
            fields["compania"] = hinted_company
        missing_required = any(not fields.get(key) for key in required_keys)
        if fast_mode and missing_required:
            # Último intento rápido sin OCR externo ni IA para evitar latencia alta.
            quick_text, quick_err = ocr_pdf_all_pages(tmp_path, use_external=False)
            if quick_text:
                quick_fields = parse_poliza_text(quick_text, source_hint=source_hint, hinted_company=hinted_company)
                quick_merged = merge_fields(fields, quick_fields)
                quick_quality = compute_ocr_quality(quick_merged, required_keys)
                if candidate_score(quick_quality) >= candidate_score(best_quality):
                    fields = quick_merged
                    best_quality = quick_quality
                    text = quick_text
                    method = "tesseract"
            elif quick_err and not err_detail:
                err_detail = quick_err
            missing_required = any(not fields.get(key) for key in required_keys)
        if (not fast_mode) and openai_available() and (missing_required or candidate_score(best_quality) < 320):
            ai_text = text or ""
            if doc_text and doc_text.strip():
                ai_text = f"{ai_text}\n\n{doc_text}".strip()
            if tmp_path and OCR_OPENAI_VISION_PAGES > 0:
                image_urls, vision_img_err = pdf_to_png_data_urls(
                    tmp_path,
                    max_pages=OCR_OPENAI_VISION_PAGES,
                    dpi=OCR_OPENAI_VISION_DPI,
                )
                if image_urls:
                    ai_vision_fields, ai_vision_err = call_openai_extract_seguro_vision(
                        image_urls,
                        text=ai_text,
                        source_hint=source_hint,
                        hinted_company=hinted_company,
                    )
                    if ai_vision_err:
                        ai_error = ai_vision_err
                    elif ai_vision_fields:
                        ai_used = True
                        candidate_fields.append(("openai_vision", normalize_extracted_fields(ai_vision_fields), 1.18))
                        merged = dict(fields)
                        for key, value in ai_vision_fields.items():
                            if value and not str(merged.get(key) or "").strip():
                                merged[key] = value
                        ai_quality = compute_ocr_quality(merged, required_keys)
                        if candidate_score(ai_quality) >= candidate_score(best_quality):
                            fields = merged
                            best_quality = ai_quality
                elif vision_img_err and not ai_error:
                    ai_error = vision_img_err
            if ai_text:
                ai_fields, ai_error_msg = call_openai_extract_seguro(
                    ai_text,
                    source_hint=source_hint,
                    hinted_company=hinted_company,
                )
                if ai_error_msg:
                    ai_error = ai_error_msg
                elif ai_fields:
                    ai_used = True
                    candidate_fields.append(("openai_text", normalize_extracted_fields(ai_fields), 0.98))
                    merged = dict(fields)
                    for key, value in ai_fields.items():
                        if value and not str(merged.get(key) or "").strip():
                            merged[key] = value
                    ai_quality = compute_ocr_quality(merged, required_keys)
                    if candidate_score(ai_quality) >= candidate_score(best_quality):
                        fields = merged
                        best_quality = ai_quality
        if (not fast_mode) and OCR_EXPERT_MODE and candidate_fields:
            expert_fields, expert_sources = blend_ocr_field_candidates(candidate_fields, required_keys)
            expert_quality = compute_ocr_quality(expert_fields, required_keys)
            if candidate_score(expert_quality) >= candidate_score(best_quality):
                fields = expert_fields
                best_quality = expert_quality
                field_sources = expert_sources
        # Post-fix final para AXA Profesional: evita valores del mediador y fechas truncadas.
        company_key = normalize_company_key(fields.get("compania") or hinted_company or "")
        text_upper = normalize_lookup_text(text)
        if ("AXA" in company_key) and ("AXA PROFESIONAL" in text_upper or "POLIZA DE SEGURO DE PROFESIONAL" in text_upper):
            def _norm_date_local(value):
                raw = str(value or "").strip()
                raw = raw.replace(".", "/").replace("-", "/")
                raw = re.sub(r"\s+", "", raw)
                return raw
            def _clean_name_local(value):
                raw = normalize_person_name(value or "")
                raw = re.sub(r"\s{2,}", " ", raw).strip(" ,;:-")
                return raw
            fields["ramo"] = "Profesional"
            pol = re.search(r"p[oó]liza\s*(?:n[º°o]|no\.?)\s*([0-9]{2})[- ]?([0-9]{7,})", text, re.IGNORECASE)
            if pol:
                fields["poliza_numero"] = f"{pol.group(1)}-{pol.group(2)}"
            dates = re.search(
                r"Datos\s+de\s+la\s+Poliza[\s\S]{0,900}?Fecha\s+efecto[\s\S]{0,200}?Fecha\s+vencimiento[\s\S]{0,220}?([0-9]{1,2}[./-][0-9]{1,2}[./-]20[0-9]{2})[\s\S]{0,120}?([0-9]{1,2}[./-][0-9]{1,2}[./-]20[0-9]{2})",
                text,
                re.IGNORECASE,
            )
            if dates:
                fields["fecha_efecto"] = _norm_date_local(dates.group(1))
                fields["fecha_vencimiento"] = _norm_date_local(dates.group(2))
            tom = re.search(r"Tomador\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{6,}?)\s+Producto", text, re.IGNORECASE)
            if tom:
                fields["tomador"] = _clean_name_local(tom.group(1))
            if not fields.get("tomador") or len((fields.get("tomador") or "").split()) < 3:
                aseg = re.search(r"Nombre\s+del\s+asegurado\s*[\r\n]+\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{8,})", text, re.IGNORECASE)
                if aseg:
                    fields["tomador"] = _clean_name_local(aseg.group(1))
            if normalize_phone(fields.get("telefono") or "") == "900909014":
                fields["telefono"] = ""
            if "fincasvelazquez" in str(fields.get("email") or "").lower():
                fields["email"] = ""
            dir_norm = normalize_lookup_text(fields.get("direccion") or "")
            if "DOMICILIO SOCIAL" in dir_norm or "MEDIADOR" in dir_norm or "PALMA DE MALLORCA" in dir_norm:
                fields["direccion"] = ""
            if not fields.get("direccion"):
                addr = re.search(r"\b(CL\s+calle[^\n]{5,160}?\d{5}\s+[A-ZÁÉÍÓÚÑ]{3,})", text, re.IGNORECASE)
                if addr:
                    fields["direccion"] = re.sub(r"\s+", " ", addr.group(1)).strip()
        if ("ALLIANZ" in company_key) and ("CERTIFICADO DE SEGURO" in text_upper or "ALLIANZ R C PYME" in text_upper or "ALLIANZ R.C.PYME" in text_upper):
            if not fields.get("ramo"):
                fields["ramo"] = "Responsabilidad civil"
            allianz_general = re.search(
                r"Datos\s+Generales([\s\S]{0,2200}?)Datos\s+del\s+Asegurado",
                text,
                re.IGNORECASE,
            )
            allianz_scope = allianz_general.group(1) if allianz_general else text
            pol = re.search(r"P[oó]liza\s*n[º°o]\s*:\s*([0-9]{6,12})", text, re.IGNORECASE)
            if pol:
                fields["poliza_numero"] = pol.group(1).strip()
            dur = re.search(
                r"Duraci[oó]n\s*:\s*Desde[\s\S]{0,220}?del\s+(\d{1,2}/\d{1,2}/\d{4})[\s\S]{0,240}?hasta[\s\S]{0,220}?del\s+(\d{1,2}/\d{1,2}/\d{4})",
                text,
                re.IGNORECASE,
            )
            if dur:
                fields["fecha_efecto"] = dur.group(1)
                fields["fecha_vencimiento"] = dur.group(2)
            m = re.search(
                r"Datos\s+Generales\s*\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s,.'\-]{5,})\s*\n\s*Tomador\s+del\s+Seguro",
                text,
                re.IGNORECASE,
            )
            if m:
                name = normalize_person_name(m.group(1))
                name = re.sub(r"\s+", " ", name).strip(" ,;:-")
                if name and len(name.split()) >= 2:
                    fields["tomador"] = name
            # Si tomador quedó como una dirección, priorizar razón social del bloque general.
            tomador_norm = normalize_lookup_text(fields.get("tomador") or "")
            looks_address = (
                " CL " in f" {tomador_norm} "
                or " CALLE " in f" {tomador_norm} "
                or " AVDA " in f" {tomador_norm} "
                or " AVD " in f" {tomador_norm} "
            )
            if looks_address or not fields.get("tomador"):
                lines = [ln.strip() for ln in allianz_scope.splitlines() if ln.strip()]
                for idx, ln in enumerate(lines):
                    if re.search(r"Tomador\s+del", ln, re.IGNORECASE):
                        if idx > 0:
                            candidate = re.sub(r"\s+", " ", lines[idx - 1]).strip(" ,;:-")
                            if candidate and len(candidate.split()) >= 2:
                                fields["tomador"] = candidate
                        break
            nif_line = re.search(r"\bNIF\s*:\s*([A-Z0-9.\-]{8,16})", allianz_scope, re.IGNORECASE)
            if nif_line:
                fields["nif"] = re.sub(r"[^A-Z0-9]", "", nif_line.group(1).upper())
                fields["dni"] = fields["nif"]
            if fields.get("fecha_efecto") and fields.get("fecha_vencimiento") and fields["fecha_efecto"] == fields["fecha_vencimiento"]:
                fields["fecha_vencimiento"] = add_year_to_date(fields["fecha_efecto"])
        doc_type = classify_seguros_document(text)
        if doc_type == "otro" and doc_text:
            doc_type = classify_seguros_document(doc_text)
        ocr_quality = compute_ocr_quality(fields, required_keys)
        if field_sources:
            ocr_quality["field_sources"] = field_sources
        cliente_match = False
        cliente_id = None
        tomador = (fields.get("tomador") or "").strip()
        nif = (fields.get("nif") or fields.get("dni") or "").strip()
        if tomador or nif:
            if nif:
                nif_norm = re.sub(r"\s+", "", nif).upper()
                row = conn.execute(
                    "SELECT id FROM clientes WHERE REPLACE(UPPER(nif), ' ', '') = ?",
                    (nif_norm,),
                ).fetchone()
                if row:
                    cliente_id = row["id"]
                    cliente_match = True
            if not cliente_match and tomador:
                nombre_norm = re.sub(r"\s+", " ", tomador).strip().upper()
                row = conn.execute(
                    "SELECT id FROM clientes WHERE TRIM(UPPER(nombre)) = ?",
                    (nombre_norm,),
                ).fetchone()
                if row:
                    cliente_id = row["id"]
                    cliente_match = True
        return {
            "fields": fields,
            "text": text,
            "language": detect_ocr_lang(),
            "method": method,
            "doc_type": doc_type,
            "ocr_quality": ocr_quality,
            "cliente_id": cliente_id,
            "cliente_match": cliente_match,
            "ai_used": ai_used,
            "ai_error": ai_error,
        }
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def enqueue_ocr_job(db_path, kind, payload):
    job_id = os.urandom(16).hex()
    now = datetime.now(timezone.utc).isoformat()
    conn = open_sqlite_conn(db_path, with_row_factory=False)
    conn.execute(
        """
        INSERT INTO ocr_jobs (id, kind, status, payload_json, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (job_id, kind, "pending", json.dumps(payload), now, now),
    )
    conn.commit()
    conn.close()
    return job_id


def fetch_next_ocr_job(conn):
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=OCR_JOB_STALE_MINUTES)).isoformat()
    row = conn.execute(
        """
        SELECT id, kind, payload_json
        FROM ocr_jobs
        WHERE status = 'pending'
           OR (status = 'processing' AND finished_at IS NULL AND started_at IS NOT NULL AND started_at < ?)
        ORDER BY created_at ASC
        LIMIT 1
        """,
        (cutoff,),
    ).fetchone()
    return row


def claim_next_ocr_job(conn):
    cutoff = (datetime.now(timezone.utc) - timedelta(minutes=OCR_JOB_STALE_MINUTES)).isoformat()
    now = datetime.now(timezone.utc).isoformat()
    try:
        conn.execute("BEGIN IMMEDIATE")
        row = conn.execute(
            """
            SELECT id, kind, payload_json
            FROM ocr_jobs
            WHERE status = 'pending'
               OR (status = 'processing' AND finished_at IS NULL AND started_at IS NOT NULL AND started_at < ?)
            ORDER BY created_at ASC
            LIMIT 1
            """,
            (cutoff,),
        ).fetchone()
        if not row:
            conn.commit()
            return None
        updated = conn.execute(
            """
            UPDATE ocr_jobs
            SET status = 'processing', error = NULL, started_at = ?, updated_at = ?
            WHERE id = ?
              AND (
                status = 'pending'
                OR (status = 'processing' AND finished_at IS NULL AND started_at IS NOT NULL AND started_at < ?)
              )
            """,
            (now, now, row["id"], cutoff),
        ).rowcount
        conn.commit()
        if updated == 1:
            return row
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    return None


def update_ocr_job(conn, job_id, status, result=None, error=None):
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        """
        UPDATE ocr_jobs
        SET status = ?, result_json = COALESCE(?, result_json), error = ?, updated_at = ?,
            started_at = CASE
                WHEN ? = 'processing' THEN ?
                WHEN started_at IS NULL THEN ?
                ELSE started_at
            END,
            finished_at = CASE WHEN ? IN ('done','error') THEN ? ELSE finished_at END
        WHERE id = ?
        """,
        (
            status,
            json.dumps(result) if result is not None else None,
            error,
            now,
            status,
            now,
            now,
            status,
            now,
            job_id,
        ),
    )


def ocr_worker_loop(jobs_db_path, main_db_path):
    while True:
        jobs_conn = None
        job_id = None
        try:
            jobs_conn = open_sqlite_conn(jobs_db_path, with_row_factory=True)
            row = claim_next_ocr_job(jobs_conn)
            if not row:
                jobs_conn.close()
                time.sleep(0.6)
                continue
            job_id = row["id"]
            kind = row["kind"]
            payload = json.loads(row["payload_json"] or "{}")
            if kind == "seguros":
                main_conn = open_sqlite_conn(main_db_path, with_row_factory=True)
                try:
                    result = process_seguros_ocr(payload, main_conn)
                finally:
                    main_conn.close()
            else:
                raise RuntimeError("Tipo OCR no soportado")
            update_ocr_job(jobs_conn, job_id, "done", result=result, error=None)
            jobs_conn.commit()
            jobs_conn.close()
        except Exception as exc:
            try:
                if jobs_conn and job_id:
                    update_ocr_job(jobs_conn, job_id, "error", result=None, error=str(exc))
                    jobs_conn.commit()
            except Exception:
                pass
            try:
                if jobs_conn:
                    jobs_conn.close()
            except Exception:
                pass
            time.sleep(1.0)

def preprocess_image_for_ocr(src_path, out_path=None):
    tmp_base = tempfile.gettempdir()
    created = False
    if not out_path:
        tmp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False, dir=tmp_base)
        out_path = tmp_file.name
        tmp_file.close()
        created = True
    magick = shutil.which("magick") or shutil.which("convert")
    if magick and os.path.exists(magick):
        try:
            run_subprocess(
                [
                    magick,
                    src_path,
                    "-colorspace",
                    "Gray",
                    "-density",
                    "300",
                    "-resize",
                    "2500x2500",
                    "-deskew",
                    "40%",
                    "-sharpen",
                    "0x1",
                    "-contrast-stretch",
                    "1%x1%",
                    out_path,
                ],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
            )
            return out_path, created
        except Exception:
            pass
    try:
        run_subprocess(
            ["sips", "-s", "format", "png", src_path, "--out", out_path],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        try:
            run_subprocess(
                ["sips", "-Z", "2400", out_path, "--out", out_path],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception:
            pass
        return out_path, created
    except Exception:
        return src_path, created

def ocr_image_file(image_path):
    lang = detect_ocr_lang()
    tesseract_cmd = (
        shutil.which("tesseract")
        or "/opt/homebrew/bin/tesseract"
        or "/usr/local/bin/tesseract"
    )
    if not tesseract_cmd or not os.path.exists(tesseract_cmd):
        return "", "tesseract no encontrado en PATH"
    env = os.environ.copy()
    if os.path.isdir(TESSDATA_DIR):
        env["TESSDATA_PREFIX"] = TESSDATA_DIR
    tmp_base = tempfile.gettempdir()
    with tempfile.TemporaryDirectory(dir=tmp_base) as tmpdir:
        processed, created = preprocess_image_for_ocr(image_path)
        variants = [processed]
        magick = shutil.which("magick") or shutil.which("convert")
        if magick and os.path.exists(magick) and os.path.exists(processed):
            for idx, extra in enumerate(
                (
                    ["-threshold", "55%", "-despeckle"],
                    ["-adaptive-threshold", "15x15+10%", "-sharpen", "0x1.2"],
                )
            ):
                out_variant = os.path.join(tmpdir, f"ocr_var_{idx}.png")
                try:
                    run_subprocess(
                        [magick, processed, *extra, out_variant],
                        check=True,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.PIPE,
                        text=True,
                    )
                    if os.path.exists(out_variant):
                        variants.append(out_variant)
                except Exception:
                    continue
        def run_tesseract(psm, source_path):
            result = run_subprocess(
                [
                    tesseract_cmd,
                    source_path,
                    "stdout",
                    "-l",
                    lang,
                    "--oem",
                    "1",
                    "--psm",
                    str(psm),
                    "-c",
                    "user_defined_dpi=300",
                    "-c",
                    "preserve_interword_spaces=1",
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                env=env,
                text=True,
            )
            return result.stdout or ""
        try:
            candidates = []
            for source_path in variants:
                for psm in OCR_TESSERACT_PSMS:
                    try:
                        candidates.append(run_tesseract(psm, source_path))
                    except subprocess.CalledProcessError:
                        continue
            if not candidates:
                return "", "tesseract: sin salida"
            best = max(candidates, key=lambda t: (len(t.strip()), sum(ch.isdigit() for ch in t)))
            return best, ""
        except subprocess.CalledProcessError as err:
            return "", f"tesseract: {err.stderr.strip()}"
        except Exception as err:
            return "", f"tesseract: {err}"
        finally:
            if created and processed and os.path.exists(processed):
                try:
                    os.unlink(processed)
                except Exception:
                    pass

def ocr_image_external(image_bytes):
    credentials_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    api_key = os.environ.get("GOOGLE_VISION_API_KEY") or os.environ.get("VISION_API_KEY")
    auth_header = None
    if credentials_path and os.path.exists(credentials_path):
        try:
            from google.oauth2 import service_account
            from google.auth.transport.requests import Request
        except Exception:
            return "", "OCR externo: instala google-auth y requests (pip install google-auth requests)"
        try:
            creds = service_account.Credentials.from_service_account_file(
                credentials_path,
                scopes=["https://www.googleapis.com/auth/cloud-platform"],
            )
            creds.refresh(Request())
            auth_header = f"Bearer {creds.token}"
        except Exception as err:
            return "", f"OCR externo: credenciales inválidas ({err})"
    elif not api_key:
        return "", "OCR externo no configurado"
    payload = {
        "requests": [
            {
                "image": {"content": base64.b64encode(image_bytes).decode("utf-8")},
                "features": [{"type": "DOCUMENT_TEXT_DETECTION"}],
            }
        ]
    }
    data = json.dumps(payload).encode("utf-8")
    if auth_header:
        url = "https://vision.googleapis.com/v1/images:annotate"
        headers = {"Content-Type": "application/json", "Authorization": auth_header}
    else:
        url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
        headers = {"Content-Type": "application/json"}
    req = urllib.request.Request(url, data=data, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            res = json.loads(resp.read().decode("utf-8"))
    except Exception as err:
        return "", f"OCR externo: {err}"
    try:
        text = res["responses"][0].get("fullTextAnnotation", {}).get("text", "")
        return text, ""
    except Exception:
        return "", "OCR externo: sin texto"

def external_ocr_available():
    credentials_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    api_key = os.environ.get("GOOGLE_VISION_API_KEY") or os.environ.get("VISION_API_KEY")
    return bool(api_key) or (credentials_path and os.path.exists(credentials_path))

def docai_available():
    processor_id = os.environ.get("DOCUMENTAI_PROCESSOR_ID") or os.environ.get("DOC_AI_PROCESSOR_ID")
    return bool(processor_id)

def normalize_field_label(value):
    value = value or ""
    value = value.lower()
    value = value.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    value = re.sub(r"[^a-z0-9 ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value

def ocr_image_docai(image_bytes, mime_type):
    credentials_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not credentials_path or not os.path.exists(credentials_path):
        return "", {}, "Document AI: credenciales no configuradas"
    processor_id = os.environ.get("DOCUMENTAI_PROCESSOR_ID") or os.environ.get("DOC_AI_PROCESSOR_ID")
    if not processor_id:
        return "", {}, "Document AI: falta DOCUMENTAI_PROCESSOR_ID"
    location = os.environ.get("DOCUMENTAI_LOCATION") or "eu"
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import Request
    except Exception:
        return "", {}, "Document AI: instala google-auth y requests"
    try:
        creds = service_account.Credentials.from_service_account_file(
            credentials_path,
            scopes=["https://www.googleapis.com/auth/cloud-platform"],
        )
        creds.refresh(Request())
        project_id = creds.project_id
    except Exception as err:
        return "", {}, f"Document AI: credenciales inválidas ({err})"
    if not project_id:
        return "", {}, "Document AI: project_id no encontrado"
    url = (
        f"https://{location}-documentai.googleapis.com/v1/projects/{project_id}"
        f"/locations/{location}/processors/{processor_id}:process"
    )
    payload = {
        "rawDocument": {
            "content": base64.b64encode(image_bytes).decode("utf-8"),
            "mimeType": mime_type or "image/jpeg",
        }
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {creds.token}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            res = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as err:
        try:
            detail = err.read().decode("utf-8")
        except Exception:
            detail = ""
        return "", {}, f"Document AI: {err} {detail}".strip()
    except Exception as err:
        return "", {}, f"Document AI: {err}"
    doc = res.get("document") or {}
    text = doc.get("text", "")
    fields = {}
    occurrences = {}
    for field in doc.get("pages", []) or []:
        for form_field in field.get("formFields", []) or []:
            name_text = ""
            value_text = ""
            name = form_field.get("fieldName", {}).get("textAnchor", {}).get("textSegments", [])
            for seg in name:
                try:
                    start = int(seg.get("startIndex", 0))
                    end = int(seg.get("endIndex", 0))
                    name_text += text[start:end]
                except Exception:
                    continue
            value = form_field.get("fieldValue", {}).get("textAnchor", {}).get("textSegments", [])
            for seg in value:
                try:
                    start = int(seg.get("startIndex", 0))
                    end = int(seg.get("endIndex", 0))
                    value_text += text[start:end]
                except Exception:
                    continue
            label = normalize_field_label(name_text)
            val = value_text.strip()
            if not label or not val:
                continue
            occurrences[label] = occurrences.get(label, 0) + 1
            key = label
            if label in (
                "nombre y apellidos",
                "nombre apellidos",
                "nombre",
                "dni",
                "nif",
                "telefono",
                "movil",
                "correo electronico",
                "email",
                "fecha nacimiento",
                "estado civil",
                "hijos",
                "profesion",
                "tipo contrato",
                "ingresos nomina",
                "ingresos",
                "nomina",
                "patrimonio alquiler",
                "prestamos",
            ):
                suffix = "1" if occurrences[label] == 1 else "2"
                key = f"{label} {suffix}"
            fields[key] = val
    return text, fields, ""

def map_docai_fields(doc_fields):
    doc_fields = doc_fields or {}
    def doc_pick(label, idx=1):
        key = f"{label} {idx}"
        return doc_fields.get(key, "") or doc_fields.get(label, "")
    fields = {}
    fields["cliente1_nombre"] = doc_pick("nombre y apellidos", 1) or doc_pick("nombre", 1)
    fields["cliente2_nombre"] = doc_pick("nombre y apellidos", 2) or doc_pick("nombre", 2)
    fields["cliente1_dni"] = doc_pick("dni", 1) or doc_pick("nif", 1)
    fields["cliente2_dni"] = doc_pick("dni", 2) or doc_pick("nif", 2)
    fields["cliente1_telefono"] = doc_pick("telefono", 1) or doc_pick("movil", 1)
    fields["cliente2_telefono"] = doc_pick("telefono", 2) or doc_pick("movil", 2)
    fields["cliente1_email"] = doc_pick("correo electronico", 1) or doc_pick("email", 1)
    fields["cliente2_email"] = doc_pick("correo electronico", 2) or doc_pick("email", 2)
    fields["cliente1_fecha_nacimiento"] = doc_pick("fecha nacimiento", 1)
    fields["cliente2_fecha_nacimiento"] = doc_pick("fecha nacimiento", 2)
    fields["cliente1_estado_civil"] = doc_pick("estado civil", 1)
    fields["cliente2_estado_civil"] = doc_pick("estado civil", 2)
    fields["cliente1_hijos"] = doc_pick("hijos", 1)
    fields["cliente2_hijos"] = doc_pick("hijos", 2)
    fields["cliente1_profesion"] = doc_pick("profesion", 1)
    fields["cliente2_profesion"] = doc_pick("profesion", 2)
    fields["cliente1_tipo_contrato"] = doc_pick("tipo contrato", 1)
    fields["cliente2_tipo_contrato"] = doc_pick("tipo contrato", 2)
    fields["cliente1_ingresos"] = doc_pick("ingresos nomina", 1) or doc_pick("ingresos", 1) or doc_pick("nomina", 1)
    fields["cliente2_ingresos"] = doc_pick("ingresos nomina", 2) or doc_pick("ingresos", 2) or doc_pick("nomina", 2)
    fields["cliente1_patrimonio"] = doc_pick("patrimonio alquiler", 1)
    fields["cliente2_patrimonio"] = doc_pick("patrimonio alquiler", 2)
    fields["cliente1_prestamos"] = doc_pick("prestamos", 1)
    fields["cliente2_prestamos"] = doc_pick("prestamos", 2)
    return fields

def map_docai_poliza_fields(doc_fields):
    doc_fields = doc_fields or {}
    def doc_pick(labels):
        for label in labels:
            key = normalize_field_label(label)
            if doc_fields.get(key):
                return doc_fields.get(key)
        return ""
    fields = {}
    fields["tomador"] = doc_pick([
        "tomador",
        "asegurado",
        "asegurado principal",
        "titular",
        "contratante",
        "nombre",
        "nombre y apellidos",
    ])
    fields["dni"] = doc_pick(["dni", "nif", "cif", "documento", "doc identificacion"])
    fields["telefono"] = doc_pick(["telefono", "móvil", "movil"])
    fields["email"] = doc_pick(["correo electronico", "email"])
    fields["direccion"] = doc_pick(["direccion", "domicilio"])
    fields["compania"] = doc_pick(["compania", "compañia", "aseguradora", "entidad aseguradora"])
    fields["ramo"] = doc_pick(["ramo", "modalidad", "producto"])
    fields["poliza_numero"] = doc_pick([
        "poliza",
        "numero poliza",
        "nº poliza",
        "número poliza",
        "certificado",
        "contrato",
    ])
    fields["fecha_efecto"] = doc_pick([
        "fecha efecto",
        "efecto",
        "inicio vigencia",
        "fecha inicio",
        "vigencia desde",
    ])
    fields["fecha_vencimiento"] = doc_pick([
        "fecha vencimiento",
        "vencimiento",
        "fin vigencia",
        "vigencia hasta",
    ])
    fields["prima_neta"] = doc_pick(["prima neta", "neta"])
    fields["prima_total"] = doc_pick(["prima total", "prima anual", "total recibo", "total"])
    return fields

def compute_ocr_quality(fields, required_keys=None):
    fields = fields or {}
    required_keys = required_keys or ()
    def is_valid_date(value):
        if not value:
            return False
        date_text = str(value).strip()
        if re.search(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", date_text):
            return True
        return bool(re.search(r"\b\d{4}-\d{2}-\d{2}\b", date_text))
    def is_valid_poliza(value):
        if not value:
            return False
        token = re.sub(r"[^A-Z0-9]", "", str(value).upper())
        return len(token) >= 6 and bool(re.search(r"\d", token))
    def is_valid_company(value):
        if not value:
            return False
        company = normalize_company_name(value)
        if len(company) < 3:
            return False
        lowered = company.lower()
        banned = ("compania", "compañia", "aseguradora", "seguro", "poliza", "póliza")
        return lowered not in banned
    def is_valid_nif_any(value):
        if not value:
            return False
        raw = normalize_nif(value)
        return bool(
            re.match(r"^[0-9]{8}[A-Z]$", raw)
            or re.match(r"^[XYZ][0-9]{7}[A-Z]$", raw)
            or re.match(r"^[ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z]$", raw)
        )
    def field_is_valid(key, value):
        if not str(value or "").strip():
            return False
        if key in ("fecha_efecto", "fecha_vencimiento", "fecha_nacimiento"):
            return is_valid_date(value)
        if key in ("poliza_numero",):
            return is_valid_poliza(value)
        if key in ("compania",):
            return is_valid_company(value)
        if key in ("dni", "nif"):
            return is_valid_nif_any(value)
        return True
    provided = [key for key, value in fields.items() if str(value or "").strip()]
    required_filled = [key for key in required_keys if str(fields.get(key) or "").strip()]
    required_valid = [key for key in required_keys if field_is_valid(key, fields.get(key))]
    invalid_required = [key for key in required_filled if key not in required_valid]
    total_required = len(required_keys)
    ratio = (len(required_valid) / total_required) if total_required else 0
    if total_required and ratio >= 0.75:
        calidad = "alta"
    elif total_required and ratio >= 0.45:
        calidad = "media"
    elif total_required:
        calidad = "baja"
    else:
        calidad = "desconocida"
    confidence = round((len(required_valid) / total_required), 3) if total_required else 0
    return {
        "calidad": calidad,
        "campos": provided,
        "required_filled": required_filled,
        "required_valid": required_valid,
        "required_invalid": invalid_required,
        "missing_required": [key for key in required_keys if key not in required_filled],
        "confidence": confidence,
    }


def is_valid_ocr_field(key, value):
    raw = str(value or "").strip()
    if not raw:
        return False
    if key in ("fecha_efecto", "fecha_vencimiento", "fecha_nacimiento"):
        return bool(
            re.search(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", raw)
            or re.search(r"\b\d{4}-\d{2}-\d{2}\b", raw)
        )
    if key == "poliza_numero":
        token = normalize_poliza_key(raw)
        if len(token) < 6 or not re.search(r"\d", token):
            return False
        if any(month in token for month in ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE")):
            return False
        return True
    if key == "compania":
        company = normalize_company_name(raw)
        if len(company) < 3:
            return False
        return normalize_lookup_text(company) not in ("COMPANIA", "ASEGURADORA", "SEGURO", "POLIZA")
    if key in ("dni", "nif"):
        nif = normalize_nif(raw)
        return bool(
            re.match(r"^[0-9]{8}[A-Z]$", nif)
            or re.match(r"^[XYZ][0-9]{7}[A-Z]$", nif)
            or re.match(r"^[ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z]$", nif)
        )
    if key in ("telefono",):
        return bool(re.match(r"^[0-9]{9}$", normalize_phone(raw) or ""))
    if key in ("email",):
        return bool(re.match(r"^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$", raw.upper()))
    if key in ("prima_neta", "prima_total"):
        return bool(re.match(r"^\d{1,3}(?:\.\d{3})*,\d{2}$|^\d+,\d{2}$|^\d+(?:\.\d+)?$", raw))
    if key == "ramo":
        if len(raw) > 48:
            return False
        if "@" in raw or "€" in raw:
            return False
        if re.search(r"\d{2,}", raw):
            return False
        return True
    return True


def normalize_extracted_fields(fields):
    fields = dict(fields or {})
    if fields.get("tomador"):
        fields["tomador"] = normalize_person_name(fields["tomador"])
    if fields.get("dni"):
        fields["dni"] = normalize_nif(fields["dni"])
    if fields.get("nif"):
        fields["nif"] = normalize_nif(fields["nif"])
    if fields.get("telefono"):
        fields["telefono"] = normalize_phone(fields["telefono"])
    if fields.get("email"):
        fields["email"] = normalize_email(fields["email"])
    if fields.get("compania"):
        fields["compania"] = normalize_company_name(fields["compania"])
    if fields.get("fecha_efecto"):
        fields["fecha_efecto"] = str(fields["fecha_efecto"]).strip()
    if fields.get("fecha_vencimiento"):
        fields["fecha_vencimiento"] = str(fields["fecha_vencimiento"]).strip()
    if fields.get("fecha_nacimiento"):
        fields["fecha_nacimiento"] = str(fields["fecha_nacimiento"]).strip()
    if fields.get("poliza_numero"):
        fields["poliza_numero"] = normalize_poliza_key(fields["poliza_numero"])
    return fields


def blend_ocr_field_candidates(candidates, required_keys=None):
    required_keys = tuple(required_keys or ())
    keys = (
        "tomador",
        "dni",
        "nif",
        "telefono",
        "email",
        "direccion",
        "compania",
        "ramo",
        "poliza_numero",
        "fecha_efecto",
        "fecha_vencimiento",
        "prima_neta",
        "prima_total",
    )
    merged = {}
    debug = {}
    for key in keys:
        ranked = []
        for source_name, source_fields, source_weight in candidates:
            val = str((source_fields or {}).get(key) or "").strip()
            if not val:
                continue
            score = float(source_weight)
            valid = is_valid_ocr_field(key, val)
            if valid:
                score += 2.0
            else:
                score -= 1.25
            if key in required_keys:
                score += 0.8
            if key in ("tomador", "direccion") and len(val) >= 8:
                score += 0.2
            if key == "ramo":
                ramo_up = normalize_lookup_text(val)
                if ramo_up in ("HOGAR", "AUTO", "RESPONSABILIDAD CIVIL", "IMPAGO ALQUILER", "ALQUILER", "COMERCIO"):
                    score += 0.6
            if key == "compania":
                if normalize_company_key(val):
                    score += 0.4
            ranked.append((score, valid, source_name, val))
        if not ranked:
            continue
        ranked.sort(key=lambda item: (item[0], item[1], len(item[3])), reverse=True)
        best = ranked[0]
        merged[key] = best[3]
        debug[key] = {
            "source": best[2],
            "score": round(best[0], 3),
            "valid": bool(best[1]),
        }
    merged = normalize_extracted_fields(merged)
    return merged, debug

def compute_fin_quality(fields):
    required = (
        "inmobiliaria_asesor",
        "fecha",
        "asesor",
        "cliente1_nombre",
        "cliente1_dni",
        "cliente1_telefono",
        "cliente1_email",
        "cliente1_ingresos",
    )
    return compute_ocr_quality(fields, required)

def openai_available():
    return bool(os.environ.get("OPENAI_API_KEY"))

def normalize_openai_model_name(value, default="gpt-4o-mini"):
    raw = str(value or "").strip()
    if not raw:
        return default
    if re.match(r"^gpt[-_]", raw, re.IGNORECASE):
        return raw.replace("_", "-").lower()
    return raw

def extract_openai_output(resp):
    if not resp:
        return ""
    if isinstance(resp, dict):
        if resp.get("output_text"):
            return resp.get("output_text")
        output = resp.get("output") or []
        parts = []
        for item in output:
            content = item.get("content") or []
            for block in content:
                if block.get("type") == "output_text":
                    parts.append(block.get("text", ""))
        if parts:
            return "\n".join(parts).strip()
    return ""

def call_openai(prompt, model=None, temperature=0.2, max_tokens=600):
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return "", "OPENAI_API_KEY no configurada"
    model = normalize_openai_model_name(model or os.environ.get("OPENAI_MODEL"))
    payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": "Eres un copiloto interno para un CRM de seguros. Responde en español.",
                    }
                ],
            },
            {"role": "user", "content": [{"type": "input_text", "text": prompt}]},
        ],
        "temperature": temperature,
        "max_output_tokens": max_tokens,
        "store": False,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            res = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as err:
        body = ""
        try:
            body = err.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        details = body or str(err)
        return "", f"OpenAI error ({err.code}): {details}"
    except Exception as err:
        return "", f"OpenAI error: {err}"
    return extract_openai_output(res), ""


def call_openai_content(user_content, model=None, temperature=0.0, max_tokens=700):
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return "", "OPENAI_API_KEY no configurada"
    model = normalize_openai_model_name(model or os.environ.get("OPENAI_MODEL"))
    payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": "Eres un extractor de datos para un CRM de seguros. Responde en JSON válido cuando se solicite.",
                    }
                ],
            },
            {"role": "user", "content": user_content},
        ],
        "temperature": temperature,
        "max_output_tokens": max_tokens,
        "store": False,
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            res = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as err:
        body = ""
        try:
            body = err.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        details = body or str(err)
        return "", f"OpenAI error ({err.code}): {details}"
    except Exception as err:
        return "", f"OpenAI error: {err}"
    return extract_openai_output(res), ""


def pdf_to_png_data_urls(pdf_path, max_pages=2, dpi=220):
    if not pdf_path or max_pages <= 0:
        return [], ""
    tmpdir = tempfile.mkdtemp(prefix="openai-vision-")
    prefix = os.path.join(tmpdir, "page")
    cmd = [
        "pdftoppm",
        "-f",
        "1",
        "-l",
        str(max_pages),
        "-r",
        str(dpi),
        "-png",
        pdf_path,
        prefix,
    ]
    try:
        run_subprocess(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    except Exception as err:
        shutil.rmtree(tmpdir, ignore_errors=True)
        return [], str(err)
    image_paths = sorted(
        os.path.join(tmpdir, name)
        for name in os.listdir(tmpdir)
        if name.startswith("page-") and name.endswith(".png")
    )
    data_urls = []
    try:
        for image_path in image_paths:
            with open(image_path, "rb") as handle:
                b64 = base64.b64encode(handle.read()).decode("ascii")
            data_urls.append(f"data:image/png;base64,{b64}")
    except Exception as err:
        return [], str(err)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
    return data_urls, ""

def call_openai_extract_fin(text, extra=""):
    prompt = (
        "Extrae datos de un asesoramiento financiero manuscrito. Responde solo JSON con claves conocidas, "
        "sin inventar. Usa los campos existentes del formulario (cliente1_nombre, cliente1_dni, cliente1_telefono, "
        "cliente1_email, cliente1_fecha_nacimiento, cliente1_estado_civil, cliente1_regimen, cliente1_hijos, "
        "cliente1_profesion, cliente1_tipo_contrato, cliente1_tiempo_contrato, cliente1_ingresos, cliente1_patrimonio, "
        "cliente1_prestamos, cliente1_prestamo_activo, cliente1_prestamo_entidad, cliente1_prestamo_resto, "
        "cliente2_nombre, cliente2_dni, cliente2_telefono, cliente2_email, cliente2_fecha_nacimiento, "
        "cliente2_estado_civil, cliente2_regimen, cliente2_hijos, cliente2_profesion, cliente2_tipo_contrato, "
        "cliente2_tiempo_contrato, cliente2_ingresos, cliente2_patrimonio, cliente2_prestamos, cliente2_prestamo_activo, "
        "cliente2_prestamo_entidad, cliente2_prestamo_resto, ingresos_conjuntos, entidades_financieras, avalistas, "
        "aportacion_cv, inmobiliaria_asesor, asesor, fecha). "
        "Si no ves un dato, deja el valor vacío.\n\n"
        f"Texto OCR:\n{text}\n\n"
        f"Instrucciones extra: {extra}"
    )
    output, err = call_openai(prompt, temperature=0.1, max_tokens=800)
    if err:
        return {}, err
    try:
        data = json.loads(output)
        if isinstance(data, dict):
            return data, ""
    except Exception:
        return {}, "OpenAI no devolvió JSON válido"
    return {}, "OpenAI no devolvió JSON"


def call_openai_extract_seguro(text, source_hint="", hinted_company=""):
    prompt = (
        "Extrae datos de una póliza de seguros desde OCR. Responde SOLO JSON válido, sin markdown ni texto adicional. "
        "No inventes datos. Si un campo no aparece claro, déjalo vacío.\n"
        "Claves permitidas: tomador, dni, nif, telefono, email, direccion, compania, ramo, "
        "poliza_numero, fecha_efecto, fecha_vencimiento, prima_neta, prima_total.\n"
        "Formato recomendado de fechas: DD/MM/AAAA o YYYY-MM-DD.\n"
        f"Pista de nombre/archivo: {source_hint or '-'}\n"
        f"Compañía sugerida por metadata: {hinted_company or '-'}\n\n"
        f"Texto OCR:\n{text}"
    )
    output, err = call_openai(prompt, temperature=0.0, max_tokens=700)
    if err:
        return {}, err
    try:
        data = json.loads(output)
    except Exception:
        return {}, "OpenAI no devolvió JSON válido"
    if not isinstance(data, dict):
        return {}, "OpenAI no devolvió JSON"
    allowed = (
        "tomador",
        "dni",
        "nif",
        "telefono",
        "email",
        "direccion",
        "compania",
        "ramo",
        "poliza_numero",
        "fecha_efecto",
        "fecha_vencimiento",
        "prima_neta",
        "prima_total",
    )
    normalized = {}
    for key in allowed:
        value = data.get(key)
        if value is None:
            continue
        normalized[key] = str(value).strip()
    if normalized.get("tomador"):
        normalized["tomador"] = normalize_person_name(normalized.get("tomador"))
    if normalized.get("dni"):
        normalized["dni"] = normalize_nif(normalized.get("dni"))
    if normalized.get("nif"):
        normalized["nif"] = normalize_nif(normalized.get("nif"))
    if normalized.get("telefono"):
        normalized["telefono"] = normalize_phone(normalized.get("telefono"))
    if normalized.get("email"):
        normalized["email"] = normalize_email(normalized.get("email"))
    if normalized.get("compania"):
        normalized["compania"] = normalize_company_name(normalized.get("compania"))
    if not normalized.get("compania") and hinted_company:
        normalized["compania"] = normalize_company_name(hinted_company)
    if normalized.get("poliza_numero"):
        normalized["poliza_numero"] = str(normalized.get("poliza_numero")).strip().upper()
    return normalized, ""


def call_openai_extract_seguro_vision(image_data_urls, text="", source_hint="", hinted_company=""):
    if not image_data_urls:
        return {}, "Sin imágenes para visión"
    prompt = (
        "Extrae datos de una póliza de seguros. Responde SOLO JSON válido, sin markdown ni texto adicional. "
        "No inventes datos. Si un campo no aparece claro, déjalo vacío.\n"
        "Claves permitidas: tomador, dni, nif, telefono, email, direccion, compania, ramo, "
        "poliza_numero, fecha_efecto, fecha_vencimiento, prima_neta, prima_total.\n"
        "Usa primero la imagen; usa el texto OCR como apoyo cuando exista.\n"
        f"Pista de nombre/archivo: {source_hint or '-'}\n"
        f"Compañía sugerida por metadata: {hinted_company or '-'}\n\n"
        f"Texto OCR auxiliar:\n{text[:12000] if text else '(vacío)'}"
    )
    content = [{"type": "input_text", "text": prompt}]
    for data_url in image_data_urls[: max(1, OCR_OPENAI_VISION_PAGES)]:
        content.append({"type": "input_image", "image_url": data_url})
    output, err = call_openai_content(content, temperature=0.0, max_tokens=900)
    if err:
        return {}, err
    try:
        data = json.loads(output)
    except Exception:
        return {}, "OpenAI visión no devolvió JSON válido"
    if not isinstance(data, dict):
        return {}, "OpenAI visión no devolvió JSON"
    normalized = {}
    for key in (
        "tomador",
        "dni",
        "nif",
        "telefono",
        "email",
        "direccion",
        "compania",
        "ramo",
        "poliza_numero",
        "fecha_efecto",
        "fecha_vencimiento",
        "prima_neta",
        "prima_total",
    ):
        value = data.get(key)
        if value is None:
            continue
        normalized[key] = str(value).strip()
    if normalized.get("tomador"):
        normalized["tomador"] = normalize_person_name(normalized.get("tomador"))
    if normalized.get("dni"):
        normalized["dni"] = normalize_nif(normalized.get("dni"))
    if normalized.get("nif"):
        normalized["nif"] = normalize_nif(normalized.get("nif"))
    if normalized.get("telefono"):
        normalized["telefono"] = normalize_phone(normalized.get("telefono"))
    if normalized.get("email"):
        normalized["email"] = normalize_email(normalized.get("email"))
    if normalized.get("compania"):
        normalized["compania"] = normalize_company_name(normalized.get("compania"))
    if not normalized.get("compania") and hinted_company:
        normalized["compania"] = normalize_company_name(hinted_company)
    if normalized.get("poliza_numero"):
        normalized["poliza_numero"] = str(normalized.get("poliza_numero")).strip().upper()
    return normalized, ""

def classify_seguros_document(text):
    cleaned = (text or "").lower()
    if not cleaned:
        return "otro"
    presupuesto_hits = [
        "presupuesto",
        "propuesta",
        "proyecto",
        "cotizacion",
        "cotización",
        "oferta",
        "simulacion",
        "simulación",
    ]
    poliza_hits = [
        "poliza",
        "póliza",
        "certificado",
        "contrato",
        "vigencia",
        "fecha de efecto",
        "fecha de vencimiento",
    ]
    score_presupuesto = sum(1 for key in presupuesto_hits if key in cleaned)
    score_poliza = sum(1 for key in poliza_hits if key in cleaned)
    if score_presupuesto > score_poliza:
        return "presupuesto"
    if score_poliza > 0:
        return "poliza"
    return "otro"

def merge_fields(base_fields, extra_fields):
    base_fields = base_fields or {}
    extra_fields = extra_fields or {}
    merged = dict(base_fields)
    for key, value in extra_fields.items():
        if not str(merged.get(key, "") or "").strip() and str(value or "").strip():
            merged[key] = value
    return merged

def merge_many_fields(*field_sets):
    merged = {}
    for fields in field_sets:
        for key, value in (fields or {}).items():
            if value is None:
                continue
            val = str(value).strip()
            if not val:
                continue
            if key not in merged or not str(merged.get(key) or "").strip():
                merged[key] = value
    return merged

def get_image_size(image_path):
    try:
        result = run_subprocess(
            ["sips", "-g", "pixelWidth", "-g", "pixelHeight", image_path],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except Exception:
        return None
    width = None
    height = None
    for line in result.stdout.splitlines():
        if "pixelWidth" in line:
            parts = re.findall(r"([0-9]+)", line)
            if parts:
                width = int(parts[-1])
        if "pixelHeight" in line:
            parts = re.findall(r"([0-9]+)", line)
            if parts:
                height = int(parts[-1])
    if width and height:
        return width, height
    return None

def crop_image_region(image_path, x, y, w, h, out_path):
    magick = shutil.which("magick") or shutil.which("convert")
    if magick and os.path.exists(magick):
        try:
            run_subprocess(
                [
                    magick,
                    image_path,
                    "-crop",
                    f"{int(w)}x{int(h)}+{int(x)}+{int(y)}",
                    "+repage",
                    out_path,
                ],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
            )
            return True
        except Exception:
            pass
    try:
        run_subprocess(
            [
                "sips",
                "-c",
                str(int(h)),
                str(int(w)),
                "--cropOffset",
                str(int(x)),
                str(int(y)),
                image_path,
                "--out",
                out_path,
            ],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
        )
        return True
    except Exception:
        return False

def prepare_image_bytes_for_vision(image_path):
    magick = shutil.which("magick") or shutil.which("convert")
    if magick and os.path.exists(magick):
        tmp = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
                tmp = tmp_file.name
            run_subprocess(
                [
                    magick,
                    image_path,
                    "-colorspace",
                    "Gray",
                    "-auto-level",
                    "-contrast-stretch",
                    "2%x2%",
                    "-sharpen",
                    "0x1",
                    "-deskew",
                    "40%",
                    tmp,
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            with open(tmp, "rb") as handle:
                return handle.read()
        except Exception:
            pass
        finally:
            if tmp and os.path.exists(tmp):
                try:
                    os.unlink(tmp)
                except Exception:
                    pass
    try:
        with open(image_path, "rb") as handle:
            return handle.read()
    except Exception:
        return b""

def asesoramiento_image_boxes(width, height):
    return {
        "header": (0.03 * width, 0.17 * height, 0.94 * width, 0.16 * height),
        "cliente1": (0.03 * width, 0.31 * height, 0.94 * width, 0.21 * height),
        "cliente2": (0.03 * width, 0.54 * height, 0.94 * width, 0.20 * height),
        "resumen": (0.03 * width, 0.74 * height, 0.94 * width, 0.18 * height),
    }

def ocr_best_block(image_path, box, use_external):
    x, y, w, h = box
    shifts = (0.0, 0.02, 0.04, -0.02)
    best = ""
    best_err = ""
    for shift in shifts:
        tmp_crop = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_file:
                tmp_crop = tmp_file.name
            y_shifted = max(0, y + shift * h * 2)
            if not crop_image_region(image_path, x, y_shifted, w, h, tmp_crop):
                continue
            text = ""
            err = ""
            if use_external:
                vision_bytes = prepare_image_bytes_for_vision(tmp_crop)
                if vision_bytes:
                    text, err = ocr_image_external(vision_bytes)
            if not text:
                text, err = ocr_image_file(tmp_crop)
            if len(text.strip()) > len(best.strip()):
                best = text
                best_err = err
        finally:
            if tmp_crop and os.path.exists(tmp_crop):
                try:
                    os.unlink(tmp_crop)
                except Exception:
                    pass
    return best, best_err


def poliza_image_boxes(width, height):
    return {
        "header_left": (0.02 * width, 0.03 * height, 0.52 * width, 0.34 * height),
        "header_right": (0.54 * width, 0.03 * height, 0.44 * width, 0.34 * height),
        "mid_full": (0.02 * width, 0.30 * height, 0.96 * width, 0.38 * height),
        "footer_full": (0.02 * width, 0.62 * height, 0.96 * width, 0.34 * height),
    }


def ocr_poliza_key_regions(path, use_external=False):
    images = []
    tmpdirs = []
    path_str = str(path or "")
    if path_str.lower().endswith(".pdf"):
        for dpi in OCR_PDF_DPI_VARIANTS:
            dpi_images, _err, tmpdir = pdftoppm_first_page(path_str, pages=1, dpi=dpi)
            if dpi_images:
                images.extend(dpi_images[:1])
            if tmpdir:
                tmpdirs.append(tmpdir)
    else:
        if os.path.exists(path_str):
            images.append(path_str)
    best_text = ""
    best_err = ""
    try:
        for image_path in images:
            size = get_image_size(image_path)
            if not size:
                continue
            width, height = size
            boxes = poliza_image_boxes(width, height)
            chunks = []
            for box in boxes.values():
                block_text, block_err = ocr_best_block(image_path, box, use_external)
                if block_text and len(block_text.strip()) >= 12:
                    chunks.append(block_text.strip())
                elif block_err and not best_err:
                    best_err = block_err
            joined = "\n".join(dict.fromkeys(chunks))
            if not joined:
                continue
            current_score = (len(joined.strip()), sum(ch.isdigit() for ch in joined))
            best_score = (len(best_text.strip()), sum(ch.isdigit() for ch in best_text))
            if current_score > best_score:
                best_text = joined
        return best_text, best_err
    finally:
        for tmpdir in tmpdirs:
            shutil.rmtree(tmpdir, ignore_errors=True)

def ocr_pdf_first_page(pdf_path):
    is_pdf = str(pdf_path).lower().endswith(".pdf")
    tmpdir = None
    tmp_generated = ""
    img_path = pdf_path
    if is_pdf:
        try:
            images, img_err, tmp_generated = pdftoppm_first_page(pdf_path, pages=1)
            if not images:
                return "", img_err or "pdftoppm: sin imagen"
            img_path = images[0]
        except Exception as err:
            return "", f"pdftoppm: {err}"
    if not os.path.exists(img_path):
        if tmp_generated:
            shutil.rmtree(tmp_generated, ignore_errors=True)
        return "", "imagen no encontrada para OCR"
    lang = detect_ocr_lang()
    tesseract_cmd = (
        shutil.which("tesseract")
        or "/opt/homebrew/bin/tesseract"
        or "/usr/local/bin/tesseract"
    )
    if not tesseract_cmd or not os.path.exists(tesseract_cmd):
        if tmp_generated:
            shutil.rmtree(tmp_generated, ignore_errors=True)
        return "", "tesseract no encontrado en PATH"
    env = os.environ.copy()
    if os.path.isdir(TESSDATA_DIR):
        env["TESSDATA_PREFIX"] = TESSDATA_DIR
    def run_tesseract(psm):
        result = run_subprocess(
            [
                tesseract_cmd,
                img_path,
                "stdout",
                "-l",
                lang,
                "--oem",
                "1",
                "--psm",
                str(psm),
                "-c",
                "user_defined_dpi=300",
                "-c",
                "preserve_interword_spaces=1",
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            env=env,
            text=True,
        )
        return result.stdout or ""
    try:
        candidates = []
        for psm in OCR_TESSERACT_PSMS:
            try:
                candidates.append(run_tesseract(psm))
            except subprocess.CalledProcessError:
                continue
        if not candidates:
            return "", "tesseract: sin salida"
        best = max(candidates, key=lambda t: (len(t.strip()), sum(ch.isdigit() for ch in t)))
        return best, ""
    except subprocess.CalledProcessError as err:
        return "", f"tesseract: {err.stderr.strip()}"
    except Exception as err:
        return "", f"tesseract: {err}"
    finally:
        if tmp_generated:
            shutil.rmtree(tmp_generated, ignore_errors=True)

def pdftotext_extract(pdf_path, pages=None):
    cmd = (
        shutil.which("pdftotext")
        or "/opt/homebrew/bin/pdftotext"
        or "/usr/local/bin/pdftotext"
    )
    if not cmd or not os.path.exists(cmd):
        return "", "pdftotext no encontrado"
    tmp_base = tempfile.gettempdir()
    with tempfile.TemporaryDirectory(dir=tmp_base) as tmpdir:
        out_txt = os.path.join(tmpdir, "out.txt")
        args = [cmd, "-layout", "-nopgbrk"]
        if pages and isinstance(pages, int):
            args.extend(["-f", "1", "-l", str(pages)])
        try:
            run_subprocess(
                [*args, pdf_path, out_txt],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
            )
        except subprocess.CalledProcessError as err:
            return "", f"pdftotext: {err.stderr.strip()}"
        except Exception as err:
            return "", f"pdftotext: {err}"
        if not os.path.exists(out_txt):
            return "", "pdftotext: sin salida"
        try:
            with open(out_txt, "r", encoding="utf-8", errors="ignore") as handle:
                return handle.read(), ""
        except Exception as err:
            return "", f"pdftotext: {err}"

def pdfinfo_page_size(pdf_path):
    cmd = (
        shutil.which("pdfinfo")
        or "/opt/homebrew/bin/pdfinfo"
        or "/usr/local/bin/pdfinfo"
    )
    if not cmd or not os.path.exists(cmd):
        return None
    try:
        result = run_subprocess(
            [cmd, pdf_path],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
    except Exception:
        return None
    for line in result.stdout.splitlines():
        if line.lower().startswith("page size"):
            parts = re.findall(r"([0-9]+(?:\.[0-9]+)?)", line)
            if len(parts) >= 2:
                try:
                    return float(parts[0]), float(parts[1])
                except Exception:
                    return None
    return None

def pdftotext_crop(pdf_path, x, y, w, h):
    cmd = (
        shutil.which("pdftotext")
        or "/opt/homebrew/bin/pdftotext"
        or "/usr/local/bin/pdftotext"
    )
    if not cmd or not os.path.exists(cmd):
        return ""
    tmp_base = tempfile.gettempdir()
    with tempfile.TemporaryDirectory(dir=tmp_base) as tmpdir:
        out_txt = os.path.join(tmpdir, "crop.txt")
        args = [
            cmd,
            "-f",
            "1",
            "-l",
            "1",
            "-layout",
            "-nopgbrk",
            "-x",
            str(int(x)),
            "-y",
            str(int(y)),
            "-W",
            str(int(w)),
            "-H",
            str(int(h)),
        ]
        try:
            run_subprocess(
                [*args, pdf_path, out_txt],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
            )
            if not os.path.exists(out_txt):
                return ""
            with open(out_txt, "r", encoding="utf-8", errors="ignore") as handle:
                return handle.read()
        except Exception:
            return ""

def pdftoppm_first_page(pdf_path, pages=None, dpi=None):
    cmd = (
        shutil.which("pdftoppm")
        or "/opt/homebrew/bin/pdftoppm"
        or "/usr/local/bin/pdftoppm"
    )
    if not cmd or not os.path.exists(cmd):
        return [], "pdftoppm no encontrado", ""
    tmp_base = tempfile.gettempdir()
    tmpdir = tempfile.mkdtemp(dir=tmp_base)
    base = os.path.join(tmpdir, "page")
    args = [cmd, "-f", "1"]
    if pages is None and OCR_PDF_MAX_PAGES > 0:
        pages = OCR_PDF_MAX_PAGES
    if pages and isinstance(pages, int):
        args.extend(["-l", str(pages)])
    render_dpi = dpi or OCR_PDF_DPI
    try:
        run_subprocess(
            [*args, "-r", str(render_dpi), "-gray", "-png", pdf_path, base],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.PIPE,
            text=True,
        )
    except subprocess.CalledProcessError as err:
        shutil.rmtree(tmpdir, ignore_errors=True)
        return [], f"pdftoppm: {err.stderr.strip()}", ""
    except Exception as err:
        shutil.rmtree(tmpdir, ignore_errors=True)
        return [], f"pdftoppm: {err}", ""
    images = sorted(
        [
            os.path.join(tmpdir, name)
            for name in os.listdir(tmpdir)
            if name.startswith("page-") and name.endswith(".png")
        ]
    )
    if not images:
        shutil.rmtree(tmpdir, ignore_errors=True)
        return [], "pdftoppm: sin imagen", ""
    return images, "", tmpdir

def extract_pdf_text(pdf_path):
    text, err = pdftotext_extract(pdf_path, pages=None)
    if text and len(text.strip()) >= 30:
        return text, "", "pdftotext"
    images, img_err, tmpdir = pdftoppm_first_page(pdf_path, pages=None)
    if images:
        combined = []
        try:
            for img_path in images:
                page_text, ocr_err = ocr_pdf_first_page(img_path)
                if page_text:
                    combined.append(page_text)
            if combined:
                return "\n".join(combined), "", "tesseract"
            return "", ocr_err, "tesseract"
        finally:
            if tmpdir:
                shutil.rmtree(tmpdir, ignore_errors=True)
    text, ocr_err = ocr_pdf_first_page(pdf_path)
    if text:
        return text, "", "tesseract"
    return "", err or img_err or ocr_err, "tesseract"

def ocr_pdf_all_pages(pdf_path, use_external=False):
    images, img_err, tmpdir = pdftoppm_first_page(pdf_path, pages=None)
    if images:
        combined = []
        last_err = ""
        try:
            for img_path in images:
                page_text = ""
                ocr_err = ""
                if use_external:
                    vision_bytes = prepare_image_bytes_for_vision(img_path)
                    if vision_bytes:
                        page_text, ocr_err = ocr_image_external(vision_bytes)
                if not page_text:
                    page_text, ocr_err = ocr_pdf_first_page(img_path)
                if page_text:
                    combined.append(page_text)
                elif ocr_err:
                    last_err = ocr_err
            if combined:
                return "\n".join(combined), ""
            return "", last_err or img_err
        finally:
            if tmpdir:
                shutil.rmtree(tmpdir, ignore_errors=True)
    text, ocr_err = ocr_pdf_first_page(pdf_path)
    if text:
        return text, ""
    return "", ocr_err or img_err

def parse_poliza_text(text, source_hint="", hinted_company=""):
    DATE_TOKEN = r"(?<!\d)(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})(?!\d)"
    cleaned = text.replace("\u00a0", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = cleaned.replace("N°", "Nº").replace("Nº", "Nº")
    cleaned = cleaned.replace("Poliza", "Póliza").replace("Poliza", "Póliza")
    def pick(patterns):
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()
            m = re.search(pat, cleaned, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""
    def normalize_ocr_digits(value):
        if not value:
            return ""
        raw = str(value).strip()
        if not raw:
            return ""
        confusable = set("OoIlSB")
        ratio = sum(1 for ch in raw if ch.isdigit() or ch in confusable) / max(len(raw), 1)
        if ratio < 0.6:
            return raw
        table = str.maketrans({"O": "0", "o": "0", "I": "1", "l": "1", "S": "5", "B": "8"})
        return raw.translate(table)
    def normalize_nif_ocr(value):
        if not value:
            return ""
        raw = str(value).strip().upper()
        if not raw:
            return ""
        raw = raw.replace(" ", "").replace("-", "").replace(".", "")
        raw = raw.replace("O", "0").replace("I", "1").replace("L", "1").replace("S", "5").replace("B", "8")
        raw = re.sub(r"[^A-Z0-9]", "", raw)
        if len(raw) > 9:
            raw = raw[:9]
        return raw
    def is_valid_nif(value):
        if not value:
            return False
        return bool(
            re.match(r"^[0-9]{8}[A-Z]$", value)
            or re.match(r"^[XYZ][0-9]{7}[A-Z]$", value)
            or re.match(r"^[ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z]$", value)
        )
    def normalize_nif_candidate(value):
        if not value:
            return ""
        candidate = normalize_nif_ocr(value)
        if is_valid_nif(candidate):
            return candidate
        if len(candidate) >= 9:
            candidate = candidate[:9]
            if is_valid_nif(candidate):
                return candidate
        return ""
    def normalize_ocr_date(value):
        if not value:
            return ""
        raw = str(value).strip()
        if not raw:
            return ""
        raw = normalize_ocr_digits(raw)
        raw = re.sub(r"[.\-]", "/", raw)
        raw = re.sub(r"(\d)\s+(\d)", r"\1/\2", raw)
        return raw
    def extract_date_range(text_value):
        if not text_value:
            return "", ""
        value = normalize_ocr_date(text_value)
        value = value.replace(" a ", " hasta ").replace(" al ", " hasta ")
        match = re.search(
            rf"(?:desde|vigencia\s+desde|periodo\s+del\s+seguro)\s*[:\-]?\s*{DATE_TOKEN}.*?"
            rf"(?:hasta|a)\s*{DATE_TOKEN}",
            value,
            re.IGNORECASE,
        )
        if match:
            return match.group(1), match.group(2)
        dates = re.findall(DATE_TOKEN, value)
        if len(dates) >= 2:
            return dates[0], dates[1]
        return "", ""
    def normalize_poliza_number(value, compania=""):
        if not value:
            return ""
        raw = normalize_ocr_digits(value)
        raw = re.sub(r"\s+", "", raw).strip()
        if not raw:
            return ""
        # Formato típico en pólizas de impago (GAGxxxxx) que a veces llega contaminado.
        gag_match = re.search(r"\bGAG[0-9O][A-Z0-9]{4,}\b", raw, re.IGNORECASE)
        if gag_match:
            token = gag_match.group(0).upper()
            suffix = token[3:]
            suffix = suffix.replace("O", "0")
            digits = re.match(r"([0-9]{5,})", suffix)
            if digits:
                return f"GAG{digits.group(1)}"
            token_digits = re.sub(r"[^0-9]", "", suffix)
            if len(token_digits) >= 5:
                return f"GAG{token_digits}"
            return token
        raw = re.sub(r"^(N[ºo]\s*)", "", raw, flags=re.IGNORECASE)
        raw = re.sub(r"^POLIZA", "", raw, flags=re.IGNORECASE)
        raw = raw.strip(":-#")
        candidates = re.split(r"[ \t,/]+", raw)
        candidates = [c for c in candidates if c]
        if not candidates:
            candidates = [raw]
        normalized_company = (compania or "").upper().strip()
        normalized_company_key = normalize_company_key(compania or "")
        if normalized_company_key == "AXA" or "AXA" in normalized_company_key:
            axa_fmt = re.search(r"\b(\d{2})[- ]?(\d{7,})\b", value or "", re.IGNORECASE)
            if axa_fmt:
                return f"{axa_fmt.group(1)}-{axa_fmt.group(2)}"
            digits = re.sub(r"[^0-9]", "", str(value or ""))
            if len(digits) >= 10:
                return f"{digits[:2]}-{digits[2:]}"
        numeric_only = {
            "LINEA DIRECTA",
            "DIRECT SEGUROS",
            "FENIX DIRECTO",
            "MAPFRE",
            "OCASO",
            "SANTA LUCIA",
            "SANTALUCIA",
        }
        alnum_structured = {
            "AXA",
            "ALLIANZ",
            "GENERALI",
            "REALE",
            "ZURICH",
            "HELVETIA",
            "CASER",
            "LIBERTY",
            "MUTUA MADRILEÑA",
            "MUTUA MADRILENA",
        }
        if normalized_company in numeric_only:
            digit_runs = re.findall(r"\d{5,}", raw)
            if digit_runs:
                return max(digit_runs, key=len)
        if normalized_company in alnum_structured:
            alnum = re.findall(r"[A-Z0-9]{6,}", raw)
            if alnum:
                return max(alnum, key=len)
        # Prefer token with most digits and length between 5 and 20
        def score_token(token):
            digits = len(re.findall(r"\d", token))
            length = len(token)
            if length < 5 or length > 24:
                return -1
            return digits * 2 + length
        best = max(candidates, key=score_token)
        return best
    def is_valid_poliza_candidate(value):
        token = normalize_poliza_key(value)
        if not token or len(token) < 6:
            return False
        if not re.search(r"\d", token):
            return False
        month_noise = (
            "ENERO",
            "FEBRERO",
            "MARZO",
            "ABRIL",
            "MAYO",
            "JUNIO",
            "JULIO",
            "AGOSTO",
            "SEPTIEMBRE",
            "OCTUBRE",
            "NOVIEMBRE",
            "DICIEMBRE",
        )
        if any(m in token for m in month_noise):
            return False
        if token in ("POLIZA", "IMPAGO", "HOGAR", "AUTO", "SEGURO"):
            return False
        return True
    def line_pick(keys):
        if not keys:
            return ""
        key_pattern = "|".join([re.escape(k) for k in keys])
        for line in text.splitlines():
            if re.search(rf"\b({key_pattern})\b", line, re.IGNORECASE):
                parts = re.split(r"[:\-]", line, maxsplit=1)
                if len(parts) > 1:
                    value = parts[1].strip()
                    if value:
                        return value
                cleaned_line = re.sub(rf".*?\b({key_pattern})\b", "", line, flags=re.IGNORECASE).strip()
                if cleaned_line:
                    return cleaned_line
        return ""
    def pick_date_range(value):
        if not value:
            return "", ""
        dates = re.findall(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", value)
        if len(dates) >= 2:
            return dates[0], dates[1]
        return "", ""
    def clean_tomador_value(value):
        raw = normalize_person_name(value)
        if not raw:
            return ""
        raw = re.sub(r"^(NOMBRE|TOMADOR|ASEGURADO|CONTRATANTE|TITULAR)\s*[:\-]?\s*", "", raw, flags=re.IGNORECASE)
        stop_match = re.search(
            r"\b("
            r"DOCUMENTO|DOC\.?|NIF|DNI|CIF|MATRICULA|MATRÍCULA|VEHICULO|VEHÍCULO|"
            r"TIPO|RIESGO|PLANTA|SITUACION|SITUACIÓN|USO|COBERTURAS|GARANTIAS|GARANTÍAS|"
            r"DECLARACION|DECLARACIÓN|BENEFICIARIO|CLAUSULAS|CLÁUSULAS"
            r")\b",
            raw,
            re.IGNORECASE,
        )
        if stop_match:
            raw = raw[: stop_match.start()].strip()
        raw = re.sub(r"\s{2,}", " ", raw).strip(" ,;:-")
        raw_upper = normalize_lookup_text(raw)
        banned_fragments = (
            "DE ESTA POLIZA",
            "SISTEMA DE REGULARIZACION",
            "LIBRO REGISTRO",
            "DECLARACION DE SALUD",
            "COBERTURAS Y GARANTIAS",
            "COBERTURAS",
            "GARANTIAS",
            "TIPO DE RIESGO",
            "VIVIENDA EN VECINDAD",
            "PERSONA FISICA O JURIDICA",
            "EL TOMADOR DEL SEGURO",
            "TITULAR DEL INTERES",
            "DATOS DE TU MEDIADOR",
            "BENEFICIARIOS OTROS",
            "INCLUIDO",
            "NO INCLUIDO",
            "UNA PERSONA FISICA",
            "UNA PERSONA JURIDICA",
            "MAYOR DE EDAD",
            "TITULAR DEL INTERES OBJETO",
            "PERCEPTOR DEL PAGO DE LA INDEMNIZACION",
            "CUANDO QUIEN SE OPONGA",
            "ARRENDATARIO",
            "ARRENDADOR",
            "DESAHUCIO",
            "RENTA VENCIDA",
            "MES DE RENTA",
            "MESES DE RENTA",
            "NOMBRE Y APELLIDOS",
            "Y APELLIDOS",
        )
        if any(fragment in raw_upper for fragment in banned_fragments):
            return ""
        if re.search(r"\buna\s+persona\s+f[ií]sica\b.*\bo\s+una\s+persona\b", raw, re.IGNORECASE):
            return ""
        legal_noise = (
            r"\b(cuando|quien|oponga|arrendatari[oa]s?|arrendador|desahucio|renta|vencida|mensualidad(?:es)?)\b"
        )
        if re.search(legal_noise, raw, re.IGNORECASE):
            return ""
        looks_company = bool(
            re.search(r"\b(SL|S\.L\.?|SLU|S\.L\.U\.?|SA|S\.A\.?|SCP|S\.C\.P\.?|CB|C\.B\.?)\b", raw, re.IGNORECASE)
        )
        if "€" in raw or (re.search(r"\d{2,}", raw) and not looks_company):
            return ""
        words = raw.split()
        if not words:
            return ""
        if len(words) == 1:
            token = normalize_lookup_text(words[0])
            allowed_single = ("SL", "S L", "SA", "S A", "SCP", "S C P", "CB", "C B")
            if token in ("DE", "DEL", "LA", "EL", "UNA", "UNO", "VENEZUELA"):
                return ""
            if len(token) < 4:
                return ""
            if token not in allowed_single and len(token) < 7:
                return ""
        if len(words) > 9:
            raw = " ".join(words[:9])
        return raw
    def looks_corporate_name(value):
        norm = normalize_lookup_text(value or "")
        if not norm:
            return False
        corporate_tokens = (
            " SL",
            " S L",
            " SA",
            " S A",
            " SLU",
            " SCP",
            " CB",
            " SOCIEDAD",
            " FINANCIACIONES",
            " INMOBILIARIA",
            " ESTUDIO",
            " ESTUDIOS",
            " COMUNIDAD",
            " CLUB",
            " CP ",
            " C P ",
        )
        return any(token in f" {norm} " for token in corporate_tokens)
    def parse_from_source_hint():
        out = {}
        raw = str(source_hint or "").strip()
        if not raw:
            return out
        name = Path(raw).name
        stem = Path(name).stem
        upper = stem.upper()
        if "SANTA LUCIA" in upper or "STA LUCIA" in upper or "SANTALUCIA" in upper:
            out["compania"] = "Santa Lucia"
        elif "FIATC" in upper or "FIACT" in upper:
            out["compania"] = "Fiatc"
        elif "IPTIQ" in upper or "GALLEN" in upper:
            out["compania"] = "iptiQ EMEA P"
        elif "ARAG" in upper:
            out["compania"] = "ARAG"
        elif "OCASO" in upper:
            out["compania"] = "Ocaso"
        elif "REALE" in upper:
            out["compania"] = "Reale"
        elif "MAPFRE" in upper:
            out["compania"] = "Mapfre"
        elif "PELAYO" in upper:
            out["compania"] = "Pelayo"
        elif "AXA" in upper:
            out["compania"] = "AXA"
        elif "ALLIANZ" in upper:
            out["compania"] = "Allianz"
        elif "ZURICH" in upper:
            out["compania"] = "Zurich"
        elif "EUROINS" in upper:
            out["compania"] = "Euroins"
        elif "MUTUA PROPIETARIOS" in upper:
            out["compania"] = "Mutua Propietarios"
        elif "OCCIDENT" in upper:
            out["compania"] = "Catalana Occidente"
        tokens = re.findall(r"[A-Z0-9][A-Z0-9/_-]{5,}", upper)
        candidate_polizas = []
        for token in tokens:
            cleaned = token.strip("_-/")
            cleaned = cleaned.replace("_", "/")
            if is_valid_poliza_candidate(cleaned):
                candidate_polizas.append(cleaned)
        if candidate_polizas:
            candidate_polizas.sort(key=lambda t: (len(re.sub(r"\D", "", t)), len(t)), reverse=True)
            out["poliza_numero"] = candidate_polizas[0]
        temp = upper
        for junk in (
            "POLIZA",
            "SEGURO",
            "IMPAGO",
            "HOGAR",
            "AUTO",
            "RC",
            "SALUD",
            "STA LUCIA",
            "SANTA LUCIA",
            "SANTALUCIA",
            "MAPFRE",
            "REALE",
            "OCASO",
            "PELAYO",
            "AXA",
            "ALLIANZ",
            "FIATC",
            "FIACT",
            "IPTIQ",
            "GALLEN",
            "ARAG",
            "ZURICH",
            "OCCIDENT",
            "CATALANA",
            "EUROINS",
            "DOCUMENTACION",
            "EMISION",
            "MANDATO",
            "SEPA",
            "CERTIFICADO",
            "CONDICIONES",
            "PARTICULARES",
            "GENERALES",
        ):
            temp = re.sub(rf"\b{re.escape(junk)}\b", " ", temp)
        temp = re.sub(r"\b\d{1,4}\b", " ", temp)
        temp = re.sub(r"[_\-]+", " ", temp)
        temp = re.sub(r"\s+", " ", temp).strip()
        parts = [p for p in temp.split(" ") if p and len(p) >= 3]
        if parts:
            # Prefer final words as they are usually the person/entity name in filenames.
            tail = " ".join(parts[-4:])
            tail = clean_tomador_value(tail)
            if tail and len(tail) >= 5:
                out["tomador"] = tail
        if " RC " in f" {upper} ":
            out["ramo"] = "Responsabilidad civil"
        elif " HOGAR " in f" {upper} ":
            out["ramo"] = "Hogar"
        elif " AUTO " in f" {upper} ":
            out["ramo"] = "Auto"
        elif " MULTIRRIESGO " in f" {upper} ":
            out["ramo"] = "Hogar"
        elif " IMPAGO " in f" {upper} ":
            out["ramo"] = "Impago alquiler"
        elif " COMERCIO " in f" {upper} ":
            out["ramo"] = "Comercio"
        return out
    def company_specific_poliza(compania, base_text):
        comp_key = normalize_company_key(compania or "")
        if not comp_key:
            return ""
        patterns_common = [
            r"(?:N[ºO]\s*(?:DE\s*)?P[ÓO]LIZA|P[ÓO]LIZA|CONTRATO|CERTIFICADO)\s*[:#]?\s*([A-Z0-9][A-Z0-9/\- ]{6,})",
        ]
        patterns = []
        if comp_key == "MAPFRE":
            patterns = [
                r"\b([0-9]{10,13}(?:\s*/\s*[0-9]{3})?)\b",
                r"\b([0-9]{4,6}[A-Z]?[0-9]{5,8})\b",
            ]
        elif comp_key == "REALE":
            patterns = [
                r"\b([0-9]{12,14}(?:\s*[/-]?\s*[0-9]{1,3})?)\b",
            ]
        elif comp_key == "OCASO":
            patterns = [
                r"\b([0-9]\s*[A-Z]-[A-Z0-9]{7,})\b",
                r"\b([A-Z0-9]{1,3}\s*[A-Z]-[A-Z0-9]{7,})\b",
            ]
        elif comp_key == "AXA":
            patterns = [
                r"\b([0-9]{6,8}(?:\s*/\s*[0-9]{1,3})?)\b",
                r"\b([A-Z0-9]{6,12}(?:\s*/\s*[0-9]{1,3})?)\b",
            ]
        elif comp_key == "ALLIANZ":
            patterns = [
                r"\b([A-Z][0-9]{6,8}\s*[-/]\s*[0-9]{3,4})\b",
                r"\b([A-Z][0-9]{6,8})\b",
            ]
        elif comp_key == "MUTUA PROPIETARIOS":
            patterns = [
                r"\bPOLIZA\s*[:#]?\s*([A-Z0-9][A-Z0-9/\- ]{6,})",
                r"\bCONTRATO\s*[:#]?\s*([A-Z0-9][A-Z0-9/\- ]{6,})",
            ]
        elif comp_key == "CATALANA OCCIDENTE":
            patterns = [
                r"\bP[ÓO]LIZA\s*[:#]?\s*([A-Z0-9][A-Z0-9/\- ]{6,})",
                r"\bCERTIFICADO\s*[:#]?\s*([A-Z0-9][A-Z0-9/\- ]{6,})",
            ]
        for pat in [*patterns_common, *patterns]:
            for target in (base_text, cleaned):
                m = re.search(pat, target, re.IGNORECASE)
                if not m:
                    continue
                candidate = normalize_poliza_number(m.group(1), compania)
                if candidate and len(candidate) >= 6 and re.search(r"\d", candidate):
                    return candidate
        return ""
    def company_specific_tomador(base_text):
        patterns = [
            r"\bPresentado\s+al\s+SR\.?/Sra\.?\s*([A-ZÁÉÍÓÚÑ ,.'\-]+?)\s+N[º°]?\s*Documento",
            r"\bNombre\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{4,}?)(?:\s+Documento|\s+Doc\.?|\s+NIF|\s+DNI|\s+CIF)",
            r"\b(?:Tomador|Asegurado(?:\s+principal)?|Contratante|Titular)\s*[:\-]\s*([^\n]+)",
        ]
        for pat in patterns:
            for target in (base_text, cleaned):
                m = re.search(pat, target, re.IGNORECASE)
                if m:
                    candidate = clean_tomador_value(m.group(1))
                    if candidate and len(candidate) >= 5:
                        return candidate
        return ""
    def company_specific_ramo(compania, base_text):
        comp_key = normalize_company_key(compania or "")
        target = normalize_lookup_text(base_text)
        if "RESPONSABILIDAD CIVIL" in target or re.search(r"\bRC\b", target):
            return "Responsabilidad civil"
        if "MULTIRRIESGO HOGAR" in target or ("HOGAR" in target and "ALQUILER" not in target):
            return "Hogar"
        if "IMPAGO" in target and "ALQUILER" in target:
            return "Impago alquiler"
        if "AUTOMOVIL" in target or "AUTOMOVILES" in target or re.search(r"\bAUTO\b", target):
            return "Auto"
        if comp_key == "MUTUA PROPIETARIOS" and ("MULTIRRIESGO" in target or "HOGAR" in target):
            return "Hogar"
        return ""
    fields = {}
    fields["tomador"] = pick([
        r"DATOS\s+DEL\s+TOMADOR\s+Y\s+PROPIETARIO\s+Nombre\s+([A-ZÁÉÍÓÚÑ ,.'\-]{5,})\s+Documento\s+ID",
        r"P[oó]liza/Spto\s+[0-9]{8,14}\s*/\s*[0-9]{1,3}\s*\n+\s*([A-ZÁÉÍÓÚÑ ,.'\-]{5,})",
        r"TOMADOR\s+([A-ZÁÉÍÓÚÑ ,.'\-]{5,}?)\s+NIF\b",
        r"Presentado\s+al\s+SR\.?/Sra\.?\s*([A-ZÁÉÍÓÚÑ ,.'\-]+?)\s+N[º°]?\s*Documento",
        r"Nombre y apellidos\s*:\s*([^\n]+)",
        r"Titular\s*de\s*la\s*p[oó]liza\s*:\s*([^\n]+)",
        r"Datos\s+del\s+asegurado\s*[:\-]?\s*([^\n]+)",
        r"Asegurado/Tomador\s*[:\-]?\s*([^\n]+)",
        r"Tomador\s*y\s+Asegurado\s*[:\-]?\s*([^\n]+)",
        r"Datos\s+Tomador\s*/\s*Conductor\s*Nombre\s*([A-ZÁÉÍÓÚÑ\s]+?)\s+Doc",
        r"Nombre\s+([A-ZÁÉÍÓÚÑ\s]+?)\s+Doc\.?\s+Identificaci[oó]n",
        r"Tomador\s*de\s*seguro\s*:\s*([^\n]+)",
        r"Tomador\s*:\s*([^\n]+)",
        r"Asegurado\s*:\s*([^\n]+)",
        r"Asegurado\s*principal\s*:\s*([^\n]+)",
        r"Titular\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ\s]+)",
        r"Nombre\s*[:\-]?\s*([A-ZÁÉÍÓÚÑ\s]+)",
        r"Asegurado\s+y\s+Tomador\s*[:\-]?\s*([^\n]+)",
        r"Contratante\s*[:\-]?\s*([^\n]+)",
    ])
    if fields["tomador"]:
        fields["tomador"] = normalize_person_name(fields["tomador"])
    fields["dni"] = pick([
        r"(?:TOMADOR|ASEGURADO)[\s\S]{0,160}?NIF\s*[:\-]?\s*([A-Z0-9]{8,9})",
        r"DOC\.?\s*ID\.?\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"Doc\.?\s*Identificaci[oó]n\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"NIF/CIF\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"DNI\s*[:\-]?\s*([0-9]{8}[A-Z])",
        r"NIF\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"CIF\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"DNI/NIF\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"Documento\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"\b([0-9]{8}[A-Z])\b",
        r"\b([ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z])\b",
        r"\b([XYZ][0-9]{7}[A-Z])\b",
    ])
    fields["telefono"] = pick([
        r"Tel[eé]fono\s*[:\-]?\s*([0-9\s]{9,})",
        r"M[oó]vil\s*[:\-]?\s*([0-9\s]{9,})",
    ])
    fields["email"] = pick([
        r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})",
    ])
    fields["direccion"] = pick([
        r"TOMADOR[\s\S]{0,220}?NIF\s*:\s*[A-Z0-9]{8,9}\s*\n\s*([^\n]+)\s*\n\s*(\d{5}\s+[A-ZÁÉÍÓÚÑ\s]+)",
        r"Direcci[oó]n\s*[:\-]?\s*([^\n]+)",
        r"Direcci[oó]n\s+([A-Z0-9ÁÉÍÓÚÑ\s,./-]+?\s+\d{5}\s+[A-ZÁÉÍÓÚÑ\s]+)",
        r"Direcci[oó]n\s+([A-Z0-9ÁÉÍÓÚÑ\s,./-]+?)(?:\s+Uso\s+|\s+Beneficiario|\s+Cl[aá]usulas|\s+Datos|\n)",
        r"Domicilio\s*[:\-]?\s*([^\n]+)",
    ])
    # REALE often provides address split into "Nombre de Vía" + "Código Postal/Población".
    via_name = pick([r"Nombre\s+de\s+V[ií]a\s*:\s*([^\n]+)"])
    via_cp = pick([r"C[oó]digo\s+Postal\s*:\s*(\d{5})"])
    via_city = pick([r"Poblaci[oó]n\s*:\s*([^\n]+)"])
    if via_city:
        via_city = via_city.split("Titularidad")[0].strip()
    via_compose = " ".join(part for part in [via_name, f"{via_cp} {via_city}".strip()] if part).strip()
    if via_compose:
        if not fields.get("direccion"):
            fields["direccion"] = via_compose
        else:
            dnorm = normalize_lookup_text(fields.get("direccion"))
            if "DE LA VIVIENDA ASEGURADA" in dnorm or "ANTONIO CHACON" in dnorm:
                fields["direccion"] = via_compose
    # Additional REALE-like extraction from TOMADOR block when address is split in lines.
    tomador_block = re.search(
        r"^\s*TOMADOR\s*$([\s\S]*?)^\s*ASEGURADO\s*$",
        text,
        re.IGNORECASE | re.MULTILINE,
    )
    if tomador_block:
        block_lines = [ln.strip() for ln in tomador_block.group(1).splitlines() if ln.strip()]
        street_line = ""
        zip_city_line = ""
        for ln in block_lines:
            if (not street_line) and re.search(r"\b(CL|C/|AVD|AVDA|PZ|PLAZA|CALLE|PS|PASEO)\b", ln, re.IGNORECASE):
                street_line = ln
            if (not zip_city_line) and re.search(r"\b\d{5}\b", ln):
                zip_city_line = ln
        if street_line:
            street_line = re.sub(r"\b[A-Z0-9]{12,}\b", "", street_line).strip(" ,;:-")
            block_addr = " ".join(part for part in [street_line, zip_city_line] if part).strip()
            dnorm = normalize_lookup_text(fields.get("direccion") or "")
            if (not fields.get("direccion")) or "ANTONIO CHACON" in dnorm or "DE LA VIVIENDA ASEGURADA" in dnorm:
                fields["direccion"] = block_addr
    fields["fecha_nacimiento"] = pick([
        r"Fecha de nacimiento\s*[:\-]?\s*([0-9]{1,2}[ /.-][0-9]{1,2}[ /.-][0-9]{2,4})",
        r"F\.?\s*nacimiento\s*[:\-]?\s*([0-9]{1,2}[ /.-][0-9]{1,2}[ /.-][0-9]{2,4})",
    ])
    if fields["fecha_nacimiento"]:
        fields["fecha_nacimiento"] = normalize_ocr_date(fields["fecha_nacimiento"])
    fields["poliza_numero"] = pick([
        r"p[oó]liza\s*(?:n[º°o]|no\.?)\s*[:#]?\s*([0-9]{2}-[0-9]{7,})",
        r"P[oó]liza/Spto\s*([0-9]{8,14})(?:\s*/\s*[0-9]{1,3})?",
        r"Referencia\s*[:#]?\s*([A-Z0-9]{8,})",
        r"N[ºo]\s*POLIZA/SPTO\.?\s*[:#]?\s*([0-9]{8,14}(?:\s*/\s*[0-9]{1,3})?)",
        r"P[oó]liza\s*/\s*Producto\s*[:#]?\s*([0-9]{6,})",
        r"P[oó]liza\s*/\s*Producto\s*[:#]?\s*([A-Z0-9][A-Z0-9/.\-\s]{5,}?)(?:\s{2,}[A-ZÁÉÍÓÚÑ]|$)",
        r"N[ºo]\s*de\s*p[oó]liza\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[ºo]\s*p[oó]liza\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[ºo]\s*de\s*P[oó]liza\s*([A-Z0-9/.\-]+)",
        r"N[ºo]\s+de\s+P[oó]liza\s+([A-Z0-9/.\-]+)",
        r"N[ºo]\s+P[oó]liza\s*[:#]?\s*([0-9]{5,})",
        r"(?:P[oó]liza|P[oó]liza\s*n[ºo]|N[ºo] P[oó]liza)\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"P[oó]liza\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"P[oó]liza\s*N[úu]m\.?\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"Poliza\s*No\.?\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[úu]m\.?\s*P[oó]liza\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[úu]mero\s*de\s*p[oó]liza\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"Certificado\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[ºo]\s*de\s*certificado\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"N[ºo]\s*de\s*contrato\s*[:#]?\s*([A-Z0-9/.\-]+)",
        r"Contrato\s*[:#]?\s*([A-Z0-9/.\-]+)",
    ])
    source_fields = parse_from_source_hint()
    fields["compania"] = hinted_company or detect_company_from_metadata(source_hint) or source_fields.get("compania") or detect_company_from_text(cleaned)
    if not fields["compania"]:
        fields["compania"] = pick([
            r"Compa[ñn]ia\s*[:\-]?\s*([A-Z0-9\s\-]+)",
            r"Aseguradora\s*[:\-]?\s*([A-Z0-9\s\-]+)",
            r"Entidad\s+aseguradora\s*[:\-]?\s*([A-Z0-9\s\-]+)",
            r"Compa[ñn]ia\s+aseguradora\s*[:\-]?\s*([A-Z0-9\s\-]+)",
        ])
    if fields["compania"]:
        normalized = re.sub(r"[^A-Z0-9 ]+", " ", fields["compania"].upper())
        normalized = re.sub(r"\s+", " ", normalized).strip()
        aliases = {
            "SANTALUCIA": "Santa Lucia",
            "SANTA LUCIA": "Santa Lucia",
            "MUTUA MADRILENA": "Mutua Madrileña",
            "MUTUA MADRILEÑA": "Mutua Madrileña",
            "LINEA DIRECTA": "Línea Directa",
            "DIRECT SEGUROS": "Direct Seguros",
            "FENIX DIRECTO": "Fénix Directo",
            "NATIONALE NEDERLANDEN": "Nationale Nederlanden",
            "CAJA RURAL": "Caja Rural",
            "PLUS ULTRA": "Plus Ultra",
            "HEL VETIA": "Helvetia",
            "SEGUROS BILBAO": "Seguros Bilbao",
            "CATALANA OCCIDENTE": "Catalana Occidente",
        }
        fields["compania"] = aliases.get(normalized, fields["compania"].strip())
    if fields["compania"]:
        fields["compania"] = normalize_company_name(fields["compania"])
    fields["fecha_efecto"] = pick([
        rf"Fecha\s+de\s+efecto\s+{DATE_TOKEN}",
        rf"Fecha\s*Efecto\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Fecha de efecto\s*:\s*{DATE_TOKEN}",
        rf"Efecto\s*:\s*{DATE_TOKEN}",
        rf"Vigencia\s*desde\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Fecha\s*inicio\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Inicio\s*vigencia\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Per[ií]odo\s*del\s*seguro\s*{DATE_TOKEN}",
        rf"Per[ií]odo\s*del\s*seguro\s*{DATE_TOKEN}\s*[0-9:]*",
        rf"Desde\s*[:\-]?\s*{DATE_TOKEN}",
    ])
    fields["fecha_vencimiento"] = pick([
        rf"Fecha\s+de\s+vencimiento\s+{DATE_TOKEN}",
        rf"Fecha de vencimiento\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Vencimiento\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Hasta\s+las\s+\d+\s+horas\s+del\s+{DATE_TOKEN}",
        rf"Fin\s*vigencia\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Vigencia\s*hasta\s*[:\-]?\s*{DATE_TOKEN}",
        rf"Hasta\s*[:\-]?\s*{DATE_TOKEN}",
    ])
    if fields["fecha_efecto"]:
        fields["fecha_efecto"] = normalize_ocr_date(fields["fecha_efecto"])
    if fields["fecha_vencimiento"]:
        fields["fecha_vencimiento"] = normalize_ocr_date(fields["fecha_vencimiento"])
    if not fields["fecha_efecto"] or not fields["fecha_vencimiento"]:
        vigencia_line = line_pick(["Vigencia", "Periodo", "Período", "Duración"])
        start_date, end_date = pick_date_range(vigencia_line)
        if start_date and not fields["fecha_efecto"]:
            fields["fecha_efecto"] = start_date
        if end_date and not fields["fecha_vencimiento"]:
            fields["fecha_vencimiento"] = end_date
    fields["ramo"] = pick([
        r"P[oó]liza\s*/\s*Producto\s*[:#]?\s*[A-Z0-9/\-]+\s*-\s*([^\n]+)",
        r"(AUTOM[ÓO]VILES?\s+PARTICULARES\s*-\s*[A-ZÁÉÍÓÚÑ ]+)",
        r"(MULTIRRIESGO\s+COMERCIOS?\s+Y\s+AUTOEMPRENDEDORES)",
        r"(MULTIRRIESGO\s+COMERCIOS?)",
        r"Condiciones\s+Particulares\s+Ocaso\s+([A-ZÁÉÍÓÚÑa-z\s]{4,})",
        r"Ramo\s*[:\-]?\s*([^\n]+)",
        r"Modalidad\s*[:\-]?\s*([^\n]+)",
        r"Modalidad\s+([A-ZÁÉÍÓÚÑa-z\s]+?)\s+Datos\s+Tomador",
        r"Producto\s*[:\-]?\s*([^\n]+)",
        r"Seguro\s*de\s*([A-ZÁÉÍÓÚÑa-z\s]+?)(?:\\.|\\n|Datos)",
    ])
    if not fields["ramo"]:
        ramo_keywords = [
            "Hogar",
            "Auto",
            "Automóvil",
            "Vida",
            "Salud",
            "Comercio",
            "RC",
            "Responsabilidad Civil",
            "Decesos",
            "Accidentes",
            "Comunidad",
            "Pymes",
            "Multirriesgo",
            "Ciber",
            "Mascotas",
            "Moto",
            "Coche",
            "Agro",
            "Construcción",
            "Transportes",
        ]
        for keyword in ramo_keywords:
            if re.search(rf"\\b{re.escape(keyword)}\\b", cleaned, re.IGNORECASE):
                fields["ramo"] = keyword
                break
    fields["prima_neta"] = pick([
        r"Prima neta\s*[:€]?\s*([0-9\.,]+)",
        r"Prima\s+neta.*?Total\s+Recibo.*?\bAnual\b[^\d]*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€",
        r"Importe\s*prima\s*neta\s*[:€]?\s*([0-9\.,]+)",
        r"Neta\s*[:€]?\s*([0-9\.,]+)",
        r"Prima\s+neta\s+anual\s*[:€]?\s*([0-9\.,]+)",
    ])
    fields["prima_total"] = pick([
        r"Prima total\s*[:€]?\s*([0-9\.,]+)",
        r"Total\s+Recibo.*?\bAnual\b.*?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€\s*$",
        r"Prima\s+neta.*?Total\s+Recibo.*?\bAnual\b.*?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€",
        r"Prima anual\s*[:€]?\s*([0-9\.,]+)",
        r"Importe\s*total\s*[:€]?\s*([0-9\.,]+)",
        r"Total\s*[:€]?\s*([0-9\.,]+)",
        r"Prima\s*total\s*anual\s*[:€]?\s*([0-9\.,]+)",
        r"Total\s+recibo\s*[:€]?\s*([0-9\.,]+)",
    ])
    # MAPFRE/others may render currency as "¤" and place amounts in the next line
    # under "PRIMA DEL SEGURO" table headers.
    amount_token = r"[0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2}"
    prima_table = re.search(r"PRIMA\s+DEL\s+SEGURO", text, re.IGNORECASE)
    if prima_table:
        window = text[prima_table.start() : prima_table.start() + 1800]
        amounts = re.findall(amount_token, window)
        if amounts:
            def _money_value(raw):
                try:
                    return float(raw.replace(".", "").replace(",", "."))
                except Exception:
                    return 0.0
            if not fields.get("prima_neta"):
                fields["prima_neta"] = amounts[0]
            if not fields.get("prima_total"):
                fields["prima_total"] = max(amounts, key=_money_value)
    if (not fields.get("prima_total")):
        primer_recibo = re.search(
            rf"Importe\s+a\s+pagar\s+del\s+primer\s+recibo\s*({amount_token})",
            text,
            re.IGNORECASE,
        )
        if primer_recibo:
            fields["prima_total"] = primer_recibo.group(1)
    # Table rows like "Del dd-mm-yyyy al dd-mm-yyyy ... Prima ... Total"
    # are usually the most reliable source for net/total amounts.
    for line in text.splitlines():
        if not re.search(r"\bDel\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s+al\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", line, re.IGNORECASE):
            continue
        amounts = re.findall(amount_token, line)
        if len(amounts) >= 2:
            fields["prima_neta"] = amounts[0]
            fields["prima_total"] = amounts[-1]
            break
    if fields["tomador"]:
        tomador = fields["tomador"].splitlines()[0].strip()
        for cut in ["Marca", "Matrícula", "Doc."]:
            if cut in tomador:
                tomador = tomador.split(cut)[0].strip()
        fields["tomador"] = clean_tomador_value(tomador)
    if fields["direccion"]:
        fields["direccion"] = re.sub(r"\s{2,}.*$", "", fields["direccion"]).strip()
        # Remove long cadastral/alphanumeric refs appended to the street line.
        fields["direccion"] = re.sub(r"\b[A-Z0-9]{12,}\b", "", fields["direccion"]).strip()
        fields["direccion"] = re.sub(r"\s{2,}", " ", fields["direccion"]).strip(" ,;:-")
        dir_upper = normalize_lookup_text(fields["direccion"])
        if any(
            token in dir_upper
            for token in (
                "DOMICILIO SOCIAL",
                "CARRETERA DE POZUELO",
                "MAJADAHONDA MADRID",
                "DATOS DE TU MEDIADOR",
            )
        ):
            fields["direccion"] = ""
        postal_match = re.search(r"\b\d{5}\s+[A-ZÁÉÍÓÚÑ\s]+\b", cleaned)
        if postal_match:
            postal_chunk = postal_match.group(0).strip()
            postal_norm = normalize_lookup_text(postal_chunk)
            dir_norm = normalize_lookup_text(fields["direccion"])
            # Skip mediator/office address fragments that leak from header blocks.
            if "ANTONIO CHACON" not in postal_norm and "DE ANDALUCIA" not in postal_norm:
                if postal_norm not in dir_norm:
                    fields["direccion"] = f"{fields['direccion']} {postal_chunk}".strip()
    if fields["telefono"]:
        fields["telefono"] = normalize_phone(fields["telefono"])
    if fields["email"]:
        fields["email"] = normalize_email(fields["email"])
        if "@" not in fields["email"] or " " in fields["email"]:
            fields["email"] = ""
    if fields["poliza_numero"]:
        if not re.search(r"\d", fields["poliza_numero"]):
            fields["poliza_numero"] = ""
    if not fields["poliza_numero"]:
        pol_match = re.search(r"N[ºo]\s*P[oó]liza\s*[:#]?\s*([0-9]{5,})", cleaned, re.IGNORECASE)
        if not pol_match:
            pol_match = re.search(r"N[ºo]\s*P[oó]liza\s*[:#]?\s*([0-9]{5,})", text, re.IGNORECASE)
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1).strip()
    if fields["dni"]:
        dni = normalize_nif_candidate(fields["dni"])
        fields["dni"] = dni
    if not fields["dni"]:
        dni_match = re.search(r"\b([0-9]{8}[A-Z])\b", text)
        if dni_match:
            fields["dni"] = normalize_nif_candidate(dni_match.group(1))
        else:
            cif_match = re.search(r"\b([ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z])\b", text)
            if cif_match:
                fields["dni"] = normalize_nif_candidate(cif_match.group(1))
    if fields["ramo"] and "@" in fields["ramo"]:
        fields["ramo"] = ""
    if fields["ramo"]:
        fields["ramo"] = fields["ramo"].splitlines()[0].strip()
        if re.match(r"^\d+[.)]?\s+", fields["ramo"]):
            fields["ramo"] = ""
        if "€" in fields["ramo"] or re.search(r"\d{2,}", fields["ramo"]):
            fields["ramo"] = ""
        if re.search(r"veh[ií]culo\s+asegurado|la\s+aseguradora|arrastrado", fields["ramo"], re.IGNORECASE):
            fields["ramo"] = ""
        if re.search(r"cobertura\s+total\s+o\s+parcial", fields["ramo"], re.IGNORECASE):
            fields["ramo"] = ""
        if len(fields["ramo"]) > 48 and "seguro" in normalize_lookup_text(fields["ramo"]):
            fields["ramo"] = ""
        ramo_inline = re.match(r"^[A-Z0-9/.\-]{5,}\s*-\s*(.+)$", fields["ramo"], re.IGNORECASE)
        if ramo_inline:
            fields["ramo"] = ramo_inline.group(1).strip()
        ramo_upper = normalize_lookup_text(fields["ramo"])
        if ramo_upper in ("DE SEGURO", "SEGURO", "TIPO DE SEGURO", "DEL SEGURO"):
            fields["ramo"] = ""
            ramo_upper = ""
        if "IMPAGO" in ramo_upper and "ALQUILER" in ramo_upper:
            fields["ramo"] = "Impago alquiler"
        elif "HOGAR" in ramo_upper and "ALQUILER" in ramo_upper:
            fields["ramo"] = "Hogar alquiler"
        elif "HOGAR" in ramo_upper:
            fields["ramo"] = "Hogar"
        elif "AUTOMOVIL" in ramo_upper or "AUTOMOVILES" in ramo_upper or "AUTO" in ramo_upper:
            fields["ramo"] = "Auto"
        elif "ALQUILER" in ramo_upper:
            fields["ramo"] = "Alquiler"
        elif "MULTIRRIESGO COMERCIOS" in ramo_upper:
            fields["ramo"] = "Comercio"
        elif "RESPONSABILIDAD CIVIL" in ramo_upper or ramo_upper == "RC":
            fields["ramo"] = "Responsabilidad civil"
    if not fields["ramo"]:
        modal_match = re.search(
            r"Modalidad\s+([A-ZÁÉÍÓÚÑa-z\s]+?)\s+Datos\s+Tomador",
            cleaned,
            re.IGNORECASE,
        )
        if modal_match:
            fields["ramo"] = modal_match.group(1).strip()
    if not fields["poliza_numero"] or not fields["ramo"] or not fields["fecha_efecto"]:
        lines = text.splitlines()
        for idx, line in enumerate(lines):
            if "Modalidad" in line and idx + 1 < len(lines):
                cols = [c for c in re.split(r"\s{2,}", lines[idx + 1].strip()) if c]
                if len(cols) >= 4:
                    if not fields["poliza_numero"] and re.search(r"\d{5,}", cols[1]):
                        fields["poliza_numero"] = re.search(r"\d{5,}", cols[1]).group(0)
                    if not fields["fecha_efecto"]:
                        fecha_match = re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", cols[2])
                        if fecha_match:
                            fields["fecha_efecto"] = fecha_match.group(0)
                    if not fields["ramo"]:
                        fields["ramo"] = cols[-1]
                break
    if not fields["fecha_efecto"] or not fields["fecha_vencimiento"]:
        start_date, end_date = extract_date_range(cleaned)
        if start_date and not fields["fecha_efecto"]:
            fields["fecha_efecto"] = start_date
        if end_date and not fields["fecha_vencimiento"]:
            fields["fecha_vencimiento"] = end_date
    if not fields["tomador"]:
        fields["tomador"] = line_pick(["Tomador", "Asegurado", "Titular", "Contratante"])
    if fields["tomador"]:
        fields["tomador"] = clean_tomador_value(fields["tomador"])
    if not fields["dni"]:
        fields["dni"] = line_pick(["DNI", "NIF", "CIF", "Documento"])
    if fields.get("dni"):
        normalized_dni = normalize_nif_candidate(fields["dni"])
        fields["dni"] = normalized_dni or ""
    if not fields.get("dni"):
        personal_ids = re.findall(r"\b(?:[0-9]{8}[A-Z]|[XYZ][0-9]{7}[A-Z])\b", text.upper())
        if personal_ids:
            fields["dni"] = personal_ids[0]
    if not fields["compania"]:
        fields["compania"] = line_pick(["Compañia", "Compania", "Aseguradora", "Entidad aseguradora"])
    if fields["compania"]:
        fields["compania"] = normalize_company_name(fields["compania"])
    if not fields["poliza_numero"]:
        fields["poliza_numero"] = line_pick(
            ["Póliza", "Poliza", "Nº póliza", "Numero de poliza", "Certificado", "Contrato"]
        )
    if fields.get("poliza_numero"):
        gag_inline = re.search(r"\bGAG[0-9]{5,}\b", str(fields["poliza_numero"]), re.IGNORECASE)
        if gag_inline:
            fields["poliza_numero"] = gag_inline.group(0).upper()
    if not fields["ramo"]:
        fields["ramo"] = line_pick(["Ramo", "Modalidad", "Producto"])
    if not fields["fecha_efecto"]:
        fields["fecha_efecto"] = line_pick(["Fecha efecto", "Efecto", "Inicio vigencia", "Fecha inicio"])
    if not fields["fecha_vencimiento"]:
        fields["fecha_vencimiento"] = line_pick(["Vencimiento", "Fin vigencia", "Vigencia hasta"])
    if not fields["prima_total"]:
        fields["prima_total"] = line_pick(["Prima total", "Total recibo", "Total"])
    if not fields["prima_neta"]:
        fields["prima_neta"] = line_pick(["Prima neta", "Neta"])
    if not fields["direccion"]:
        fields["direccion"] = line_pick(["Direccion", "Domicilio"])
    if fields["direccion"]:
        fields["direccion"] = re.sub(r"\s{2,}.*$", "", fields["direccion"]).strip()
        dir_upper2 = normalize_lookup_text(fields["direccion"])
        if any(
            token in dir_upper2
            for token in (
                "DOMICILIO SOCIAL",
                "CARRETERA DE POZUELO",
                "MAJADAHONDA MADRID",
                "DATOS DE TU MEDIADOR",
            )
        ):
            fields["direccion"] = ""
    if not fields["telefono"]:
        fields["telefono"] = line_pick(["Telefono", "Teléfono", "Movil", "Móvil", "Tfno", "Tlf"])
    if fields["telefono"]:
        fields["telefono"] = normalize_phone(fields["telefono"])
    if not fields["email"]:
        fields["email"] = line_pick(["Email", "Correo", "Correo electronico", "Correo electrónico"])
    if fields["email"]:
        fields["email"] = normalize_email(fields["email"])
    if fields["email"] and normalize_lookup_text(fields["email"]) in ("OCASO OCASO ES",):
        preferred = pick([r"Medio\/s\s+de\s+Contacto\s*:\s*([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})"])
        fields["email"] = normalize_email(preferred) if preferred else ""
    money_token = re.compile(r"^\d{1,3}(?:\.\d{3})*,\d{2}$|^\d+,\d{2}$|^\d+(?:\.\d+)?$")
    if fields["prima_neta"] and not money_token.match(str(fields["prima_neta"]).strip()):
        fallback = re.search(
            r"Prima\s+neta.*?Total\s+Recibo.*?\bAnual\b[^\d]*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        fields["prima_neta"] = fallback.group(1) if fallback else ""
    if fields["prima_total"] and not money_token.match(str(fields["prima_total"]).strip()):
        fallback = re.search(
            r"Prima\s+neta.*?Total\s+Recibo.*?\bAnual\b.*?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        fields["prima_total"] = fallback.group(1) if fallback else ""
    anual_row = re.search(r"\bAnual\b(.{0,4000})", text, re.IGNORECASE | re.DOTALL)
    if anual_row:
        amounts = re.findall(r"([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€", anual_row.group(1))
        if amounts:
            if (not fields.get("prima_neta")) or (not money_token.match(str(fields.get("prima_neta")).strip())):
                fields["prima_neta"] = amounts[0]
            if (not fields.get("prima_total")) or (not money_token.match(str(fields.get("prima_total")).strip())):
                fields["prima_total"] = amounts[-1]
            if fields.get("prima_total") == fields.get("prima_neta") and len(amounts) >= 2:
                fields["prima_total"] = amounts[-1]
    if not fields.get("tomador") or len((fields.get("tomador") or "").split()) <= 1:
        better_tomador = company_specific_tomador(text)
        if better_tomador:
            fields["tomador"] = better_tomador
    if fields.get("tomador"):
        fields["tomador"] = clean_tomador_value(fields["tomador"])
    if not fields.get("poliza_numero") or len(normalize_poliza_key(fields.get("poliza_numero"))) < 6:
        better_poliza = company_specific_poliza(fields.get("compania"), text)
        if not better_poliza and source_hint:
            better_poliza = company_specific_poliza(fields.get("compania"), source_hint)
        if not better_poliza and source_fields.get("poliza_numero"):
            better_poliza = source_fields.get("poliza_numero")
        if better_poliza:
            fields["poliza_numero"] = better_poliza
    if (not fields.get("tomador") or len((fields.get("tomador") or "").split()) <= 1) and source_fields.get("tomador"):
        fields["tomador"] = source_fields.get("tomador")
    if (not fields.get("ramo")) and source_fields.get("ramo"):
        fields["ramo"] = source_fields.get("ramo")
    if not fields.get("ramo"):
        mapfre_auto = re.search(r"AUTOM[ÓO]VILES?\s+PARTICULARES", text, re.IGNORECASE)
        if mapfre_auto:
            fields["ramo"] = "Auto"
    if not fields.get("ramo"):
        better_ramo = company_specific_ramo(fields.get("compania"), text)
        if not better_ramo and source_hint:
            better_ramo = company_specific_ramo(fields.get("compania"), source_hint)
        if better_ramo:
            fields["ramo"] = better_ramo
    if fields.get("compania"):
        # Prefer company from filename when OCR text gives a conflicting generic label.
        source_company = source_fields.get("compania") or hinted_company or detect_company_from_metadata(source_hint)
        if source_company:
            src_key = normalize_company_key(source_company)
            cur_key = normalize_company_key(fields.get("compania"))
            if src_key and cur_key and src_key != cur_key:
                fields["compania"] = normalize_company_name(source_company)

    # Reglas específicas Reale Oficinas / Comercio.
    is_reale = normalize_company_key(fields.get("compania") or "") == "REALE"
    upper_text = normalize_lookup_text(text)
    if is_reale and ("REALE OFICINAS" in upper_text or "DESCRIPCION DEL RIESGO COMERCIO U OFICINA" in upper_text):
        fields["ramo"] = "Comercio"
        tomador_block_match = re.search(
            r"TOMADOR([\s\S]{0,1200}?)EFECTO DEL SEGURO",
            text,
            re.IGNORECASE,
        )
        tomador_block = tomador_block_match.group(1) if tomador_block_match else ""
        if tomador_block:
            candidate_name = ""
            for raw_line in tomador_block.splitlines():
                line = re.sub(r"\s+", " ", raw_line).strip(" ,;:-")
                if not line:
                    continue
                line_up = normalize_lookup_text(line)
                if line_up in ("ASEGURADO", "TOMADOR"):
                    continue
                if "CIF" in line_up or "NIF" in line_up:
                    continue
                if re.search(r"^\d{5}\b", line):
                    continue
                if "MALAGA" in line_up and len(line.split()) <= 2:
                    continue
                if len(line) >= 4:
                    candidate_name = line
                    break
            candidate_name = normalize_person_name(candidate_name).strip(" ,;:-")
            if candidate_name:
                fields["tomador"] = candidate_name
            cif_match = re.search(r"\b([ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z])\b", tomador_block, re.IGNORECASE)
            if cif_match:
                fields["dni"] = cif_match.group(1).upper()
                fields["nif"] = fields["dni"]
            addr_line = ""
            cp_city = ""
            for raw_line in tomador_block.splitlines():
                line = re.sub(r"\s+", " ", raw_line).strip(" ,;:-")
                if not line:
                    continue
                if re.search(r"\b(CL|CALLE|AVD|AVDA|AVENIDA|PLAZA|PZA|CTRA|CAMINO)\b", normalize_lookup_text(line)):
                    addr_line = line
                if re.match(r"^\d{5}\b", line):
                    cp_city = line
                    break
            if addr_line:
                fields["direccion"] = f"{addr_line} {cp_city}".strip() if cp_city else addr_line
            block_mail = re.search(r"\b([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})\b", tomador_block, re.IGNORECASE)
            if block_mail:
                fields["email"] = normalize_email(block_mail.group(1))
        primas_block_match = re.search(
            r"IMPORTE DEL RECIBO Y M[ÉE]TODO DE PAGO([\s\S]{0,900}?)DESCRIPCI[ÓO]N DEL RIESGO",
            text,
            re.IGNORECASE,
        )
        primas_block = primas_block_match.group(1) if primas_block_match else ""
        if primas_block:
            prima_match = re.search(r"\b([0-9]{1,3},[0-9]{2})\b", primas_block)
            if prima_match:
                fields["prima_neta"] = prima_match.group(1)
            total_match = re.search(r"Total\s*[\r\n]+\s*([0-9]{1,3},[0-9]{2})\s*€", primas_block, re.IGNORECASE)
            if total_match:
                fields["prima_total"] = total_match.group(1)
            elif "prima_total" not in fields or not fields.get("prima_total"):
                amounts = re.findall(r"\b([0-9]{1,3},[0-9]{2})\b", primas_block)
                if amounts:
                    fields["prima_total"] = max(amounts, key=parse_money_value)
        fields["telefono"] = ""

    # Reglas especificas AXA Profesional (evita capturar datos del mediador).
    axa_company_key = normalize_company_key(fields.get("compania") or "")
    is_axa = axa_company_key == "AXA" or "AXA" in axa_company_key
    if is_axa:
        if "AXA PROFESIONAL" in upper_text or "AXA PROF" in upper_text or "POLIZA DE SEGURO DE PROFESIONAL" in upper_text:
            fields["ramo"] = "Profesional"
        axa_poliza = re.search(r"p[oó]liza\s*(?:n[º°o]|no\.?)\s*([0-9]{2})[- ]?([0-9]{7,})", text, re.IGNORECASE)
        if axa_poliza:
            fields["poliza_numero"] = f"{axa_poliza.group(1)}-{axa_poliza.group(2)}"
        axa_tomador_info = re.search(
            r"Tomador\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{6,}?)\s+Producto",
            text,
            re.IGNORECASE,
        )
        if axa_tomador_info:
            fields["tomador"] = clean_tomador_value(axa_tomador_info.group(1))
        if not fields.get("tomador") or len((fields.get("tomador") or "").split()) < 3:
            axa_asegurado_name = re.search(
                r"Nombre\s+del\s+asegurado\s*[\r\n]+\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{8,})",
                text,
                re.IGNORECASE,
            )
            if axa_asegurado_name:
                fields["tomador"] = clean_tomador_value(axa_asegurado_name.group(1))
        axa_policy_block = re.search(
            r"Datos\s+de\s+la\s+P[oó]liza([\s\S]{0,900}?)Asegurado",
            text,
            re.IGNORECASE,
        )
        axa_policy_scope = axa_policy_block.group(1) if axa_policy_block else text
        axa_policy_dates = re.search(
            r"Fecha\s+efecto[\s\S]{0,120}?Fecha\s+vencimiento[\s\S]{0,160}?([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})[\s\S]{0,80}?([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})",
            axa_policy_scope,
            re.IGNORECASE,
        )
        if axa_policy_dates:
            fields["fecha_efecto"] = normalize_ocr_date(axa_policy_dates.group(1))
            fields["fecha_vencimiento"] = normalize_ocr_date(axa_policy_dates.group(2))
        if (fields.get("fecha_efecto") or "").endswith("/20") or (fields.get("fecha_vencimiento") or "").endswith("/20"):
            axa_policy_dates_global = re.search(
                r"Fecha\s+efecto[\s\S]{0,220}?([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})[\s\S]{0,220}?Fecha\s+vencimiento[\s\S]{0,220}?([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})",
                text,
                re.IGNORECASE,
            )
            if axa_policy_dates_global:
                fields["fecha_efecto"] = normalize_ocr_date(axa_policy_dates_global.group(1))
                fields["fecha_vencimiento"] = normalize_ocr_date(axa_policy_dates_global.group(2))
        axa_effect = re.search(
            r"Fecha\s+efecto[^\d]*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})",
            axa_policy_scope,
            re.IGNORECASE,
        )
        if axa_effect:
            fields["fecha_efecto"] = normalize_ocr_date(axa_effect.group(1))
        axa_due = re.search(
            r"Fecha\s+vencimiento[^\d]*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{4})",
            axa_policy_scope,
            re.IGNORECASE,
        )
        if axa_due:
            fields["fecha_vencimiento"] = normalize_ocr_date(axa_due.group(1))

        axa_tomador_block_match = re.search(
            r"Datos\s+del\s+Tomador([\s\S]{0,2200}?)Datos\s+de\s+la\s+P[oó]liza",
            text,
            re.IGNORECASE,
        )
        axa_tomador_block = axa_tomador_block_match.group(1) if axa_tomador_block_match else ""
        if axa_tomador_block:
            axa_tomador_name = re.search(
                r"Tomador\s+del\s+seguro[\s\S]{0,140}?([A-ZÁÉÍÓÚÑ]{2,}(?:\s+[A-ZÁÉÍÓÚÑ]{2,}){1,5})",
                axa_tomador_block,
                re.IGNORECASE,
            )
            if axa_tomador_name:
                candidate_name = clean_tomador_value(axa_tomador_name.group(1))
                if "SEGUROS Y REASEGUROS" not in normalize_lookup_text(candidate_name):
                    fields["tomador"] = candidate_name
            axa_tomador_nif = re.search(
                r"\b([0-9]{8}[A-Z]|[XYZ][0-9]{7}[A-Z])\b",
                axa_tomador_block,
                re.IGNORECASE,
            )
            if axa_tomador_nif:
                nif_val = normalize_nif_candidate(axa_tomador_nif.group(1))
                if nif_val:
                    fields["dni"] = nif_val
                    fields["nif"] = nif_val
            axa_tomador_phone = re.search(
                r"Tel[eé]fono\s+m[oó]vil\s*[\r\n]+\s*([0-9 ]{9,})",
                axa_tomador_block,
                re.IGNORECASE,
            )
            if axa_tomador_phone:
                fields["telefono"] = normalize_phone(axa_tomador_phone.group(1))
            axa_tomador_dir = re.search(
                r"Direcci[oó]n\s*[\r\n]+\s*([^\n]+)\n\s*(\d{5}\s+[A-ZÁÉÍÓÚÑ][^\n]*)\n\s*Tel[eé]fono\s+m[oó]vil",
                axa_tomador_block,
                re.IGNORECASE,
            )
            if axa_tomador_dir:
                fields["direccion"] = f"{axa_tomador_dir.group(1).strip()} {axa_tomador_dir.group(2).strip()}".strip()
        if not fields.get("direccion") or "MEDIADOR" in normalize_lookup_text(fields.get("direccion") or ""):
            axa_addr_fallback = re.search(
                r"\b(CL\s+[^\n]{5,140}?\d{5}\s+[A-ZÁÉÍÓÚÑ]{3,})",
                text,
                re.IGNORECASE,
            )
            if axa_addr_fallback:
                candidate_addr = re.sub(r"\s+", " ", axa_addr_fallback.group(1)).strip()
                if "ILDEFONSO" not in normalize_lookup_text(candidate_addr) and "MEDIADOR" not in normalize_lookup_text(candidate_addr):
                    fields["direccion"] = candidate_addr

        # En AXA el email suele ser del mediador, no del tomador.
        if fields.get("email") and "fincasvelazquez" in str(fields.get("email") or "").lower():
            fields["email"] = ""
        if normalize_phone(fields.get("telefono") or "") == "900909014":
            fields["telefono"] = ""
        if "MEDIADOR" in normalize_lookup_text(fields.get("tomador") or ""):
            fields["tomador"] = ""
        if "MEDIADOR" in normalize_lookup_text(fields.get("direccion") or ""):
            fields["direccion"] = ""

    # Reglas específicas MAPFRE RC (evita texto legal y prioriza bloque de tomador real).
    mapfre_company_key = normalize_company_key(fields.get("compania") or "")
    is_mapfre = mapfre_company_key == "MAPFRE" or "MAPFRE" in mapfre_company_key
    if is_mapfre and "DATOS DEL TOMADOR Y ASEGURADO" in upper_text:
        mapfre_block_match = re.search(
            r"DATOS\s+DEL\s+TOMADOR\s+Y\s+ASEGURADO([\s\S]{0,2200}?)PRIMA\s+DEL\s+SEGURO",
            text,
            re.IGNORECASE,
        )
        mapfre_block = mapfre_block_match.group(1) if mapfre_block_match else text
        mapfre_name = re.search(
            r"Nombre\s+([^\n]+?)\s+Documento\s+ID\s+([A-Z0-9]{8,12})",
            mapfre_block,
            re.IGNORECASE,
        )
        if mapfre_name:
            name_main = re.sub(r"\s+", " ", mapfre_name.group(1)).strip(" ,;:-")
            name_extra_match = re.search(
                r"Documento\s+ID\s+[A-Z0-9]{8,12}\s*\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]{3,})\s*\n\s*Direcci[oó]n",
                mapfre_block,
                re.IGNORECASE,
            )
            if name_extra_match:
                name_extra = re.sub(r"\s+", " ", name_extra_match.group(1)).strip(" ,;:-")
                if name_extra and normalize_lookup_text(name_extra) not in ("LOCALIDAD", "PROVINCIA"):
                    name_main = f"{name_main} {name_extra}".strip()
            if name_main:
                fields["tomador"] = clean_tomador_value(name_main)
            mapfre_doc = normalize_nif_candidate(mapfre_name.group(2))
            if mapfre_doc:
                fields["dni"] = mapfre_doc
                fields["nif"] = mapfre_doc
        mapfre_mobile = re.search(r"Tel[eé]fono\s+m[oó]vil\s*([+0-9 ]{9,})", mapfre_block, re.IGNORECASE)
        if mapfre_mobile:
            fields["telefono"] = normalize_phone(mapfre_mobile.group(1))
        mapfre_addr = re.search(
            r"Direcci[oó]n\s+([\s\S]{0,140}?)\s+C[oó]digo\s+postal\s+(\d{5})",
            mapfre_block,
            re.IGNORECASE,
        )
        if mapfre_addr:
            addr_main = re.sub(r"\s+", " ", mapfre_addr.group(1)).strip(" ,;:-")
            addr_main = re.sub(r"Tel[eé]fono\s+fijo.*$", "", addr_main, flags=re.IGNORECASE).strip(" ,;:-")
            addr_main = re.sub(r"Tel[eé]fono\s+m[oó]vil.*$", "", addr_main, flags=re.IGNORECASE).strip(" ,;:-")
            cp = mapfre_addr.group(2)
            loc = ""
            prov = ""
            loc_m = re.search(r"Localidad\s+([A-ZÁÉÍÓÚÑ ]{3,})", mapfre_block, re.IGNORECASE)
            if loc_m:
                loc = re.sub(r"\s+", " ", loc_m.group(1)).strip()
            prov_m = re.search(r"Provincia\s+([A-ZÁÉÍÓÚÑ ]{3,})", mapfre_block, re.IGNORECASE)
            if prov_m:
                prov = re.sub(r"\s+", " ", prov_m.group(1)).strip()
            fields["direccion"] = " ".join([x for x in (addr_main, cp, loc, prov) if x]).strip()
        if normalize_phone(fields.get("telefono") or "") == "918365365":
            fields["telefono"] = ""
        if fields.get("tomador") and "A LA FECHA DE RENOVACION" in normalize_lookup_text(fields["tomador"]):
            fields["tomador"] = ""
        if fields.get("email"):
            fields["email"] = ""

    # Reglas específicas Zurich Accidentes (prioriza bloque "Datos del tomador").
    zurich_company_key = normalize_company_key(fields.get("compania") or "")
    is_zurich = zurich_company_key == "ZURICH" or "ZURICH" in zurich_company_key
    if is_zurich and ("SEGURO DE ACCIDENTES" in upper_text or "ZURICH ACCIDENTES" in upper_text):
        if "ACCIDENTE" in upper_text:
            fields["ramo"] = "Accidentes"
        zurich_block_match = re.search(
            r"Datos\s+del\s+tomador([\s\S]{0,2200}?)(?:Garant[ií]as|Coberturas|Titular/Entidad|$)",
            text,
            re.IGNORECASE,
        )
        zurich_block = zurich_block_match.group(1) if zurich_block_match else text
        z_name = re.search(r"Nombre\s+o\s+raz[oó]n\s+social\s*:\s*([^\n]+)", zurich_block, re.IGNORECASE)
        if z_name:
            name = clean_tomador_value(z_name.group(1))
            if name and len(name.split()) >= 2:
                fields["tomador"] = name
        z_nif = re.search(r"NIF/CIF\s*:\s*([A-Z0-9.\- ]{8,16})", zurich_block, re.IGNORECASE)
        if z_nif:
            nif_clean = normalize_nif_candidate(z_nif.group(1))
            if nif_clean:
                fields["dni"] = nif_clean
                fields["nif"] = nif_clean
        z_addr = re.search(r"Direcci[oó]n\s*:\s*([^\n]+)", zurich_block, re.IGNORECASE)
        if z_addr:
            addr = re.sub(r"\s+", " ", z_addr.group(1)).strip(" ,;:-")
            if addr and addr != "-":
                fields["direccion"] = addr
        z_phone = re.search(r"Tel[eé]fono\s*:\s*([^\n]+)", zurich_block, re.IGNORECASE)
        if z_phone:
            phone = normalize_phone(z_phone.group(1))
            fields["telefono"] = phone
        z_vig = re.search(
            r"Vigencia\s*:\s*desde[\s\S]{0,120}?(\d{1,2}/\d{1,2}/\d{4})[\s\S]{0,120}?hasta[\s\S]{0,120}?(\d{1,2}/\d{1,2}/\d{4})",
            zurich_block,
            re.IGNORECASE,
        )
        if z_vig:
            fields["fecha_efecto"] = normalize_ocr_date(z_vig.group(1))
            fields["fecha_vencimiento"] = normalize_ocr_date(z_vig.group(2))
        # Correo/teléfono de siniestros o mediador no deben quedar como contacto del tomador.
        mail_norm = normalize_email(fields.get("email") or "")
        if mail_norm in ("aperturas@zurich.com", "miguelangelperez@grupomodernia.es"):
            fields["email"] = ""
        phone_norm = normalize_phone(fields.get("telefono") or "")
        if phone_norm in ("913755755", "934165046", "951394365"):
            fields["telefono"] = ""

    # Reglas específicas Allianz RC PYME (certificado de seguro).
    allianz_company_key = normalize_company_key(fields.get("compania") or "")
    is_allianz = allianz_company_key == "ALLIANZ" or "ALLIANZ" in allianz_company_key
    if is_allianz and ("ALLIANZ R.C.PYME" in upper_text or "CERTIFICADO DE SEGURO" in upper_text):
        if "RESPONSABILIDAD CIVIL" in upper_text or "R C PYME" in upper_text or "R.C.PYME" in upper_text:
            fields["ramo"] = "Responsabilidad civil"
        cover_name = re.search(
            r"\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s,.'\-]{8,})\s*\n\s*P[oó]liza\s+[0-9]{6,}",
            text,
            re.IGNORECASE,
        )
        if cover_name:
            candidate_cover = clean_tomador_value(cover_name.group(1))
            if candidate_cover and len(candidate_cover.split()) >= 2:
                fields["tomador"] = candidate_cover
        allianz_block_match = re.search(
            r"Datos\s+Generales([\s\S]{0,2200}?)Datos\s+del\s+Asegurado",
            text,
            re.IGNORECASE,
        )
        allianz_block = allianz_block_match.group(1) if allianz_block_match else text
        allianz_lines = [ln.strip() for ln in allianz_block.splitlines() if ln.strip()]
        for idx, ln in enumerate(allianz_lines):
            if re.search(r"Tomador\s+del\s+Seguro\s*:", ln, re.IGNORECASE):
                if idx > 0:
                    prev = clean_tomador_value(allianz_lines[idx - 1])
                    if prev and len(prev.split()) >= 2 and "POLIZA" not in normalize_lookup_text(prev):
                        fields["tomador"] = prev
                inline = re.split(r":", ln, maxsplit=1)
                if len(inline) > 1:
                    maybe_addr = inline[1].strip()
                    if maybe_addr and not fields.get("direccion"):
                        fields["direccion"] = maybe_addr
                break
        a_name_header = re.search(
            r"Datos\s+Generales\s*\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s,.'\-]{8,})\s*\n\s*Tomador\s+del\s+Seguro",
            text,
            re.IGNORECASE,
        )
        if a_name_header:
            name_header = clean_tomador_value(a_name_header.group(1))
            if name_header and len(name_header.split()) >= 2:
                fields["tomador"] = name_header
        if not fields.get("tomador"):
            a_name_prev = re.search(
                r"\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s,.'\-]{6,})\s*\n\s*Tomador\s+del\s+Seguro\s*:",
                allianz_block,
                re.IGNORECASE,
            )
            if a_name_prev:
                name_prev = clean_tomador_value(a_name_prev.group(1))
                if name_prev and len(name_prev.split()) >= 2:
                    fields["tomador"] = name_prev
        a_tom = re.search(
            r"Tomador\s+del\s+Seguro\s*:\s*([\s\S]{0,240}?)NIF\s*:\s*([A-Z0-9.\-]{8,16})",
            allianz_block,
            re.IGNORECASE,
        )
        if a_tom:
            tom_raw = re.sub(r"\s+", " ", a_tom.group(1)).strip(" ,;:-")
            if tom_raw:
                fields["tomador"] = clean_tomador_value(tom_raw)
            nif_val = normalize_nif_candidate(a_tom.group(2))
            if nif_val:
                fields["dni"] = nif_val
                fields["nif"] = nif_val
        a_pol = re.search(r"P[oó]liza\s*n[º°o]\s*:\s*([0-9]{6,12})", allianz_block, re.IGNORECASE)
        if a_pol:
            fields["poliza_numero"] = a_pol.group(1).strip()
        a_dur = re.search(
            r"Duraci[oó]n\s*:\s*Desde[\s\S]{0,100}?(\d{1,2}/\d{1,2}/\d{4})[\s\S]{0,120}?hasta[\s\S]{0,80}?(\d{1,2}/\d{1,2}/\d{4})",
            allianz_block,
            re.IGNORECASE,
        )
        if a_dur:
            fields["fecha_efecto"] = normalize_ocr_date(a_dur.group(1))
            fields["fecha_vencimiento"] = normalize_ocr_date(a_dur.group(2))
        a_dur_global = re.search(
            r"Duraci[oó]n\s*:\s*Desde[\s\S]{0,220}?del\s+(\d{1,2}/\d{1,2}/\d{4})[\s\S]{0,240}?hasta[\s\S]{0,220}?del\s+(\d{1,2}/\d{1,2}/\d{4})",
            text,
            re.IGNORECASE,
        )
        if a_dur_global:
            fields["fecha_efecto"] = normalize_ocr_date(a_dur_global.group(1))
            fields["fecha_vencimiento"] = normalize_ocr_date(a_dur_global.group(2))
        # Evitar contacto del mediador en ficha de cliente.
        if fields.get("email") and (
            "grupomodernia.es" in str(fields.get("email") or "").lower()
            or "allianz" in str(fields.get("email") or "").lower()
        ):
            fields["email"] = ""
        phone_norm = normalize_phone(fields.get("telefono") or "")
        if phone_norm in ("951394365", "651075059", "900300250"):
            fields["telefono"] = ""

    # Reglas específicas DKV Salud (solicitud de seguro).
    dkv_company_key = normalize_company_key(fields.get("compania") or "")
    is_dkv = dkv_company_key == "DKV" or "DKV" in dkv_company_key
    if is_dkv and ("SOLICITUD DE SEGURO" in upper_text or "DKV INTEGRAL" in upper_text):
        fields["ramo"] = "Salud"
        dkv_block_match = re.search(
            r"Tomador\s+del\s+seguro([\s\S]{0,1200}?)Datos\s+bancarios",
            text,
            re.IGNORECASE,
        )
        dkv_block = dkv_block_match.group(1) if dkv_block_match else text
        dkv_row = re.search(
            r"\n\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s.'\-]{2,})\s{2,}([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s.'\-]{2,})\s{2,}([A-Z0-9.\-]{8,12})\s{2,}(\d{1,2}/\d{1,2}/\d{4})",
            dkv_block,
            re.IGNORECASE,
        )
        if dkv_row:
            apellidos = re.sub(r"\s+", " ", dkv_row.group(1)).strip(" ,;:-")
            nombre = re.sub(r"\s+", " ", dkv_row.group(2)).strip(" ,;:-")
            if nombre and apellidos:
                fields["tomador"] = f"{nombre} {apellidos}".strip()
            dkv_nif = normalize_nif_candidate(dkv_row.group(3))
            if dkv_nif:
                fields["dni"] = dkv_nif
                fields["nif"] = dkv_nif
            if not fields.get("fecha_nacimiento"):
                fields["fecha_nacimiento"] = normalize_ocr_date(dkv_row.group(4))
        dkv_addr = re.search(
            r"Domicilio\s+C\.?P\.?\s+Localidad\s+Provincia[\s\S]{0,180}?\n\s*([^\n]+?)\s+(\d{4,5})\s+([A-ZÁÉÍÓÚÑa-zñ ]{3,})\s+([A-ZÁÉÍÓÚÑa-zñ ]{3,})",
            dkv_block,
            re.IGNORECASE,
        )
        if dkv_addr:
            calle = re.sub(r"\s+", " ", dkv_addr.group(1)).strip(" ,;:-")
            cp = dkv_addr.group(2).strip()
            loc = re.sub(r"\s+", " ", dkv_addr.group(3)).strip(" ,;:-")
            prov = re.sub(r"\s+", " ", dkv_addr.group(4)).strip(" ,;:-")
            fields["direccion"] = " ".join([x for x in (calle, cp, loc, prov) if x]).strip()
        dkv_mobile = re.search(r"\b([67][0-9]{8})\b", dkv_block)
        if dkv_mobile:
            fields["telefono"] = normalize_phone(dkv_mobile.group(1))
        dkv_email = re.search(r"\b([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})\b", dkv_block, re.IGNORECASE)
        if dkv_email:
            fields["email"] = normalize_email(dkv_email.group(1))
        if not fields.get("poliza_numero"):
            dkv_doc = re.search(
                r"Sucursal[\s\S]{0,200}?Documento[\s\S]{0,220}?\b([0-9]{10,15})\b",
                text,
                re.IGNORECASE,
            )
            if dkv_doc:
                fields["poliza_numero"] = dkv_doc.group(1)
        if not fields.get("poliza_numero"):
            dkv_doc2 = re.search(r"\bBanco\s+([0-9]{10,15})\b", text, re.IGNORECASE)
            if dkv_doc2:
                fields["poliza_numero"] = dkv_doc2.group(1)
        dkv_eff = re.search(r"Fec\.\s*efecto[\s\S]{0,120}?(\d{1,2}/\d{1,2}/\d{4})", text, re.IGNORECASE)
        if dkv_eff:
            fields["fecha_efecto"] = normalize_ocr_date(dkv_eff.group(1))
            if not fields.get("fecha_vencimiento"):
                fields["fecha_vencimiento"] = add_year_to_date(fields["fecha_efecto"])
        if fields.get("fecha_efecto") and fields.get("fecha_vencimiento") and fields["fecha_efecto"] == fields["fecha_vencimiento"]:
            fields["fecha_vencimiento"] = add_year_to_date(fields["fecha_efecto"])

    # Reglas específicas Euroins Auto (evita capturar datos del mediador / DAS).
    source_upper = normalize_lookup_text(str(source_hint or ""))
    is_euroins_doc = ("EUROINS" in upper_text) or ("EUROINS" in source_upper)
    if is_euroins_doc:
        fields["compania"] = "Euroins"
        euroins_block_match = re.search(
            r"COMPAÑ[IÍ]A([\s\S]{0,1800}?)VEHICULO ASEGURADO",
            text,
            re.IGNORECASE,
        )
        euroins_block = euroins_block_match.group(1) if euroins_block_match else text
        if "SEGURO DE AUTOMOVIL" in upper_text:
            fields["ramo"] = "Auto"
        euroins_tomador = re.search(r"Nombre y apellidos:\s*([^\n]+)", euroins_block, re.IGNORECASE)
        if euroins_tomador:
            cand = clean_tomador_value(euroins_tomador.group(1))
            if cand:
                fields["tomador"] = cand
        euroins_nif = re.search(r"\b([0-9]{8}[A-Z]|[XYZ][0-9]{7}[A-Z])\b", euroins_block, re.IGNORECASE)
        if euroins_nif:
            fields["dni"] = normalize_nif_candidate(euroins_nif.group(1)) or ""
        euroins_phone_candidates = re.findall(r"\b([679][0-9]{8})\b", euroins_block)
        if euroins_phone_candidates:
            fields["telefono"] = normalize_phone(euroins_phone_candidates[-1]) or ""
        periodo_match = re.search(
            r"PERIODO DE VIGENCIA([\s\S]{0,300}?)SEGURO DE AUTOM[ÓO]VIL",
            text,
            re.IGNORECASE,
        )
        if periodo_match:
            periodo_dates = re.findall(r"\d{1,2}/\d{1,2}/\d{2,4}", periodo_match.group(1))
            if len(periodo_dates) >= 2:
                fields["fecha_efecto"] = periodo_dates[0]
                fields["fecha_vencimiento"] = periodo_dates[1]
        primas_window_match = re.search(r"PAGO DE PRIMAS([\s\S]{0,700}?)PERIODO DE VIGENCIA", text, re.IGNORECASE)
        primas_window = primas_window_match.group(1) if primas_window_match else ""
        primas_values = re.findall(r"\b([0-9]+[.,][0-9]{2})\b", primas_window)
        if primas_values:
            fields["prima_neta"] = primas_values[0]
            fields["prima_total"] = primas_values[-1]
        mail_norm = normalize_email(fields.get("email") or "")
        if mail_norm.endswith("@das.es") or "gallen" in mail_norm:
            fields["email"] = ""
        phone_norm = normalize_phone(fields.get("telefono") or "")
        if phone_norm in ("910609386",):
            fields["telefono"] = ""
    for key in ("fecha_efecto", "fecha_vencimiento", "fecha_nacimiento"):
        if fields.get(key):
            fields[key] = normalize_ocr_date(fields[key])
    if fields.get("poliza_numero"):
        fields["poliza_numero"] = normalize_poliza_number(fields["poliza_numero"], fields.get("compania"))
        if not is_valid_poliza_candidate(fields.get("poliza_numero")):
            fields["poliza_numero"] = ""
    if fields["dni"] and not fields.get("nif"):
        fields["nif"] = fields["dni"]
    if fields.get("nif"):
        fields["nif"] = normalize_nif_candidate(fields["nif"]) or fields["nif"]

    # Final ramo cleanup after all fallback assignments.
    if fields.get("ramo"):
        ramo = str(fields["ramo"]).splitlines()[0].strip()
        if "€" in ramo or re.search(r"\d{2,}", ramo):
            ramo = ""
        ramo_upper = normalize_lookup_text(ramo)
        if (
            len(ramo) > 48
            or "informacion general previa" in ramo_upper
            or "distribuidor de seguros" in ramo_upper
            or "ley de ordenacion" in ramo_upper
            or "aseguradora" in ramo_upper
        ):
            ramo = ""
        if "HOGAR" in ramo_upper:
            ramo = "Hogar"
        elif "AUTOMOVIL" in ramo_upper or "AUTOMOVILES" in ramo_upper or "AUTO" in ramo_upper:
            ramo = "Auto"
        elif "RESPONSABILIDAD CIVIL" in ramo_upper or ramo_upper == "RC":
            ramo = "Responsabilidad civil"
        fields["ramo"] = ramo
    if (not fields.get("ramo")) and source_fields.get("ramo"):
        fields["ramo"] = source_fields.get("ramo")
    if source_fields.get("ramo"):
        source_ramo = str(source_fields.get("ramo") or "").strip()
        current_ramo = str(fields.get("ramo") or "").strip()
        if source_ramo and (
            not current_ramo
            or current_ramo in ("Alquiler", "Hogar alquiler")
            or (
                source_ramo == "Responsabilidad civil"
                and current_ramo not in ("Responsabilidad civil",)
            )
        ):
            fields["ramo"] = source_ramo
    # Reglas finales para pólizas de impago (GAG/FIATC/iptiQ) con mucho texto legal.
    source_upper = normalize_lookup_text(str(source_hint or ""))
    current_poliza_norm = normalize_poliza_key(fields.get("poliza_numero") or "")
    if (
        "IMPAGO" in source_upper
        or current_poliza_norm.startswith("GAG")
        or "GAG" in source_upper
    ):
        if not fields.get("ramo") or normalize_lookup_text(fields.get("ramo") or "") in ("DE SEGURO", "SEGURO", "ALQUILER"):
            fields["ramo"] = "Impago alquiler"
        # En impago, a menudo OCR captura el contacto del mediador en lugar del asegurado.
        if normalize_email(fields.get("email") or "") in ("info@fincasvelazquez.es", "fiatc@fiatc.es"):
            fields["email"] = ""
        phone_norm = normalize_phone(fields.get("telefono") or "")
        if phone_norm in ("910609386",):
            fields["telefono"] = ""
    if fields.get("tomador"):
        fields["tomador"] = clean_tomador_value(fields["tomador"])
    # Saneo final para pólizas de impago con texto legal dominante (FIATC/iptiQ).
    if fields.get("email"):
        mail_norm = normalize_lookup_text(fields["email"])
        if (
            ("@" not in fields["email"])
            or mail_norm.endswith("FIATC ES")
            or mail_norm.endswith("IPTIQ COM")
            or "POLIZA Y SE CONSIDERARA ACEPTADA" in mail_norm
        ):
            fields["email"] = ""
    if fields.get("direccion"):
        dir_norm = normalize_lookup_text(fields["direccion"])
        if (
            "DIRECCION GENERAL DE SEGUROS" in dir_norm
            or "DEPENDIENTE DE LA DIRECCION GENERAL" in dir_norm
            or "FONDOS DE PENSIONES" in dir_norm
            or "SOCIAL DE LA ENTIDAD" in dir_norm
            or "AVENIDA DIAGONAL" in dir_norm
            or ("BARCELONA" in dir_norm and "MALAGA" in dir_norm)
            or "TELEFONO FIJO" in dir_norm
            or "TELEFONO MOVIL" in dir_norm
        ):
            fields["direccion"] = ""
    if fields.get("dni"):
        dni_norm = normalize_nif_candidate(fields.get("dni"))
        tomador_norm = normalize_lookup_text(fields.get("tomador") or "")
        looks_company_id = bool(re.match(r"^[ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z]$", dni_norm or ""))
        looks_person_name = bool(
            tomador_norm
            and not looks_corporate_name(fields.get("tomador") or "")
            and len((fields.get("tomador") or "").split()) >= 2
        )
        if looks_company_id and looks_person_name:
            fields["dni"] = ""
            if fields.get("nif") == dni_norm:
                fields["nif"] = ""
    if fields.get("fecha_efecto") and fields.get("fecha_vencimiento"):
        efecto = parse_iso_date(fields.get("fecha_efecto"))
        venc = parse_iso_date(fields.get("fecha_vencimiento"))
        if efecto and venc and venc < efecto:
            fields["fecha_vencimiento"] = add_year_to_date(fields.get("fecha_efecto"))
        elif (
            efecto
            and venc
            and venc == efecto
            and (
                normalize_poliza_key(fields.get("poliza_numero") or "").startswith("GAG")
                or normalize_lookup_text(fields.get("ramo") or "") == "IMPAGO ALQUILER"
            )
        ):
            fields["fecha_vencimiento"] = add_year_to_date(fields.get("fecha_efecto"))
    if fields.get("dni"):
        dni_checked = normalize_nif_candidate(fields.get("dni"))
        if not is_valid_nif(dni_checked):
            fields["dni"] = ""
            if fields.get("nif") and not is_valid_nif(normalize_nif_candidate(fields.get("nif"))):
                fields["nif"] = ""
    if is_euroins_doc:
        # En Euroins, el identificador puede venir como NIF textual aunque no pase el checksum clásico.
        if not fields.get("dni"):
            raw_id = re.search(r"\b([0-9]{8}[A-Z]|[XYZ][0-9]{7}[A-Z])\b", text, re.IGNORECASE)
            if raw_id:
                fields["dni"] = raw_id.group(1).upper()
                if not fields.get("nif"):
                    fields["nif"] = fields["dni"]
        range_dates = re.search(
            r"Desde[^\n]*?(\d{1,2}/\d{1,2}/\d{2,4})[\s\S]{0,120}?Hasta[^\n]*?(\d{1,2}/\d{1,2}/\d{2,4})",
            text,
            re.IGNORECASE,
        )
        if range_dates:
            fields["fecha_efecto"] = normalize_ocr_date(range_dates.group(1))
            fields["fecha_vencimiento"] = normalize_ocr_date(range_dates.group(2))
    if is_reale and ("REALE OFICINAS" in upper_text or "DESCRIPCION DEL RIESGO COMERCIO U OFICINA" in upper_text):
        if not fields.get("dni"):
            reale_block_match = re.search(
                r"TOMADOR([\s\S]{0,1200}?)EFECTO DEL SEGURO",
                text,
                re.IGNORECASE,
            )
            reale_scope = reale_block_match.group(1) if reale_block_match else text
            reale_cif = re.search(r"\b([ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z])\b", reale_scope, re.IGNORECASE)
            if reale_cif:
                fields["dni"] = reale_cif.group(1).upper()
                fields["nif"] = fields["dni"]
    # Reglas específicas FIATC Auto (evitar captura de texto legal en ramo/tomador/fechas).
    is_fiatc_auto = (
        normalize_company_key(fields.get("compania") or "") == "FIATC"
        and ("SEGURO AUTOM" in upper_text or "POLIZA AUTOM" in upper_text)
    )
    if is_fiatc_auto:
        fields["compania"] = "Fiatc"
        fields["ramo"] = "Auto"
        tom_match = re.search(
            r"SUSCRITA\s+ENTRE\s+([\s\S]{0,180}?)\s+Y\s+[\s\S]{0,120}?FIATC",
            text,
            re.IGNORECASE,
        )
        if tom_match:
            tom = normalize_person_name(tom_match.group(1)).strip(" ,;:-")
            if tom:
                fields["tomador"] = tom
        pol_match = re.search(
            r"N[ºo°]\s*P[ÓO]LIZA\s*[:#]?\s*([0-9]{3,5}[-/][0-9]{6,10}[-/][0-9]{1,4})",
            text,
            re.IGNORECASE,
        )
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1)
        if not fields.get("poliza_numero"):
            loose_pol = re.search(r"\b([0-9]{4}[-/][0-9]{7}[-/][0-9]{1,3})\b", text)
            if loose_pol:
                fields["poliza_numero"] = loose_pol.group(1)
        efecto_match = re.search(
            r"FECHA\s+EFECTO\s*[:#]?\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
            text,
            re.IGNORECASE,
        )
        if efecto_match:
            fields["fecha_efecto"] = normalize_ocr_date(efecto_match.group(1))
        # Normalmente no trae "fecha vencimiento" clara en cabecera; derivamos +1 año.
        if fields.get("fecha_efecto"):
            fields["fecha_vencimiento"] = add_year_to_date(fields.get("fecha_efecto"))
        # Evitar contaminación con datos de contacto corporativos FIATC.
        if normalize_email(fields.get("email") or "").endswith("@fiatc.es"):
            fields["email"] = ""
        fields["telefono"] = ""
    # Reglas específicas FIATC Impago Alquiler.
    is_fiatc_impago = (
        normalize_company_key(fields.get("compania") or "") == "FIATC"
        and ("IMPAGO DE ALQUILER" in upper_text or "CONTINGENCIAS / ALQUILERES" in upper_text)
    )
    if is_fiatc_impago:
        fields["compania"] = "Fiatc"
        fields["ramo"] = "Impago alquiler"
        tom_match = re.search(
            r"SUSCRITA\s+ENTRE\s*([\s\S]{0,180}?)\s+Y\s+[\s\S]{0,120}?FIATC",
            text,
            re.IGNORECASE,
        )
        if tom_match:
            tom = normalize_person_name(tom_match.group(1)).strip(" ,;:-")
            if tom:
                fields["tomador"] = tom
        pol_match = re.search(
            r"N.?.?.?\s*P.?.?LIZA\s*([0-9]{4}[-/][0-9]{7}[-/][0-9]{1,3})",
            text,
            re.IGNORECASE,
        )
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1)
        eff_match = re.search(
            r"FECHA\s+EFECTO\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
            text,
            re.IGNORECASE,
        )
        if eff_match:
            fields["fecha_efecto"] = normalize_ocr_date(eff_match.group(1))
        if fields.get("fecha_efecto"):
            fields["fecha_vencimiento"] = add_year_to_date(fields.get("fecha_efecto"))
        nif_match = re.search(r"\bNIF\s*[:\-]?\s*([A-Z0-9]{8,9})\b", text, re.IGNORECASE)
        if nif_match:
            fields["dni"] = normalize_nif_candidate(nif_match.group(1)) or fields.get("dni") or ""
            if fields.get("dni"):
                fields["nif"] = fields["dni"]
        if normalize_email(fields.get("email") or "").endswith("@fiatc.es"):
            fields["email"] = ""
        fields["telefono"] = ""
    # Reglas específicas FIATC Comunidad (Propiedad de edificios / Multirriesgo comunidades).
    is_fiatc_comunidad = (
        normalize_company_key(fields.get("compania") or "") == "FIATC"
        and ("PROPIEDAD DE EDIFICIOS" in upper_text or "MULTIRRIESGO COMUNIDADES" in upper_text)
    )
    if is_fiatc_comunidad:
        fields["compania"] = "Fiatc"
        fields["ramo"] = "Comunidad"
        tom_match = re.search(
            r"SUSCRITA\s+ENTRE\s*([\s\S]{0,180}?)\s+Y\s+[\s\S]{0,120}?FIATC",
            text,
            re.IGNORECASE,
        )
        if tom_match:
            tom = normalize_person_name(tom_match.group(1)).strip(" ,;:-")
            if tom:
                fields["tomador"] = tom
        pol_match = re.search(
            r"N.?.?.?\s*P.?.?LIZA\s*([0-9]{4}[-/][0-9]{7}[-/][0-9]{1,3})",
            text,
            re.IGNORECASE,
        )
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1)
        eff_match = re.search(
            r"FECHA\s+EFECTO\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
            text,
            re.IGNORECASE,
        )
        if eff_match:
            fields["fecha_efecto"] = normalize_ocr_date(eff_match.group(1))
        if fields.get("fecha_efecto"):
            fields["fecha_vencimiento"] = add_year_to_date(fields.get("fecha_efecto"))
        prima_block = re.search(
            r"PRIMA\s+NETA[\s\S]{0,420}?PRIMA\s+TOTAL[\s\S]{0,320}",
            text,
            re.IGNORECASE,
        )
        if prima_block:
            nums = re.findall(r"([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})", prima_block.group(0))
            if nums:
                fields["prima_neta"] = nums[0]
                fields["prima_total"] = nums[-1]
        if not fields.get("prima_total"):
            prima_total = re.search(r"Prima\s+total\s+anual[\s\r\n]+([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})", text, re.IGNORECASE)
            if prima_total:
                fields["prima_total"] = prima_total.group(1)
        if normalize_email(fields.get("email") or "").endswith("@fiatc.es"):
            fields["email"] = ""
        fields["telefono"] = ""
    # Reglas específicas OCCIDENT Auto (evitar contaminación por texto legal/corporativo).
    is_occident_auto = (
        ("OCCIDENT" in upper_text or normalize_company_key(fields.get("compania") or "") in ("OCCIDENT", "CATALANAOCCIDENTE"))
        and ("SEGURO DE AUTOMOVIL" in upper_text or "AUTOMOVIL" in upper_text)
    )
    if is_occident_auto:
        fields["compania"] = "Occident"
        fields["ramo"] = "Auto"
        pol_match = re.search(
            r"N\.\s*[ºo]\s*de\s*p[oó]liza\s*[:#]?\s*([A-Z0-9.\-]{8,})",
            text,
            re.IGNORECASE,
        ) or re.search(
            r"N[ºo°]\s*de\s*p[oó]liza\s*[:#]?\s*([A-Z0-9.\-]{8,})",
            text,
            re.IGNORECASE,
        )
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1).strip()
        tomador_match = re.search(r"Tomador del seguro\s*([\s\S]{0,260})", text, re.IGNORECASE)
        if tomador_match:
            block = tomador_match.group(1)
            lines = [re.sub(r"\s+", " ", ln).strip(" ,;:-") for ln in block.splitlines() if ln.strip()]
            candidate_index = -1
            if lines:
                candidate = ""
                for idx, ln in enumerate(lines):
                    ln_up = normalize_lookup_text(ln)
                    if any(
                        token in ln_up
                        for token in (
                            "OCCIDENT",
                            "REASEGUROS",
                            "DOMICILIO SOCIAL",
                            "MENDEZ ALVARO",
                            "MADRID",
                            "OFICINA EMISORA",
                            "CORREDOR",
                            "ZONA SURESTE",
                            "FINCAS VELAZQUEZ",
                            "TELEFONO",
                            "FECHA DE EFECTO",
                        )
                    ):
                        continue
                    if re.match(r"^\d{5}\b", ln):
                        continue
                    candidate = ln
                    candidate_index = idx
                    break
                if candidate:
                    fields["tomador"] = candidate
            nif_match = re.search(r"\bNIF\s*([A-Z0-9]{8,9})\b", block, re.IGNORECASE)
            if nif_match:
                fields["dni"] = nif_match.group(1).upper()
                fields["nif"] = fields["dni"]
            address_lines = []
            for idx, ln in enumerate(lines):
                if candidate_index >= 0 and idx <= candidate_index:
                    continue
                ln_up = normalize_lookup_text(ln)
                if "OFICINA EMISORA" in ln_up or "CORREDOR" in ln_up:
                    break
                if re.search(r"\b(AVENIDA|CALLE|C/|CL\b|PLAZA|PZA|CTRA|CAMINO)\b", ln_up):
                    address_lines.append(ln)
                    continue
                if re.match(r"^\d{5}\b", ln) or ln_up in ("MALAGA", "MURCIA", "MADRID"):
                    ln_up = normalize_lookup_text(ln)
                    address_lines.append(ln)
            if address_lines:
                fields["direccion"] = " ".join(address_lines[:3]).strip()
        month_map = {
            "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
            "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
            "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
        }
        eff_match = re.search(
            r"toma efecto[^\n]*?d[ií]a\s+([0-9]{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+([0-9]{4})",
            text,
            re.IGNORECASE,
        )
        exp_match = re.search(
            r"Fecha de vencimiento\s*:\s*([0-9]{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+([0-9]{4})",
            text,
            re.IGNORECASE,
        )
        if eff_match:
            d, mtxt, y = eff_match.groups()
            mm = month_map.get(normalize_lookup_text(mtxt).lower(), "")
            if mm:
                fields["fecha_efecto"] = f"{int(d):02d}/{mm}/{y}"
        if exp_match:
            d, mtxt, y = exp_match.groups()
            mm = month_map.get(normalize_lookup_text(mtxt).lower(), "")
            if mm:
                fields["fecha_vencimiento"] = f"{int(d):02d}/{mm}/{y}"
        prima_neta = re.search(r"Prima\s+neta\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€", text, re.IGNORECASE)
        prima_total_poliza = re.search(
            r"Prima\s+total\s+p[oó]liza\s*:\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            text,
            re.IGNORECASE,
        )
        prima_total_recibo = re.search(r"Prima\s+total\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€", text, re.IGNORECASE)
        if prima_neta:
            fields["prima_neta"] = prima_neta.group(1)
        if prima_total_poliza:
            fields["prima_total"] = prima_total_poliza.group(1)
        elif prima_total_recibo:
            fields["prima_total"] = prima_total_recibo.group(1)
        if not fields.get("prima_neta"):
            desglose = re.search(
                r"Desglose del recibo de prima([\s\S]{0,450})Prima total p[oó]liza",
                text,
                re.IGNORECASE,
            )
            if desglose:
                nums = re.findall(r"([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s*€", desglose.group(1))
                if nums:
                    fields["prima_neta"] = nums[0]
        if normalize_email(fields.get("email") or "").endswith("@gco.com"):
            fields["email"] = ""
        fields["telefono"] = ""
    # Reglas específicas Santa Lucía Hogar.
    is_santalucia_hogar = (
        normalize_company_key(fields.get("compania") or "") in ("SANTALUCIA", "SANTALUCIA")
        and ("SEGURO DE HOGAR" in upper_text or "HOGAR COMPLETO" in upper_text)
    ) or ("SANTA LUCIA" in upper_text and "HOGAR" in upper_text)
    if is_santalucia_hogar:
        fields["compania"] = "Santa Lucia"
        fields["ramo"] = "Hogar"
        pol_match = re.search(
            r"P[ÓO]LIZA\s+N[ÚU]MERO\s*[\r\n]+\s*([A-Z0-9\-]{5,})",
            text,
            re.IGNORECASE,
        ) or re.search(
            r"P[ÓO]LIZA\s+N[ÚU]MERO\s*[:#]?\s*([A-Z0-9\-]{5,})",
            text,
            re.IGNORECASE,
        )
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1).strip()
        tom_match = re.search(
            r"DATOS\s+DEL\s+TOMADOR\s+DEL\s+SEGURO\s*[\r\n]+\s*([^\n]+)",
            text,
            re.IGNORECASE,
        )
        if tom_match:
            tom = normalize_person_name(tom_match.group(1)).strip(" ,;:-")
            if tom:
                fields["tomador"] = tom
        nif_match = re.search(r"DATOS\s+DEL\s+TOMADOR[\s\S]{0,220}?NIF\s*:\s*([A-Z0-9]{8,9})", text, re.IGNORECASE)
        if nif_match:
            fields["dni"] = nif_match.group(1).upper()
            fields["nif"] = fields["dni"]
        month_map = {
            "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
            "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
            "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
        }
        eff_match = re.search(
            r"DESDE\s+LAS\s+[0-9]{2}:[0-9]{2}\s+HORAS\s+DEL\s+DIA\s+([0-9]{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+([0-9]{4})",
            text,
            re.IGNORECASE,
        )
        exp_match = re.search(
            r"HASTA\s+LAS\s+[0-9]{2}:[0-9]{2}\s+HORAS\s+DEL\s+DIA\s+([0-9]{1,2})\s+de\s+([a-záéíóú]+)\s+de\s+([0-9]{4})",
            text,
            re.IGNORECASE,
        )
        if eff_match:
            d, mtxt, y = eff_match.groups()
            mm = month_map.get(normalize_lookup_text(mtxt).lower(), "")
            if mm:
                fields["fecha_efecto"] = f"{int(d):02d}/{mm}/{y}"
        if exp_match:
            d, mtxt, y = exp_match.groups()
            mm = month_map.get(normalize_lookup_text(mtxt).lower(), "")
            if mm:
                fields["fecha_vencimiento"] = f"{int(d):02d}/{mm}/{y}"
        prima_tarifa = re.search(
            r"PRIMA\s+TARIFA\s*:\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            text,
            re.IGNORECASE,
        )
        prima_total = re.search(
            r"PRIMA\s+TOTAL\s*:\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            text,
            re.IGNORECASE,
        )
        if prima_tarifa:
            fields["prima_neta"] = prima_tarifa.group(1)
        if prima_total:
            fields["prima_total"] = prima_total.group(1)
        risk_addr = re.search(
            r"SITUACI[ÓO]N\s+DE\s+LA\s+VIVIENDA[\s\S]{0,180}?[\r\n]+\s*([^\n]+)",
            text,
            re.IGNORECASE,
        )
        if risk_addr:
            fields["direccion"] = normalize_person_name(risk_addr.group(1)).strip(" ,;:-")
        if normalize_email(fields.get("email") or "").endswith("@santalucia.es"):
            fields["email"] = ""
        fields["telefono"] = ""
    # Reglas específicas pólizas EXSEL/Lloyd's RC Profesional (BASWZ...).
    is_lloyds_exsel = any(token in upper_text for token in ("BASWZ", "EXSEL UNDERWRITING", "LLOYD"))
    if is_lloyds_exsel:
        fields["compania"] = "Lloyd's"
        fields["ramo"] = "Responsabilidad civil"
        pol_match = re.search(r"P[oó]liza\s*(?:n[ºo]\s*)?:\s*([A-Z0-9]{10,})", text, re.IGNORECASE)
        if pol_match:
            fields["poliza_numero"] = pol_match.group(1).upper()
        tom_match = re.search(r"TOMADOR:\s*([^\n]+)", text, re.IGNORECASE)
        if tom_match:
            fields["tomador"] = normalize_person_name(tom_match.group(1)).strip(" ,;:-")
        cif_match = re.search(r"CIF/NIF:\s*([A-Z][0-9]{8})", text, re.IGNORECASE)
        if cif_match:
            fields["dni"] = cif_match.group(1).upper()
            fields["nif"] = fields["dni"]
        phone_match = re.search(r"\+34\s*([679][0-9\s]{8,})", text, re.IGNORECASE)
        if phone_match:
            fields["telefono"] = normalize_phone(phone_match.group(1)) or fields.get("telefono")
        mail_match = re.search(r"\b([A-Z0-9._%+\-]+@(?:LLOYDS|EXSEL)[A-Z0-9.\-]*\.[A-Z]{2,})\b", text, re.IGNORECASE)
        if mail_match:
            fields["email"] = normalize_email(mail_match.group(1))
    if "ramo" in fields:
        fields["ramo"] = canonicalize_ramo(fields.get("ramo"))
    return fields

def parse_asesoramiento_block(block):
    block_clean = re.sub(r"\s+", " ", block.replace("\u00a0", " "))
    def pick(patterns):
        for pat in patterns:
            m = re.search(pat, block, re.IGNORECASE)
            if m:
                return m.group(1).strip()
            m = re.search(pat, block_clean, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""
    data = {}
    data["nombre"] = pick([
        r"Nombre\s+y\s+apellidos\s*[:\-]?\s*([^\n]+)",
        r"Nombre\s+apellidos\s*[:\-]?\s*([^\n]+)",
        r"Nombre\s*[:\-]?\s*([^\n]+)",
    ])
    data["dni"] = pick([
        r"DNI\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"NIF\s*[:\-]?\s*([A-Z0-9\-]+)",
    ])
    data["telefono"] = pick([
        r"Tel[eé]fono\s*[:\-]?\s*([0-9\s]{9,})",
        r"M[oó]vil\s*[:\-]?\s*([0-9\s]{9,})",
        r"Tfno\.?\s*[:\-]?\s*([0-9\s]{9,})",
        r"Tlf\.?\s*[:\-]?\s*([0-9\s]{9,})",
    ])
    data["email"] = pick([
        r"Correo\s*Electr[oó]nico\s*[:\-]?\s*([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})",
    ])
    data["fecha_nacimiento"] = pick([
        r"Fecha\s*Nacimiento\s*[:\-]?\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
    ])
    data["estado_civil"] = pick([
        r"Estado\s*Civil\s*[:\-]?\s*([^\n]+)",
        r"Est\.?\s*Civil\s*[:\-]?\s*([^\n]+)",
    ])
    data["hijos"] = pick([
        r"Hijos\s*[:\-]?\s*([0-9]+)",
    ])
    data["profesion"] = pick([
        r"Profesi[oó]n\s*[:\-]?\s*([^\n]+)",
        r"Prof\.?\s*[:\-]?\s*([^\n]+)",
    ])
    data["tipo_contrato"] = pick([
        r"Tipo\s*Contrato\s*[:\-]?\s*([^\n]+)",
        r"Contrato\s*[:\-]?\s*([^\n]+)",
    ])
    data["ingresos"] = pick([
        r"Ingresos/N[oó]mina\s*[:\-]?\s*([0-9\.,]+)",
        r"Ingresos\s*[:\-]?\s*([0-9\.,]+)",
        r"N[oó]mina\s*[:\-]?\s*([0-9\.,]+)",
        r"Sueldo\s*[:\-]?\s*([0-9\.,]+)",
    ])
    data["patrimonio"] = pick([
        r"Patrimonio-?Alquiler\s*[:\-]?\s*([^\n]+)",
    ])
    data["prestamos"] = pick([
        r"PR[EÉ]STAMOS.*?\s*[:\-]?\s*([^\n]+)",
        r"PR[EÉ]STAMOS\s*[:\-]?\s*([^\n]+)",
    ])
    return data

def parse_asesoramiento_template(pdf_path):
    size = pdfinfo_page_size(pdf_path)
    if not size:
        return {}
    width, height = size
    boxes = {
        "cliente1": (0.05 * width, 0.28 * height, 0.9 * width, 0.22 * height),
        "cliente2": (0.05 * width, 0.53 * height, 0.9 * width, 0.22 * height),
        "resumen": (0.05 * width, 0.76 * height, 0.9 * width, 0.20 * height),
        "header": (0.05 * width, 0.16 * height, 0.9 * width, 0.10 * height),
    }
    fields = {}
    header = pdftotext_crop(pdf_path, *boxes["header"])
    if header:
        fields["inmobiliaria_asesor"] = parse_asesoramiento_text(header).get("inmobiliaria_asesor", "")
        fields["fecha"] = parse_asesoramiento_text(header).get("fecha", "")
    cliente1_text = pdftotext_crop(pdf_path, *boxes["cliente1"])
    if cliente1_text:
        data1 = parse_asesoramiento_block(cliente1_text)
        fields["cliente1_nombre"] = data1.get("nombre", "")
        fields["cliente1_dni"] = data1.get("dni", "")
        fields["cliente1_telefono"] = data1.get("telefono", "")
        fields["cliente1_email"] = data1.get("email", "")
        fields["cliente1_fecha_nacimiento"] = data1.get("fecha_nacimiento", "")
        fields["cliente1_estado_civil"] = data1.get("estado_civil", "")
        fields["cliente1_hijos"] = data1.get("hijos", "")
        fields["cliente1_profesion"] = data1.get("profesion", "")
        fields["cliente1_tipo_contrato"] = data1.get("tipo_contrato", "")
        fields["cliente1_ingresos"] = data1.get("ingresos", "")
        fields["cliente1_patrimonio"] = data1.get("patrimonio", "")
        fields["cliente1_prestamos"] = data1.get("prestamos", "")
    cliente2_text = pdftotext_crop(pdf_path, *boxes["cliente2"])
    if cliente2_text:
        data2 = parse_asesoramiento_block(cliente2_text)
        fields["cliente2_nombre"] = data2.get("nombre", "")
        fields["cliente2_dni"] = data2.get("dni", "")
        fields["cliente2_telefono"] = data2.get("telefono", "")
        fields["cliente2_email"] = data2.get("email", "")
        fields["cliente2_fecha_nacimiento"] = data2.get("fecha_nacimiento", "")
        fields["cliente2_estado_civil"] = data2.get("estado_civil", "")
        fields["cliente2_hijos"] = data2.get("hijos", "")
        fields["cliente2_profesion"] = data2.get("profesion", "")
        fields["cliente2_tipo_contrato"] = data2.get("tipo_contrato", "")
        fields["cliente2_ingresos"] = data2.get("ingresos", "")
        fields["cliente2_patrimonio"] = data2.get("patrimonio", "")
        fields["cliente2_prestamos"] = data2.get("prestamos", "")
    resumen = pdftotext_crop(pdf_path, *boxes["resumen"])
    if resumen:
        resumen_fields = parse_asesoramiento_text(resumen)
        for key in ("ingresos_conjuntos", "entidades_financieras", "avalistas", "aportacion_cv"):
            if resumen_fields.get(key):
                fields[key] = resumen_fields.get(key)
    return fields

def parse_asesoramiento_template_image(image_path):
    size = get_image_size(image_path)
    if not size:
        return {}
    width, height = size
    base_boxes = asesoramiento_image_boxes(width, height)
    shifts = (-0.015, 0.0, 0.015)
    fields = {}
    tmp_base = tempfile.gettempdir()
    with tempfile.TemporaryDirectory(dir=tmp_base) as tmpdir:
        processed, created = preprocess_image_for_ocr(image_path)
        for dy in shifts:
            for key, (x, y, w, h) in base_boxes.items():
                out_path = os.path.join(tmpdir, f"{key}_{dy:+.3f}.png")
                ok = crop_image_region(processed, x, y + dy * height, w, h, out_path)
                if not ok:
                    continue
                block_text, _ = ocr_image_file(out_path)
                if not block_text:
                    continue
                if key == "header":
                    header_fields = parse_asesoramiento_text(block_text)
                    for field in ("inmobiliaria_asesor", "fecha"):
                        if header_fields.get(field) and not fields.get(field):
                            fields[field] = header_fields.get(field)
                elif key == "cliente1":
                    data1 = parse_asesoramiento_block(block_text)
                    for field, target in (
                        ("nombre", "cliente1_nombre"),
                        ("dni", "cliente1_dni"),
                        ("telefono", "cliente1_telefono"),
                        ("email", "cliente1_email"),
                        ("fecha_nacimiento", "cliente1_fecha_nacimiento"),
                        ("estado_civil", "cliente1_estado_civil"),
                        ("hijos", "cliente1_hijos"),
                        ("profesion", "cliente1_profesion"),
                        ("tipo_contrato", "cliente1_tipo_contrato"),
                        ("ingresos", "cliente1_ingresos"),
                        ("patrimonio", "cliente1_patrimonio"),
                        ("prestamos", "cliente1_prestamos"),
                    ):
                        if data1.get(field) and not fields.get(target):
                            fields[target] = data1.get(field)
                elif key == "cliente2":
                    data2 = parse_asesoramiento_block(block_text)
                    for field, target in (
                        ("nombre", "cliente2_nombre"),
                        ("dni", "cliente2_dni"),
                        ("telefono", "cliente2_telefono"),
                        ("email", "cliente2_email"),
                        ("fecha_nacimiento", "cliente2_fecha_nacimiento"),
                        ("estado_civil", "cliente2_estado_civil"),
                        ("hijos", "cliente2_hijos"),
                        ("profesion", "cliente2_profesion"),
                        ("tipo_contrato", "cliente2_tipo_contrato"),
                        ("ingresos", "cliente2_ingresos"),
                        ("patrimonio", "cliente2_patrimonio"),
                        ("prestamos", "cliente2_prestamos"),
                    ):
                        if data2.get(field) and not fields.get(target):
                            fields[target] = data2.get(field)
                elif key == "resumen":
                    resumen_fields = parse_asesoramiento_text(block_text)
                    for field in ("ingresos_conjuntos", "entidades_financieras", "avalistas", "aportacion_cv"):
                        if resumen_fields.get(field) and not fields.get(field):
                            fields[field] = resumen_fields.get(field)
        if created and processed and os.path.exists(processed):
            try:
                os.unlink(processed)
            except Exception:
                pass
    return fields

def parse_asesoramiento_text(text):
    cleaned = text.replace("\u00a0", " ")
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = cleaned.replace("D.N.I.", "DNI").replace("DNI.", "DNI")
    cleaned = cleaned.replace("Nómina", "Nomina").replace("NÓMINA", "NOMINA")
    def pick(patterns, source_text=None, source_clean=None):
        source_text = source_text or text
        source_clean = source_clean or cleaned
        for pat in patterns:
            m = re.search(pat, source_text, re.IGNORECASE)
            if m:
                return m.group(1).strip()
            m = re.search(pat, source_clean, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""

    fields = {}
    fields["inmobiliaria_asesor"] = pick([
        r"Inmobiliaria/?Asesor\s*[:\-]?\s*([^\n]+)",
        r"INMOBILIARIA/ASESOR\s*[:\-]?\s*([^\n]+)",
    ])
    fields["fecha"] = pick([
        r"Fecha\s*[:\-]?\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})",
    ])

    parts = re.split(r"(CLIENTE\s*1|CLIENTE\s*2)", text, flags=re.IGNORECASE)
    blocks = {"1": "", "2": ""}
    current = ""
    for part in parts:
        if re.match(r"CLIENTE\s*1", part, re.IGNORECASE):
            current = "1"
            continue
        if re.match(r"CLIENTE\s*2", part, re.IGNORECASE):
            current = "2"
            continue
        if current in blocks:
            blocks[current] += part

    if blocks["1"]:
        data1 = parse_asesoramiento_block(blocks["1"])
        fields["cliente1_nombre"] = data1.get("nombre", "")
        fields["cliente1_dni"] = data1.get("dni", "")
        fields["cliente1_telefono"] = data1.get("telefono", "")
        fields["cliente1_email"] = data1.get("email", "")
        fields["cliente1_fecha_nacimiento"] = data1.get("fecha_nacimiento", "")
        fields["cliente1_estado_civil"] = data1.get("estado_civil", "")
        fields["cliente1_hijos"] = data1.get("hijos", "")
        fields["cliente1_profesion"] = data1.get("profesion", "")
        fields["cliente1_tipo_contrato"] = data1.get("tipo_contrato", "")
        fields["cliente1_ingresos"] = data1.get("ingresos", "")
        fields["cliente1_patrimonio"] = data1.get("patrimonio", "")
        fields["cliente1_prestamos"] = data1.get("prestamos", "")

    if blocks["2"]:
        data2 = parse_asesoramiento_block(blocks["2"])
        fields["cliente2_nombre"] = data2.get("nombre", "")
        fields["cliente2_dni"] = data2.get("dni", "")
        fields["cliente2_telefono"] = data2.get("telefono", "")
        fields["cliente2_email"] = data2.get("email", "")
        fields["cliente2_fecha_nacimiento"] = data2.get("fecha_nacimiento", "")
        fields["cliente2_estado_civil"] = data2.get("estado_civil", "")
        fields["cliente2_hijos"] = data2.get("hijos", "")
        fields["cliente2_profesion"] = data2.get("profesion", "")
        fields["cliente2_tipo_contrato"] = data2.get("tipo_contrato", "")
        fields["cliente2_ingresos"] = data2.get("ingresos", "")
        fields["cliente2_patrimonio"] = data2.get("patrimonio", "")
        fields["cliente2_prestamos"] = data2.get("prestamos", "")

    if not fields.get("cliente1_nombre"):
        fields["cliente1_nombre"] = pick([
            r"Cliente\s*1.*?Nombre\s+y\s+apellidos\s*[:\-]?\s*([^\n]+)",
        ])
    if not fields.get("cliente2_nombre"):
        fields["cliente2_nombre"] = pick([
            r"Cliente\s*2.*?Nombre\s+y\s+apellidos\s*[:\-]?\s*([^\n]+)",
        ])

    fields["ingresos_conjuntos"] = pick([
        r"Ingresos\s+conjuntos\s*[:\-]?\s*([0-9\.,]+)",
        r"Ingresos\s+conjunto\s*[:\-]?\s*([0-9\.,]+)",
        r"Ingresos\s*[:\-]?\s*([0-9\.,]+)",
    ])
    fields["entidades_financieras"] = pick([
        r"Entidades\s+Financieras\s*[:\-]?\s*([^\n]+)",
        r"Entidad(?:es)?\s+Financiera(?:s)?\s*[:\-]?\s*([^\n]+)",
    ])
    fields["avalistas"] = pick([
        r"Posibles\s+Avalistas/Cotitulares\s*[:\-]?\s*([^\n]+)",
    ])
    fields["aportacion_cv"] = pick([
        r"Aportaci[oó]n\s+para\s+CV\s*[:\-]?\s*([0-9\.,]+)",
        r"Aportaci[oó]n\s*para\s*CV\s*[:\-]?\s*([0-9\.,]+)",
        r"Aportaci[oó]n\s*[:\-]?\s*([0-9\.,]+)",
        r"Aportaci[oó]n\s*CV\s*[:\-]?\s*([0-9\.,]+)",
    ])

    phones = re.findall(r"\b[6-9][0-9]{8}\b", cleaned)
    emails = re.findall(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", cleaned, re.IGNORECASE)
    dnis = re.findall(r"\b[0-9]{8}[A-Z]\b", cleaned)
    cifs = re.findall(r"\b[ABCDEFGHJNPQRSUVW][0-9]{7}[0-9A-Z]\b", cleaned)
    docs = dnis + cifs

    if phones:
        if not fields.get("cliente1_telefono"):
            fields["cliente1_telefono"] = phones[0]
        if len(phones) > 1 and not fields.get("cliente2_telefono"):
            fields["cliente2_telefono"] = phones[1]
    if emails:
        if not fields.get("cliente1_email"):
            fields["cliente1_email"] = emails[0]
        if len(emails) > 1 and not fields.get("cliente2_email"):
            fields["cliente2_email"] = emails[1]
    if docs:
        if not fields.get("cliente1_dni"):
            fields["cliente1_dni"] = docs[0]
        if len(docs) > 1 and not fields.get("cliente2_dni"):
            fields["cliente2_dni"] = docs[1]

    if not fields.get("entidades_financieras"):
        bancos = [
            "BBVA",
            "ING",
            "Santander",
            "CaixaBank",
            "Sabadell",
            "Unicaja",
            "Abanca",
            "Ibercaja",
            "Kutxabank",
            "Cajamar",
            "Bankinter",
        ]
        encontrados = []
        for bank in bancos:
            if re.search(rf"\\b{re.escape(bank)}\\b", cleaned, re.IGNORECASE):
                encontrados.append(bank)
        if encontrados:
            fields["entidades_financieras"] = ", ".join(sorted(set(encontrados)))

    return fields

def ensure_cliente_for_seguro(conn, empresa_id, tomador, nif, now, extra=None):
    if not tomador:
        return None
    tomador = normalize_person_name(tomador)
    nif = (nif or "").strip()
    extra = extra or {}
    cliente = None
    nif_norm = normalize_nif(nif)
    if nif_norm:
        cliente = conn.execute(
            """
            SELECT id FROM clientes
            WHERE REPLACE(REPLACE(REPLACE(UPPER(nif), ' ', ''), '-', ''), '.', '') = ?
            """,
            (nif_norm,),
        ).fetchone()
    if not cliente:
        nombre_norm = normalize_person_name(tomador).upper()
        cliente = conn.execute(
            "SELECT id FROM clientes WHERE TRIM(UPPER(nombre)) = ?",
            (nombre_norm,),
        ).fetchone()
    if not cliente:
        tipo_persona = None
        if nif_norm:
            if re.match(r"^[0-9]{8}[A-Z]$", nif_norm):
                tipo_persona = "Física"
            elif re.match(r"^[A-Z][0-9]{7}[0-9A-Z]$", nif_norm):
                tipo_persona = "Jurídica"
        cliente_id = os.urandom(16).hex()
        conn.execute(
            """
            INSERT INTO clientes (
              id, nombre, tipo_persona, nif, telefono, email, fecha_nacimiento, direccion, estado, created_at, updated_at
            ) VALUES (
              ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
            )
            """,
            (
                cliente_id,
                tomador,
                tipo_persona,
                nif_norm or None,
                extra.get("telefono"),
                extra.get("email"),
                extra.get("fecha_nacimiento"),
                extra.get("direccion"),
                "Activo",
                now,
                now,
            ),
        )
    else:
        cliente_id = cliente["id"]
        updates = {}
        if nif_norm:
            updates["nif"] = nif_norm
        for key in ("telefono", "email", "fecha_nacimiento", "direccion"):
            value = extra.get(key)
            if value:
                updates[key] = value
        if updates:
            set_clause = ", ".join([f"{key} = COALESCE({key}, ?)" for key in updates])
            values = list(updates.values()) + [now, cliente_id]
            conn.execute(
                f"UPDATE clientes SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
    link = conn.execute(
        """
        SELECT id FROM clientes_empresas
        WHERE cliente_id = ? AND empresa_id = ? AND servicio = ?
        """,
        (cliente_id, empresa_id, "seguros"),
    ).fetchone()
    if not link:
        conn.execute(
            """
            INSERT INTO clientes_empresas (
              id, cliente_id, empresa_id, servicio, estado,
              fecha_inicio, fecha_fin, created_at, updated_at
            ) VALUES (
              ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
            )
            """,
            (
                os.urandom(16).hex(),
                cliente_id,
                empresa_id,
                "seguros",
                "Activo",
                None,
                None,
                now,
                now,
            ),
        )
    return cliente_id


def ensure_cliente_servicio_link(conn, cliente_id, empresa_id, servicio, now, estado="Activo"):
    if not cliente_id or not empresa_id or not servicio:
        return
    existing = conn.execute(
        """
        SELECT id FROM clientes_empresas
        WHERE cliente_id = ? AND empresa_id = ? AND LOWER(servicio) = LOWER(?)
        LIMIT 1
        """,
        (cliente_id, empresa_id, servicio),
    ).fetchone()
    if existing:
        return
    conn.execute(
        """
        INSERT INTO clientes_empresas (
          id, cliente_id, empresa_id, servicio, estado, fecha_inicio, fecha_fin, created_at, updated_at
        ) VALUES (
          ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
        )
        """,
        (
            os.urandom(16).hex(),
            cliente_id,
            empresa_id,
            servicio,
            estado,
            None,
            None,
            now,
            now,
        ),
    )


def autolink_uploaded_seguros_for_cliente(conn, cliente_id, empresa_id, now):
    if not cliente_id or not empresa_id:
        return {"linked": 0, "docs": 0}
    cliente = conn.execute(
        "SELECT id, nombre FROM clientes WHERE id = ?",
        (cliente_id,),
    ).fetchone()
    if not cliente:
        return {"linked": 0, "docs": 0}
    wanted_tokens = {t for t in normalize_lookup_text(cliente["nombre"] or "").split(" ") if len(t) >= 3}
    if not wanted_tokens:
        return {"linked": 0, "docs": 0}
    rows = conn.execute(
        f"""
        SELECT *
        FROM seguros
        WHERE empresa_id = ?
          AND ({uploaded_policy_filter()})
          AND (
            cliente_id IS NULL OR TRIM(cliente_id) = '' OR cliente_id = ?
          )
          AND tomador IS NOT NULL AND TRIM(tomador) <> ''
        ORDER BY COALESCE(fecha_efecto, created_at) DESC
        """,
        (empresa_id, cliente_id),
    ).fetchall()
    to_link = []
    for row in rows:
        cand_tokens = {t for t in normalize_lookup_text(row["tomador"] or "").split(" ") if len(t) >= 3}
        if not cand_tokens:
            continue
        if wanted_tokens.issubset(cand_tokens):
            to_link.append(row["id"])
            continue
        overlap = len(wanted_tokens.intersection(cand_tokens))
        if overlap >= 2 and (overlap / max(1, len(wanted_tokens))) >= 0.8:
            to_link.append(row["id"])
    linked = 0
    if to_link:
        conn.executemany(
            "UPDATE seguros SET cliente_id = ?, updated_at = datetime(?) WHERE id = ?",
            [(cliente_id, now, sid) for sid in to_link],
        )
        linked = len(to_link)
    docs = 0
    linked_rows = conn.execute(
        f"""
        SELECT *
        FROM seguros
        WHERE empresa_id = ?
          AND cliente_id = ?
          AND ({uploaded_policy_filter()})
        """,
        (empresa_id, cliente_id),
    ).fetchall()
    for row in linked_rows:
        doc_id = ensure_seguro_doc_link(conn, row, now)
        if doc_id:
            docs += 1
    if linked_rows:
        ensure_cliente_servicio_link(conn, cliente_id, empresa_id, "seguros", now)
    return {"linked": linked, "docs": docs}


def ensure_seguro_doc_link(conn, seguro_row, now, calidad_ocr=None, campos_ocr=""):
    if not seguro_row:
        return None
    cliente_id = seguro_row["cliente_id"]
    empresa_id = seguro_row["empresa_id"]
    poliza_key = (seguro_row["poliza_key"] or "").strip()
    poliza_url = (seguro_row["poliza_url"] or "").strip()
    if not cliente_id or not empresa_id or (not poliza_key and not poliza_url):
        return None
    where = ["cliente_id = ?", "empresa_id = ?", "LOWER(COALESCE(referencia_tipo, '')) = 'seguros'"]
    values = [cliente_id, empresa_id]
    key_or_url = []
    if poliza_key:
        key_or_url.append("doc_key = ?")
        values.append(poliza_key)
    if poliza_url:
        key_or_url.append("doc_url = ?")
        values.append(poliza_url)
    if key_or_url:
        where.append(f"({' OR '.join(key_or_url)})")
    exists = conn.execute(
        f"SELECT id FROM gestoria_docs WHERE {' AND '.join(where)} LIMIT 1",
        values,
    ).fetchone()
    if exists:
        conn.execute(
            """
            UPDATE gestoria_docs
            SET referencia_id = COALESCE(?, referencia_id),
                nombre = COALESCE(NULLIF(?, ''), nombre),
                tipo = COALESCE(NULLIF(?, ''), tipo),
                fecha = COALESCE(NULLIF(?, ''), fecha),
                estado = COALESCE(NULLIF(?, ''), estado),
                notas = COALESCE(NULLIF(?, ''), notas),
                doc_key = COALESCE(NULLIF(?, ''), doc_key),
                doc_url = COALESCE(NULLIF(?, ''), doc_url),
                calidad_ocr = COALESCE(NULLIF(?, ''), calidad_ocr),
                campos_ocr = COALESCE(NULLIF(?, ''), campos_ocr),
                updated_at = datetime(?)
            WHERE id = ?
            """,
            (
                seguro_row["id"],
                seguro_row["poliza_numero"] or seguro_row["tomador"] or "Póliza seguro",
                "Seguros",
                seguro_row["fecha_efecto"] or seguro_row["mes_creacion"] or "",
                "Recibido" if (poliza_key or poliza_url) else (seguro_row["estado"] or "En vigor"),
                " · ".join([value for value in (seguro_row["compania"], seguro_row["ramo"]) if value]),
                poliza_key or "",
                poliza_url or "",
                str(calidad_ocr or ""),
                str(campos_ocr or ""),
                now,
                exists["id"],
            ),
        )
        return exists["id"]
    # Fallback: si ya existe doc por referencia de la póliza, actualizarlo.
    by_ref = conn.execute(
        """
        SELECT id FROM gestoria_docs
        WHERE referencia_tipo = 'seguros' AND referencia_id = ?
        LIMIT 1
        """,
        (seguro_row["id"],),
    ).fetchone()
    if by_ref:
        conn.execute(
            """
            UPDATE gestoria_docs
            SET cliente_id = ?,
                empresa_id = ?,
                doc_key = COALESCE(NULLIF(?, ''), doc_key),
                doc_url = COALESCE(NULLIF(?, ''), doc_url),
                estado = COALESCE(NULLIF(?, ''), estado),
                calidad_ocr = COALESCE(NULLIF(?, ''), calidad_ocr),
                campos_ocr = COALESCE(NULLIF(?, ''), campos_ocr),
                updated_at = datetime(?)
            WHERE id = ?
            """,
            (
                cliente_id,
                empresa_id,
                poliza_key or "",
                poliza_url or "",
                "Recibido" if (poliza_key or poliza_url) else "",
                str(calidad_ocr or ""),
                str(campos_ocr or ""),
                now,
                by_ref["id"],
            ),
        )
        return by_ref["id"]
    nombre_doc = seguro_row["poliza_numero"] or seguro_row["tomador"] or "Póliza seguro"
    notas_doc = " · ".join([value for value in (seguro_row["compania"], seguro_row["ramo"]) if value])
    doc_id = os.urandom(16).hex()
    conn.execute(
        """
        INSERT INTO gestoria_docs (
          id, empresa_id, cliente_id, referencia_tipo, referencia_id,
          nombre, tipo, fecha, estado, notas, doc_key, doc_url,
          calidad_ocr, campos_ocr, created_at, updated_at
        ) VALUES (
          ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
        )
        """,
        (
            doc_id,
            empresa_id,
            cliente_id,
            "seguros",
            seguro_row["id"],
            nombre_doc,
            "Seguros",
            seguro_row["fecha_efecto"] or seguro_row["mes_creacion"],
            seguro_row["estado"] or "En vigor",
            notas_doc,
            poliza_key or None,
            poliza_url or None,
            calidad_ocr,
            campos_ocr or "",
            now,
            now,
        ),
    )
    return doc_id


def build_cliente_ficha_payload(conn, cliente_id, services_filter=None):
    cliente = conn.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,)).fetchone()
    if not cliente:
        return None
    services_filter = services_filter or []
    if services_filter and not cliente_has_servicio(conn, cliente_id, services_filter):
        return {"error": "Cliente no disponible para este servicio"}

    empresas_query = """
        SELECT ce.id AS rel_id, ce.empresa_id, e.nombre AS empresa, ce.servicio, ce.estado,
               ce.fecha_inicio, ce.fecha_fin
        FROM clientes_empresas ce
        LEFT JOIN empresas e ON e.id = ce.empresa_id
        WHERE ce.cliente_id = ?
    """
    values = [cliente_id]
    if services_filter:
        placeholders = ",".join(["?"] * len(services_filter))
        empresas_query += f" AND LOWER(ce.servicio) IN ({placeholders})"
        values.extend(services_filter)
    empresas_query += " ORDER BY e.nombre"
    empresas = [dict(r) for r in conn.execute(empresas_query, values).fetchall()]

    service_keys = []
    allowed_keys = {"gestoria", "seguros", "inmobiliaria", "financiaciones"}
    for row in empresas:
        service_key = normalize_service_key(row.get("servicio"))
        active = is_active_service_state(row.get("estado"), row.get("fecha_fin"))
        row["servicio_key"] = service_key
        row["is_active"] = active
        if not active:
            continue
        if service_key == "hipotecas":
            service_key = "financiaciones"
        if service_key in allowed_keys and service_key not in service_keys:
            service_keys.append(service_key)
    empresa_ids = [row["empresa_id"] for row in empresas if row.get("empresa_id")]

    seguros_rows = [
        dict(r)
        for r in conn.execute(
            """
            SELECT id, cliente_id, empresa_id, compania, ramo, poliza_numero, fecha_efecto, fecha_vencimiento,
                   estado, prima_neta, prima_total, tomador, estado_renovacion, renovacion_fecha,
                   nueva_poliza_ref, colaborador, produccion, mes_creacion,
                   poliza_key, poliza_url, fecha_baja, motivo_baja
            FROM seguros
            WHERE cliente_id = ?
            ORDER BY COALESCE(fecha_efecto, created_at) DESC
            """,
            (cliente_id,),
        ).fetchall()
    ]
    for row in seguros_rows:
        row.update(compute_seguro_display(row))

    docs_rows = [
        dict(r)
        for r in conn.execute(
            """
            SELECT id, empresa_id, cliente_id, referencia_tipo, referencia_id, nombre, tipo, fecha,
                   estado, notas, doc_key, doc_url
            FROM gestoria_docs
            WHERE cliente_id = ?
            ORDER BY created_at DESC
            """,
            (cliente_id,),
        ).fetchall()
    ]
    docs_by_service = {"seguros": [], "gestoria": [], "financiaciones": [], "inmobiliaria": [], "otros": []}
    for row in docs_rows:
        key = normalize_service_key(row.get("referencia_tipo") or row.get("tipo"))
        if key not in docs_by_service:
            key = "otros"
        docs_by_service[key].append(row)

    facturas = []
    if empresa_ids:
        placeholders = ",".join(["?"] * len(empresa_ids))
        raw_facturas = conn.execute(
            f"""
            SELECT id, empresa_id, cliente_id, cliente_ids_json, fecha, concepto, gestion, tipo, importe, notas
            FROM gestoria_contabilidad
            WHERE empresa_id IN ({placeholders})
            ORDER BY fecha DESC, created_at DESC
            LIMIT 1000
            """,
            [*empresa_ids],
        ).fetchall()
        for raw_row in raw_facturas:
            row = dict(raw_row)
            asignados = parse_cliente_ids_payload(row.get("cliente_ids_json"))
            if not asignados and row.get("cliente_id"):
                asignados = [str(row.get("cliente_id")).strip()]
            if cliente_id not in asignados:
                continue
            reparto = 1.0 / max(1, len(asignados))
            importe_original = parse_money_value(row.get("importe"))
            row["importe_asignado"] = round(importe_original * reparto, 2)
            row["clientes_reparto"] = asignados
            facturas.append(row)
        facturas = facturas[:500]
    fact_total = sum(
        parse_money_value(row.get("importe_asignado") if row.get("importe_asignado") not in (None, "") else row.get("importe"))
        for row in facturas
        if normalize_lookup_text(row.get("tipo")) != "GASTO"
    )

    trabajos = [
        dict(r)
        for r in conn.execute(
            """
            SELECT id, empresa_id, cliente_id, tipo_trabajo, estado, fecha_inicio, fecha_fin, responsable, importe, notas
            FROM gestoria_trabajos
            WHERE cliente_id = ?
            ORDER BY COALESCE(fecha_fin, fecha_inicio, created_at) DESC
            LIMIT 500
            """,
            (cliente_id,),
        ).fetchall()
    ]

    acciones = [
        dict(r)
        for r in conn.execute(
            """
            SELECT id, empresa_id, cliente_id, fecha, hora, tipo, responsable, estado, notas, servicio, recordatorio_min
            FROM acciones
            WHERE cliente_id = ?
            ORDER BY fecha DESC, hora DESC
            LIMIT 500
            """,
            (cliente_id,),
        ).fetchall()
    ]

    historico = []
    for row in trabajos:
        historico.append(
            {
                "fecha": row.get("fecha_fin") or row.get("fecha_inicio") or "",
                "servicio": "gestoria",
                "concepto": row.get("tipo_trabajo") or "Trabajo",
                "estado": row.get("estado") or "-",
                "importe": row.get("importe"),
            }
        )
    for row in seguros_rows:
        historico.append(
            {
                "fecha": row.get("vencimiento_display") or row.get("fecha_efecto") or "",
                "servicio": "seguros",
                "concepto": f"{row.get('compania') or 'Seguro'} {row.get('poliza_numero') or ''}".strip(),
                "estado": row.get("estado_display") or row.get("estado") or "-",
                "importe": row.get("prima_total"),
            }
        )
    historico.sort(key=lambda r: str(r.get("fecha") or ""), reverse=True)

    pendientes_trabajos = sum(
        1
        for row in trabajos
        if "FINAL" not in normalize_lookup_text(row.get("estado"))
        and "CANCEL" not in normalize_lookup_text(row.get("estado"))
    )
    pendientes_acciones = sum(
        1
        for row in acciones
        if all(
            token not in normalize_lookup_text(row.get("estado"))
            for token in ("HECHO", "FINAL", "CANCEL")
        )
    )
    today = datetime.now(timezone.utc).date()
    citas_programadas = [
        row
        for row in acciones
        if parse_iso_date(row.get("fecha")) and parse_iso_date(row.get("fecha")) >= today
    ]
    primas_total = sum(parse_money_value(row.get("prima_total")) for row in seguros_rows)
    realizado = sum(
        parse_money_value(row.get("importe"))
        for row in trabajos
        if "FINAL" in normalize_lookup_text(row.get("estado"))
    )
    cobrado = fact_total
    rentabilidad = cobrado - realizado

    profesionales = {
        "gestoria": {},
        "seguros": {
            "polizas_total": len(seguros_rows),
            "companias": sorted(
                {str(row.get("compania") or "").strip() for row in seguros_rows if str(row.get("compania") or "").strip()}
            ),
        },
    }
    gestoria_row = conn.execute(
        """
        SELECT tipo_cliente, mod_fiscal, mod_laboral, mod_contable, mod_renta, mod_registro, mod_trafico, mod_puntuales, renta_detalles
        FROM cliente_gestoria
        WHERE cliente_id = ?
        """,
        (cliente_id,),
    ).fetchone()
    if gestoria_row:
        profesionales["gestoria"] = dict(gestoria_row)

    return {
        "cliente": dict(cliente),
        "datos_personales": dict(cliente),
        "empresas": empresas,
        "servicios_activos": service_keys,
        "datos_profesionales_por_servicio": profesionales,
        "servicios": {
            "seguros": seguros_rows,
            "gestoria_trabajos": trabajos,
            "acciones": acciones,
        },
        "documentacion": {
            "all": docs_rows,
            "by_service": docs_by_service,
        },
        "facturas": facturas,
        "historico": historico[:500],
        "dashboard": {
            "rentabilidad": {
                "realizado": realizado,
                "cobrado": cobrado,
                "margen": rentabilidad,
            },
            "primas_total": primas_total,
            "tareas_pendientes": pendientes_trabajos + pendientes_acciones,
            "citas_programadas": len(citas_programadas),
            "proxima_cita": min(
                [row["fecha"] for row in citas_programadas if row.get("fecha")],
                default="",
            ),
            "series": {
                "rentabilidad": [
                    {"label": "Realizado", "value": realizado},
                    {"label": "Cobrado", "value": cobrado},
                    {"label": "Margen", "value": rentabilidad},
                ],
                "actividad": [
                    {"label": "Primas", "value": primas_total},
                    {"label": "Pendientes", "value": pendientes_trabajos + pendientes_acciones},
                    {"label": "Citas", "value": len(citas_programadas)},
                ],
            },
        },
    }

def ensure_cliente_for_financiacion(conn, empresa_id, nombre, nif, now, extra=None):
    if not nombre:
        return None
    nombre = str(nombre).strip()
    nif = (nif or "").strip()
    extra = extra or {}
    cliente = None
    if nif:
        cliente = conn.execute(
            "SELECT id FROM clientes WHERE nif = ?",
            (nif,),
        ).fetchone()
    if not cliente:
        cliente = conn.execute(
            "SELECT id FROM clientes WHERE nombre = ?",
            (nombre,),
        ).fetchone()
    if not cliente:
        tipo_persona = None
        if nif:
            if re.match(r"^[0-9]{8}[A-Z]$", nif):
                tipo_persona = "Física"
            elif re.match(r"^[A-Z][0-9]{7}[0-9A-Z]$", nif):
                tipo_persona = "Jurídica"
        cliente_id = os.urandom(16).hex()
        conn.execute(
            """
            INSERT INTO clientes (
              id, nombre, tipo_persona, nif, telefono, email, fecha_nacimiento, direccion, estado, created_at, updated_at
            ) VALUES (
              ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
            )
            """,
            (
                cliente_id,
                nombre,
                tipo_persona,
                nif or None,
                extra.get("telefono"),
                extra.get("email"),
                extra.get("fecha_nacimiento"),
                extra.get("direccion"),
                "Activo",
                now,
                now,
            ),
        )
    else:
        cliente_id = cliente["id"]
        updates = {}
        if nif:
            updates["nif"] = nif
        for key in ("telefono", "email", "fecha_nacimiento", "direccion"):
            value = extra.get(key)
            if value:
                updates[key] = value
        if updates:
            set_clause = ", ".join([f"{key} = COALESCE({key}, ?)" for key in updates])
            values = list(updates.values()) + [now, cliente_id]
            conn.execute(
                f"UPDATE clientes SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
    link = conn.execute(
        """
        SELECT id FROM clientes_empresas
        WHERE cliente_id = ? AND empresa_id = ? AND servicio = ?
        """,
        (cliente_id, empresa_id, "financiaciones"),
    ).fetchone()
    if not link:
        conn.execute(
            """
            INSERT INTO clientes_empresas (
              id, cliente_id, empresa_id, servicio, estado,
              fecha_inicio, fecha_fin, created_at, updated_at
            ) VALUES (
              ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
            )
            """,
            (
                os.urandom(16).hex(),
                cliente_id,
                empresa_id,
                "financiaciones",
                "Activo",
                None,
                None,
                now,
                now,
            ),
        )
    return cliente_id

TABLES = [
    "movimientos",
    "seguros",
    "gestoria",
    "captaciones",
    "inmuebles",
    "demandas",
    "visitas",
    "hipotecas",
    "alquileres",
    "inversores",
    "inversure_operaciones",
]


def get_db(db_path):
    return open_sqlite_conn(db_path, with_row_factory=True)


def open_sqlite_conn(db_path, with_row_factory=False):
    conn = sqlite3.connect(db_path, timeout=90)
    if with_row_factory:
        conn.row_factory = sqlite3.Row
    # Reduce bloqueos en escenarios multi-hilo (web + OCR worker).
    try:
        conn.execute("PRAGMA journal_mode=WAL")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA busy_timeout=90000")
    except Exception:
        pass
    try:
        conn.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        pass
    return conn


def ensure_tables(db_path):
    conn = open_sqlite_conn(db_path, with_row_factory=False)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ocr_jobs (
          id TEXT PRIMARY KEY,
          kind TEXT,
          status TEXT,
          payload_json TEXT,
          result_json TEXT,
          error TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL,
          started_at TEXT,
          finished_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cliente_gestoria (
          id TEXT PRIMARY KEY,
          cliente_id TEXT UNIQUE,
          tipo_cliente TEXT,
          mod_fiscal INTEGER,
          mod_laboral INTEGER,
          mod_contable INTEGER,
          mod_renta INTEGER,
          mod_registro INTEGER,
          mod_trafico INTEGER,
          mod_puntuales INTEGER,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    cliente_gestoria_cols = [row[1] for row in conn.execute("PRAGMA table_info(cliente_gestoria)").fetchall()]
    if "mod_renta" not in cliente_gestoria_cols:
        conn.execute("ALTER TABLE cliente_gestoria ADD COLUMN mod_renta INTEGER")
    if "renta_detalles" not in cliente_gestoria_cols:
        conn.execute("ALTER TABLE cliente_gestoria ADD COLUMN renta_detalles TEXT")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_modelos (
          id TEXT PRIMARY KEY,
          cliente_id TEXT,
          modelo TEXT,
          periodicidad TEXT,
          proxima_fecha TEXT,
          responsable TEXT,
          estado TEXT,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_trabajos (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          cliente_id TEXT,
          tipo_trabajo TEXT,
          estado TEXT,
          fecha_inicio TEXT,
          fecha_fin TEXT,
          responsable TEXT,
          importe REAL,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    try:
        cap_cols = [row[1] for row in conn.execute("PRAGMA table_info(captaciones)").fetchall()]
        if "codigo_postal" not in cap_cols:
            conn.execute("ALTER TABLE captaciones ADD COLUMN codigo_postal TEXT")
        if "poblacion" not in cap_cols:
            conn.execute("ALTER TABLE captaciones ADD COLUMN poblacion TEXT")
        if "provincia" not in cap_cols:
            conn.execute("ALTER TABLE captaciones ADD COLUMN provincia TEXT")
    except sqlite3.Error:
        pass
    try:
        inm_cols = [row[1] for row in conn.execute("PRAGMA table_info(inmuebles)").fetchall()]
        if "codigo_postal" not in inm_cols:
            conn.execute("ALTER TABLE inmuebles ADD COLUMN codigo_postal TEXT")
        if "poblacion" not in inm_cols:
            conn.execute("ALTER TABLE inmuebles ADD COLUMN poblacion TEXT")
        if "provincia" not in inm_cols:
            conn.execute("ALTER TABLE inmuebles ADD COLUMN provincia TEXT")
    except sqlite3.Error:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_docs (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          cliente_id TEXT,
          referencia_tipo TEXT,
          referencia_id TEXT,
          nombre TEXT,
          tipo TEXT,
          fecha TEXT,
          estado TEXT,
          notas TEXT,
          doc_key TEXT,
          doc_url TEXT,
          calidad_ocr TEXT,
          campos_ocr TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    try:
        docs_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_docs)").fetchall()]
        if "doc_key" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN doc_key TEXT")
        if "doc_url" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN doc_url TEXT")
        if "referencia_tipo" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN referencia_tipo TEXT")
        if "referencia_id" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN referencia_id TEXT")
        if "calidad_ocr" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN calidad_ocr TEXT")
        if "campos_ocr" not in docs_cols:
            conn.execute("ALTER TABLE gestoria_docs ADD COLUMN campos_ocr TEXT")
    except sqlite3.Error:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS asesoramientos_financiacion (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          origen TEXT,
          inmobiliaria_asesor TEXT,
          asesor TEXT,
          fecha TEXT,
          estado TEXT,
          cliente1_id TEXT,
          cliente1_nombre TEXT,
          cliente1_dni TEXT,
          cliente1_telefono TEXT,
          cliente1_email TEXT,
          cliente1_fecha_nacimiento TEXT,
          cliente1_estado_civil TEXT,
          cliente1_regimen TEXT,
          cliente1_hijos TEXT,
          cliente1_profesion TEXT,
          cliente1_tipo_contrato TEXT,
          cliente1_tiempo_contrato TEXT,
          cliente1_ingresos REAL,
          cliente1_patrimonio TEXT,
          cliente1_prestamos TEXT,
          cliente1_prestamo_activo TEXT,
          cliente1_prestamo_entidad TEXT,
          cliente1_prestamo_resto REAL,
          cliente2_id TEXT,
          cliente2_nombre TEXT,
          cliente2_dni TEXT,
          cliente2_telefono TEXT,
          cliente2_email TEXT,
          cliente2_fecha_nacimiento TEXT,
          cliente2_estado_civil TEXT,
          cliente2_regimen TEXT,
          cliente2_hijos TEXT,
          cliente2_profesion TEXT,
          cliente2_tipo_contrato TEXT,
          cliente2_tiempo_contrato TEXT,
          cliente2_ingresos REAL,
          cliente2_patrimonio TEXT,
          cliente2_prestamos TEXT,
          cliente2_prestamo_activo TEXT,
          cliente2_prestamo_entidad TEXT,
          cliente2_prestamo_resto REAL,
          ingresos_conjuntos REAL,
          entidades_financieras TEXT,
          avalistas TEXT,
          aportacion_cv REAL,
          notas TEXT,
          notas_ocr TEXT,
          calidad_ocr TEXT,
          campos_ocr TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    ases_cols = [row[1] for row in conn.execute("PRAGMA table_info(asesoramientos_financiacion)").fetchall()]
    if "cliente1_tiempo_contrato" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente1_tiempo_contrato TEXT")
    if "cliente2_tiempo_contrato" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente2_tiempo_contrato TEXT")
    if "cliente1_regimen" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente1_regimen TEXT")
    if "cliente2_regimen" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente2_regimen TEXT")
    if "cliente1_prestamo_activo" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente1_prestamo_activo TEXT")
    if "cliente1_prestamo_entidad" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente1_prestamo_entidad TEXT")
    if "cliente1_prestamo_resto" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente1_prestamo_resto REAL")
    if "cliente2_prestamo_activo" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente2_prestamo_activo TEXT")
    if "cliente2_prestamo_entidad" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente2_prestamo_entidad TEXT")
    if "cliente2_prestamo_resto" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN cliente2_prestamo_resto REAL")
    if "calidad_ocr" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN calidad_ocr TEXT")
    if "campos_ocr" not in ases_cols:
        conn.execute("ALTER TABLE asesoramientos_financiacion ADD COLUMN campos_ocr TEXT")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_contabilidad (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          cliente_id TEXT,
          cliente_ids_json TEXT,
          seguro_id TEXT,
          poliza_numero TEXT,
          fecha TEXT,
          concepto TEXT,
          gestion TEXT,
          tipo TEXT,
          importe REAL,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_conta_config (
          id TEXT PRIMARY KEY,
          cliente_id TEXT UNIQUE,
          periodo TEXT,
          fecha_inicio TEXT,
          responsable TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_conta_tasks (
          id TEXT PRIMARY KEY,
          cliente_id TEXT,
          periodo TEXT,
          tarea TEXT,
          estado TEXT,
          fecha_limite TEXT,
          responsable TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_terceros (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          nif TEXT,
          nombre TEXT,
          tipo TEXT,
          cuenta_contable TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_facturas (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          cliente_id TEXT,
          tercero_id TEXT,
          tipo TEXT,
          numero TEXT,
          fecha_emision TEXT,
          descripcion TEXT,
          base_imponible REAL,
          cuota_iva REAL,
          cuota_irpf REAL,
          total REAL,
          iva_pct REAL,
          estado_ocr TEXT,
          doc_key TEXT,
          raw_text TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_asientos (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          cliente_id TEXT,
          factura_id TEXT,
          fecha TEXT,
          concepto TEXT,
          diario TEXT,
          referencia TEXT,
          total_debe REAL,
          total_haber REAL,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestoria_asiento_lineas (
          id TEXT PRIMARY KEY,
          asiento_id TEXT,
          tercero_id TEXT,
          cuenta TEXT,
          descripcion TEXT,
          debe REAL,
          haber REAL,
          impuesto_tipo TEXT,
          impuesto_pct REAL,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS inmueble_checklist (
          id TEXT PRIMARY KEY,
          inmueble_id TEXT NOT NULL,
          etapa TEXT,
          tarea TEXT,
          estado TEXT,
          responsable TEXT,
          fecha_limite TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_ofertas (
          id TEXT PRIMARY KEY,
          cliente_id TEXT,
          ramo TEXT,
          compania TEXT,
          propuesta TEXT,
          estado TEXT,
          motivo TEXT,
          fecha TEXT,
          responsable TEXT,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_preferencias (
          id TEXT PRIMARY KEY,
          cliente_id TEXT UNIQUE,
          prioridad_precio INTEGER DEFAULT 0,
          prioridad_compania INTEGER DEFAULT 0,
          prioridad_coberturas INTEGER DEFAULT 0,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_referidos (
          id TEXT PRIMARY KEY,
          cliente_id TEXT,
          referido_por TEXT,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_campanas (
          id TEXT PRIMARY KEY,
          compania TEXT,
          nombre TEXT,
          ramo TEXT,
          origen TEXT,
          fecha_inicio TEXT,
          fecha_fin TEXT,
          descripcion TEXT,
          url TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    campanas_cols = [row[1] for row in conn.execute("PRAGMA table_info(seguros_campanas)").fetchall()]
    if "origen" not in campanas_cols:
        conn.execute("ALTER TABLE seguros_campanas ADD COLUMN origen TEXT")
    try:
        seguros_cols = [row[1] for row in conn.execute("PRAGMA table_info(seguros)").fetchall()]
        if "poliza_key" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN poliza_key TEXT")
        if "poliza_url" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN poliza_url TEXT")
        if "comision" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN comision REAL")
        if "porcentaje" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN porcentaje REAL")
        if "produccion" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN produccion REAL")
        if "colaborador" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN colaborador TEXT")
        if "cliente_id" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN cliente_id TEXT")
        if "fecha_baja" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN fecha_baja TEXT")
        if "motivo_baja" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN motivo_baja TEXT")
        if "estado_poliza" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN estado_poliza TEXT")
        if "poliza_origen_id" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN poliza_origen_id TEXT")
        if "poliza_sustituta_id" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN poliza_sustituta_id TEXT")
        if "version_grupo" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN version_grupo TEXT")
        if "tipo_vigencia" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN tipo_vigencia TEXT")
        if "datos_ramo_json" not in seguros_cols:
            conn.execute("ALTER TABLE seguros ADD COLUMN datos_ramo_json TEXT")
    except sqlite3.Error:
        pass
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_comisiones (
          id TEXT PRIMARY KEY,
          compania TEXT,
          ramo TEXT,
          porcentaje REAL,
          vigencia_desde TEXT,
          vigencia_hasta TEXT,
          notas TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_checklist (
          id TEXT PRIMARY KEY,
          poliza_id TEXT NOT NULL,
          tarea TEXT,
          estado TEXT,
          responsable TEXT,
          fecha_limite TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL,
          FOREIGN KEY (poliza_id) REFERENCES seguros(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_eventos (
          id TEXT PRIMARY KEY,
          seguro_id TEXT,
          cliente_id TEXT,
          empresa_id TEXT,
          tipo TEXT,
          fecha TEXT,
          motivo TEXT,
          payload_json TEXT,
          created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_reclamaciones (
          id TEXT PRIMARY KEY,
          seguro_id TEXT,
          cliente_id TEXT,
          empresa_id TEXT,
          estado TEXT,
          canal TEXT,
          fecha_apertura TEXT,
          fecha_cierre TEXT,
          asunto TEXT,
          detalle TEXT,
          resolucion TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS seguros_ipid_log (
          id TEXT PRIMARY KEY,
          seguro_id TEXT,
          cliente_id TEXT,
          empresa_id TEXT,
          documento_key TEXT,
          documento_url TEXT,
          fecha_entrega TEXT,
          metodo TEXT,
          usuario TEXT,
          created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS fin_checklist (
          id TEXT PRIMARY KEY,
          asesoramiento_id TEXT NOT NULL,
          tarea TEXT,
          estado TEXT,
          responsable TEXT,
          fecha_limite TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL,
          FOREIGN KEY (asesoramiento_id) REFERENCES asesoramientos_financiacion(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS auditoria (
          id TEXT PRIMARY KEY,
          empresa_id TEXT,
          entidad TEXT,
          entidad_id TEXT,
          accion TEXT,
          usuario TEXT,
          detalles TEXT,
          created_at TEXT NOT NULL,
          FOREIGN KEY (empresa_id) REFERENCES empresas(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS postal_catalogo (
          codigo_postal TEXT,
          poblacion TEXT,
          provincia TEXT
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_postal_cp ON postal_catalogo(codigo_postal)"
    )
    ensure_usuarios_schema(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS cnae_catalogo (
          codigo TEXT PRIMARY KEY,
          descripcion TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS iae_catalogo (
          codigo TEXT PRIMARY KEY,
          descripcion TEXT NOT NULL
        )
        """
    )
    cols = [row[1] for row in conn.execute("PRAGMA table_info(clientes)").fetchall()]
    if "tipo_persona" not in cols:
        conn.execute("ALTER TABLE clientes ADD COLUMN tipo_persona TEXT")
    if "codigo_postal" not in cols:
        conn.execute("ALTER TABLE clientes ADD COLUMN codigo_postal TEXT")
    if "poblacion" not in cols:
        conn.execute("ALTER TABLE clientes ADD COLUMN poblacion TEXT")
    if "provincia" not in cols:
        conn.execute("ALTER TABLE clientes ADD COLUMN provincia TEXT")
    modelo_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_modelos)").fetchall()]
    if "responsable" not in modelo_cols:
        conn.execute("ALTER TABLE gestoria_modelos ADD COLUMN responsable TEXT")
    trabajo_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_trabajos)").fetchall()]
    if "responsable" not in trabajo_cols:
        conn.execute("ALTER TABLE gestoria_trabajos ADD COLUMN responsable TEXT")
    if "sla_dias" not in trabajo_cols:
        conn.execute("ALTER TABLE gestoria_trabajos ADD COLUMN sla_dias INTEGER")
    acciones_cols = [row[1] for row in conn.execute("PRAGMA table_info(acciones)").fetchall()]
    if "responsable" not in acciones_cols:
        conn.execute("ALTER TABLE acciones ADD COLUMN responsable TEXT")
    if "recordatorio_min" not in acciones_cols:
        conn.execute("ALTER TABLE acciones ADD COLUMN recordatorio_min INTEGER")
    if "inmueble_id" not in acciones_cols:
        conn.execute("ALTER TABLE acciones ADD COLUMN inmueble_id TEXT")
    inm_cols = [row[1] for row in conn.execute("PRAGMA table_info(inmuebles)").fetchall()]
    if "valor_referencia" not in inm_cols:
        conn.execute("ALTER TABLE inmuebles ADD COLUMN valor_referencia REAL")
    conta_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_contabilidad)").fetchall()]
    if "gestion" not in conta_cols:
        conn.execute("ALTER TABLE gestoria_contabilidad ADD COLUMN gestion TEXT")
    if "seguro_id" not in conta_cols:
        conn.execute("ALTER TABLE gestoria_contabilidad ADD COLUMN seguro_id TEXT")
    if "poliza_numero" not in conta_cols:
        conn.execute("ALTER TABLE gestoria_contabilidad ADD COLUMN poliza_numero TEXT")
    if "cliente_ids_json" not in conta_cols:
        conn.execute("ALTER TABLE gestoria_contabilidad ADD COLUMN cliente_ids_json TEXT")
    terceros_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_terceros)").fetchall()]
    if "cuenta_contable" not in terceros_cols:
        conn.execute("ALTER TABLE gestoria_terceros ADD COLUMN cuenta_contable TEXT")
    facturas_cols = [row[1] for row in conn.execute("PRAGMA table_info(gestoria_facturas)").fetchall()]
    if "iva_pct" not in facturas_cols:
        conn.execute("ALTER TABLE gestoria_facturas ADD COLUMN iva_pct REAL")
    if "estado_ocr" not in facturas_cols:
        conn.execute("ALTER TABLE gestoria_facturas ADD COLUMN estado_ocr TEXT")
    if "doc_key" not in facturas_cols:
        conn.execute("ALTER TABLE gestoria_facturas ADD COLUMN doc_key TEXT")
    if "raw_text" not in facturas_cols:
        conn.execute("ALTER TABLE gestoria_facturas ADD COLUMN raw_text TEXT")
    load_postal_catalog(conn)
    conn.commit()
    conn.close()


def ensure_usuarios_schema(conn):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS usuarios (
          id TEXT PRIMARY KEY,
          nombre TEXT NOT NULL,
          apellido TEXT,
          usuario TEXT UNIQUE,
          email TEXT UNIQUE,
          servicio TEXT,
          rol TEXT,
          password_hash TEXT,
          activo INTEGER DEFAULT 1,
          invite_token TEXT,
          invite_expires_at TEXT,
          invite_sent_at TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        )
        """
    )
    user_cols = [row[1] for row in conn.execute("PRAGMA table_info(usuarios)").fetchall()]
    if "apellido" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN apellido TEXT")
    if "usuario" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN usuario TEXT")
    if "email" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN email TEXT")
    if "servicio" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN servicio TEXT")
    if "password_hash" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN password_hash TEXT")
    if "invite_token" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN invite_token TEXT")
    if "invite_expires_at" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN invite_expires_at TEXT")
    if "invite_sent_at" not in user_cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN invite_sent_at TEXT")
    users_count = conn.execute("SELECT COUNT(*) AS total FROM usuarios").fetchone()
    total_users = 0
    if users_count:
        try:
            total_users = users_count["total"]
        except (TypeError, KeyError, IndexError):
            total_users = users_count[0]
    if total_users == 0:
        conn.execute(
            "INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, activo, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))",
            (os.urandom(16).hex(), "Administrador", "General", "admin", "admin@liv.local", "Administración", "Administrador", 1),
        )


def ensure_ocr_tables(db_path):
    conn = open_sqlite_conn(db_path, with_row_factory=False)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS ocr_jobs (
          id TEXT PRIMARY KEY,
          kind TEXT,
          status TEXT,
          payload_json TEXT,
          result_json TEXT,
          error TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL,
          started_at TEXT,
          finished_at TEXT
        )
        """
    )
    conn.commit()
    conn.close()


def json_response(handler, data, status=200, cookies=None, extra_headers=None):
    payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    for key, value in (extra_headers or []):
        handler.send_header(key, value)
    for cookie_value in (cookies or []):
        handler.send_header("Set-Cookie", cookie_value)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)


def send_file(handler, path):
    if not path.exists() or not path.is_file():
        handler.send_error(404, "Not found")
        return

    content_type = "text/plain"
    if path.suffix == ".html":
        content_type = "text/html; charset=utf-8"
    elif path.suffix == ".css":
        content_type = "text/css; charset=utf-8"
    elif path.suffix == ".js":
        content_type = "application/javascript; charset=utf-8"
    elif path.suffix in (".png", ".jpg", ".jpeg", ".gif"):
        content_type = f"image/{path.suffix.lstrip('.')}"
    elif path.suffix == ".svg":
        content_type = "image/svg+xml"

    try:
        data = path.read_bytes()
        handler.send_response(200)
        handler.send_header("Content-Type", content_type)
        handler.send_header("Content-Length", str(len(data)))
        handler.end_headers()
        handler.wfile.write(data)
    except Exception:
        handler.send_response(500)
        handler.send_header("Content-Type", "text/plain; charset=utf-8")
        handler.end_headers()
        handler.wfile.write(b"Failed to read file")


class Handler(BaseHTTPRequestHandler):
    db_path = DB_DEFAULT
    ocr_db_path = OCR_DB_DEFAULT

    def log_message(self, format, *args):
        return

    def _track_conn(self, conn):
        if conn is None:
            return
        if not hasattr(self, "_tracked_conns"):
            self._tracked_conns = []
        self._tracked_conns.append(conn)

    def _close_tracked_conns(self):
        conns = getattr(self, "_tracked_conns", [])
        while conns:
            conn = conns.pop()
            try:
                conn.close()
            except Exception:
                pass

    def finish(self):
        try:
            super().finish()
        finally:
            self._close_tracked_conns()

    def _parse_cookies(self):
        raw = self.headers.get("Cookie", "") or ""
        result = {}
        for part in raw.split(";"):
            if "=" not in part:
                continue
            key, value = part.split("=", 1)
            key = key.strip()
            if not key:
                continue
            result[key] = value.strip()
        return result

    def _session_cookie_secure(self):
        forced = os.environ.get("APP_SESSION_COOKIE_SECURE", "").strip().lower()
        if forced in ("1", "true", "yes", "on"):
            return True
        if forced in ("0", "false", "no", "off"):
            return False
        proto = (self.headers.get("X-Forwarded-Proto") or "").strip().lower()
        return proto == "https" or bool(os.environ.get("RENDER"))

    def _build_session_cookie(self, value, max_age=None):
        parts = [f"{SESSION_COOKIE_NAME}={value}", "Path=/", "HttpOnly", "SameSite=Lax"]
        if max_age is not None:
            parts.append(f"Max-Age={int(max_age)}")
        if self._session_cookie_secure():
            parts.append("Secure")
        return "; ".join(parts)

    def _current_session(self):
        token = self._parse_cookies().get(SESSION_COOKIE_NAME, "")
        return get_auth_session(token)

    def _auth_user_payload(self, session):
        if not session:
            return None
        full_name = " ".join(x for x in [session.get("nombre"), session.get("apellido")] if x).strip()
        return {
            "id": session.get("user_id"),
            "usuario": session.get("usuario") or "",
            "nombre": session.get("nombre") or "",
            "apellido": session.get("apellido") or "",
            "nombre_completo": full_name or (session.get("usuario") or ""),
            "rol": session.get("rol") or "",
            "email": session.get("email") or "",
            "servicio": session.get("servicio") or "",
        }

    def _external_base_url(self):
        configured = (os.environ.get("APP_BASE_URL") or "").strip().rstrip("/")
        if configured:
            return configured
        proto = (self.headers.get("X-Forwarded-Proto") or "").strip() or ("https" if os.environ.get("RENDER") else "http")
        host = (self.headers.get("X-Forwarded-Host") or self.headers.get("Host") or "").strip()
        if host:
            return f"{proto}://{host}"
        return "http://localhost:8000"

    def _require_api_auth(self):
        session = self._current_session()
        if session:
            self.auth_session = session
            return True
        json_response(self, {"error": "No autenticado"}, status=401)
        return False

    def _auth_allowed_services(self):
        session = getattr(self, "auth_session", None) or self._current_session()
        if not session:
            return None
        rol = normalize_service_key(session.get("rol") or "")
        servicio_raw = str(session.get("servicio") or "")
        services = set()
        for item in parse_services_param(servicio_raw):
            key = normalize_service_key(item)
            if key:
                services.add(key)
        expanded = set(services)
        if any(item in {"direccion", "administracion"} for item in expanded):
            return None
        servicio_key = normalize_service_key(servicio_raw)
        if servicio_key in {"direccion", "administracion"}:
            return None
        # If admin role has explicit services configured, keep service scoping.
        if rol in {"administrador", "direccion", "administracion"} and not expanded:
            return None
        if "gestoria" in expanded:
            expanded.add("administracion fincas")
        if "administracion fincas" in expanded:
            expanded.add("gestoria")
        if "financiaciones" in expanded:
            expanded.add("hipotecas")
        if "hipotecas" in expanded:
            expanded.add("financiaciones")
        return expanded

    def _service_from_tabla(self, tabla):
        table_service_map = {
            "seguros": "seguros",
            "seguros_ofertas": "seguros",
            "seguros_preferencias": "seguros",
            "seguros_referidos": "seguros",
            "seguros_campanas": "seguros",
            "seguros_comisiones": "seguros",
            "seguros_checklist": "seguros",
            "seguros_reclamaciones": "seguros",
            "gestoria": "gestoria",
            "gestoria_docs": "gestoria",
            "gestoria_trabajos": "gestoria",
            "gestoria_contabilidad": "gestoria",
            "gestoria_modelos": "gestoria",
            "gestoria_conta_tasks": "gestoria",
            "cliente_gestoria": "gestoria",
            "hipotecas": "financiaciones",
            "asesoramientos_financiacion": "financiaciones",
            "fin_checklist": "financiaciones",
            "captaciones": "inmobiliaria",
            "inmuebles": "inmobiliaria",
            "demandas": "inmobiliaria",
            "visitas": "inmobiliaria",
            "movimientos": "inmobiliaria",
            "alquileres": "inmobiliaria",
        }
        return table_service_map.get(str(tabla or "").strip().lower(), "")

    def _resolve_required_service(self, path, params=None, payload=None):
        params = params or {}
        payload = payload or {}
        if path.startswith("/api/seguros"):
            return "seguros"
        if path.startswith("/api/gestoria") or path.startswith("/api/cliente_gestoria"):
            return "gestoria"
        if path.startswith("/api/fin_") or path.startswith("/api/hipotecas"):
            return "financiaciones"
        if path.startswith("/api/capt") or path.startswith("/api/inmueble") or path.startswith("/api/demandas") or path.startswith("/api/visitas"):
            return "inmobiliaria"
        if path in {"/api/acciones", "/api/acciones_update"}:
            raw = payload.get("servicio") or (params.get("servicio", [""])[0] if params else "")
            return normalize_service_key(raw)
        if path in {"/api/clientes_link", "/api/cliente_empresa_update"}:
            raw = payload.get("servicio") or ""
            return normalize_service_key(raw)
        if path == "/api/tabla":
            tabla = params.get("tabla", [""])[0] if params else ""
            return self._service_from_tabla(tabla)
        return ""

    def _enforce_service_access(self, path, params=None, payload=None):
        public_paths = {
            "/api/login",
            "/api/logout",
            "/api/auth_set_password",
            "/api/me",
            "/api/auth_invite_status",
            "/api/usuarios",
            "/api/usuarios_update",
            "/api/usuarios_delete",
            "/api/usuarios_invitar",
            "/api/health",
        }
        if path in public_paths:
            return True
        allowed = self._auth_allowed_services()
        if allowed is None:
            return True
        required = self._resolve_required_service(path, params=params, payload=payload)
        if not required:
            return True
        required = normalize_service_key(required)
        if required in {"administracion fincas", "gestoria"}:
            if "gestoria" in allowed or "administracion fincas" in allowed:
                return True
        if required in {"financiaciones", "hipotecas"}:
            if "financiaciones" in allowed or "hipotecas" in allowed:
                return True
        if required in allowed:
            return True
        json_response(self, {"error": "Sin permisos para este servicio"}, status=403)
        return False

    def do_HEAD(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path in ("/health", "/api/health"):
            self.send_response(200)
            self.end_headers()
            return
        self.send_response(200)
        self.end_headers()

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path in ("/health", "/api/health"):
            self.send_response(200)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write(b"ok")
            return
        if parsed.path.startswith("/api/"):
            if parsed.path not in AUTH_PUBLIC_GET_ENDPOINTS and not self._require_api_auth():
                return
            get_params = urllib.parse.parse_qs(parsed.query)
            if not self._enforce_service_access(parsed.path, params=get_params):
                return
            self.handle_api(parsed)
            return

        if parsed.path == "/" or parsed.path == "":
            send_file(self, ROOT / "index.html")
            return

        if parsed.path.startswith("/assets/"):
            rel = parsed.path.replace("/assets/", "")
            send_file(self, ASSETS / rel)
            return
        if parsed.path.startswith("/uploads/"):
            rel = parsed.path.replace("/uploads/", "")
            send_file(self, UPLOADS / rel)
            return

        rel_path = parsed.path.lstrip("/")
        file_path = ROOT / rel_path
        send_file(self, file_path)

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path not in (
            "/api/movimientos",
            "/api/hipotecas",
            "/api/hipotecas/firmar",
            "/api/gestoria",
            "/api/gestoria_trabajos",
            "/api/gestoria_trabajos_update",
            "/api/gestoria_trabajos_delete",
            "/api/gestoria_docs",
            "/api/gestoria_docs_update",
            "/api/gestoria_docs_delete",
            "/api/gestoria_contabilidad",
            "/api/gestoria_contabilidad_update",
            "/api/gestoria_contabilidad_delete",
            "/api/gestoria_factura_ocr",
            "/api/gestoria_update",
            "/api/seguros_ocr",
            "/api/seguros_ocr_async",
            "/api/fin_asesoramiento_ocr",
            "/api/fin_asesoramiento_ocr_guided",
            "/api/fin_asesoramiento_ocr_auto",
            "/api/seguros",
            "/api/seguros_update",
            "/api/seguros_cambio_compania",
            "/api/seguros_delete",
            "/api/seguros_poliza_accion",
            "/api/seguros_enrich",
            "/api/seguros_reclamacion",
            "/api/seguros_reclamacion_update",
            "/api/seguros_reclamacion_delete",
            "/api/seguros_ipid_register",
            "/api/fin_asesoramientos",
            "/api/fin_asesoramientos_update",
            "/api/fin_asesoramientos_convert",
            "/api/seguros_ofertas",
            "/api/seguros_ofertas_update",
            "/api/seguros_ofertas_delete",
            "/api/seguros_preferencias",
            "/api/seguros_referidos",
            "/api/seguros_campanas",
            "/api/seguros_campanas_update",
            "/api/seguros_campanas_delete",
            "/api/seguros_comisiones",
            "/api/seguros_comisiones_update",
            "/api/seguros_comisiones_delete",
            "/api/seguros_checklist_generate",
            "/api/seguros_checklist_update",
            "/api/fin_checklist_generate",
            "/api/fin_checklist_update",
            "/api/ai_seguros_copilot",
            "/api/ai_fin_copilot",
            "/api/s3_presign",
            "/api/s3_multipart_start",
            "/api/s3_multipart_presign",
            "/api/s3_multipart_complete",
            "/api/s3_multipart_abort",
            "/api/clientes",
            "/api/clientes_link",
            "/api/clientes_link_delete",
            "/api/cliente_update",
            "/api/cliente_empresa_update",
            "/api/acciones",
            "/api/acciones_update",
            "/api/cliente_gestoria_update",
            "/api/gestoria_modelos",
            "/api/gestoria_modelos_update",
            "/api/gestoria_modelos_delete",
            "/api/cliente_profesional",
            "/api/cliente_profesional_update",
            "/api/cliente_profesional_delete",
            "/api/captaciones",
            "/api/captaciones_update",
            "/api/captacion_update",
            "/api/inmueble_update",
            "/api/inmueble_propietarios_update",
            "/api/inmueble_docs",
            "/api/inmueble_checklist_generate",
            "/api/inmueble_checklist_update",
            "/api/demandas",
            "/api/visitas",
            "/api/usuarios",
            "/api/usuarios_update",
            "/api/usuarios_delete",
            "/api/usuarios_invitar",
            "/api/login",
            "/api/logout",
            "/api/auth_set_password",
        ):
            json_response(self, {"error": "Endpoint no valido"}, status=404)
            return

        content_length = int(self.headers.get("Content-Length", 0))
        body = self.rfile.read(content_length or 0)
        try:
            payload = json.loads(body.decode("utf-8"))
        except json.JSONDecodeError:
            json_response(self, {"error": "JSON invalido"}, status=400)
            return

        if parsed.path not in AUTH_PUBLIC_POST_ENDPOINTS and not self._require_api_auth():
            return
        if not self._enforce_service_access(parsed.path, payload=payload):
            return

        if parsed.path == "/api/logout":
            token = self._parse_cookies().get(SESSION_COOKIE_NAME, "")
            delete_auth_session(token)
            json_response(
                self,
                {"ok": True},
                cookies=[self._build_session_cookie("", max_age=0)],
            )
            return

        if parsed.path == "/api/login":
            usuario_raw = str(payload.get("usuario") or payload.get("email") or "").strip()
            password = str(payload.get("password") or "")
            if not usuario_raw or not password:
                json_response(self, {"error": "usuario y contraseña requeridos"}, status=400)
                return
            conn = get_db(self.db_path)
            self._track_conn(conn)
            ensure_usuarios_schema(conn)
            conn.commit()
            row = conn.execute(
                """
                SELECT id, nombre, apellido, usuario, email, servicio, rol, activo, password_hash
                FROM usuarios
                WHERE activo = 1
                  AND (
                    LOWER(COALESCE(usuario, '')) = LOWER(?)
                    OR LOWER(COALESCE(email, '')) = LOWER(?)
                  )
                LIMIT 1
                """,
                (usuario_raw, usuario_raw),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Usuario o contraseña incorrectos"}, status=401)
                return
            first_password_set = False
            stored_hash = row["password_hash"]
            if stored_hash:
                if not verify_password(password, stored_hash):
                    json_response(self, {"error": "Usuario o contraseña incorrectos"}, status=401)
                    return
            else:
                if not AUTH_ALLOW_FIRST_PASSWORD_SET:
                    json_response(self, {"error": "Usuario sin contraseña inicializada"}, status=403)
                    return
                conn.execute(
                    "UPDATE usuarios SET password_hash = ?, updated_at = datetime('now') WHERE id = ?",
                    (hash_password(password), row["id"]),
                )
                conn.commit()
                first_password_set = True
                row = conn.execute(
                    """
                    SELECT id, nombre, apellido, usuario, email, servicio, rol, activo, password_hash
                    FROM usuarios WHERE id = ?
                    """,
                    (row["id"],),
                ).fetchone()
            session = create_auth_session(row)
            json_response(
                self,
                {
                    "ok": True,
                    "user": self._auth_user_payload(session),
                    "first_password_set": first_password_set,
                },
                cookies=[self._build_session_cookie(session["token"], max_age=APP_SESSION_TTL_SECONDS)],
            )
            return

        if parsed.path == "/api/auth_set_password":
            token = str(payload.get("token") or "").strip()
            password = str(payload.get("password") or "")
            if not token or not password:
                json_response(self, {"error": "token y password requeridos"}, status=400)
                return
            if len(password) < 8:
                json_response(self, {"error": "La contraseña debe tener al menos 8 caracteres"}, status=400)
                return
            conn = get_db(self.db_path)
            self._track_conn(conn)
            ensure_usuarios_schema(conn)
            conn.commit()
            row = conn.execute(
                """
                SELECT id, activo, invite_expires_at
                FROM usuarios
                WHERE invite_token = ?
                LIMIT 1
                """,
                (token,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Invitación inválida"}, status=404)
                return
            if not row["activo"]:
                json_response(self, {"error": "Usuario inactivo"}, status=403)
                return
            expires_raw = str(row["invite_expires_at"] or "").strip()
            if expires_raw:
                try:
                    dt = datetime.fromisoformat(expires_raw.replace("Z", ""))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    if dt < datetime.now(timezone.utc):
                        json_response(self, {"error": "Invitación caducada"}, status=410)
                        return
                except Exception:
                    pass
            conn.execute(
                """
                UPDATE usuarios
                SET password_hash = ?, invite_token = NULL, invite_expires_at = NULL, updated_at = datetime('now')
                WHERE id = ?
                """,
                (hash_password(password), row["id"]),
            )
            conn.commit()
            json_response(self, {"ok": True})
            return

        empresa_nombre = payload.get("empresa_nombre")
        if parsed.path not in (
            "/api/hipotecas/firmar",
            "/api/clientes",
            "/api/clientes_link",
            "/api/clientes_link_delete",
            "/api/inmueble_update",
            "/api/inmueble_propietarios_update",
            "/api/captacion_update",
            "/api/cliente_update",
            "/api/cliente_empresa_update",
            "/api/acciones",
            "/api/acciones_update",
            "/api/cliente_profesional",
            "/api/cliente_profesional_update",
            "/api/cliente_profesional_delete",
            "/api/usuarios",
            "/api/usuarios_update",
            "/api/usuarios_delete",
            "/api/usuarios_invitar",
            "/api/auth_set_password",
            "/api/cliente_gestoria_update",
            "/api/inmueble_docs",
            "/api/inmueble_checklist_generate",
            "/api/inmueble_checklist_update",
            "/api/gestoria_modelos",
            "/api/gestoria_trabajos",
            "/api/gestoria_docs",
            "/api/gestoria_contabilidad",
            "/api/gestoria_trabajos_update",
            "/api/gestoria_trabajos_delete",
            "/api/gestoria_docs_update",
            "/api/gestoria_docs_delete",
            "/api/gestoria_contabilidad_update",
            "/api/gestoria_contabilidad_delete",
            "/api/fin_asesoramientos",
            "/api/fin_asesoramientos_update",
            "/api/fin_asesoramientos_convert",
            "/api/gestoria_conta_config",
            "/api/gestoria_conta_tasks_bulk",
            "/api/gestoria_conta_task_update",
            "/api/seguros_ocr",
            "/api/seguros_ocr_async",
            "/api/fin_asesoramiento_ocr",
            "/api/fin_asesoramiento_ocr_guided",
            "/api/fin_asesoramiento_ocr_auto",
            "/api/seguros_delete",
            "/api/seguros_cambio_compania",
            "/api/seguros_update",
            "/api/seguros_poliza_accion",
            "/api/seguros_ofertas",
            "/api/seguros_ofertas_update",
            "/api/seguros_ofertas_delete",
            "/api/seguros_preferencias",
            "/api/seguros_referidos",
            "/api/seguros_campanas",
            "/api/seguros_campanas_update",
            "/api/seguros_campanas_delete",
            "/api/seguros_comisiones",
            "/api/seguros_comisiones_update",
            "/api/seguros_comisiones_delete",
            "/api/seguros_reclamacion",
            "/api/seguros_reclamacion_update",
            "/api/seguros_reclamacion_delete",
            "/api/seguros_ipid_register",
            "/api/fin_checklist_generate",
            "/api/fin_checklist_update",
            "/api/ai_seguros_copilot",
            "/api/ai_fin_copilot",
            "/api/s3_presign",
            "/api/s3_multipart_start",
            "/api/s3_multipart_presign",
            "/api/s3_multipart_complete",
            "/api/s3_multipart_abort",
        ):
            if not empresa_nombre:
                json_response(self, {"error": "empresa_nombre requerido"}, status=400)
                return

        if empresa_nombre == "Inmovere Gestión AIE":
            try:
                anio = int(payload.get("anio"))
            except (TypeError, ValueError):
                json_response(self, {"error": "anio invalido"}, status=400)
                return
            mes = str(payload.get("mes", "")).strip().lower()
            meses = [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ]
            mes_idx = meses.index(mes) if mes in meses else -1
            if anio < 2026 or (anio == 2026 and mes_idx < 1):
                json_response(self, {"error": "Solo desde Febrero 2026"}, status=400)
                return

        conn = get_db(self.db_path)
        self._track_conn(conn)
        if parsed.path in (
            "/api/login",
            "/api/auth_set_password",
            "/api/usuarios",
            "/api/usuarios_update",
            "/api/usuarios_delete",
            "/api/usuarios_invitar",
        ):
            ensure_usuarios_schema(conn)
            conn.commit()
        empresa = None
        if parsed.path not in (
            "/api/hipotecas/firmar",
            "/api/clientes",
            "/api/clientes_link",
            "/api/clientes_link_delete",
            "/api/cliente_update",
            "/api/cliente_empresa_update",
            "/api/cliente_gestoria_update",
            "/api/cliente_profesional",
            "/api/cliente_profesional_update",
            "/api/cliente_profesional_delete",
            "/api/usuarios",
            "/api/usuarios_update",
            "/api/usuarios_delete",
            "/api/usuarios_invitar",
            "/api/auth_set_password",
            "/api/acciones_update",
            "/api/gestoria_modelos",
            "/api/gestoria_modelos_update",
            "/api/gestoria_modelos_delete",
            "/api/gestoria_trabajos_update",
            "/api/gestoria_trabajos_delete",
            "/api/gestoria_docs_update",
            "/api/gestoria_docs_delete",
            "/api/seguros_delete",
            "/api/seguros_cambio_compania",
            "/api/seguros_update",
            "/api/seguros_poliza_accion",
            "/api/gestoria_contabilidad_update",
            "/api/gestoria_contabilidad_delete",
            "/api/auditoria",
            "/api/acciones",
            "/api/seguros_ocr",
            "/api/seguros_ocr_async",
            "/api/fin_asesoramiento_ocr",
            "/api/fin_asesoramiento_ocr_guided",
            "/api/fin_asesoramiento_ocr_auto",
            "/api/s3_presign",
            "/api/s3_multipart_start",
            "/api/s3_multipart_presign",
            "/api/s3_multipart_complete",
            "/api/s3_multipart_abort",
            "/api/inmueble_checklist_generate",
            "/api/inmueble_checklist_update",
            "/api/seguros_reclamacion",
            "/api/seguros_reclamacion_update",
            "/api/seguros_reclamacion_delete",
            "/api/seguros_ipid_register",
            "/api/ai_seguros_copilot",
        ):
            empresa = conn.execute(
                "SELECT id FROM empresas WHERE nombre = ?",
                (empresa_nombre,),
            ).fetchone()
            if not empresa:
                json_response(self, {"error": "Empresa no encontrada"}, status=400)
                return

        now = "now"
        def audit(entidad, entidad_id, accion, detalles=None, usuario=None):
            conn.execute(
                """
                INSERT INTO auditoria (
                  id, empresa_id, entidad, entidad_id, accion, usuario, detalles, created_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"] if empresa else None,
                    entidad,
                    entidad_id,
                    accion,
                    usuario or "Sistema",
                    detalles,
                    now,
                ),
            )
        if parsed.path == "/api/s3_presign":
            filename = payload.get("filename") or "archivo.pdf"
            content_type = payload.get("content_type") or "application/pdf"
            prefix = payload.get("prefix") or "seguros"
            client = s3_client()
            if not client:
                bucket, region = s3_config()
                missing = []
                if not bucket:
                    missing.append("AWS_S3_BUCKET")
                if not region:
                    missing.append("AWS_REGION")
                if not S3_BOTO3_AVAILABLE:
                    missing.append("boto3")
                detail = f" (faltan: {', '.join(missing)})" if missing else ""
                json_response(self, {"error": f"S3 no configurado{detail}"}, status=400)
                return
            bucket, region = s3_config()
            key = s3_safe_key(prefix, filename)
            try:
                url = client.generate_presigned_url(
                    "put_object",
                    Params={
                        "Bucket": bucket,
                        "Key": key,
                        "ContentType": content_type,
                    },
                    ExpiresIn=900,
                )
            except Exception:
                json_response(self, {"error": "No se pudo firmar la subida"}, status=500)
                return
            public_url = f"https://{bucket}.s3.{region}.amazonaws.com/{key}"
            json_response(self, {"url": url, "key": key, "public_url": public_url})
            return
        if parsed.path == "/api/s3_multipart_start":
            filename = payload.get("filename") or "archivo.pdf"
            content_type = payload.get("content_type") or "application/pdf"
            prefix = payload.get("prefix") or "docs"
            file_size = int(payload.get("size") or 0)
            client = s3_client()
            if not client:
                bucket, region = s3_config()
                missing = []
                if not bucket:
                    missing.append("AWS_S3_BUCKET")
                if not region:
                    missing.append("AWS_REGION")
                if not S3_BOTO3_AVAILABLE:
                    missing.append("boto3")
                detail = f" (faltan: {', '.join(missing)})" if missing else ""
                json_response(self, {"error": f"S3 no configurado{detail}"}, status=400)
                return
            bucket, region = s3_config()
            key = s3_safe_key(prefix, filename)
            part_size = max(5 * 1024 * 1024, int(payload.get("part_size") or 8 * 1024 * 1024))
            max_parts = 10000
            if file_size and (file_size / part_size) > max_parts:
                part_size = int((file_size / max_parts) + 1)
            try:
                created = client.create_multipart_upload(
                    Bucket=bucket,
                    Key=key,
                    ContentType=content_type,
                )
            except Exception:
                json_response(self, {"error": "No se pudo iniciar subida multipart"}, status=500)
                return
            upload_id = created.get("UploadId")
            if not upload_id:
                json_response(self, {"error": "UploadId no disponible"}, status=500)
                return
            public_url = f"https://{bucket}.s3.{region}.amazonaws.com/{key}"
            json_response(
                self,
                {
                    "key": key,
                    "upload_id": upload_id,
                    "part_size": part_size,
                    "public_url": public_url,
                },
            )
            return
        if parsed.path == "/api/s3_multipart_presign":
            key = (payload.get("key") or "").strip()
            upload_id = (payload.get("upload_id") or "").strip()
            try:
                part_number = int(payload.get("part_number") or 0)
            except Exception:
                part_number = 0
            if not key or not upload_id or part_number <= 0 or part_number > 10000:
                json_response(self, {"error": "key, upload_id y part_number válidos requeridos"}, status=400)
                return
            client = s3_client()
            bucket, region = s3_config()
            if not client or not bucket or not region:
                json_response(self, {"error": "S3 no configurado"}, status=400)
                return
            try:
                url = client.generate_presigned_url(
                    "upload_part",
                    Params={
                        "Bucket": bucket,
                        "Key": key,
                        "UploadId": upload_id,
                        "PartNumber": part_number,
                    },
                    ExpiresIn=900,
                )
            except Exception:
                json_response(self, {"error": "No se pudo firmar la parte"}, status=500)
                return
            json_response(self, {"url": url, "part_number": part_number})
            return
        if parsed.path == "/api/s3_multipart_complete":
            key = (payload.get("key") or "").strip()
            upload_id = (payload.get("upload_id") or "").strip()
            if not key or not upload_id:
                json_response(self, {"error": "key y upload_id requeridos"}, status=400)
                return
            client = s3_client()
            bucket, region = s3_config()
            if not client or not bucket or not region:
                json_response(self, {"error": "S3 no configurado"}, status=400)
                return
            try:
                parts = []
                kwargs = {"Bucket": bucket, "Key": key, "UploadId": upload_id}
                while True:
                    listed = client.list_parts(**kwargs)
                    for p in listed.get("Parts", []) or []:
                        if p.get("PartNumber") and p.get("ETag"):
                            parts.append({"PartNumber": p["PartNumber"], "ETag": p["ETag"]})
                    if listed.get("IsTruncated"):
                        kwargs["PartNumberMarker"] = listed.get("NextPartNumberMarker")
                    else:
                        break
                if not parts:
                    json_response(self, {"error": "No hay partes subidas para completar"}, status=400)
                    return
                parts.sort(key=lambda p: p["PartNumber"])
                client.complete_multipart_upload(
                    Bucket=bucket,
                    Key=key,
                    UploadId=upload_id,
                    MultipartUpload={"Parts": parts},
                )
            except Exception:
                json_response(self, {"error": "No se pudo completar la subida multipart"}, status=500)
                return
            public_url = f"https://{bucket}.s3.{region}.amazonaws.com/{key}"
            json_response(self, {"ok": True, "key": key, "public_url": public_url})
            return
        if parsed.path == "/api/s3_multipart_abort":
            key = (payload.get("key") or "").strip()
            upload_id = (payload.get("upload_id") or "").strip()
            if not key or not upload_id:
                json_response(self, {"error": "key y upload_id requeridos"}, status=400)
                return
            client = s3_client()
            bucket, region = s3_config()
            if not client or not bucket or not region:
                json_response(self, {"error": "S3 no configurado"}, status=400)
                return
            try:
                client.abort_multipart_upload(
                    Bucket=bucket,
                    Key=key,
                    UploadId=upload_id,
                )
            except Exception:
                json_response(self, {"error": "No se pudo abortar la subida multipart"}, status=500)
                return
            json_response(self, {"ok": True})
            return
        if parsed.path == "/api/movimientos":
            conn.execute(
                """
                INSERT INTO movimientos (
                  id, empresa_id, concepto, pisos_vendidos, comision, asesor,
                  anio, mes, sl, tipo, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("concepto"),
                    payload.get("pisos_vendidos"),
                    payload.get("comision"),
                    payload.get("asesor"),
                    payload.get("anio"),
                    payload.get("mes"),
                    payload.get("sl"),
                    payload.get("tipo"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/gestoria":
            conn.execute(
                """
                INSERT INTO gestoria (
                  id, empresa_id, cliente, fecha, cuota, precio, tipo, perfil,
                  estado, fecha_baja,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("cliente"),
                    payload.get("fecha"),
                    payload.get("cuota"),
                    payload.get("precio"),
                    payload.get("tipo"),
                    payload.get("perfil"),
                    payload.get("estado"),
                    payload.get("fecha_baja"),
                    now,
                    now,
                ),
            )
            audit("gestoria_cliente", payload.get("cliente"), "crear", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_trabajos":
            conn.execute(
                """
                INSERT INTO gestoria_trabajos (
                  id, empresa_id, cliente_id, tipo_trabajo, estado,
                  fecha_inicio, fecha_fin, sla_dias, responsable, importe, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("cliente_id"),
                    payload.get("tipo_trabajo"),
                    payload.get("estado"),
                    payload.get("fecha_inicio"),
                    payload.get("fecha_fin"),
                    payload.get("sla_dias"),
                    payload.get("responsable"),
                    payload.get("importe"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
            audit("gestoria_trabajo", payload.get("cliente_id"), "crear", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_trabajos_update":
            trabajo_id = payload.get("id")
            if not trabajo_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = (
                "tipo_trabajo",
                "estado",
                "fecha_inicio",
                "fecha_fin",
                "sla_dias",
                "responsable",
                "importe",
                "notas",
            )
            updates = []
            values = []
            for field in allowed:
                if field in payload:
                    updates.append(f"{field} = ?")
                    values.append(payload.get(field))
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            conn.execute(
                f"""
                UPDATE gestoria_trabajos
                SET {", ".join(updates)}, updated_at = datetime(?)
                WHERE id = ?
                """,
                (*values, now, trabajo_id),
            )
            audit("gestoria_trabajo", trabajo_id, "actualizar", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_trabajos_delete":
            trabajo_id = payload.get("id")
            if not trabajo_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM gestoria_trabajos WHERE id = ?", (trabajo_id,))
            audit("gestoria_trabajo", trabajo_id, "eliminar", None, payload.get("usuario"))
        elif parsed.path == "/api/gestoria_docs":
            estado_doc = payload.get("estado")
            if payload.get("doc_key") or payload.get("doc_url"):
                estado_norm = normalize_lookup_text(estado_doc)
                if not estado_norm or estado_norm == "PENDIENTE":
                    estado_doc = "Recibido"
            conn.execute(
                """
                INSERT INTO gestoria_docs (
                  id, empresa_id, cliente_id, referencia_tipo, referencia_id,
                  nombre, tipo, fecha, estado, notas, doc_key, doc_url,
                  calidad_ocr, campos_ocr, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("cliente_id"),
                    payload.get("referencia_tipo"),
                    payload.get("referencia_id"),
                    payload.get("nombre"),
                    payload.get("tipo"),
                    payload.get("fecha"),
                    estado_doc,
                    payload.get("notas"),
                    payload.get("doc_key"),
                    payload.get("doc_url"),
                    payload.get("calidad_ocr"),
                    payload.get("campos_ocr"),
                    now,
                    now,
                ),
            )
            audit("gestoria_doc", payload.get("cliente_id"), "crear", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_docs_update":
            doc_id = payload.get("id")
            if not doc_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = (
                "nombre",
                "referencia_tipo",
                "referencia_id",
                "tipo",
                "fecha",
                "estado",
                "notas",
                "doc_key",
                "doc_url",
                "calidad_ocr",
                "campos_ocr",
            )
            updates = []
            values = []
            for field in allowed:
                if field in payload:
                    updates.append(f"{field} = ?")
                    values.append(payload.get(field))
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            conn.execute(
                f"""
                UPDATE gestoria_docs
                SET {", ".join(updates)}, updated_at = datetime(?)
                WHERE id = ?
                """,
                (*values, now, doc_id),
            )
            audit("gestoria_doc", doc_id, "actualizar", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_docs_delete":
            doc_id = payload.get("id")
            if not doc_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM gestoria_docs WHERE id = ?", (doc_id,))
            audit("gestoria_doc", doc_id, "eliminar", None, payload.get("usuario"))
        elif parsed.path == "/api/ocr_job":
            json_response(self, {"error": "Usa GET"}, status=405)
            return
        elif parsed.path == "/api/usuarios":
            nombre = payload.get("nombre")
            apellido = payload.get("apellido")
            usuario = payload.get("usuario")
            email = normalize_email(payload.get("email"))
            servicio = payload.get("servicio")
            password = payload.get("password")
            if not nombre:
                json_response(self, {"error": "nombre requerido"}, status=400)
                return
            if not apellido:
                json_response(self, {"error": "apellido requerido"}, status=400)
                return
            if not usuario:
                json_response(self, {"error": "usuario requerido"}, status=400)
                return
            if not email:
                json_response(self, {"error": "email requerido"}, status=400)
                return
            if not servicio:
                json_response(self, {"error": "servicio requerido"}, status=400)
                return
            password_hash = hash_password(password) if password else None
            conn.execute(
                """
                INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, password_hash, activo, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?))
                """,
                (
                    os.urandom(16).hex(),
                    nombre,
                    apellido,
                    usuario,
                    email,
                    servicio,
                    payload.get("rol"),
                    password_hash,
                    int(payload.get("activo") or 1),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/usuarios_invitar":
            user_id = str(payload.get("id") or "").strip()
            if not user_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute(
                """
                SELECT id, nombre, apellido, usuario, email, activo
                FROM usuarios
                WHERE id = ?
                LIMIT 1
                """,
                (user_id,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Usuario no encontrado"}, status=404)
                return
            if not row["activo"]:
                json_response(self, {"error": "Usuario inactivo"}, status=400)
                return
            email = normalize_email(row["email"] or "")
            if not email:
                json_response(self, {"error": "El usuario no tiene email válido"}, status=400)
                return
            token = secrets.token_urlsafe(32)
            expires_at = (datetime.now(timezone.utc) + timedelta(seconds=AUTH_INVITE_TTL_SECONDS)).isoformat()
            conn.execute(
                """
                UPDATE usuarios
                SET invite_token = ?, invite_expires_at = ?, invite_sent_at = datetime('now'), updated_at = datetime('now')
                WHERE id = ?
                """,
                (token, expires_at, user_id),
            )
            invite_link = f"{self._external_base_url()}/?activar_token={urllib.parse.quote(token)}"
            sent = False
            mail_error = None
            try:
                subject = "Activa tu acceso al CRM"
                full_name = " ".join(x for x in [row["nombre"] or "", row["apellido"] or ""] if x).strip()
                greeting = full_name or (row["usuario"] or "usuario")
                text_body = (
                    f"Hola {greeting},\n\n"
                    "Te han invitado a acceder al CRM.\n"
                    "Pulsa este enlace para validar tu acceso y definir tu contraseña:\n\n"
                    f"{invite_link}\n\n"
                    f"Este enlace caduca en {int(AUTH_INVITE_TTL_SECONDS/3600)} horas.\n"
                )
                html_body = (
                    f"<p>Hola {greeting},</p>"
                    "<p>Te han invitado a acceder al CRM.</p>"
                    "<p><a href=\"%s\">Pulsa aquí para validar tu acceso y definir tu contraseña</a></p>"
                    "<p>Si el botón no funciona, copia este enlace:</p>"
                    f"<p>{invite_link}</p>"
                ) % invite_link
                send_mail_smtp(subject, email, text_body, html_body=html_body)
                sent = True
            except Exception as exc:
                mail_error = str(exc)
            json_response(
                self,
                {
                    "ok": True,
                    "sent": sent,
                    "invite_link": invite_link,
                    "mail_error": mail_error,
                },
            )
            return
        elif parsed.path == "/api/usuarios_update":
            user_id = payload.get("id")
            if not user_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("nombre", "apellido", "usuario", "email", "servicio", "rol", "activo", "password")
            updates = []
            values = []
            for field in allowed:
                if field in payload:
                    if field == "password":
                        updates.append("password_hash = ?")
                        values.append(hash_password(payload.get("password")))
                        updates.append("invite_token = NULL")
                        updates.append("invite_expires_at = NULL")
                    else:
                        updates.append(f"{field} = ?")
                        values.append(payload.get(field))
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            conn.execute(
                f"""
                UPDATE usuarios
                SET {", ".join(updates)}, updated_at = datetime(?)
                WHERE id = ?
                """,
                (*values, now, user_id),
            )
        elif parsed.path == "/api/usuarios_delete":
            user_id = payload.get("id")
            if not user_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM usuarios WHERE id = ?", (user_id,))
        elif parsed.path == "/api/gestoria_contabilidad":
            seguro_id = (payload.get("seguro_id") or "").strip()
            poliza_numero = ""
            cliente_ids = parse_cliente_ids_payload(payload.get("cliente_ids_json"))
            if not cliente_ids:
                cliente_ids = parse_cliente_ids_payload(payload.get("cliente_id"))
            cliente_id = cliente_ids[0] if cliente_ids else None
            if seguro_id:
                poliza_numero, cliente_seguro_id = resolve_seguro_contabilidad_link(conn, seguro_id)
                if cliente_seguro_id and not cliente_ids:
                    cliente_ids = [cliente_seguro_id]
                    cliente_id = cliente_seguro_id
            cliente_ids_json = json.dumps(cliente_ids, ensure_ascii=False) if cliente_ids else None
            conn.execute(
                """
                INSERT INTO gestoria_contabilidad (
                  id, empresa_id, cliente_id, cliente_ids_json, seguro_id, poliza_numero, fecha, concepto, gestion, tipo, importe, notas,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    cliente_id,
                    cliente_ids_json,
                    seguro_id or None,
                    poliza_numero,
                    payload.get("fecha"),
                    payload.get("concepto"),
                    payload.get("gestion"),
                    payload.get("tipo"),
                    payload.get("importe"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/gestoria_contabilidad_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = (
                "fecha",
                "concepto",
                "gestion",
                "tipo",
                "importe",
                "notas",
                "cliente_id",
                "cliente_ids_json",
                "seguro_id",
            )
            updates = []
            values = []
            cliente_ids = None
            for field in allowed:
                if field in payload:
                    if field == "seguro_id":
                        updates.append(f"{field} = ?")
                        values.append((payload.get(field) or "").strip() or None)
                    elif field == "cliente_id":
                        cliente_ids = parse_cliente_ids_payload(payload.get(field))
                        updates.append("cliente_id = ?")
                        values.append(cliente_ids[0] if cliente_ids else None)
                        updates.append("cliente_ids_json = ?")
                        values.append(json.dumps(cliente_ids, ensure_ascii=False) if cliente_ids else None)
                    elif field == "cliente_ids_json":
                        cliente_ids = parse_cliente_ids_payload(payload.get(field))
                        updates.append("cliente_ids_json = ?")
                        values.append(json.dumps(cliente_ids, ensure_ascii=False) if cliente_ids else None)
                        updates.append("cliente_id = ?")
                        values.append(cliente_ids[0] if cliente_ids else None)
                    else:
                        updates.append(f"{field} = ?")
                        values.append(payload.get(field))
            if "seguro_id" in payload:
                seguro_id = (payload.get("seguro_id") or "").strip()
                poliza_numero = ""
                cliente_seguro_id = None
                if seguro_id:
                    poliza_numero, cliente_seguro_id = resolve_seguro_contabilidad_link(conn, seguro_id)
                updates.append("poliza_numero = ?")
                values.append(poliza_numero)
                if cliente_seguro_id and not cliente_ids:
                    cliente_ids = [cliente_seguro_id]
                    updates.append("cliente_ids_json = ?")
                    values.append(json.dumps(cliente_ids, ensure_ascii=False))
                    updates.append("cliente_id = ?")
                    values.append(cliente_seguro_id)
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            conn.execute(
                f"""
                UPDATE gestoria_contabilidad
                SET {", ".join(updates)}, updated_at = datetime(?)
                WHERE id = ?
                """,
                (*values, now, record_id),
            )
        elif parsed.path == "/api/gestoria_factura_ocr":
            cliente_id = (payload.get("cliente_id") or "").strip() or None
            tipo_factura = (payload.get("tipo_factura") or payload.get("tipo") or "").strip().lower()
            if tipo_factura not in ("compra", "venta"):
                tipo_factura = "compra"
            try:
                doc_bytes, mime, source_hint = decode_document_payload(payload)
            except ValueError as exc:
                json_response(self, {"error": str(exc)}, status=400)
                return
            tmp_path = None
            text = ""
            err_detail = ""
            method = "tesseract"
            try:
                suffix = ".pdf"
                if mime.startswith("image/"):
                    ext = mime.split("/", 1)[1] or "jpg"
                    suffix = f".{ext}"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_file:
                    tmp_file.write(doc_bytes)
                    tmp_path = tmp_file.name
                if mime.startswith("image/"):
                    if external_ocr_available():
                        text, err_detail = ocr_image_external(doc_bytes)
                        method = "vision" if text else "tesseract"
                    if not text:
                        text, err_detail = ocr_image_file(tmp_path)
                        method = "tesseract"
                    if docai_available():
                        doc_text, _doc_fields, doc_err = ocr_image_docai(doc_bytes, mime)
                        if doc_text and len(doc_text) > len(text or ""):
                            text = doc_text
                            method = "docai"
                        elif doc_err and not err_detail:
                            err_detail = doc_err
                else:
                    text, err_detail, method = extract_pdf_text(tmp_path)
                    if not text:
                        text, page_err = ocr_pdf_all_pages(tmp_path, use_external=external_ocr_available())
                        if text:
                            method = "ocr_all_pages"
                        elif page_err and not err_detail:
                            err_detail = page_err
                if not text:
                    json_response(self, {"error": err_detail or "No se pudo extraer texto de la factura"}, status=400)
                    return
                parsed_factura = parse_invoice_text(text)
                if not parsed_factura:
                    json_response(self, {"error": "No se pudieron extraer datos de factura"}, status=400)
                    return
                parsed_factura["tipo"] = tipo_factura
                for key_src, key_dst in (
                    ("numero", "numero"),
                    ("fecha", "fecha"),
                    ("nif", "nif"),
                    ("tercero", "tercero"),
                ):
                    incoming = str(payload.get(key_src) or "").strip()
                    if incoming:
                        parsed_factura[key_dst] = incoming
                for num_key in ("base_imponible", "cuota_iva", "cuota_irpf", "total", "iva_pct"):
                    incoming = payload.get(num_key)
                    if incoming not in (None, ""):
                        parsed_factura[num_key] = round(parse_decimal_eu(incoming), 2)
                if not parsed_factura.get("fecha"):
                    parsed_factura["fecha"] = datetime.now().strftime("%Y-%m-%d")
                third_type = "cliente" if tipo_factura == "venta" else ("proveedor" if parsed_factura.get("numero") else "acreedor")
                tercero_id, counterpart_account = ensure_gestoria_tercero(
                    conn,
                    empresa["id"],
                    parsed_factura.get("nif"),
                    parsed_factura.get("tercero"),
                    third_type,
                    now,
                )
                factura_id = os.urandom(16).hex()
                doc_key = (payload.get("s3_key") or "").strip()
                conn.execute(
                    """
                    INSERT INTO gestoria_facturas (
                      id, empresa_id, cliente_id, tercero_id, tipo, numero, fecha_emision, descripcion,
                      base_imponible, cuota_iva, cuota_irpf, total, iva_pct, estado_ocr, doc_key, raw_text,
                      created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        factura_id,
                        empresa["id"],
                        cliente_id,
                        tercero_id,
                        tipo_factura,
                        parsed_factura.get("numero"),
                        parsed_factura.get("fecha"),
                        parsed_factura.get("descripcion"),
                        parsed_factura.get("base_imponible") or 0.0,
                        parsed_factura.get("cuota_iva") or 0.0,
                        parsed_factura.get("cuota_irpf") or 0.0,
                        parsed_factura.get("total") or 0.0,
                        parsed_factura.get("iva_pct") or 0.0,
                        "ok",
                        doc_key,
                        parsed_factura.get("raw_text") or text,
                        now,
                        now,
                    ),
                )
                lines, total_debe, total_haber = build_invoice_asiento(parsed_factura, counterpart_account)
                asiento_id = os.urandom(16).hex()
                referencia = parsed_factura.get("numero") or factura_id
                concepto = parsed_factura.get("descripcion") or "Factura OCR"
                conn.execute(
                    """
                    INSERT INTO gestoria_asientos (
                      id, empresa_id, cliente_id, factura_id, fecha, concepto, diario, referencia,
                      total_debe, total_haber, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        asiento_id,
                        empresa["id"],
                        cliente_id,
                        factura_id,
                        parsed_factura.get("fecha"),
                        concepto,
                        "FACT",
                        referencia,
                        total_debe,
                        total_haber,
                        now,
                        now,
                    ),
                )
                for item in lines:
                    conn.execute(
                        """
                        INSERT INTO gestoria_asiento_lineas (
                          id, asiento_id, tercero_id, cuenta, descripcion, debe, haber,
                          impuesto_tipo, impuesto_pct, created_at, updated_at
                        ) VALUES (
                          ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                        )
                        """,
                        (
                            os.urandom(16).hex(),
                            asiento_id,
                            tercero_id,
                            item.get("cuenta"),
                            item.get("descripcion"),
                            item.get("debe") or 0.0,
                            item.get("haber") or 0.0,
                            item.get("impuesto"),
                            item.get("porcentaje"),
                            now,
                            now,
                        ),
                    )
                conn.execute(
                    """
                    INSERT INTO gestoria_contabilidad (
                      id, empresa_id, cliente_id, cliente_ids_json, fecha, concepto, gestion, tipo, importe, notas, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        empresa["id"],
                        cliente_id,
                        json.dumps([cliente_id], ensure_ascii=False) if cliente_id else None,
                        parsed_factura.get("fecha"),
                        concepto,
                        "Contable",
                        "Ingreso" if tipo_factura == "venta" else "Gasto",
                        parsed_factura.get("total") or 0.0,
                        f"Factura OCR {referencia} · asiento {asiento_id}",
                        now,
                        now,
                    ),
                )
                json_response(
                    self,
                    {
                        "ok": True,
                        "factura_id": factura_id,
                        "asiento_id": asiento_id,
                        "ocr_method": method,
                        "parsed": parsed_factura,
                        "lineas": lines,
                        "totales": {"debe": total_debe, "haber": total_haber},
                    },
                )
                return
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
        elif parsed.path == "/api/gestoria_contabilidad_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM gestoria_contabilidad WHERE id = ?", (record_id,))
        elif parsed.path == "/api/gestoria_conta_config":
            cliente_id = payload.get("cliente_id")
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            periodo = payload.get("periodo")
            fecha_inicio = payload.get("fecha_inicio")
            responsable = payload.get("responsable")
            existing = conn.execute(
                "SELECT id FROM gestoria_conta_config WHERE cliente_id = ?",
                (cliente_id,),
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE gestoria_conta_config
                    SET periodo = ?, fecha_inicio = ?, responsable = ?, updated_at = datetime(?)
                    WHERE cliente_id = ?
                    """,
                    (periodo, fecha_inicio, responsable, now, cliente_id),
                )
                record_id = existing["id"]
            else:
                record_id = os.urandom(16).hex()
                conn.execute(
                    """
                    INSERT INTO gestoria_conta_config (
                      id, cliente_id, periodo, fecha_inicio, responsable, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (record_id, cliente_id, periodo, fecha_inicio, responsable, now, now),
                )
            audit("gestoria_conta_config", record_id, "Guardar configuración contable", usuario=payload.get("usuario"))
        elif parsed.path == "/api/gestoria_conta_tasks_bulk":
            cliente_id = payload.get("cliente_id")
            periodo = payload.get("periodo")
            tareas = payload.get("tareas", [])
            if not cliente_id or not periodo or not isinstance(tareas, list):
                json_response(self, {"error": "cliente_id, periodo y tareas requeridos"}, status=400)
                return
            conn.execute(
                "DELETE FROM gestoria_conta_tasks WHERE cliente_id = ? AND periodo = ?",
                (cliente_id, periodo),
            )
            for tarea in tareas:
                conn.execute(
                    """
                    INSERT INTO gestoria_conta_tasks (
                      id, cliente_id, periodo, tarea, estado, fecha_limite, responsable, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        cliente_id,
                        periodo,
                        tarea.get("tarea"),
                        tarea.get("estado") or "Pendiente",
                        tarea.get("fecha_limite"),
                        tarea.get("responsable"),
                        now,
                        now,
                    ),
                )
            audit(
                "gestoria_conta_tasks",
                cliente_id,
                f"Crear checklist contable ({periodo})",
                usuario=payload.get("usuario"),
            )
        elif parsed.path == "/api/gestoria_conta_task_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            fields = ["tarea", "estado", "fecha_limite", "responsable"]
            updates = {f: payload.get(f) for f in fields if f in payload}
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{k} = ?" for k in updates])
            values = list(updates.values())
            values.append(now)
            values.append(record_id)
            conn.execute(
                f"UPDATE gestoria_conta_tasks SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            audit("gestoria_conta_tasks", record_id, "Actualizar tarea contable", usuario=payload.get("usuario"))
        elif parsed.path == "/api/seguros_ocr":
            try:
                result = process_seguros_ocr(payload, conn)
                json_response(self, result)
                return
            except Exception as exc:
                json_response(self, {"error": "No se pudo procesar el PDF.", "detail": str(exc)}, status=400)
                return
        elif parsed.path == "/api/seguros_ocr_async":
            try:
                if not (payload.get("file_base64") or payload.get("data") or payload.get("s3_key")):
                    json_response(self, {"error": "Archivo requerido"}, status=400)
                    return
                job_id = enqueue_ocr_job(Handler.ocr_db_path, "seguros", payload)
                json_response(self, {"job_id": job_id})
                return
            except Exception as exc:
                json_response(self, {"error": "No se pudo encolar OCR", "detail": str(exc)}, status=400)
                return
        elif parsed.path == "/api/fin_asesoramiento_ocr":
            data_uri = payload.get("file_base64") or payload.get("data")
            if not data_uri:
                json_response(self, {"error": "Archivo requerido"}, status=400)
                return
            use_external = bool(payload.get("use_external"))
            ocr_mode = (payload.get("ocr_mode") or "").strip().lower()
            if ocr_mode == "handwritten":
                use_external = True
                ocr_mode = "hybrid"
            if external_ocr_available() and not use_external:
                use_external = True
            if not ocr_mode:
                if external_ocr_available() and docai_available():
                    ocr_mode = "hybrid"
                elif docai_available():
                    ocr_mode = "docai"
            if external_ocr_available() and not use_external:
                use_external = True
            if not ocr_mode:
                if external_ocr_available() and docai_available():
                    ocr_mode = "hybrid"
                elif docai_available():
                    ocr_mode = "docai"
            mime = ""
            if "," in data_uri:
                header, data_uri = data_uri.split(",", 1)
                if header.startswith("data:") and ";base64" in header:
                    mime = header.split(":", 1)[1].split(";", 1)[0]
            try:
                pdf_bytes = base64.b64decode(data_uri)
            except Exception:
                json_response(self, {"error": "Base64 invalido"}, status=400)
                return
            tmp_path = None
            external_error = ""
            external_used = False
            try:
                suffix = ".pdf"
                if mime.startswith("image/"):
                    ext = mime.split("/", 1)[1]
                    suffix = f".{ext}"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp_file:
                    tmp_file.write(pdf_bytes)
                    tmp_path = tmp_file.name
                if mime.startswith("image/"):
                    text = ""
                    err_detail = ""
                    method = "tesseract"
                    if ocr_mode == "docai":
                        text, doc_fields, err_detail = ocr_image_docai(pdf_bytes, mime)
                        method = "docai"
                        if doc_fields:
                            fields = map_docai_fields(doc_fields)
                        else:
                            fields = parse_asesoramiento_text(text) if text else {}
                    elif ocr_mode == "hybrid":
                        vision_text = ""
                        vision_err = ""
                        if use_external:
                            vision_text, vision_err = ocr_image_external(pdf_bytes)
                            external_error = vision_err or ""
                            external_used = bool(vision_text)
                        if not vision_text:
                            vision_text, vision_err = ocr_image_file(tmp_path)
                        vision_fields = parse_asesoramiento_text(vision_text) if vision_text else {}
                        template_fields = parse_asesoramiento_template_image(tmp_path)
                        vision_fields = merge_fields(vision_fields, template_fields)
                        doc_text, doc_fields, doc_err = ocr_image_docai(pdf_bytes, mime)
                        if doc_err:
                            err_detail = doc_err
                        doc_mapped = map_docai_fields(doc_fields)
                        fields = merge_fields(vision_fields, doc_mapped)
                        text = "\n".join([t for t in (vision_text, doc_text) if t])
                        method = "hybrid"
                    elif use_external:
                        text, err_detail = ocr_image_external(pdf_bytes)
                        external_error = err_detail or ""
                        method = "vision"
                        external_used = bool(text)
                    if not text:
                        text, err_detail = ocr_image_file(tmp_path)
                        method = "tesseract"
                    if "template_fields" not in locals():
                        template_fields = parse_asesoramiento_template_image(tmp_path)
                    if "fields" not in locals():
                        fields = parse_asesoramiento_text(text) if text else {}
                    if openai_available() and text:
                        ai_fields, ai_err = call_openai_extract_fin(text)
                        if ai_fields:
                            fields = merge_many_fields(fields, ai_fields)
                    for key, value in template_fields.items():
                        if not str(fields.get(key, "") or "").strip() and str(value or "").strip():
                            fields[key] = value
                    if openai_available() and text:
                        ai_fields, ai_err = call_openai_extract_fin(text)
                        if ai_fields:
                            fields = merge_many_fields(fields, ai_fields)
                else:
                    text, err_detail, method = extract_pdf_text(tmp_path)
                    if not text:
                        json_response(
                            self,
                            {
                                "error": "No se pudo extraer texto.",
                                "detail": err_detail or "Verifica tesseract y spa.traineddata.",
                                "language": detect_ocr_lang(),
                                "method": method,
                            },
                            status=400,
                        )
                        return
                    fields = parse_asesoramiento_text(text)
                    template_fields = parse_asesoramiento_template(tmp_path)
                    for key, value in template_fields.items():
                        if not str(fields.get(key, "") or "").strip() and str(value or "").strip():
                            fields[key] = value
                    filled = sum(1 for value in fields.values() if str(value or "").strip())
                    if filled < 8 or use_external:
                        ocr_text, ocr_err = ocr_pdf_all_pages(tmp_path, use_external=use_external)
                        if ocr_text:
                            ocr_fields = parse_asesoramiento_text(ocr_text)
                            for key, value in ocr_fields.items():
                                if not str(fields.get(key, "") or "").strip() and str(value or "").strip():
                                    fields[key] = value
                            text = ocr_text
                            method = "vision" if use_external else "tesseract"
                            external_used = external_used or bool(use_external)
                        elif ocr_err:
                            err_detail = ocr_err
                    doc_text = ""
                    if docai_available() and (use_external or filled < 10):
                        doc_text, doc_fields, doc_err = ocr_image_docai(pdf_bytes, "application/pdf")
                        if doc_err and not err_detail:
                            err_detail = doc_err
                        doc_mapped = map_docai_fields(doc_fields)
                        for key, value in doc_mapped.items():
                            if not str(fields.get(key, "") or "").strip() and str(value or "").strip():
                                fields[key] = value
                    if doc_text:
                        text = text or doc_text
                        method = "hybrid" if use_external else "docai"
                        external_used = external_used or bool(doc_text)
                    if openai_available() and text:
                        ai_fields, ai_err = call_openai_extract_fin(text)
                        if ai_fields:
                            fields = merge_many_fields(fields, ai_fields)
                if not text and not any(str(value or "").strip() for value in fields.values()):
                    json_response(
                        self,
                        {
                            "error": "No se pudo extraer texto.",
                            "detail": err_detail or "Verifica tesseract y spa.traineddata.",
                            "language": detect_ocr_lang(),
                            "method": method,
                        },
                        status=400,
                    )
                    return
                fin_quality = compute_fin_quality(fields)
                json_response(
                    self,
                    {
                        "fields": fields,
                        "text": text,
                        "language": detect_ocr_lang(),
                        "method": method,
                        "external_error": external_error,
                        "external_used": external_used,
                        "ocr_quality": fin_quality,
                    },
                )
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        elif parsed.path == "/api/fin_asesoramiento_ocr_guided":
            sections = payload.get("sections") or {}
            use_external = bool(payload.get("use_external"))
            ocr_mode = (payload.get("ocr_mode") or "").strip().lower()
            if ocr_mode == "handwritten":
                use_external = True
            allowed = {
                "header": "Cabecera",
                "cliente1": "Cliente 1",
                "cliente2": "Cliente 2",
                "resumen": "Resumen",
            }
            if not any(sections.get(key) for key in allowed):
                json_response(self, {"error": "Archivos requeridos"}, status=400)
                return
            fields = {}
            texts = []
            external_error = ""
            external_used = False
            for key in ("header", "cliente1", "cliente2", "resumen"):
                data_uri = sections.get(key)
                if not data_uri:
                    continue
                mime = ""
                if "," in data_uri:
                    header, data_uri = data_uri.split(",", 1)
                    if header.startswith("data:") and ";base64" in header:
                        mime = header.split(":", 1)[1].split(";", 1)[0]
                if not mime.startswith("image/"):
                    json_response(self, {"error": f"{allowed[key]} debe ser imagen"}, status=400)
                    return
                try:
                    image_bytes = base64.b64decode(data_uri)
                except Exception:
                    json_response(self, {"error": f"{allowed[key]} base64 invalido"}, status=400)
                    return
                text = ""
                err_detail = ""
                method = "tesseract"
                if use_external:
                    text, err_detail = ocr_image_external(image_bytes)
                    external_error = err_detail or external_error
                    method = "vision"
                    external_used = external_used or bool(text)
                if not text:
                    tmp_path = None
                    try:
                        ext = mime.split("/", 1)[1]
                        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp_file:
                            tmp_file.write(image_bytes)
                            tmp_path = tmp_file.name
                        text, err_detail = ocr_image_file(tmp_path)
                        method = "tesseract"
                    finally:
                        if tmp_path and os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                if not text:
                    continue
                texts.append(f"[{allowed[key]}]\n{text}")
                if key == "header":
                    header_fields = parse_asesoramiento_text(text)
                    for field in ("inmobiliaria_asesor", "fecha", "asesor"):
                        if header_fields.get(field) and not fields.get(field):
                            fields[field] = header_fields.get(field)
                elif key == "cliente1":
                    data1 = parse_asesoramiento_block(text)
                    for field, target in (
                        ("nombre", "cliente1_nombre"),
                        ("dni", "cliente1_dni"),
                        ("telefono", "cliente1_telefono"),
                        ("email", "cliente1_email"),
                        ("fecha_nacimiento", "cliente1_fecha_nacimiento"),
                        ("estado_civil", "cliente1_estado_civil"),
                        ("hijos", "cliente1_hijos"),
                        ("profesion", "cliente1_profesion"),
                        ("tipo_contrato", "cliente1_tipo_contrato"),
                        ("ingresos", "cliente1_ingresos"),
                        ("patrimonio", "cliente1_patrimonio"),
                        ("prestamos", "cliente1_prestamos"),
                    ):
                        if data1.get(field) and not fields.get(target):
                            fields[target] = data1.get(field)
                elif key == "cliente2":
                    data2 = parse_asesoramiento_block(text)
                    for field, target in (
                        ("nombre", "cliente2_nombre"),
                        ("dni", "cliente2_dni"),
                        ("telefono", "cliente2_telefono"),
                        ("email", "cliente2_email"),
                        ("fecha_nacimiento", "cliente2_fecha_nacimiento"),
                        ("estado_civil", "cliente2_estado_civil"),
                        ("hijos", "cliente2_hijos"),
                        ("profesion", "cliente2_profesion"),
                        ("tipo_contrato", "cliente2_tipo_contrato"),
                        ("ingresos", "cliente2_ingresos"),
                        ("patrimonio", "cliente2_patrimonio"),
                        ("prestamos", "cliente2_prestamos"),
                    ):
                        if data2.get(field) and not fields.get(target):
                            fields[target] = data2.get(field)
                elif key == "resumen":
                    resumen_fields = parse_asesoramiento_text(text)
                    for field in ("ingresos_conjuntos", "entidades_financieras", "avalistas", "aportacion_cv"):
                        if resumen_fields.get(field) and not fields.get(field):
                            fields[field] = resumen_fields.get(field)
                if not any(str(value or "").strip() for value in fields.values()):
                    json_response(
                        self,
                        {
                            "error": "No se pudieron detectar campos.",
                            "detail": "Sube recortes más cercanos a cada bloque.",
                        },
                        status=400,
                    )
                    return
                if openai_available() and texts:
                    ai_fields, ai_err = call_openai_extract_fin("\n\n".join(texts))
                    if ai_fields:
                        fields = merge_many_fields(fields, ai_fields)
                fin_quality = compute_fin_quality(fields)
            json_response(
                self,
                {
                    "fields": fields,
                    "text": "\n\n".join(texts).strip(),
                    "language": detect_ocr_lang(),
                    "method": "guided",
                    "external_error": external_error,
                    "external_used": external_used,
                    "ocr_quality": fin_quality,
                },
            )
        elif parsed.path == "/api/fin_asesoramiento_ocr_auto":
            data_uri = payload.get("file_base64") or payload.get("data")
            if not data_uri:
                json_response(self, {"error": "Archivo requerido"}, status=400)
                return
            use_external = bool(payload.get("use_external"))
            ocr_mode = (payload.get("ocr_mode") or "").strip().lower()
            if ocr_mode == "handwritten":
                use_external = True
                ocr_mode = "hybrid"
            external_error = ""
            external_used = False
            mime = ""
            if "," in data_uri:
                header, data_uri = data_uri.split(",", 1)
                if header.startswith("data:") and ";base64" in header:
                    mime = header.split(":", 1)[1].split(";", 1)[0]
            if not mime.startswith("image/"):
                json_response(self, {"error": "La imagen debe ser JPG o PNG."}, status=400)
                return
            try:
                image_bytes = base64.b64decode(data_uri)
            except Exception:
                json_response(self, {"error": "Base64 invalido"}, status=400)
                return
            tmp_path = None
            try:
                ext = mime.split("/", 1)[1]
                with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp_file:
                    tmp_file.write(image_bytes)
                    tmp_path = tmp_file.name
                if ocr_mode == "docai":
                    text, doc_fields, err_detail = ocr_image_docai(image_bytes, mime)
                    fields = map_docai_fields(doc_fields) if doc_fields else parse_asesoramiento_text(text)
                    fin_quality = compute_fin_quality(fields)
                    json_response(
                        self,
                        {
                            "fields": fields,
                            "text": text,
                            "language": detect_ocr_lang(),
                            "method": "docai",
                            "external_error": err_detail,
                            "external_used": True if text else False,
                            "ocr_quality": fin_quality,
                        },
                    )
                    return
                if ocr_mode == "hybrid":
                    vision_text = ""
                    vision_err = ""
                    if use_external:
                        vision_text, vision_err = ocr_image_external(image_bytes)
                        external_error = vision_err or ""
                        external_used = bool(vision_text)
                    if not vision_text:
                        vision_text, _ = ocr_image_file(tmp_path)
                    doc_text, doc_fields, doc_err = ocr_image_docai(image_bytes, mime)
                    if doc_err:
                        external_error = doc_err
                    fields = map_docai_fields(doc_fields)
                    vision_fields = parse_asesoramiento_text(vision_text) if vision_text else {}
                    fields = merge_fields(vision_fields, fields)
                    fin_quality = compute_fin_quality(fields)
                    json_response(
                        self,
                        {
                            "fields": fields,
                            "text": "\n".join([t for t in (vision_text, doc_text) if t]),
                            "language": detect_ocr_lang(),
                            "method": "hybrid",
                            "external_error": external_error,
                            "external_used": external_used,
                            "ocr_quality": fin_quality,
                        },
                    )
                    return
                size = get_image_size(tmp_path)
                if not size:
                    json_response(self, {"error": "No se pudo leer la imagen."}, status=400)
                    return
                width, height = size
                boxes = asesoramiento_image_boxes(width, height)
                fields = {}
                texts = []
                external_error = ""
                external_used = False
                for key in ("header", "cliente1", "cliente2", "resumen"):
                    block_text, err = ocr_best_block(tmp_path, boxes[key], use_external)
                    if err:
                        external_error = external_error or err
                    if not block_text:
                        continue
                    texts.append(f"[{key}]\n{block_text}")
                    if use_external and block_text:
                        external_used = True
                    if key == "header":
                        header_fields = parse_asesoramiento_text(block_text)
                        for field in ("inmobiliaria_asesor", "fecha", "asesor"):
                            if header_fields.get(field) and not fields.get(field):
                                fields[field] = header_fields.get(field)
                    elif key == "cliente1":
                        data1 = parse_asesoramiento_block(block_text)
                        for field, target in (
                            ("nombre", "cliente1_nombre"),
                            ("dni", "cliente1_dni"),
                            ("telefono", "cliente1_telefono"),
                            ("email", "cliente1_email"),
                            ("fecha_nacimiento", "cliente1_fecha_nacimiento"),
                            ("estado_civil", "cliente1_estado_civil"),
                            ("hijos", "cliente1_hijos"),
                            ("profesion", "cliente1_profesion"),
                            ("tipo_contrato", "cliente1_tipo_contrato"),
                            ("ingresos", "cliente1_ingresos"),
                            ("patrimonio", "cliente1_patrimonio"),
                            ("prestamos", "cliente1_prestamos"),
                        ):
                            if data1.get(field) and not fields.get(target):
                                fields[target] = data1.get(field)
                    elif key == "cliente2":
                        data2 = parse_asesoramiento_block(block_text)
                        for field, target in (
                            ("nombre", "cliente2_nombre"),
                            ("dni", "cliente2_dni"),
                            ("telefono", "cliente2_telefono"),
                            ("email", "cliente2_email"),
                            ("fecha_nacimiento", "cliente2_fecha_nacimiento"),
                            ("estado_civil", "cliente2_estado_civil"),
                            ("hijos", "cliente2_hijos"),
                            ("profesion", "cliente2_profesion"),
                            ("tipo_contrato", "cliente2_tipo_contrato"),
                            ("ingresos", "cliente2_ingresos"),
                            ("patrimonio", "cliente2_patrimonio"),
                            ("prestamos", "cliente2_prestamos"),
                        ):
                            if data2.get(field) and not fields.get(target):
                                fields[target] = data2.get(field)
                    elif key == "resumen":
                        resumen_fields = parse_asesoramiento_text(block_text)
                        for field in ("ingresos_conjuntos", "entidades_financieras", "avalistas", "aportacion_cv"):
                            if resumen_fields.get(field) and not fields.get(field):
                                fields[field] = resumen_fields.get(field)
                if not any(str(value or "").strip() for value in fields.values()):
                    json_response(
                        self,
                        {"error": "No se pudieron detectar campos.", "detail": "Prueba con recortes manuales."},
                        status=400,
                    )
                    return
                fin_quality = compute_fin_quality(fields)
                json_response(
                    self,
                    {
                        "fields": fields,
                        "text": "\n\n".join(texts).strip(),
                        "language": detect_ocr_lang(),
                        "method": "auto",
                        "external_error": external_error,
                        "external_used": external_used,
                        "ocr_quality": fin_quality,
                    },
                )
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        elif parsed.path == "/api/seguros":
            poliza_id = os.urandom(16).hex()
            cliente_id = payload.get("cliente_id")
            if cliente_id:
                exists = conn.execute(
                    "SELECT id FROM clientes WHERE id = ?",
                    (cliente_id,),
                ).fetchone()
                if not exists:
                    cliente_id = None
            if not cliente_id:
                cliente_id = ensure_cliente_for_seguro(
                    conn,
                    empresa["id"],
                    payload.get("tomador"),
                    payload.get("nif") or payload.get("dni"),
                    now,
                    {
                        "telefono": payload.get("telefono"),
                        "email": payload.get("email"),
                        "fecha_nacimiento": payload.get("fecha_nacimiento"),
                        "direccion": payload.get("direccion"),
                    },
                )
            else:
                ensure_cliente_servicio_link(conn, cliente_id, empresa["id"], "seguros", now)
            poliza_key = payload.get("poliza_key") or ""
            poliza_url = payload.get("poliza_url") or ""
            ocr_quality = payload.get("ocr_quality") or {}
            calidad_ocr = ocr_quality.get("calidad") if isinstance(ocr_quality, dict) else payload.get("calidad_ocr")
            campos_ocr = ""
            if isinstance(ocr_quality, dict):
                campos_list = ocr_quality.get("campos") or []
                if isinstance(campos_list, list):
                    campos_ocr = ",".join(campos_list)
                else:
                    campos_ocr = str(campos_list)
            else:
                campos_ocr = payload.get("campos_ocr") or ""
            # Deduplicación suave: si existe póliza con mismo nº y compañía, actualizar campos vacíos
            dup_id = find_existing_seguro_id(
                conn,
                empresa["id"],
                payload.get("poliza_numero"),
                payload.get("compania"),
            )

            if dup_id:
                # Enriquecer la póliza existente con campos vacíos
                row = conn.execute("SELECT * FROM seguros WHERE id = ?", (dup_id,)).fetchone()
                updates = {}
                for key in (
                    "cliente_id",
                    "estado",
                    "fecha_efecto",
                    "fecha_vencimiento",
                    "estado_renovacion",
                    "renovacion_fecha",
                    "nueva_poliza_ref",
                    "poliza_numero",
                    "poliza_key",
                    "poliza_url",
                    "tomador",
                    "compania",
                    "ramo",
                    "tipo_vigencia",
                    "prima_neta",
                    "prima_total",
                    "comision",
                    "produccion",
                    "colaborador",
                ):
                    incoming = payload.get(key)
                    if key == "ramo":
                        incoming = canonicalize_ramo(incoming)
                    if key == "tipo_vigencia":
                        incoming = infer_tipo_vigencia(payload.get("ramo"), incoming)
                    if key == "cliente_id" and not str(incoming or "").strip():
                        incoming = cliente_id
                    if incoming in (None, ""):
                        continue
                    current = row[key] if key in row.keys() else None
                    if key == "cliente_id":
                        # En altas manuales, si ya existe la póliza, priorizamos el vínculo
                        # explícito al cliente actual para evitar "alta correcta" sin reflejo en ficha.
                        if str(incoming).strip() and str(current or "").strip() != str(incoming).strip():
                            updates[key] = incoming
                        continue
                    if current is None or str(current).strip() == "":
                        updates[key] = incoming
                if updates:
                    set_clause = ", ".join([f"{key} = ?" for key in updates])
                    values = list(updates.values()) + [now, dup_id]
                    conn.execute(
                        f"UPDATE seguros SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                        values,
                    )
                poliza_id = dup_id
                row = conn.execute("SELECT * FROM seguros WHERE id = ?", (poliza_id,)).fetchone()
                if row and row["cliente_id"]:
                    ensure_cliente_servicio_link(conn, row["cliente_id"], row["empresa_id"], "seguros", now)
            else:
                conn.execute(
                    """
                    INSERT INTO seguros (
                      id, empresa_id, cliente_id, mes_creacion, fecha_efecto, fecha_vencimiento,
                      tomador, compania, ramo, poliza_numero, prima_neta,
                      prima_total, comision, produccion, colaborador, estado,
                      estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                      poliza_key, poliza_url, estado_poliza, version_grupo, tipo_vigencia,
                      created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        poliza_id,
                        empresa["id"],
                        cliente_id,
                        payload.get("mes_creacion"),
                        payload.get("fecha_efecto"),
                        payload.get("fecha_vencimiento"),
                        payload.get("tomador"),
                        payload.get("compania"),
                        canonicalize_ramo(payload.get("ramo")),
                        payload.get("poliza_numero"),
                        payload.get("prima_neta"),
                        payload.get("prima_total"),
                        payload.get("comision"),
                        payload.get("produccion"),
                        payload.get("colaborador"),
                        payload.get("estado"),
                        payload.get("estado_renovacion"),
                        payload.get("renovacion_fecha"),
                        payload.get("nueva_poliza_ref"),
                        payload.get("poliza_key"),
                        payload.get("poliza_url"),
                        "activa",
                        poliza_id,
                        infer_tipo_vigencia(payload.get("ramo"), payload.get("tipo_vigencia")),
                        now,
                        now,
                    ),
                )
            doc_id = None
            contabilidad_id = None
            poliza_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (poliza_id,)).fetchone()
            if poliza_row:
                doc_id = ensure_seguro_doc_link(conn, poliza_row, now, calidad_ocr=calidad_ocr, campos_ocr=campos_ocr)
                log_seguro_event(conn, poliza_row, "alta", now, payload={"origen": "api/seguros"})
                contabilidad_id = upsert_seguro_comision_contabilidad(conn, poliza_row, now, movimiento="emision")
            # Crear acción si faltan campos obligatorios
            missing = []
            for key in ("tomador", "poliza_numero", "compania", "fecha_efecto"):
                if poliza_row and not str(poliza_row[key] or "").strip():
                    missing.append(key)
            if missing:
                notas = f"Completar datos póliza ({', '.join(missing)}). Poliza ID: {poliza_id}"
                if poliza_row and poliza_row["poliza_numero"]:
                    notas = f"Completar datos póliza {poliza_row['poliza_numero']} ({', '.join(missing)}). Poliza ID: {poliza_id}"
                exists = conn.execute(
                    """
                    SELECT id FROM acciones
                    WHERE servicio = 'Seguros'
                      AND cliente_id = ?
                      AND tipo = 'Completar datos póliza'
                      AND notas LIKE ?
                    LIMIT 1
                    """,
                    (cliente_id, f"%{poliza_id}%"),
                ).fetchone()
                if not exists:
                    today = datetime.now().strftime("%Y-%m-%d")
                    conn.execute(
                        """
                        INSERT INTO acciones (
                          id, empresa_id, servicio, cliente_id, inmueble_id, cliente_nombre,
                          fecha, hora, tipo, responsable, estado, notas, recordatorio_min, created_at, updated_at
                        ) VALUES (
                          ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                        )
                        """,
                        (
                            os.urandom(16).hex(),
                            empresa["id"],
                            "Seguros",
                            cliente_id,
                            None,
                            poliza_row["tomador"] if poliza_row else "",
                            today,
                            None,
                            "Completar datos póliza",
                            None,
                            "Pendiente",
                            notas,
                            None,
                            now,
                            now,
                        ),
                    )
            conn.commit()
            json_response(
                self,
                {
                    "ok": True,
                    "id": poliza_id,
                    "cliente_id": cliente_id,
                    "doc_id": doc_id,
                    "contabilidad_id": contabilidad_id,
                    "ocr_quality": ocr_quality,
                    "duplicate_of": dup_id,
                },
            )
            return
        elif parsed.path == "/api/fin_asesoramientos":
            cliente1_id = ensure_cliente_for_financiacion(
                conn,
                empresa["id"],
                payload.get("cliente1_nombre"),
                payload.get("cliente1_dni"),
                now,
                {
                    "telefono": payload.get("cliente1_telefono"),
                    "email": payload.get("cliente1_email"),
                    "fecha_nacimiento": payload.get("cliente1_fecha_nacimiento"),
                },
            )
            cliente2_id = ensure_cliente_for_financiacion(
                conn,
                empresa["id"],
                payload.get("cliente2_nombre"),
                payload.get("cliente2_dni"),
                now,
                {
                    "telefono": payload.get("cliente2_telefono"),
                    "email": payload.get("cliente2_email"),
                    "fecha_nacimiento": payload.get("cliente2_fecha_nacimiento"),
                },
            )
            dup_id = None
            dni1 = normalize_fin_nif(payload.get("cliente1_dni"))
            dni2 = normalize_fin_nif(payload.get("cliente2_dni"))
            nombre1 = normalize_person_name(payload.get("cliente1_nombre")).lower()
            telefono1 = normalize_phone(payload.get("cliente1_telefono"))
            fecha = (payload.get("fecha") or "").strip()
            candidates = conn.execute(
                "SELECT * FROM asesoramientos_financiacion WHERE empresa_id = ?",
                (empresa["id"],),
            ).fetchall()
            for row in candidates:
                score = 0
                row_dni1 = normalize_fin_nif(row["cliente1_dni"])
                row_dni2 = normalize_fin_nif(row["cliente2_dni"])
                if dni1 and (dni1 == row_dni1 or dni1 == row_dni2):
                    score += 3
                if dni2 and (dni2 == row_dni2 or dni2 == row_dni1):
                    score += 2
                if fecha and row["fecha"] and row["fecha"] == fecha:
                    score += 1
                if nombre1 and normalize_person_name(row["cliente1_nombre"]).lower() == nombre1:
                    score += 1
                if telefono1 and normalize_phone(row["cliente1_telefono"]) == telefono1:
                    score += 1
                if score >= 3:
                    dup_id = row["id"]
                    break

            allowed_fields = (
                "origen",
                "inmobiliaria_asesor",
                "asesor",
                "fecha",
                "estado",
                "cliente1_id",
                "cliente1_nombre",
                "cliente1_dni",
                "cliente1_telefono",
                "cliente1_email",
                "cliente1_fecha_nacimiento",
                "cliente1_estado_civil",
                "cliente1_regimen",
                "cliente1_hijos",
                "cliente1_profesion",
                "cliente1_tipo_contrato",
                "cliente1_tiempo_contrato",
                "cliente1_ingresos",
                "cliente1_patrimonio",
                "cliente1_prestamos",
                "cliente1_prestamo_activo",
                "cliente1_prestamo_entidad",
                "cliente1_prestamo_resto",
                "cliente2_id",
                "cliente2_nombre",
                "cliente2_dni",
                "cliente2_telefono",
                "cliente2_email",
                "cliente2_fecha_nacimiento",
                "cliente2_estado_civil",
                "cliente2_regimen",
                "cliente2_hijos",
                "cliente2_profesion",
                "cliente2_tipo_contrato",
                "cliente2_tiempo_contrato",
                "cliente2_ingresos",
                "cliente2_patrimonio",
                "cliente2_prestamos",
                "cliente2_prestamo_activo",
                "cliente2_prestamo_entidad",
                "cliente2_prestamo_resto",
                "ingresos_conjuntos",
                "entidades_financieras",
                "avalistas",
                "aportacion_cv",
                "notas",
                "notas_ocr",
                "calidad_ocr",
                "campos_ocr",
            )
            if dup_id:
                row = conn.execute(
                    "SELECT * FROM asesoramientos_financiacion WHERE id = ?",
                    (dup_id,),
                ).fetchone()
                updates = {}
                for key in allowed_fields:
                    incoming = payload.get(key)
                    if incoming in (None, ""):
                        continue
                    current = row[key] if key in row.keys() else None
                    if current is None or str(current).strip() == "":
                        updates[key] = incoming
                if cliente1_id:
                    updates["cliente1_id"] = cliente1_id
                if cliente2_id:
                    updates["cliente2_id"] = cliente2_id
                if updates:
                    set_clause = ", ".join([f"{key} = ?" for key in updates])
                    values = list(updates.values()) + [now, dup_id]
                    conn.execute(
                        f"UPDATE asesoramientos_financiacion SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                        values,
                    )
                asesoramiento_id = dup_id
            else:
                asesoramiento_id = os.urandom(16).hex()
                conn.execute(
                    """
                    INSERT INTO asesoramientos_financiacion (
                      id, empresa_id, origen, inmobiliaria_asesor, asesor, fecha, estado,
                      cliente1_id, cliente1_nombre, cliente1_dni, cliente1_telefono, cliente1_email,
                      cliente1_fecha_nacimiento, cliente1_estado_civil, cliente1_regimen, cliente1_hijos, cliente1_profesion,
                      cliente1_tipo_contrato, cliente1_tiempo_contrato, cliente1_ingresos, cliente1_patrimonio, cliente1_prestamos,
                      cliente1_prestamo_activo, cliente1_prestamo_entidad, cliente1_prestamo_resto,
                      cliente2_id, cliente2_nombre, cliente2_dni, cliente2_telefono, cliente2_email,
                      cliente2_fecha_nacimiento, cliente2_estado_civil, cliente2_regimen, cliente2_hijos, cliente2_profesion,
                      cliente2_tipo_contrato, cliente2_tiempo_contrato, cliente2_ingresos, cliente2_patrimonio, cliente2_prestamos,
                      cliente2_prestamo_activo, cliente2_prestamo_entidad, cliente2_prestamo_resto,
                      ingresos_conjuntos, entidades_financieras, avalistas, aportacion_cv,
                      notas, notas_ocr, calidad_ocr, campos_ocr, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        asesoramiento_id,
                        empresa["id"],
                        payload.get("origen"),
                        payload.get("inmobiliaria_asesor"),
                        payload.get("asesor"),
                        payload.get("fecha"),
                        payload.get("estado") or "En estudio",
                        cliente1_id,
                        payload.get("cliente1_nombre"),
                        payload.get("cliente1_dni"),
                        payload.get("cliente1_telefono"),
                        payload.get("cliente1_email"),
                        payload.get("cliente1_fecha_nacimiento"),
                        payload.get("cliente1_estado_civil"),
                        payload.get("cliente1_regimen"),
                        payload.get("cliente1_hijos"),
                        payload.get("cliente1_profesion"),
                        payload.get("cliente1_tipo_contrato"),
                        payload.get("cliente1_tiempo_contrato"),
                        payload.get("cliente1_ingresos"),
                        payload.get("cliente1_patrimonio"),
                        payload.get("cliente1_prestamos"),
                        payload.get("cliente1_prestamo_activo"),
                        payload.get("cliente1_prestamo_entidad"),
                        payload.get("cliente1_prestamo_resto"),
                        cliente2_id,
                        payload.get("cliente2_nombre"),
                        payload.get("cliente2_dni"),
                        payload.get("cliente2_telefono"),
                        payload.get("cliente2_email"),
                        payload.get("cliente2_fecha_nacimiento"),
                        payload.get("cliente2_estado_civil"),
                        payload.get("cliente2_regimen"),
                        payload.get("cliente2_hijos"),
                        payload.get("cliente2_profesion"),
                        payload.get("cliente2_tipo_contrato"),
                        payload.get("cliente2_tiempo_contrato"),
                        payload.get("cliente2_ingresos"),
                        payload.get("cliente2_patrimonio"),
                        payload.get("cliente2_prestamos"),
                        payload.get("cliente2_prestamo_activo"),
                        payload.get("cliente2_prestamo_entidad"),
                        payload.get("cliente2_prestamo_resto"),
                        payload.get("ingresos_conjuntos"),
                        payload.get("entidades_financieras"),
                        payload.get("avalistas"),
                        payload.get("aportacion_cv"),
                        payload.get("notas"),
                        payload.get("notas_ocr"),
                        payload.get("calidad_ocr"),
                        payload.get("campos_ocr"),
                        now,
                        now,
                    ),
                )
            row = conn.execute(
                "SELECT * FROM asesoramientos_financiacion WHERE id = ?",
                (asesoramiento_id,),
            ).fetchone()
            missing = fin_missing_fields(row)
            fin_sync_missing_action(
                conn,
                empresa["id"],
                asesoramiento_id,
                cliente1_id or row["cliente1_id"],
                row["cliente1_nombre"] if row else "",
                missing,
                now,
            )
            json_response(
                self,
                {
                    "ok": True,
                    "id": asesoramiento_id,
                    "duplicate_of": dup_id,
                    "missing": missing,
                },
            )
            return
        elif parsed.path == "/api/fin_asesoramientos_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            cliente1_id = ensure_cliente_for_financiacion(
                conn,
                empresa["id"],
                payload.get("cliente1_nombre"),
                payload.get("cliente1_dni"),
                now,
                {
                    "telefono": payload.get("cliente1_telefono"),
                    "email": payload.get("cliente1_email"),
                    "fecha_nacimiento": payload.get("cliente1_fecha_nacimiento"),
                },
            )
            cliente2_id = ensure_cliente_for_financiacion(
                conn,
                empresa["id"],
                payload.get("cliente2_nombre"),
                payload.get("cliente2_dni"),
                now,
                {
                    "telefono": payload.get("cliente2_telefono"),
                    "email": payload.get("cliente2_email"),
                    "fecha_nacimiento": payload.get("cliente2_fecha_nacimiento"),
                },
            )
            allowed = (
                "origen",
                "inmobiliaria_asesor",
                "asesor",
                "fecha",
                "estado",
                "cliente1_id",
                "cliente1_nombre",
                "cliente1_dni",
                "cliente1_telefono",
                "cliente1_email",
                "cliente1_fecha_nacimiento",
                "cliente1_estado_civil",
                "cliente1_regimen",
                "cliente1_hijos",
                "cliente1_profesion",
                "cliente1_tipo_contrato",
                "cliente1_tiempo_contrato",
                "cliente1_ingresos",
                "cliente1_patrimonio",
                "cliente1_prestamos",
                "cliente1_prestamo_activo",
                "cliente1_prestamo_entidad",
                "cliente1_prestamo_resto",
                "cliente2_id",
                "cliente2_nombre",
                "cliente2_dni",
                "cliente2_telefono",
                "cliente2_email",
                "cliente2_fecha_nacimiento",
                "cliente2_estado_civil",
                "cliente2_regimen",
                "cliente2_hijos",
                "cliente2_profesion",
                "cliente2_tipo_contrato",
                "cliente2_tiempo_contrato",
                "cliente2_ingresos",
                "cliente2_patrimonio",
                "cliente2_prestamos",
                "cliente2_prestamo_activo",
                "cliente2_prestamo_entidad",
                "cliente2_prestamo_resto",
                "ingresos_conjuntos",
                "entidades_financieras",
                "avalistas",
                "aportacion_cv",
                "notas",
                "notas_ocr",
                "calidad_ocr",
                "campos_ocr",
            )
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if cliente1_id:
                updates["cliente1_id"] = cliente1_id
            if cliente2_id:
                updates["cliente2_id"] = cliente2_id
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE asesoramientos_financiacion SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            row = conn.execute(
                "SELECT * FROM asesoramientos_financiacion WHERE id = ?",
                (record_id,),
            ).fetchone()
            if row:
                missing = fin_missing_fields(row)
                fin_sync_missing_action(
                    conn,
                    empresa["id"],
                    record_id,
                    row["cliente1_id"],
                    row["cliente1_nombre"],
                    missing,
                    now,
                )
        elif parsed.path == "/api/fin_asesoramientos_convert":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute(
                "SELECT * FROM asesoramientos_financiacion WHERE id = ?",
                (record_id,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            cliente_nombre = row["cliente1_nombre"] or ""
            if row["cliente2_nombre"]:
                cliente_nombre = f"{cliente_nombre} / {row['cliente2_nombre']}".strip(" /")
            fecha = row["fecha"] or ""
            try:
                anio = int(fecha.split("/")[-1]) if "/" in fecha else int(fecha.split("-")[0])
            except Exception:
                anio = None
            hipoteca_id = os.urandom(16).hex()
            conn.execute(
                """
                INSERT INTO hipotecas (
                  id, empresa_id, cliente, banco, precio, importe_hipoteca, porcentaje,
                  entrada, comision, oficina, fecha_encargo, encargo, tipo_hipoteca,
                  fecha_firma, cesion, comision_juan, comision_modernia, inmobiliaria_compra,
                  asesor, estado, anio, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    hipoteca_id,
                    empresa["id"],
                    cliente_nombre,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    fecha,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    row["inmobiliaria_asesor"],
                    row["asesor"],
                    "Pendiente",
                    anio,
                    now,
                    now,
                ),
            )
            conn.execute(
                "UPDATE asesoramientos_financiacion SET estado = ?, updated_at = datetime(?) WHERE id = ?",
                ("Convertido", now, record_id),
            )
            json_response(self, {"hipoteca_id": hipoteca_id})
        elif parsed.path == "/api/gestoria_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            updates = {}
            for key in ("estado", "fecha_baja"):
                if key in payload:
                    updates[key] = payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE gestoria SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
        elif parsed.path == "/api/seguros_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            current_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if not current_row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            updates = {}
            for key in (
                "cliente_id",
                "estado",
                "compania",
                "ramo",
                "tomador",
                "prima_neta",
                "prima_total",
                "comision",
                "porcentaje",
                "produccion",
                "colaborador",
                "fecha_efecto",
                "fecha_vencimiento",
                "fecha_baja",
                "motivo_baja",
                "estado_renovacion",
                "renovacion_fecha",
                "nueva_poliza_ref",
                "estado_poliza",
                "poliza_origen_id",
                "poliza_sustituta_id",
                "version_grupo",
                "poliza_numero",
                "poliza_key",
                "poliza_url",
                "tipo_vigencia",
                "datos_ramo_json",
            ):
                if key in payload:
                    updates[key] = payload.get(key)
            if "ramo" in updates:
                updates["ramo"] = canonicalize_ramo(updates.get("ramo"))
            if "tipo_vigencia" in updates or "ramo" in updates:
                updates["tipo_vigencia"] = infer_tipo_vigencia(
                    updates.get("ramo", current_row["ramo"]),
                    updates.get("tipo_vigencia", current_row["tipo_vigencia"] if "tipo_vigencia" in current_row.keys() else None),
                )
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            incoming_cliente_id = updates.get("cliente_id")
            if incoming_cliente_id:
                cliente_exists = conn.execute(
                    "SELECT id FROM clientes WHERE id = ?",
                    (incoming_cliente_id,),
                ).fetchone()
                if not cliente_exists:
                    json_response(self, {"error": "cliente_id no válido"}, status=400)
                    return
            poliza_candidate = updates.get("poliza_numero", current_row["poliza_numero"])
            compania_candidate = updates.get("compania", current_row["compania"])
            dup_id = find_existing_seguro_id(
                conn,
                current_row["empresa_id"],
                poliza_candidate,
                compania_candidate,
                exclude_id=record_id,
            )
            if dup_id:
                json_response(
                    self,
                    {
                        "error": "Ya existe una póliza con ese número y compañía",
                        "duplicate_of": dup_id,
                    },
                    status=409,
                )
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE seguros SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if row and row["cliente_id"]:
                ensure_cliente_servicio_link(conn, row["cliente_id"], row["empresa_id"], "seguros", now)
            if row:
                ensure_seguro_doc_link(conn, row, now)
                log_seguro_event(conn, row, "actualizacion", now, payload={"campos": sorted(list(updates.keys()))})
                upsert_seguro_comision_contabilidad(conn, row, now, movimiento="emision")
            if row:
                missing = []
                for key in ("tomador", "poliza_numero", "compania", "fecha_efecto"):
                    if not str(row[key] or "").strip():
                        missing.append(key)
                if not missing:
                    conn.execute(
                        """
                        UPDATE acciones
                        SET estado = 'Hecho', updated_at = datetime(?)
                        WHERE servicio = 'Seguros'
                          AND tipo = 'Completar datos póliza'
                          AND notas LIKE ?
                        """,
                        (now, f"%{record_id}%"),
                    )
            conn.commit()
            json_response(
                self,
                {
                    "ok": True,
                    "id": record_id,
                    "ramo": row["ramo"] if row else "",
                    "estado": row["estado"] if row else "",
                },
            )
            return
        elif parsed.path == "/api/seguros_cambio_compania":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            nueva_compania = (payload.get("nueva_compania") or payload.get("compania") or "").strip()
            if not nueva_compania:
                json_response(self, {"error": "nueva_compania requerida"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            fecha_cambio = (payload.get("fecha_cambio") or payload.get("fecha") or now[:10]).strip()
            nueva_poliza = (payload.get("nueva_poliza_numero") or payload.get("poliza_numero") or "").strip()
            if not nueva_poliza:
                nueva_poliza = (payload.get("nueva_poliza_ref") or row["nueva_poliza_ref"] or row["poliza_numero"] or "").strip()
            nueva_fecha_efecto = payload.get("nueva_fecha_efecto") or payload.get("fecha_efecto") or row["fecha_efecto"]
            nueva_fecha_venc = payload.get("nueva_fecha_vencimiento") or payload.get("fecha_vencimiento") or row["fecha_vencimiento"]
            nuevo_estado = payload.get("nuevo_estado") or "En vigor"
            nuevo_cliente_id = payload.get("cliente_id") or row["cliente_id"]
            if nuevo_cliente_id:
                cliente_exists = conn.execute(
                    "SELECT id FROM clientes WHERE id = ?",
                    (nuevo_cliente_id,),
                ).fetchone()
                if not cliente_exists:
                    json_response(self, {"error": "cliente_id no válido"}, status=400)
                    return
            version_grupo = (row["version_grupo"] or row["id"] or "").strip() or row["id"]
            old_policy = row["poliza_numero"]
            new_id = os.urandom(16).hex()
            dup_id = find_existing_seguro_id(
                conn,
                row["empresa_id"],
                nueva_poliza,
                nueva_compania,
            )
            if dup_id:
                json_response(
                    self,
                    {
                        "error": "Ya existe una póliza con ese número y compañía",
                        "duplicate_of": dup_id,
                    },
                    status=409,
                )
                return
            conn.execute(
                """
                UPDATE seguros
                SET
                  estado = 'Sustituida',
                  estado_poliza = 'sustituida',
                  fecha_baja = ?,
                  motivo_baja = ?,
                  estado_renovacion = 'Cambio compañía',
                  renovacion_fecha = ?,
                  nueva_poliza_ref = ?,
                  poliza_sustituta_id = ?,
                  version_grupo = ?,
                  updated_at = datetime(?)
                WHERE id = ?
                """,
                (
                    fecha_cambio,
                    payload.get("motivo_baja") or "cambio_compania",
                    fecha_cambio,
                    nueva_poliza or "",
                    new_id,
                    version_grupo,
                    now,
                    record_id,
                ),
            )
            conn.execute(
                """
                INSERT INTO seguros (
                  id, empresa_id, cliente_id, mes_creacion, fecha_efecto, fecha_vencimiento,
                  tomador, compania, ramo, poliza_numero, prima_neta, prima_total,
                  comision, produccion, colaborador, estado,
                  estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                  poliza_key, poliza_url, fecha_baja, motivo_baja,
                  estado_poliza, poliza_origen_id, poliza_sustituta_id, version_grupo, tipo_vigencia,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    new_id,
                    row["empresa_id"],
                    nuevo_cliente_id,
                    row["mes_creacion"],
                    nueva_fecha_efecto,
                    nueva_fecha_venc,
                    row["tomador"],
                    nueva_compania,
                    row["ramo"],
                    nueva_poliza or row["poliza_numero"],
                    row["prima_neta"],
                    row["prima_total"],
                    row["comision"],
                    row["produccion"],
                    row["colaborador"],
                    nuevo_estado,
                    "Alta por cambio compañía",
                    fecha_cambio,
                    old_policy or "",
                    payload.get("poliza_key") or "",
                    payload.get("poliza_url") or "",
                    None,
                    None,
                    "activa",
                    record_id,
                    None,
                    version_grupo,
                    infer_tipo_vigencia(row["ramo"], row["tipo_vigencia"] if "tipo_vigencia" in row.keys() else None),
                    now,
                    now,
                ),
            )
            conn.execute(
                "UPDATE seguros SET poliza_sustituta_id = ?, updated_at = datetime(?) WHERE id = ?",
                (new_id, now, record_id),
            )
            new_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (new_id,)).fetchone()
            if new_row and new_row["cliente_id"]:
                ensure_cliente_servicio_link(conn, new_row["cliente_id"], new_row["empresa_id"], "seguros", now)
                ensure_seguro_doc_link(conn, new_row, now)
            old_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if old_row:
                log_seguro_event(
                    conn,
                    old_row,
                    "cambio_compania_salida",
                    now,
                    motivo=payload.get("motivo_baja") or "cambio_compania",
                    payload={"new_id": new_id, "nueva_compania": nueva_compania, "nueva_poliza": nueva_poliza},
                )
            if new_row:
                log_seguro_event(
                    conn,
                    new_row,
                    "cambio_compania_entrada",
                    now,
                    payload={"old_id": record_id, "compania_origen": row["compania"], "poliza_origen": old_policy},
                )
                upsert_seguro_comision_contabilidad(conn, new_row, now, movimiento="emision")
            json_response(
                self,
                {
                    "ok": True,
                    "old_id": record_id,
                    "new_id": new_id,
                    "version_grupo": version_grupo,
                },
            )
            conn.commit()
            return
        elif parsed.path == "/api/seguros_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            log_seguro_event(conn, row, "eliminacion", now, payload={"origen": "api/seguros_delete"})
            conn.execute("DELETE FROM seguros_checklist WHERE poliza_id = ?", (record_id,))
            conn.execute(
                """
                DELETE FROM acciones
                WHERE servicio = 'Seguros'
                  AND notas LIKE ?
                """,
                (f"%{record_id}%",),
            )
            conn.execute(
                """
                DELETE FROM gestoria_docs
                WHERE referencia_tipo = 'seguros'
                  AND referencia_id = ?
                """,
                (record_id,),
            )
            conn.execute("DELETE FROM seguros WHERE id = ?", (record_id,))
            json_response(
                self,
                {
                    "ok": True,
                    "id": record_id,
                    "cliente_id": row["cliente_id"],
                    "empresa_id": row["empresa_id"],
                },
            )
            conn.commit()
            return
        elif parsed.path == "/api/seguros_poliza_accion":
            record_id = payload.get("id")
            action = normalize_lookup_text(payload.get("accion") or payload.get("action"))
            if not record_id or not action:
                json_response(self, {"error": "id y accion requeridos"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            if action in ("RENOVAR", "RENEW"):
                fecha_renovacion = (payload.get("fecha_renovacion") or payload.get("fecha") or now[:10]).strip()
                nueva_fecha_venc = (payload.get("nueva_fecha_vencimiento") or payload.get("fecha_vencimiento") or "").strip()
                set_parts = [
                    "estado_renovacion = ?",
                    "renovacion_fecha = ?",
                    "estado_poliza = 'activa'",
                    "fecha_vencimiento = COALESCE(NULLIF(?, ''), fecha_vencimiento)",
                ]
                set_values = ["Renovada manual", fecha_renovacion, nueva_fecha_venc]
                if "comision" in payload:
                    set_parts.append("comision = ?")
                    set_values.append(parse_money_value(payload.get("comision")))
                conn.execute(
                    f"""
                    UPDATE seguros
                    SET {", ".join(set_parts)}, updated_at = datetime(?)
                    WHERE id = ?
                    """,
                    (*set_values, now, record_id),
                )
                row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
                log_seguro_event(conn, row, "renovacion_manual", now, payload={"fecha": fecha_renovacion})
                contabilidad_id = upsert_seguro_comision_contabilidad(
                    conn,
                    row,
                    now,
                    movimiento="renovacion",
                    fecha=fecha_renovacion,
                )
                json_response(self, {"ok": True, "id": record_id, "accion": "renovar", "contabilidad_id": contabilidad_id})
                conn.commit()
                return
            if action in ("ANULAR", "CANCELAR", "CANCEL"):
                fecha_baja = (payload.get("fecha_baja") or payload.get("fecha") or now[:10]).strip()
                motivo_baja = (payload.get("motivo_baja") or payload.get("motivo") or "otros").strip()
                conn.execute(
                    """
                    UPDATE seguros
                    SET estado = 'Anulada',
                        estado_poliza = 'anulada',
                        fecha_baja = ?,
                        motivo_baja = ?,
                        updated_at = datetime(?)
                    WHERE id = ?
                    """,
                    (fecha_baja, motivo_baja, now, record_id),
                )
                row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
                log_seguro_event(conn, row, "anulacion", now, motivo=motivo_baja, payload={"fecha_baja": fecha_baja})
                json_response(self, {"ok": True, "id": record_id, "accion": "anular"})
                conn.commit()
                return
            json_response(self, {"error": "Accion no soportada"}, status=400)
            return
        elif parsed.path == "/api/seguros_reclamacion":
            seguro_id = payload.get("seguro_id")
            cliente_id = payload.get("cliente_id")
            empresa_id = payload.get("empresa_id")
            if seguro_id:
                row = conn.execute("SELECT id, cliente_id, empresa_id FROM seguros WHERE id = ?", (seguro_id,)).fetchone()
                if not row:
                    json_response(self, {"error": "seguro_id no encontrado"}, status=404)
                    return
                cliente_id = cliente_id or row["cliente_id"]
                empresa_id = empresa_id or row["empresa_id"]
            if not cliente_id or not empresa_id:
                json_response(self, {"error": "cliente_id y empresa_id requeridos"}, status=400)
                return
            rec_id = os.urandom(16).hex()
            estado = (payload.get("estado") or "abierta").strip()
            fecha_apertura = (payload.get("fecha_apertura") or payload.get("fecha") or now[:10]).strip()
            conn.execute(
                """
                INSERT INTO seguros_reclamaciones (
                  id, seguro_id, cliente_id, empresa_id, estado, canal, fecha_apertura, fecha_cierre,
                  asunto, detalle, resolucion, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    rec_id,
                    seguro_id,
                    cliente_id,
                    empresa_id,
                    estado,
                    payload.get("canal"),
                    fecha_apertura,
                    payload.get("fecha_cierre"),
                    payload.get("asunto"),
                    payload.get("detalle"),
                    payload.get("resolucion"),
                    now,
                    now,
                ),
            )
            if seguro_id:
                seguro_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (seguro_id,)).fetchone()
                log_seguro_event(conn, seguro_row, "reclamacion_alta", now, payload={"reclamacion_id": rec_id, "estado": estado})
            json_response(self, {"ok": True, "id": rec_id})
            conn.commit()
            return
        elif parsed.path == "/api/seguros_reclamacion_update":
            rec_id = payload.get("id")
            if not rec_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros_reclamaciones WHERE id = ?", (rec_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Reclamación no encontrada"}, status=404)
                return
            updates = {}
            for key in ("estado", "canal", "fecha_apertura", "fecha_cierre", "asunto", "detalle", "resolucion"):
                if key in payload:
                    updates[key] = payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, rec_id]
            conn.execute(
                f"UPDATE seguros_reclamaciones SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            if row["seguro_id"]:
                seguro_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (row["seguro_id"],)).fetchone()
                log_seguro_event(conn, seguro_row, "reclamacion_update", now, payload={"reclamacion_id": rec_id, "campos": sorted(list(updates.keys()))})
            json_response(self, {"ok": True, "id": rec_id})
            conn.commit()
            return
        elif parsed.path == "/api/seguros_reclamacion_delete":
            rec_id = payload.get("id")
            if not rec_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros_reclamaciones WHERE id = ?", (rec_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Reclamación no encontrada"}, status=404)
                return
            conn.execute("DELETE FROM seguros_reclamaciones WHERE id = ?", (rec_id,))
            if row["seguro_id"]:
                seguro_row = conn.execute("SELECT * FROM seguros WHERE id = ?", (row["seguro_id"],)).fetchone()
                log_seguro_event(conn, seguro_row, "reclamacion_delete", now, payload={"reclamacion_id": rec_id})
            json_response(self, {"ok": True, "id": rec_id})
            conn.commit()
            return
        elif parsed.path == "/api/seguros_ipid_register":
            seguro_id = payload.get("seguro_id")
            if not seguro_id:
                json_response(self, {"error": "seguro_id requerido"}, status=400)
                return
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (seguro_id,)).fetchone()
            if not row:
                json_response(self, {"error": "Póliza no encontrada"}, status=404)
                return
            ipid_id = os.urandom(16).hex()
            fecha_entrega = (payload.get("fecha_entrega") or payload.get("fecha") or now[:10]).strip()
            conn.execute(
                """
                INSERT INTO seguros_ipid_log (
                  id, seguro_id, cliente_id, empresa_id, documento_key, documento_url,
                  fecha_entrega, metodo, usuario, created_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?)
                )
                """,
                (
                    ipid_id,
                    seguro_id,
                    row["cliente_id"],
                    row["empresa_id"],
                    payload.get("documento_key") or row["poliza_key"],
                    payload.get("documento_url") or row["poliza_url"],
                    fecha_entrega,
                    payload.get("metodo") or "digital",
                    payload.get("usuario") or "Sistema",
                    now,
                ),
            )
            log_seguro_event(conn, row, "ipid_entregado", now, payload={"ipid_id": ipid_id, "fecha": fecha_entrega})
            json_response(self, {"ok": True, "id": ipid_id})
            conn.commit()
            return
        elif parsed.path == "/api/seguros_enrich":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute(
                "SELECT * FROM seguros WHERE id = ?",
                (record_id,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Registro no encontrado"}, status=404)
                return
            allowed = (
                "cliente_id",
                "tomador",
                "compania",
                "ramo",
                "poliza_numero",
                "prima_neta",
                "prima_total",
                "fecha_efecto",
                "fecha_vencimiento",
                "poliza_key",
                "poliza_url",
            )
            updates = {}
            for key in allowed:
                incoming = payload.get(key)
                if key == "ramo":
                    incoming = canonicalize_ramo(incoming)
                if incoming in (None, ""):
                    continue
                current = row[key] if key in row.keys() else None
                if current is None or str(current).strip() == "":
                    updates[key] = incoming
            if updates:
                set_clause = ", ".join([f"{key} = ?" for key in updates])
                values = list(updates.values()) + [now, record_id]
                conn.execute(
                    f"UPDATE seguros SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                    values,
                )
            # Captura inteligente por ramo desde OCR: persiste señales comerciales
            # aunque no existan columnas dedicadas en la tabla.
            smart_allowed = (
                "direccion_riesgo",
                "codigo_postal",
                "fecha_nacimiento_asegurado",
                "fecha_nacimiento_conductor",
                "fecha_carnet",
                "matricula",
                "marca_modelo",
                "anio_matriculacion",
                "uso_vehiculo",
                "garaje",
                "tipo_vivienda",
                "metros2",
                "anio_construccion",
                "continente",
                "contenido",
                "profesion",
                "fumador",
                "capital_asegurado",
                "beneficiarios",
                "deporte_riesgo",
                "actividad",
                "facturacion_anual",
                "empleados",
                "superficie",
                "medidas_seguridad",
                "notas_comerciales",
            )
            raw_existing = row["datos_ramo_json"] if "datos_ramo_json" in row.keys() else ""
            try:
                existing_smart = json.loads(raw_existing) if raw_existing else {}
                if not isinstance(existing_smart, dict):
                    existing_smart = {}
            except Exception:
                existing_smart = {}
            incoming_smart = {}
            raw_smart_payload = payload.get("datos_ramo_json")
            if raw_smart_payload:
                try:
                    parsed_smart = (
                        json.loads(raw_smart_payload)
                        if isinstance(raw_smart_payload, str)
                        else raw_smart_payload
                    )
                    if isinstance(parsed_smart, dict):
                        incoming_smart.update(
                            {
                                str(k): v
                                for k, v in parsed_smart.items()
                                if str(v or "").strip() != ""
                            }
                        )
                except Exception:
                    pass
            for key in smart_allowed:
                val = payload.get(key)
                if val is None:
                    continue
                text = str(val).strip()
                if not text:
                    continue
                incoming_smart[key] = val
            if incoming_smart:
                merged = dict(existing_smart)
                merged.update(incoming_smart)
                conn.execute(
                    "UPDATE seguros SET datos_ramo_json = ?, updated_at = datetime(?) WHERE id = ?",
                    (json.dumps(merged, ensure_ascii=False), now, record_id),
                )
            row = conn.execute("SELECT * FROM seguros WHERE id = ?", (record_id,)).fetchone()
            if row:
                if row["cliente_id"]:
                    ensure_cliente_servicio_link(conn, row["cliente_id"], row["empresa_id"], "seguros", now)
                ensure_seguro_doc_link(conn, row, now)
                missing = []
                for key in ("tomador", "poliza_numero", "compania", "fecha_efecto"):
                    if not str(row[key] or "").strip():
                        missing.append(key)
                if not missing:
                    conn.execute(
                        """
                        UPDATE acciones
                        SET estado = 'Hecho', updated_at = datetime(?)
                        WHERE servicio = 'Seguros'
                          AND tipo = 'Completar datos póliza'
                          AND notas LIKE ?
                        """,
                        (now, f"%{record_id}%"),
                    )
        elif parsed.path == "/api/seguros_ofertas":
            conn.execute(
                """
                INSERT INTO seguros_ofertas (
                  id, cliente_id, ramo, compania, propuesta, estado, motivo,
                  fecha, responsable, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    payload.get("cliente_id"),
                    canonicalize_ramo(payload.get("ramo")),
                    payload.get("compania"),
                    payload.get("propuesta"),
                    payload.get("estado"),
                    payload.get("motivo"),
                    payload.get("fecha"),
                    payload.get("responsable"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/seguros_ofertas_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("ramo", "compania", "propuesta", "estado", "motivo", "fecha", "responsable", "notas", "cliente_id")
            updates = {}
            for key in allowed:
                if key not in payload:
                    continue
                updates[key] = canonicalize_ramo(payload.get(key)) if key == "ramo" else payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE seguros_ofertas SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
        elif parsed.path == "/api/seguros_ofertas_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM seguros_ofertas WHERE id = ?", (record_id,))
        elif parsed.path == "/api/seguros_preferencias":
            cliente_id = payload.get("cliente_id")
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            prefs = (
                "prioridad_precio",
                "prioridad_compania",
                "prioridad_coberturas",
                "notas",
            )
            updates = {key: payload.get(key) for key in prefs if key in payload}
            existing = conn.execute(
                "SELECT id FROM seguros_preferencias WHERE cliente_id = ?",
                (cliente_id,),
            ).fetchone()
            if existing:
                if updates:
                    set_clause = ", ".join([f"{key} = ?" for key in updates])
                    values = list(updates.values()) + [now, cliente_id]
                    conn.execute(
                        f"UPDATE seguros_preferencias SET {set_clause}, updated_at = datetime(?) WHERE cliente_id = ?",
                        values,
                    )
            else:
                conn.execute(
                    """
                    INSERT INTO seguros_preferencias (
                      id, cliente_id, prioridad_precio, prioridad_compania, prioridad_coberturas,
                      notas, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        cliente_id,
                        payload.get("prioridad_precio"),
                        payload.get("prioridad_compania"),
                        payload.get("prioridad_coberturas"),
                        payload.get("notas"),
                        now,
                        now,
                    ),
                )
        elif parsed.path == "/api/seguros_referidos":
            conn.execute(
                """
                INSERT INTO seguros_referidos (
                  id, cliente_id, referido_por, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    payload.get("cliente_id"),
                    payload.get("referido_por"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/seguros_campanas":
            conn.execute(
                """
                INSERT INTO seguros_campanas (
                  id, compania, nombre, ramo, origen, fecha_inicio, fecha_fin, descripcion, url,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    payload.get("compania"),
                    payload.get("nombre"),
                    canonicalize_ramo(payload.get("ramo")),
                    payload.get("origen"),
                    payload.get("fecha_inicio"),
                    payload.get("fecha_fin"),
                    payload.get("descripcion"),
                    payload.get("url"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/seguros_campanas_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("compania", "nombre", "ramo", "origen", "fecha_inicio", "fecha_fin", "descripcion", "url")
            updates = {}
            for key in allowed:
                if key not in payload:
                    continue
                updates[key] = canonicalize_ramo(payload.get(key)) if key == "ramo" else payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE seguros_campanas SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
        elif parsed.path == "/api/seguros_campanas_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM seguros_campanas WHERE id = ?", (record_id,))
        elif parsed.path == "/api/seguros_comisiones":
            conn.execute(
                """
                INSERT INTO seguros_comisiones (
                  id, compania, ramo, porcentaje, vigencia_desde, vigencia_hasta, notas,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    payload.get("compania"),
                    canonicalize_ramo(payload.get("ramo")),
                    payload.get("porcentaje"),
                    payload.get("vigencia_desde"),
                    payload.get("vigencia_hasta"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/seguros_comisiones_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("compania", "ramo", "porcentaje", "vigencia_desde", "vigencia_hasta", "notas")
            updates = {}
            for key in allowed:
                if key not in payload:
                    continue
                updates[key] = canonicalize_ramo(payload.get(key)) if key == "ramo" else payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE seguros_comisiones SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
        elif parsed.path == "/api/seguros_comisiones_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM seguros_comisiones WHERE id = ?", (record_id,))
        elif parsed.path == "/api/seguros_checklist_generate":
            poliza_id = payload.get("poliza_id")
            if not poliza_id:
                json_response(self, {"error": "poliza_id requerido"}, status=400)
                return
            conn.execute("DELETE FROM seguros_checklist WHERE poliza_id = ?", (poliza_id,))
            tasks = [
                "Póliza firmada",
                "Documento identidad",
                "Recibo emitido",
                "Pago prima",
                "Suplementos/Anexos",
            ]
            for tarea in tasks:
                conn.execute(
                    """
                    INSERT INTO seguros_checklist (
                      id, poliza_id, tarea, estado, responsable, fecha_limite, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        poliza_id,
                        tarea,
                        "Pendiente",
                        payload.get("responsable"),
                        payload.get("fecha_limite"),
                        now,
                        now,
                    ),
                )
            json_response(self, {"ok": True})
            return
        elif parsed.path == "/api/seguros_checklist_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            updates = {}
            for key in ("estado", "responsable", "fecha_limite"):
                if key in payload:
                    updates[key] = payload.get(key)
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE seguros_checklist SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            return
        elif parsed.path == "/api/fin_checklist_generate":
            asesoramiento_id = payload.get("asesoramiento_id")
            if not asesoramiento_id:
                json_response(self, {"error": "asesoramiento_id requerido"}, status=400)
                return
            conn.execute(
                "DELETE FROM fin_checklist WHERE asesoramiento_id = ?",
                (asesoramiento_id,),
            )
            tasks = [
                "DNI clientes",
                "Vida laboral",
                "Nóminas últimos meses",
                "Declaración de la renta",
                "Preaprobación bancaria",
                "Tasación",
            ]
            for tarea in tasks:
                conn.execute(
                    """
                    INSERT INTO fin_checklist (
                      id, asesoramiento_id, tarea, estado, responsable, fecha_limite, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        asesoramiento_id,
                        tarea,
                        "Pendiente",
                        payload.get("responsable"),
                        payload.get("fecha_limite"),
                        now,
                        now,
                    ),
                )
            json_response(self, {"ok": True})
            return
        elif parsed.path == "/api/fin_checklist_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            updates = {}
            for key in ("estado", "responsable", "fecha_limite"):
                if key in payload:
                    updates[key] = payload.get(key)
            if not updates:
                json_response(self, {"error": "sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE fin_checklist SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            return
        elif parsed.path == "/api/ai_seguros_copilot":
            if not openai_available():
                json_response(self, {"error": "OPENAI_API_KEY no configurada"}, status=400)
                return
            poliza_id = payload.get("poliza_id")
            if not poliza_id:
                json_response(self, {"error": "poliza_id requerido"}, status=400)
                return
            task = (payload.get("task") or "resumen").strip().lower()
            extra = (payload.get("extra") or "").strip()
            poliza = conn.execute("SELECT * FROM seguros WHERE id = ?", (poliza_id,)).fetchone()
            if not poliza:
                json_response(self, {"error": "Póliza no encontrada"}, status=404)
                return
            poliza_data = dict(poliza)
            cliente = None
            if poliza_data.get("cliente_id"):
                cliente = conn.execute(
                    "SELECT id, nombre, nif, telefono, email FROM clientes WHERE id = ?",
                    (poliza_data["cliente_id"],),
                ).fetchone()
            context = {
                "poliza": poliza_data,
                "cliente": dict(cliente) if cliente else {},
            }
            if task == "email_renovacion":
                prompt = (
                    "Redacta un email breve de renovación de póliza. Incluye nombre del cliente, "
                    "número de póliza, compañía y fecha de vencimiento si está disponible. "
                    "Tono profesional y cercano. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            elif task == "faltantes":
                prompt = (
                    "Lista los campos obligatorios faltantes de la póliza (tomador, poliza_numero, compania, fecha_efecto). "
                    "Si no falta ninguno, indícalo. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            else:
                prompt = (
                    "Genera un resumen claro de la póliza: tomador, compañía, ramo, fechas, prima y estado. "
                    "Si falta algún dato, indícalo. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            output, err = call_openai(prompt)
            if err:
                json_response(self, {"error": err}, status=400)
                return
            json_response(self, {"output": output})
            return
        elif parsed.path == "/api/ai_fin_copilot":
            if not openai_available():
                json_response(self, {"error": "OPENAI_API_KEY no configurada"}, status=400)
                return
            asesoramiento_id = payload.get("asesoramiento_id")
            if not asesoramiento_id:
                json_response(self, {"error": "asesoramiento_id requerido"}, status=400)
                return
            task = (payload.get("task") or "resumen").strip().lower()
            extra = (payload.get("extra") or "").strip()
            ases = conn.execute(
                "SELECT * FROM asesoramientos_financiacion WHERE id = ?",
                (asesoramiento_id,),
            ).fetchone()
            if not ases:
                json_response(self, {"error": "Asesoramiento no encontrado"}, status=404)
                return
            cliente1 = None
            cliente2 = None
            if ases.get("cliente1_id"):
                cliente1 = conn.execute(
                    "SELECT id, nombre, nif, telefono, email FROM clientes WHERE id = ?",
                    (ases["cliente1_id"],),
                ).fetchone()
            if ases.get("cliente2_id"):
                cliente2 = conn.execute(
                    "SELECT id, nombre, nif, telefono, email FROM clientes WHERE id = ?",
                    (ases["cliente2_id"],),
                ).fetchone()
            context = {
                "asesoramiento": dict(ases),
                "cliente1": dict(cliente1) if cliente1 else {},
                "cliente2": dict(cliente2) if cliente2 else {},
            }
            if task == "faltantes":
                prompt = (
                    "Lista los campos obligatorios faltantes del asesoramiento financiero. "
                    "Si no falta ninguno, indícalo. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            elif task == "documentacion":
                prompt = (
                    "Genera un checklist breve de documentación necesaria para este asesoramiento "
                    "según el perfil del cliente. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            else:
                prompt = (
                    "Genera un resumen claro del asesoramiento: clientes, ingresos, estado y próximos pasos. "
                    "Si falta algún dato, indícalo. No inventes datos.\n\n"
                    f"Contexto: {json.dumps(context, ensure_ascii=False)}\n"
                    f"Instrucciones extra: {extra}"
                )
            output, err = call_openai(prompt)
            if err:
                json_response(self, {"error": err}, status=400)
                return
            json_response(self, {"output": output})
            return
        elif parsed.path == "/api/captaciones":
            inmueble_id = os.urandom(16).hex()
            propietarios = payload.get("propietarios") or []
            if isinstance(propietarios, str):
                propietarios = [p for p in propietarios.split(",") if p]
            conn.execute(
                """
                INSERT INTO captaciones (
                  id, empresa_id, inmueble_id, propietario, tipo_inmueble, direccion, codigo_postal, poblacion, provincia, zona, m2,
                  habitaciones, banos, precio_objetivo, precio_valoracion, urgencia,
                  motivo, canal, etapa, probabilidad, proxima_accion, fecha_contacto,
                  asesor, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    inmueble_id,
                    payload.get("propietario"),
                    payload.get("tipo_inmueble"),
                    payload.get("direccion"),
                    payload.get("codigo_postal"),
                    payload.get("poblacion"),
                    payload.get("provincia"),
                    payload.get("zona"),
                    payload.get("m2"),
                    payload.get("habitaciones"),
                    payload.get("banos"),
                    payload.get("precio_objetivo"),
                    payload.get("precio_valoracion"),
                    payload.get("urgencia"),
                    payload.get("motivo"),
                    payload.get("canal"),
                    payload.get("etapa"),
                    payload.get("probabilidad"),
                    payload.get("proxima_accion"),
                    payload.get("fecha_contacto"),
                    payload.get("asesor"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
            conn.execute(
                """
                INSERT INTO inmuebles (
                  id, empresa_id, referencia, direccion, codigo_postal, poblacion, provincia, zona, tipo_inmueble,
                  m2, habitaciones, banos, precio_objetivo, precio_valoracion,
                  valor_referencia, estado, lat, lon, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    inmueble_id,
                    empresa["id"],
                    payload.get("referencia"),
                    payload.get("direccion"),
                    payload.get("codigo_postal"),
                    payload.get("poblacion"),
                    payload.get("provincia"),
                    payload.get("zona"),
                    payload.get("tipo_inmueble"),
                    payload.get("m2"),
                    payload.get("habitaciones"),
                    payload.get("banos"),
                    payload.get("precio_objetivo"),
                    payload.get("precio_valoracion"),
                    payload.get("valor_referencia"),
                    payload.get("etapa"),
                    payload.get("lat"),
                    payload.get("lon"),
                    now,
                    now,
                ),
            )
            for cliente_id in propietarios:
                conn.execute(
                    """
                    INSERT INTO inmueble_propietarios (
                      id, inmueble_id, cliente_id, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (os.urandom(16).hex(), inmueble_id, cliente_id, now, now),
                )
        elif parsed.path == "/api/captaciones_update":
            record_id = payload.get("id")
            etapa = payload.get("etapa")
            if not record_id or not etapa:
                json_response(self, {"error": "id y etapa requeridos"}, status=400)
                return
            conn.execute(
                """
                UPDATE captaciones
                SET etapa = ?, updated_at = datetime(?)
                WHERE id = ?
                """,
                (etapa, now, record_id),
            )
        elif parsed.path == "/api/captacion_update":
            inmueble_id = payload.get("inmueble_id")
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            allowed = (
                "propietario",
                "tipo_inmueble",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "zona",
                "m2",
                "habitaciones",
                "banos",
                "precio_objetivo",
                "precio_valoracion",
                "urgencia",
                "motivo",
                "canal",
                "etapa",
                "probabilidad",
                "proxima_accion",
                "fecha_contacto",
                "asesor",
                "notas",
            )
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, inmueble_id]
            conn.execute(
                f"UPDATE captaciones SET {set_clause}, updated_at = datetime(?) WHERE inmueble_id = ?",
                values,
            )
            shared = (
                "tipo_inmueble",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "zona",
                "m2",
                "habitaciones",
                "banos",
                "precio_objetivo",
                "precio_valoracion",
            )
            inm_updates = {key: updates[key] for key in shared if key in updates}
            if "etapa" in updates:
                inm_updates["estado"] = updates["etapa"]
            if inm_updates:
                inm_set = ", ".join([f"{key} = ?" for key in inm_updates])
                inm_values = list(inm_updates.values()) + [now, inmueble_id]
                conn.execute(
                    f"UPDATE inmuebles SET {inm_set}, updated_at = datetime(?) WHERE id = ?",
                    inm_values,
                )
        elif parsed.path == "/api/inmueble_update":
            inmueble_id = payload.get("inmueble_id")
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            allowed = (
                "referencia",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "zona",
                "tipo_inmueble",
                "m2",
                "habitaciones",
                "banos",
                "precio_objetivo",
                "precio_valoracion",
                "valor_referencia",
                "estado",
                "lat",
                "lon",
            )
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, inmueble_id]
            conn.execute(
                f"UPDATE inmuebles SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            shared = (
                "tipo_inmueble",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "zona",
                "m2",
                "habitaciones",
                "banos",
                "precio_objetivo",
                "precio_valoracion",
            )
            cap_updates = {key: updates[key] for key in shared if key in updates}
            if "estado" in updates:
                cap_updates["etapa"] = updates["estado"]
            if cap_updates:
                cap_set = ", ".join([f"{key} = ?" for key in cap_updates])
                cap_values = list(cap_updates.values()) + [now, inmueble_id]
                conn.execute(
                    f"UPDATE captaciones SET {cap_set}, updated_at = datetime(?) WHERE inmueble_id = ?",
                    cap_values,
                )
        elif parsed.path == "/api/inmueble_checklist_generate":
            inmueble_id = payload.get("inmueble_id")
            etapa = payload.get("etapa")
            tareas = payload.get("tareas", [])
            if not inmueble_id or not etapa or not isinstance(tareas, list):
                json_response(self, {"error": "inmueble_id, etapa y tareas requeridos"}, status=400)
                return
            conn.execute(
                "DELETE FROM inmueble_checklist WHERE inmueble_id = ? AND etapa = ?",
                (inmueble_id, etapa),
            )
            for tarea in tareas:
                conn.execute(
                    """
                    INSERT INTO inmueble_checklist (
                      id, inmueble_id, etapa, tarea, estado, responsable, fecha_limite, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        inmueble_id,
                        etapa,
                        tarea.get("tarea"),
                        tarea.get("estado") or "Pendiente",
                        tarea.get("responsable"),
                        tarea.get("fecha_limite"),
                        now,
                        now,
                    ),
                )
            audit("inmueble_checklist", inmueble_id, f"Generar checklist {etapa}", usuario=payload.get("usuario"))
        elif parsed.path == "/api/inmueble_checklist_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("tarea", "estado", "responsable", "fecha_limite")
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE inmueble_checklist SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            audit("inmueble_checklist", record_id, "Actualizar tarea", usuario=payload.get("usuario"))
        elif parsed.path == "/api/inmueble_propietarios_update":
            inmueble_id = payload.get("inmueble_id")
            cliente_ids = payload.get("cliente_ids", [])
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            if not isinstance(cliente_ids, list):
                json_response(self, {"error": "cliente_ids invalido"}, status=400)
                return
            conn.execute(
                "DELETE FROM inmueble_propietarios WHERE inmueble_id = ?",
                (inmueble_id,),
            )
            for cliente_id in cliente_ids:
                exists = conn.execute(
                    "SELECT id FROM clientes WHERE id = ?",
                    (cliente_id,),
                ).fetchone()
                if not exists:
                    continue
                conn.execute(
                    """
                    INSERT INTO inmueble_propietarios (
                      id, inmueble_id, cliente_id, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (os.urandom(16).hex(), inmueble_id, cliente_id, now, now),
                )
        elif parsed.path == "/api/inmueble_docs":
            inmueble_id = payload.get("inmueble_id")
            nombre = payload.get("nombre") or ""
            tipo = payload.get("tipo") or ""
            data_uri = payload.get("file_base64") or payload.get("data") or ""
            if not inmueble_id or not data_uri:
                json_response(self, {"error": "inmueble_id y archivo requeridos"}, status=400)
                return
            if "," in data_uri:
                data_uri = data_uri.split(",", 1)[1]
            try:
                file_bytes = base64.b64decode(data_uri)
            except Exception:
                json_response(self, {"error": "Base64 invalido"}, status=400)
                return
            ext = ""
            if nombre and "." in nombre:
                ext = "." + nombre.split(".")[-1].lower()
            if not ext:
                ext = ".pdf"
            folder = UPLOADS / "inmuebles"
            folder.mkdir(parents=True, exist_ok=True)
            doc_id = os.urandom(16).hex()
            filename = f"{doc_id}{ext}"
            file_path = folder / filename
            file_path.write_bytes(file_bytes)
            url = f"/uploads/inmuebles/{filename}"
            conn.execute(
                """
                INSERT INTO inmueble_docs (
                  id, inmueble_id, nombre, url, tipo, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (doc_id, inmueble_id, nombre, url, tipo, now, now),
            )
            audit("inmueble_docs", doc_id, "Subir documento", usuario=payload.get("usuario"))
        elif parsed.path == "/api/demandas":
            conn.execute(
                """
                INSERT INTO demandas (
                  id, empresa_id, cliente_id, tipo, zona, precio_max, m2_min,
                  habitaciones_min, banos_min, estado, prioridad, notas,
                  created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("cliente_id"),
                    payload.get("tipo"),
                    payload.get("zona"),
                    payload.get("precio_max"),
                    payload.get("m2_min"),
                    payload.get("habitaciones_min"),
                    payload.get("banos_min"),
                    payload.get("estado"),
                    payload.get("prioridad"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/visitas":
            conn.execute(
                """
                INSERT INTO visitas (
                  id, empresa_id, inmueble_id, demanda_id, fecha, hora, estado,
                  asesor, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    payload.get("inmueble_id"),
                    payload.get("demanda_id"),
                    payload.get("fecha"),
                    payload.get("hora"),
                    payload.get("estado"),
                    payload.get("asesor"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/clientes":
            nombre = payload.get("nombre")
            if not nombre:
                json_response(self, {"error": "nombre requerido"}, status=400)
                return
            nombre_norm = re.sub(r"\s+", " ", str(nombre)).strip()
            nif = (payload.get("nif") or "").strip()
            dup = None
            if nif:
                nif_norm = re.sub(r"\s+", "", nif).upper()
                dup = conn.execute(
                    "SELECT id FROM clientes WHERE REPLACE(UPPER(nif), ' ', '') = ?",
                    (nif_norm,),
                ).fetchone()
            if not dup:
                dup = conn.execute(
                    "SELECT id FROM clientes WHERE TRIM(UPPER(nombre)) = ?",
                    (nombre_norm.upper(),),
                ).fetchone()
            if dup:
                json_response(self, {"error": "Cliente duplicado", "id": dup["id"]}, status=409)
                return
            cliente_id = payload.get("id") or os.urandom(16).hex()
            conn.execute(
                """
                INSERT INTO clientes (
                  id, nombre, tipo_persona, nif, telefono, email, fecha_nacimiento, direccion,
                  codigo_postal, poblacion, provincia, tipo, perfil, estado, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    cliente_id,
                    nombre,
                    payload.get("tipo_persona"),
                    payload.get("nif"),
                    payload.get("telefono"),
                    payload.get("email"),
                    payload.get("fecha_nacimiento"),
                    payload.get("direccion"),
                    payload.get("codigo_postal"),
                    payload.get("poblacion"),
                    payload.get("provincia"),
                    payload.get("tipo"),
                    payload.get("perfil"),
                    payload.get("estado"),
                    now,
                    now,
                ),
            )
            conn.commit()
            json_response(self, {"ok": True, "id": cliente_id})
            return
        elif parsed.path == "/api/clientes_link":
            cliente_id = payload.get("cliente_id")
            empresa_id = payload.get("empresa_id")
            if not cliente_id or not empresa_id:
                json_response(self, {"error": "cliente_id y empresa_id requeridos"}, status=400)
                return
            cliente_exists = conn.execute(
                "SELECT id FROM clientes WHERE id = ?",
                (cliente_id,),
            ).fetchone()
            empresa_exists = conn.execute(
                "SELECT id FROM empresas WHERE id = ?",
                (empresa_id,),
            ).fetchone()
            if not cliente_exists or not empresa_exists:
                json_response(self, {"error": "Cliente o empresa no encontrados"}, status=400)
                return
            servicio = (payload.get("servicio") or "").strip()
            if not servicio:
                json_response(self, {"error": "servicio requerido"}, status=400)
                return
            sync = {"linked": 0, "docs": 0}
            sync_warning = None
            link_done = False
            for attempt in range(10):
                try:
                    existing = conn.execute(
                        """
                        SELECT id
                        FROM clientes_empresas
                        WHERE cliente_id = ?
                          AND empresa_id = ?
                          AND LOWER(servicio) = LOWER(?)
                        LIMIT 1
                        """,
                        (cliente_id, empresa_id, servicio),
                    ).fetchone()
                    if existing:
                        conn.execute(
                            """
                            UPDATE clientes_empresas
                            SET estado = COALESCE(?, estado),
                                fecha_inicio = COALESCE(?, fecha_inicio),
                                fecha_fin = COALESCE(?, fecha_fin),
                                updated_at = datetime(?)
                            WHERE id = ?
                            """,
                            (
                                payload.get("estado"),
                                payload.get("fecha_inicio"),
                                payload.get("fecha_fin"),
                                now,
                                existing["id"],
                            ),
                        )
                    else:
                        conn.execute(
                            """
                            INSERT INTO clientes_empresas (
                              id, cliente_id, empresa_id, servicio, estado,
                              fecha_inicio, fecha_fin, created_at, updated_at
                            ) VALUES (
                              ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                            )
                            """,
                            (
                                os.urandom(16).hex(),
                                cliente_id,
                                empresa_id,
                                servicio,
                                payload.get("estado"),
                                payload.get("fecha_inicio"),
                                payload.get("fecha_fin"),
                                now,
                                now,
                            ),
                        )
                    link_done = True
                    break
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        json_response(self, {"error": f"db_error: {str(exc)}"}, status=500)
                        return
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    if attempt >= 9:
                        json_response(self, {"error": "database is locked"}, status=503)
                        return
                    time.sleep(0.35 * (attempt + 1))
            if not link_done:
                json_response(self, {"error": "No se pudo vincular el servicio"}, status=503)
                return
            # Sincronización secundaria en best-effort para no bloquear la operación principal.
            if normalize_service_key(servicio) == "seguros":
                try:
                    sync = autolink_uploaded_seguros_for_cliente(conn, cliente_id, empresa_id, now)
                except sqlite3.OperationalError as exc:
                    if "database is locked" in str(exc).lower():
                        sync_warning = "sync_deferred_database_locked"
                    else:
                        raise
                except Exception as exc:
                    sync_warning = f"sync_error: {type(exc).__name__}"
            response = {
                "ok": True,
                "cliente_id": cliente_id,
                "empresa_id": empresa_id,
                "servicio": servicio,
                "sync": sync,
            }
            if sync_warning:
                response["warning"] = sync_warning
            conn.commit()
            json_response(self, response)
            return
        elif parsed.path == "/api/clientes_link_delete":
            rel_id = payload.get("id")
            if not rel_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            row = conn.execute(
                "SELECT * FROM clientes_empresas WHERE id = ?",
                (rel_id,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Vínculo no encontrado"}, status=404)
                return
            conn.execute("DELETE FROM clientes_empresas WHERE id = ?", (rel_id,))
            json_response(
                self,
                {
                    "ok": True,
                    "id": rel_id,
                    "cliente_id": row["cliente_id"],
                    "empresa_id": row["empresa_id"],
                    "servicio": row["servicio"],
                },
            )
            conn.commit()
            return
        elif parsed.path == "/api/cliente_update":
            cliente_id = payload.get("id")
            if not cliente_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = (
                "nombre",
                "tipo_persona",
                "nif",
                "telefono",
                "email",
                "fecha_nacimiento",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "tipo",
                "perfil",
                "estado",
            )
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            sync = {"linked": 0, "docs": 0}
            sync_warning = None
            updated = False
            for attempt in range(10):
                try:
                    set_clause = ", ".join([f"{key} = ?" for key in updates])
                    values = list(updates.values()) + [now, cliente_id]
                    conn.execute(
                        f"UPDATE clientes SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                        values,
                    )
                    # Si el cliente ya tiene servicio seguros, reintentar autovinculación por nombre.
                    try:
                        rels = conn.execute(
                            """
                            SELECT empresa_id
                            FROM clientes_empresas
                            WHERE cliente_id = ?
                              AND LOWER(servicio) = 'seguros'
                            """,
                            (cliente_id,),
                        ).fetchall()
                        for rel in rels:
                            partial = autolink_uploaded_seguros_for_cliente(conn, cliente_id, rel["empresa_id"], now)
                            sync["linked"] += partial.get("linked", 0)
                            sync["docs"] += partial.get("docs", 0)
                    except Exception as exc:
                        # No bloqueamos el guardado de datos personales por un fallo en la sincronización secundaria.
                        sync_warning = f"sync_error: {type(exc).__name__}"
                    updated = True
                    break
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        json_response(self, {"error": f"db_error: {str(exc)}"}, status=500)
                        return
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    if attempt >= 9:
                        json_response(self, {"error": "database is locked"}, status=503)
                        return
                    time.sleep(0.35 * (attempt + 1))
            if not updated:
                json_response(self, {"error": "No se pudo guardar el cliente"}, status=503)
                return
            response = {"ok": True, "id": cliente_id, "sync": sync}
            if sync_warning:
                response["warning"] = sync_warning
            conn.commit()
            json_response(self, response)
            return
        elif parsed.path == "/api/cliente_empresa_update":
            rel_id = payload.get("id")
            if not rel_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("servicio", "estado", "fecha_inicio", "fecha_fin")
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            sync = {"linked": 0, "docs": 0}
            updated = False
            for attempt in range(10):
                try:
                    set_clause = ", ".join([f"{key} = ?" for key in updates])
                    values = list(updates.values()) + [now, rel_id]
                    conn.execute(
                        f"UPDATE clientes_empresas SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                        values,
                    )
                    rel = conn.execute(
                        "SELECT cliente_id, empresa_id, servicio FROM clientes_empresas WHERE id = ?",
                        (rel_id,),
                    ).fetchone()
                    if rel and normalize_service_key(rel["servicio"]) == "seguros":
                        sync = autolink_uploaded_seguros_for_cliente(conn, rel["cliente_id"], rel["empresa_id"], now)
                    updated = True
                    break
                except sqlite3.OperationalError as exc:
                    if "database is locked" not in str(exc).lower():
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        json_response(self, {"error": f"db_error: {str(exc)}"}, status=500)
                        return
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    if attempt >= 9:
                        json_response(self, {"error": "database is locked"}, status=503)
                        return
                    time.sleep(0.35 * (attempt + 1))
            if not updated:
                json_response(self, {"error": "No se pudo guardar el vínculo"}, status=503)
                return
            conn.commit()
            json_response(self, {"ok": True, "id": rel_id, "sync": sync})
            return
        elif parsed.path == "/api/cliente_gestoria_update":
            cliente_id = payload.get("cliente_id")
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            allowed = (
                "tipo_cliente",
                "mod_fiscal",
                "mod_laboral",
                "mod_contable",
                "mod_renta",
                "mod_registro",
                "mod_trafico",
                "mod_puntuales",
                "renta_detalles",
            )
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            existing = conn.execute(
                "SELECT id FROM cliente_gestoria WHERE cliente_id = ?",
                (cliente_id,),
            ).fetchone()
            if existing:
                set_clause = ", ".join([f"{key} = ?" for key in updates])
                values = list(updates.values()) + [now, cliente_id]
                conn.execute(
                    f"UPDATE cliente_gestoria SET {set_clause}, updated_at = datetime(?) WHERE cliente_id = ?",
                    values,
                )
            else:
                new_id = os.urandom(16).hex()
                conn.execute(
                    """
                    INSERT INTO cliente_gestoria (
                      id, cliente_id, tipo_cliente, mod_fiscal, mod_laboral, mod_contable,
                      mod_renta, mod_registro, mod_trafico, mod_puntuales, renta_detalles, created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        new_id,
                        cliente_id,
                        updates.get("tipo_cliente"),
                        updates.get("mod_fiscal"),
                        updates.get("mod_laboral"),
                        updates.get("mod_contable"),
                        updates.get("mod_renta"),
                        updates.get("mod_registro"),
                        updates.get("mod_trafico"),
                        updates.get("mod_puntuales"),
                        updates.get("renta_detalles"),
                        now,
                        now,
                    ),
                )
        elif parsed.path == "/api/acciones":
            servicio = payload.get("servicio")
            if not servicio:
                json_response(self, {"error": "servicio requerido"}, status=400)
                return
            conn.execute(
                """
                INSERT INTO acciones (
                  id, empresa_id, servicio, cliente_id, inmueble_id, cliente_nombre,
                  fecha, hora, tipo, responsable, estado, notas, recordatorio_min, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    os.urandom(16).hex(),
                    empresa["id"],
                    servicio,
                    payload.get("cliente_id"),
                    payload.get("inmueble_id"),
                    payload.get("cliente_nombre"),
                    payload.get("fecha"),
                    payload.get("hora"),
                    payload.get("tipo"),
                    payload.get("responsable"),
                    payload.get("estado"),
                    payload.get("notas"),
                    payload.get("recordatorio_min"),
                    now,
                    now,
                ),
            )
        elif parsed.path == "/api/acciones_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            updates = {}
            for key in (
                "fecha",
                "hora",
                "tipo",
                "responsable",
                "estado",
                "notas",
                "cliente_id",
                "cliente_nombre",
                "inmueble_id",
                "recordatorio_min",
            ):
                if key in payload:
                    updates[key] = payload.get(key)
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE acciones SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
        elif parsed.path == "/api/cliente_profesional":
            cliente_id = payload.get("cliente_id")
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            new_id = os.urandom(16).hex()
            principal = 1 if str(payload.get("principal", "0")) in ("1", "true", "True") else 0
            conn.execute(
                """
                INSERT INTO cliente_profesional (
                  id, cliente_id, cnae, iae, actividad, iban, principal, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    new_id,
                    cliente_id,
                    payload.get("cnae"),
                    payload.get("iae"),
                    payload.get("actividad"),
                    payload.get("iban"),
                    principal,
                    now,
                    now,
                ),
            )
            if principal:
                conn.execute(
                    """
                    UPDATE cliente_profesional
                    SET principal = 0, updated_at = datetime(?)
                    WHERE cliente_id = ? AND id != ?
                    """,
                    (now, cliente_id, new_id),
                )
        elif parsed.path == "/api/cliente_profesional_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("cnae", "iae", "actividad", "iban", "principal")
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            if "principal" in updates:
                updates["principal"] = 1 if str(updates["principal"]) in ("1", "true", "True") else 0
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE cliente_profesional SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            if "principal" in updates and updates["principal"]:
                cliente_id = conn.execute(
                    "SELECT cliente_id FROM cliente_profesional WHERE id = ?",
                    (record_id,),
                ).fetchone()
                if cliente_id:
                    conn.execute(
                        """
                        UPDATE cliente_profesional
                        SET principal = 0, updated_at = datetime(?)
                        WHERE cliente_id = ? AND id != ?
                        """,
                        (now, cliente_id["cliente_id"], record_id),
                    )
        elif parsed.path == "/api/cliente_profesional_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM cliente_profesional WHERE id = ?", (record_id,))
        elif parsed.path == "/api/gestoria_modelos":
            cliente_id = payload.get("cliente_id")
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            new_id = os.urandom(16).hex()
            conn.execute(
                """
                INSERT INTO gestoria_modelos (
                  id, cliente_id, modelo, periodicidad, proxima_fecha, responsable, estado, notas, created_at, updated_at
                ) VALUES (
                  ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                )
                """,
                (
                    new_id,
                    cliente_id,
                    payload.get("modelo"),
                    payload.get("periodicidad"),
                    payload.get("proxima_fecha"),
                    payload.get("responsable"),
                    payload.get("estado"),
                    payload.get("notas"),
                    now,
                    now,
                ),
            )
            audit("gestoria_modelo", new_id, "crear", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_modelos_update":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            allowed = ("modelo", "periodicidad", "proxima_fecha", "responsable", "estado", "notas")
            updates = {key: payload.get(key) for key in allowed if key in payload}
            if not updates:
                json_response(self, {"error": "Sin cambios"}, status=400)
                return
            set_clause = ", ".join([f"{key} = ?" for key in updates])
            values = list(updates.values()) + [now, record_id]
            conn.execute(
                f"UPDATE gestoria_modelos SET {set_clause}, updated_at = datetime(?) WHERE id = ?",
                values,
            )
            audit("gestoria_modelo", record_id, "actualizar", json.dumps(payload), payload.get("usuario"))
        elif parsed.path == "/api/gestoria_modelos_delete":
            record_id = payload.get("id")
            if not record_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            conn.execute("DELETE FROM gestoria_modelos WHERE id = ?", (record_id,))
            audit("gestoria_modelo", record_id, "eliminar", None, payload.get("usuario"))
        elif parsed.path == "/api/hipotecas":
            # try to update existing encargo/estudio instead of creating duplicates
            cliente = payload.get("cliente")
            fecha_encargo = payload.get("fecha_encargo")
            precio = payload.get("precio")
            importe_hipoteca = payload.get("importe_hipoteca")
            estado_busqueda = ("estudio", "encargo", "pendiente")
            where = "empresa_id = ? AND LOWER(TRIM(estado)) IN (?, ?, ?) AND LOWER(TRIM(cliente)) = LOWER(TRIM(?))"
            values = [empresa["id"], *estado_busqueda, cliente]
            if fecha_encargo:
                where += " AND fecha_encargo = ?"
                values.append(fecha_encargo)
            if precio:
                where += " AND precio = ?"
                values.append(precio)
            if importe_hipoteca:
                where += " AND importe_hipoteca = ?"
                values.append(importe_hipoteca)
            existing = conn.execute(
                f"SELECT id FROM hipotecas WHERE {where} LIMIT 1",
                values,
            ).fetchone()

            if existing:
                conn.execute(
                    """
                    UPDATE hipotecas SET
                      banco = COALESCE(?, banco),
                      precio = COALESCE(?, precio),
                      importe_hipoteca = COALESCE(?, importe_hipoteca),
                      porcentaje = COALESCE(?, porcentaje),
                      entrada = COALESCE(?, entrada),
                      comision = COALESCE(?, comision),
                      oficina = COALESCE(?, oficina),
                      fecha_encargo = COALESCE(?, fecha_encargo),
                      encargo = COALESCE(?, encargo),
                      tipo_hipoteca = COALESCE(?, tipo_hipoteca),
                      fecha_firma = COALESCE(?, fecha_firma),
                      cesion = COALESCE(?, cesion),
                      comision_juan = COALESCE(?, comision_juan),
                      comision_modernia = COALESCE(?, comision_modernia),
                      inmobiliaria_compra = COALESCE(?, inmobiliaria_compra),
                      asesor = COALESCE(?, asesor),
                      estado = COALESCE(?, estado),
                      anio = COALESCE(?, anio),
                      updated_at = datetime(?)
                    WHERE id = ?
                    """,
                    (
                        payload.get("banco"),
                        precio,
                        importe_hipoteca,
                        payload.get("porcentaje"),
                        payload.get("entrada"),
                        payload.get("comision"),
                        payload.get("oficina"),
                        payload.get("fecha_encargo"),
                        payload.get("encargo"),
                        payload.get("tipo_hipoteca"),
                        payload.get("fecha_firma"),
                        payload.get("cesion"),
                        payload.get("comision_juan"),
                        payload.get("comision_modernia"),
                        payload.get("inmobiliaria_compra"),
                        payload.get("asesor"),
                        payload.get("estado"),
                        payload.get("anio"),
                        now,
                        existing["id"],
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO hipotecas (
                      id, empresa_id, cliente, banco, precio, importe_hipoteca,
                      porcentaje, entrada, comision, oficina, fecha_encargo,
                      encargo, tipo_hipoteca, fecha_firma, cesion, comision_juan,
                      comision_modernia, inmobiliaria_compra, asesor, estado, anio,
                      created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime(?), datetime(?)
                    )
                    """,
                    (
                        os.urandom(16).hex(),
                        empresa["id"],
                        payload.get("cliente"),
                        payload.get("banco"),
                        payload.get("precio"),
                        payload.get("importe_hipoteca"),
                        payload.get("porcentaje"),
                        payload.get("entrada"),
                        payload.get("comision"),
                        payload.get("oficina"),
                        payload.get("fecha_encargo"),
                        payload.get("encargo"),
                        payload.get("tipo_hipoteca"),
                        payload.get("fecha_firma"),
                        payload.get("cesion"),
                        payload.get("comision_juan"),
                        payload.get("comision_modernia"),
                        payload.get("inmobiliaria_compra"),
                        payload.get("asesor"),
                        payload.get("estado"),
                        payload.get("anio"),
                        now,
                        now,
                    ),
                )
        else:
            hipoteca_id = payload.get("id")
            fecha_firma = payload.get("fecha_firma")
            estado = payload.get("estado", "FIRMADA")
            if not hipoteca_id or not fecha_firma:
                json_response(self, {"error": "id y fecha_firma requeridos"}, status=400)
                return
            conn.execute(
                """
                UPDATE hipotecas SET
                  estado = ?,
                  fecha_firma = ?,
                  anio = COALESCE(?, anio),
                  updated_at = datetime(?)
                WHERE id = ?
                """,
                (estado, fecha_firma, payload.get("anio"), now, hipoteca_id),
            )
        conn.commit()
        json_response(self, {"ok": True})
    def handle_api(self, parsed):
        path = parsed.path
        params = urllib.parse.parse_qs(parsed.query)
        allowed_services = self._auth_allowed_services()
        if allowed_services is not None and path in {
            "/api/clientes",
            "/api/clientes_list",
            "/api/clientes_stats",
            "/api/cliente",
            "/api/cliente_lookup",
            "/api/acciones",
        }:
            if not allowed_services:
                json_response(self, {"error": "Usuario sin servicios asignados"}, status=403)
                return
            current = (params.get("servicio", [""])[0] or "").strip()
            if not current:
                params["servicio"] = [",".join(sorted(allowed_services))]
        conn = get_db(self.db_path)
        self._track_conn(conn)
        if path in ("/api/me", "/api/auth_invite_status", "/api/usuarios"):
            try:
                ensure_usuarios_schema(conn)
                conn.commit()
            except Exception:
                pass

        if path == "/api/me":
            session = self._current_session()
            if not session:
                json_response(self, {"error": "No autenticado"}, status=401)
                return
            user = conn.execute(
                """
                SELECT id, nombre, apellido, usuario, email, servicio, rol, activo
                FROM usuarios
                WHERE id = ? AND activo = 1
                """,
                (session.get("user_id"),),
            ).fetchone()
            if not user:
                delete_auth_session(session.get("token"))
                json_response(
                    self,
                    {"error": "Sesión inválida"},
                    status=401,
                    cookies=[self._build_session_cookie("", max_age=0)],
                )
                return
            refreshed_session = dict(session)
            refreshed_session.update(
                {
                    "nombre": user["nombre"] or "",
                    "apellido": user["apellido"] or "",
                    "usuario": user["usuario"] or "",
                    "email": user["email"] or "",
                    "servicio": user["servicio"] or "",
                    "rol": user["rol"] or "",
                }
            )
            with AUTH_SESSIONS_LOCK:
                token = refreshed_session.get("token")
                if token in AUTH_SESSIONS:
                    AUTH_SESSIONS[token].update(refreshed_session)
            json_response(self, {"ok": True, "user": self._auth_user_payload(refreshed_session)})
            return

        if path == "/api/auth_invite_status":
            token = (params.get("token", [""])[0] or "").strip()
            if not token:
                json_response(self, {"error": "token requerido"}, status=400)
                return
            row = conn.execute(
                """
                SELECT id, nombre, apellido, usuario, email, activo, invite_expires_at
                FROM usuarios
                WHERE invite_token = ?
                LIMIT 1
                """,
                (token,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "Invitación inválida"}, status=404)
                return
            expires_raw = str(row["invite_expires_at"] or "").strip()
            expired = False
            if expires_raw:
                try:
                    dt = datetime.fromisoformat(expires_raw.replace("Z", ""))
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    expired = dt < datetime.now(timezone.utc)
                except Exception:
                    expired = False
            json_response(
                self,
                {
                    "ok": True,
                    "valid": bool(row["activo"]) and bool(token) and not expired,
                    "expired": expired,
                    "user": {
                        "id": row["id"],
                        "nombre": row["nombre"] or "",
                        "apellido": row["apellido"] or "",
                        "usuario": row["usuario"] or "",
                        "email": row["email"] or "",
                    },
                },
            )
            return

        if path == "/api/empresas":
            rows = conn.execute(
                "SELECT id, nombre FROM empresas ORDER BY nombre"
            ).fetchall()
            json_response(self, [dict(r) for r in rows])
            return

        if path == "/api/years":
            years = set()
            tables = [
                "movimientos",
                "hipotecas",
                "alquileres",
                "seguros",
                "captaciones",
                "gestoria_contabilidad",
                "gestoria_trabajos",
                "gestoria_modelos",
                "asesoramientos_financiacion",
                "inmuebles",
                "demandas",
                "visitas",
            ]
            for table in tables:
                try:
                    cols = {
                        row["name"]
                        for row in conn.execute(f"PRAGMA table_info({table})").fetchall()
                    }
                except sqlite3.Error:
                    continue
                if "anio" in cols:
                    try:
                        for row in conn.execute(
                            f"SELECT DISTINCT anio AS y FROM {table} WHERE anio IS NOT NULL"
                        ).fetchall():
                            if row["y"] is not None:
                                years.add(str(row["y"]))
                    except sqlite3.Error:
                        pass
                elif "fecha" in cols:
                    try:
                        for row in conn.execute(
                            f"SELECT DISTINCT strftime('%Y', fecha) AS y FROM {table} WHERE fecha IS NOT NULL"
                        ).fetchall():
                            if row["y"]:
                                years.add(str(row["y"]))
                    except sqlite3.Error:
                        pass
            json_response(self, {"years": sorted(years)})
            return

        if path == "/api/s3_url":
            key = params.get("key", [""])[0]
            if not key:
                json_response(self, {"error": "key requerido"}, status=400)
                return
            client = s3_client()
            if not client:
                bucket, region = s3_config()
                missing = []
                if not bucket:
                    missing.append("AWS_S3_BUCKET")
                if not region:
                    missing.append("AWS_REGION")
                if not S3_BOTO3_AVAILABLE:
                    missing.append("boto3")
                detail = f" (faltan: {', '.join(missing)})" if missing else ""
                json_response(self, {"error": f"S3 no configurado{detail}"}, status=400)
                return
            bucket, _region = s3_config()
            try:
                url = client.generate_presigned_url(
                    "get_object",
                    Params={"Bucket": bucket, "Key": key},
                    ExpiresIn=3600,
                )
            except Exception:
                json_response(self, {"error": "No se pudo firmar el archivo"}, status=500)
                return
            json_response(self, {"url": url})
            return

        if path == "/api/tablas":
            json_response(self, TABLES)
            return

        if path == "/api/resumen":
            rows = conn.execute(
                """
                SELECT e.nombre AS empresa,
                  (SELECT COUNT(*) FROM movimientos m WHERE m.empresa_id = e.id) AS bdt,
                  (SELECT COUNT(*) FROM seguros s WHERE s.empresa_id = e.id) AS seguros,
                  (SELECT COUNT(*) FROM gestoria g WHERE g.empresa_id = e.id) AS gestoria,
                  (SELECT COUNT(*) FROM hipotecas h WHERE h.empresa_id = e.id) AS hipotecas,
                  (SELECT COUNT(*) FROM alquileres a WHERE a.empresa_id = e.id) AS alquileres,
                  (SELECT COUNT(*) FROM inversores i WHERE i.empresa_id = e.id) AS inversores,
                  (SELECT COUNT(*) FROM inversure_operaciones io WHERE io.empresa_id = e.id) AS inversure_ops
                FROM empresas e
                ORDER BY e.nombre
                """
            ).fetchall()
            json_response(self, [dict(r) for r in rows])
            return

        if path == "/api/clientes_stats":
            servicio = (params.get("servicio", [""])[0] or "").strip()
            services = parse_services_param(servicio)
            source = (params.get("source", [""])[0] or "").strip().lower()
            empresa_id = (params.get("empresa_id", [""])[0] or "").strip()
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            normalized_services = [normalize_service_key(s) for s in services]
            if source == "seguros" and ("seguros" in normalized_services or not normalized_services):
                where = ["s.cliente_id IS NOT NULL"]
                values = []
                if empresa_id:
                    where.append("s.empresa_id = ?")
                    values.append(empresa_id)
                where.append(f"({uploaded_policy_filter('s')} OR ? = 0)")
                values.append(1 if uploaded_only else 0)
                total = conn.execute(
                    f"""
                    SELECT COUNT(DISTINCT c.id) AS total
                    FROM clientes c
                    JOIN seguros s ON s.cliente_id = c.id
                    WHERE {' AND '.join(where)}
                    """,
                    values,
                ).fetchone()
                json_response(self, {"total": total["total"] if total else 0})
                return
            if services:
                placeholders = ",".join(["?"] * len(services))
                total = conn.execute(
                    f"""
                    SELECT COUNT(DISTINCT c.id) AS total
                    FROM clientes c
                    JOIN clientes_empresas ce ON ce.cliente_id = c.id
                    WHERE LOWER(ce.servicio) IN ({placeholders})
                    """,
                    services,
                ).fetchone()
            else:
                total = conn.execute("SELECT COUNT(*) AS total FROM clientes").fetchone()
            json_response(self, {"total": total["total"] if total else 0})
            return

        if path == "/api/postal_lookup":
            cp_raw = params.get("cp", [""])[0]
            cp = normalize_postal_code(cp_raw)
            if not cp:
                json_response(self, {"error": "cp requerido"}, status=400)
                return
            rows = conn.execute(
                "SELECT poblacion, provincia FROM postal_catalogo WHERE codigo_postal = ?",
                (cp,),
            ).fetchall()
            opciones = []
            poblaciones = []
            provincia = ""
            for row in rows:
                poblacion = (row["poblacion"] or "").strip()
                prov = (row["provincia"] or "").strip()
                if not provincia and prov:
                    provincia = prov
                if poblacion:
                    poblaciones.append(poblacion)
                    opciones.append({"poblacion": poblacion, "provincia": prov or provincia})
            if not provincia:
                provincia = POSTAL_PROVINCES.get(cp[:2], "")
            unique_poblaciones = sorted(set(poblaciones))
            poblacion_value = unique_poblaciones[0] if len(unique_poblaciones) == 1 else ""
            json_response(
                self,
                {
                    "codigo_postal": cp,
                    "poblacion": poblacion_value,
                    "provincia": provincia,
                    "opciones": opciones,
                },
            )
            return

        if path == "/api/clientes_list":
            servicio = (params.get("servicio", [""])[0] or "").strip()
            services = parse_services_param(servicio)
            source = (params.get("source", [""])[0] or "").strip().lower()
            empresa_id = (params.get("empresa_id", [""])[0] or "").strip()
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            normalized_services = [normalize_service_key(s) for s in services]
            if source == "seguros" and ("seguros" in normalized_services or not normalized_services):
                where = ["s.cliente_id IS NOT NULL"]
                values = []
                if empresa_id:
                    where.append("s.empresa_id = ?")
                    values.append(empresa_id)
                where.append(f"({uploaded_policy_filter('s')} OR ? = 0)")
                values.append(1 if uploaded_only else 0)
                rows = conn.execute(
                    f"""
                    SELECT DISTINCT c.id, c.nombre
                    FROM clientes c
                    JOIN seguros s ON s.cliente_id = c.id
                    WHERE {' AND '.join(where)}
                    ORDER BY c.nombre
                    """,
                    values,
                ).fetchall()
                json_response(self, [dict(r) for r in rows])
                return
            if services:
                placeholders = ",".join(["?"] * len(services))
                rows = conn.execute(
                    f"""
                    SELECT DISTINCT c.id, c.nombre
                    FROM clientes c
                    JOIN clientes_empresas ce ON ce.cliente_id = c.id
                    WHERE LOWER(ce.servicio) IN ({placeholders})
                    ORDER BY c.nombre
                    """,
                    services,
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT id, nombre FROM clientes ORDER BY nombre"
                ).fetchall()
            json_response(self, [dict(r) for r in rows])
            return

        if path == "/api/fin_inmobiliarias":
            rows = conn.execute(
                """
                SELECT DISTINCT TRIM(inmobiliaria_compra) AS nombre
                FROM hipotecas
                WHERE inmobiliaria_compra IS NOT NULL AND TRIM(inmobiliaria_compra) <> ''
                ORDER BY nombre
                """
            ).fetchall()
            json_response(self, {"items": [row["nombre"] for row in rows if row["nombre"]]})
            return

        if path == "/api/debug_db_path":
            resolved = Path(self.db_path).expanduser().resolve()
            ocr_resolved = Path(self.ocr_db_path).expanduser().resolve()
            exists = resolved.exists()
            stat = resolved.stat() if exists else None
            ocr_exists = ocr_resolved.exists()
            ocr_stat = ocr_resolved.stat() if ocr_exists else None
            json_response(
                self,
                {
                    "db_path": str(resolved),
                    "exists": exists,
                    "size_bytes": stat.st_size if stat else 0,
                    "ocr_db_path": str(ocr_resolved),
                    "ocr_exists": ocr_exists,
                    "ocr_size_bytes": ocr_stat.st_size if ocr_stat else 0,
                    "mtime": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat()
                    if stat
                    else "",
                },
            )
            return

        if path == "/api/cliente_ficha":
            cliente_id = params.get("id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            servicio = (params.get("servicio", [""])[0] or "").strip()
            services = parse_services_param(servicio)
            ficha = build_cliente_ficha_payload(conn, cliente_id, services)
            if not ficha:
                json_response(self, {"error": "Cliente no encontrado"}, status=404)
                return
            if ficha.get("error"):
                json_response(self, {"error": ficha["error"]}, status=404)
                return
            json_response(self, ficha)
            return

        if path == "/api/cliente":
            # Compatibilidad con frontend antiguo.
            cliente_id = params.get("id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            servicio = (params.get("servicio", [""])[0] or "").strip()
            services = parse_services_param(servicio)
            ficha = build_cliente_ficha_payload(conn, cliente_id, services)
            if not ficha:
                json_response(self, {"error": "Cliente no encontrado"}, status=404)
                return
            if ficha.get("error"):
                json_response(self, {"error": ficha["error"]}, status=404)
                return
            json_response(
                self,
                {
                    "cliente": ficha.get("cliente", {}),
                    "empresas": ficha.get("empresas", []),
                    "servicios": ficha.get("empresas", []),
                },
            )
            return

        if path == "/api/cliente_lookup":
            nif = (params.get("nif", [""])[0] or "").strip()
            if not nif:
                json_response(self, {"error": "nif requerido"}, status=400)
                return
            servicio = (params.get("servicio", [""])[0] or "").strip()
            services = parse_services_param(servicio)
            nif_norm = normalize_nif(nif)
            row = conn.execute(
                """
                SELECT id, nombre, nif, telefono, email
                FROM clientes
                WHERE REPLACE(REPLACE(REPLACE(UPPER(nif), ' ', ''), '-', ''), '.', '') = ?
                """,
                (nif_norm,),
            ).fetchone()
            if not row:
                json_response(self, {"found": False})
                return
            has_service = True
            if services:
                has_service = cliente_has_servicio(conn, row["id"], services)
            if not has_service:
                json_response(
                    self,
                    {
                        "found": True,
                        "restricted": True,
                        "cliente": {"id": row["id"]},
                        "has_servicio": False,
                    },
                )
                return
            servicios = conn.execute(
                """
                SELECT ce.servicio, ce.estado, ce.empresa_id, e.nombre AS empresa
                FROM clientes_empresas ce
                LEFT JOIN empresas e ON e.id = ce.empresa_id
                WHERE ce.cliente_id = ?
                """,
                (row["id"],),
            ).fetchall()
            json_response(
                self,
                {
                    "found": True,
                    "cliente": dict(row),
                    "servicios": [dict(r) for r in servicios],
                    "has_servicio": True,
                },
            )
            return

        if path == "/api/ocr_job":
            job_id = params.get("id", [""])[0]
            if not job_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            ocr_conn = get_db(self.ocr_db_path)
            self._track_conn(ocr_conn)
            row = ocr_conn.execute(
                """
                SELECT id, kind, status, result_json, error, created_at, started_at, finished_at
                FROM ocr_jobs
                WHERE id = ?
                """,
                (job_id,),
            ).fetchone()
            if not row:
                json_response(self, {"error": "job no encontrado"}, status=404)
                return
            result = None
            if row["result_json"]:
                try:
                    result = json.loads(row["result_json"])
                except Exception:
                    result = None
            json_response(
                self,
                {
                    "id": row["id"],
                    "kind": row["kind"],
                    "status": row["status"],
                    "result": result,
                    "error": row["error"],
                    "created_at": row["created_at"],
                    "started_at": row["started_at"],
                    "finished_at": row["finished_at"],
                },
            )
            return

        if path == "/api/cliente_gestoria":
            cliente_id = params.get("cliente_id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            row = conn.execute(
                """
                SELECT tipo_cliente, mod_fiscal, mod_laboral, mod_contable,
                       mod_renta, mod_registro, mod_trafico, mod_puntuales, renta_detalles
                FROM cliente_gestoria
                WHERE cliente_id = ?
                """,
                (cliente_id,),
            ).fetchone()
            json_response(self, {"row": dict(row) if row else {}})
            return

        if path == "/api/acciones":
            empresa_id = params.get("empresa_id", [""])[0]
            servicio = params.get("servicio", [""])[0]
            cliente_id = params.get("cliente_id", [""])[0]
            inmueble_id = params.get("inmueble_id", [""])[0]
            if not servicio:
                json_response(self, {"error": "servicio requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT
                  a.id, a.cliente_id, a.fecha, a.hora,
                  COALESCE(c.nombre, a.cliente_nombre) AS cliente,
                  a.tipo, a.responsable, a.estado, a.notas, a.servicio, a.recordatorio_min, a.inmueble_id
                FROM acciones a
                LEFT JOIN clientes c ON c.id = a.cliente_id
                WHERE a.servicio = ?
                  AND (? = '' OR a.empresa_id = ?)
                  AND (? = '' OR a.cliente_id = ?)
                  AND (? = '' OR a.inmueble_id = ?)
                ORDER BY a.fecha DESC, a.hora DESC
                LIMIT 300
                """,
                (servicio, empresa_id, empresa_id, cliente_id, cliente_id, inmueble_id, inmueble_id),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/fin_asesoramientos":
            empresa_id = params.get("empresa_id", [""])[0]
            q = params.get("q", [""])[0].strip().lower()
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT *
                FROM asesoramientos_financiacion
                WHERE empresa_id = ?
                ORDER BY created_at DESC
                """,
                (empresa_id,),
            ).fetchall()
            data = []
            for row in rows:
                row_dict = dict(row)
                missing = fin_missing_fields(row_dict)
                row_dict["missing_fields"] = missing
                row_dict["missing_count"] = len(missing)
                data.append(row_dict)
            if q:
                filtered = []
                for row in data:
                    hay = " ".join([str(row.get(key, "") or "") for key in (
                        "cliente1_nombre",
                        "cliente2_nombre",
                        "cliente1_dni",
                        "cliente2_dni",
                        "inmobiliaria_asesor",
                        "asesor",
                        "estado",
                        "entidades_financieras",
                    )]).lower()
                    if q in hay:
                        filtered.append(row)
                data = filtered
            json_response(self, {"rows": data})
            return

        if path == "/api/seguros_ofertas":
            cliente_id = params.get("cliente_id", [""])[0]
            where = []
            values = []
            if cliente_id:
                where.append("o.cliente_id = ?")
                values.append(cliente_id)
            where_clause = f"WHERE {' AND '.join(where)}" if where else ""
            rows = conn.execute(
                f"""
                SELECT o.id, o.cliente_id, o.ramo, o.compania, o.propuesta, o.estado,
                       o.motivo, o.fecha, o.responsable, o.notas,
                       COALESCE(c.nombre, '') AS cliente
                FROM seguros_ofertas o
                LEFT JOIN clientes c ON c.id = o.cliente_id
                {where_clause}
                ORDER BY o.fecha DESC, o.created_at DESC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_preferencias":
            cliente_id = params.get("cliente_id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            row = conn.execute(
                """
                SELECT cliente_id, prioridad_precio, prioridad_compania, prioridad_coberturas, notas
                FROM seguros_preferencias
                WHERE cliente_id = ?
                """,
                (cliente_id,),
            ).fetchone()
            json_response(self, {"row": dict(row) if row else {}})
            return

        if path == "/api/seguros_referidos":
            rows = conn.execute(
                """
                SELECT r.id, r.cliente_id, r.referido_por, r.notas,
                       COALESCE(c.nombre, '') AS cliente
                FROM seguros_referidos r
                LEFT JOIN clientes c ON c.id = r.cliente_id
                ORDER BY r.created_at DESC
                """
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_campanas":
            rows = conn.execute(
                """
                SELECT id, compania, nombre, ramo, origen, fecha_inicio, fecha_fin, descripcion, url
                FROM seguros_campanas
                ORDER BY fecha_inicio DESC, created_at DESC
                """
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_comisiones":
            rows = conn.execute(
                """
                SELECT id, compania, ramo, porcentaje, vigencia_desde, vigencia_hasta, notas
                FROM seguros_comisiones
                ORDER BY compania, ramo
                """
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_cliente":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            tomador = (params.get("tomador", [""])[0] or "").strip()
            autolink = (params.get("autolink", ["0"])[0] or "").strip() in ("1", "true", "yes")
            now = datetime.now(timezone.utc).isoformat()
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            uploaded_filter_sql = uploaded_policy_filter()
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            where = ["cliente_id = ?"]
            values = [cliente_id]
            if empresa_id:
                where.append("empresa_id = ?")
                values.append(empresa_id)
            if uploaded_only:
                where.append(uploaded_policy_filter())
            rows = conn.execute(
                f"""
                SELECT
                  id, cliente_id, empresa_id, compania, ramo, poliza_numero,
                  fecha_efecto, fecha_vencimiento, estado, prima_neta, prima_total,
                  tomador, estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                  colaborador, produccion, mes_creacion, poliza_key, poliza_url,
                  fecha_baja, motivo_baja, estado_poliza, poliza_origen_id, poliza_sustituta_id, version_grupo, tipo_vigencia
                FROM seguros
                WHERE {' AND '.join(where)}
                ORDER BY COALESCE(fecha_efecto, created_at) DESC
                """,
                values,
            ).fetchall()
            if not rows and tomador:
                tomador_rows = conn.execute(
                    f"""
                    SELECT
                      id, cliente_id, empresa_id, compania, ramo, poliza_numero,
                      fecha_efecto, fecha_vencimiento, estado, prima_neta, prima_total,
                      tomador, estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                      colaborador, produccion, mes_creacion, poliza_key, poliza_url,
                      fecha_baja, motivo_baja, estado_poliza, poliza_origen_id, poliza_sustituta_id, version_grupo, tipo_vigencia
                    FROM seguros
                    WHERE UPPER(TRIM(tomador)) = UPPER(TRIM(?))
                      AND (? = '' OR empresa_id = ?)
                      AND ({uploaded_filter_sql} OR ? = 0)
                    ORDER BY COALESCE(fecha_efecto, created_at) DESC
                    """,
                    (tomador, empresa_id, empresa_id, 1 if uploaded_only else 0),
                ).fetchall()
                if not tomador_rows:
                    # Fallback robusto: coincidencia por tokens normalizados de nombre (orden flexible).
                    wanted_tokens = [t for t in normalize_lookup_text(tomador).split(" ") if len(t) >= 3]
                    if wanted_tokens:
                        candidates = conn.execute(
                            f"""
                            SELECT
                              id, cliente_id, empresa_id, compania, ramo, poliza_numero,
                              fecha_efecto, fecha_vencimiento, estado, prima_neta, prima_total,
                              tomador, estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                              colaborador, produccion, mes_creacion, poliza_key, poliza_url,
                              fecha_baja, motivo_baja, estado_poliza, poliza_origen_id, poliza_sustituta_id, version_grupo, tipo_vigencia
                            FROM seguros
                            WHERE tomador IS NOT NULL AND TRIM(tomador) <> ''
                              AND (? = '' OR empresa_id = ?)
                              AND ({uploaded_filter_sql} OR ? = 0)
                            ORDER BY COALESCE(fecha_efecto, created_at) DESC
                            """,
                            (empresa_id, empresa_id, 1 if uploaded_only else 0),
                        ).fetchall()
                        matched = []
                        wanted_set = set(wanted_tokens)
                        for cand in candidates:
                            cand_tokens = {
                                t for t in normalize_lookup_text(cand["tomador"] or "").split(" ") if len(t) >= 3
                            }
                            if not cand_tokens:
                                continue
                            # Aceptamos cuando todos los tokens buscados están presentes.
                            if wanted_set.issubset(cand_tokens):
                                matched.append(cand)
                                continue
                            # O cuando el solape es muy alto para evitar falsos negativos.
                            overlap = len(wanted_set.intersection(cand_tokens))
                            if wanted_set and (overlap / max(1, len(wanted_set))) >= 0.8 and overlap >= 2:
                                matched.append(cand)
                        tomador_rows = matched
                if tomador_rows and autolink:
                    distinct_cliente_ids = {
                        str(r["cliente_id"] or "").strip()
                        for r in tomador_rows
                        if str(r["cliente_id"] or "").strip()
                    }
                    # Solo relink automático si la coincidencia es clara.
                    can_relink = len(tomador_rows) == 1 or len(distinct_cliente_ids) <= 1
                    if can_relink:
                        conn.executemany(
                            "UPDATE seguros SET cliente_id = ?, updated_at = datetime(?) WHERE id = ?",
                            [(cliente_id, now, r["id"]) for r in tomador_rows],
                        )
                        if empresa_id:
                            ensure_cliente_servicio_link(conn, cliente_id, empresa_id, "seguros", now)
                        rows = conn.execute(
                            f"""
                            SELECT
                              id, cliente_id, empresa_id, compania, ramo, poliza_numero,
                              fecha_efecto, fecha_vencimiento, estado, prima_neta, prima_total,
                              tomador, estado_renovacion, renovacion_fecha, nueva_poliza_ref,
                              colaborador, produccion, mes_creacion, poliza_key, poliza_url,
                              fecha_baja, motivo_baja, estado_poliza, poliza_origen_id, poliza_sustituta_id, version_grupo, tipo_vigencia
                            FROM seguros
                            WHERE cliente_id = ?
                              AND (? = '' OR empresa_id = ?)
                              AND ({uploaded_filter_sql} OR ? = 0)
                            ORDER BY COALESCE(fecha_efecto, created_at) DESC
                            """,
                            (cliente_id, empresa_id, empresa_id, 1 if uploaded_only else 0),
                        ).fetchall()
                if not rows:
                    rows = [
                        {
                            "id": r["id"],
                            "cliente_id": r["cliente_id"],
                            "empresa_id": r["empresa_id"],
                            "compania": r["compania"],
                            "ramo": r["ramo"],
                            "poliza_numero": r["poliza_numero"],
                            "fecha_efecto": r["fecha_efecto"],
                            "fecha_vencimiento": r["fecha_vencimiento"],
                            "estado": r["estado"],
                            "prima_neta": r["prima_neta"],
                            "prima_total": r["prima_total"],
                            "tomador": r["tomador"],
                            "estado_renovacion": r["estado_renovacion"],
                            "renovacion_fecha": r["renovacion_fecha"],
                            "nueva_poliza_ref": r["nueva_poliza_ref"],
                            "colaborador": r["colaborador"],
                            "produccion": r["produccion"],
                            "mes_creacion": r["mes_creacion"],
                            "poliza_key": r["poliza_key"],
                            "poliza_url": r["poliza_url"],
                            "fecha_baja": r["fecha_baja"],
                            "motivo_baja": r["motivo_baja"],
                            "estado_poliza": r["estado_poliza"],
                            "poliza_origen_id": r["poliza_origen_id"],
                            "poliza_sustituta_id": r["poliza_sustituta_id"],
                            "version_grupo": r["version_grupo"],
                            "tipo_vigencia": r["tipo_vigencia"],
                        }
                        for r in tomador_rows
                    ]
            rows_dict = [dict(r) for r in rows]
            for row in rows_dict:
                if row.get("cliente_id"):
                    ensure_cliente_servicio_link(conn, row.get("cliente_id"), row.get("empresa_id"), "seguros", now)
                    ensure_seguro_doc_link(conn, row, now)
            json_response(self, {"rows": rows_dict})
            return

        if path == "/api/seguros_insights":
            empresa_id = params.get("empresa_id", [""])[0]
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            in_vigor_expr = in_vigor_policy_filter()
            compania_expr = "LOWER(TRIM(compania))"
            exclude_sin_seguro = f"({compania_expr} IS NULL OR {compania_expr} = '' OR {compania_expr} != 'sin seguro')"
            uploaded_clause = uploaded_policy_filter()
            por_ramo_raw = conn.execute(
                f"""
                SELECT ramo
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {in_vigor_expr}
                  AND {exclude_sin_seguro}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchall()
            por_ramo_map = {}
            for row in por_ramo_raw:
                ramo_label = canonicalize_ramo(row["ramo"]) or "Sin ramo"
                por_ramo_map[ramo_label] = por_ramo_map.get(ramo_label, 0) + 1
            por_ramo = [
                {"ramo": ramo, "total": total}
                for ramo, total in sorted(por_ramo_map.items(), key=lambda item: item[1], reverse=True)
            ]
            por_compania = conn.execute(
                f"""
                SELECT compania, COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {in_vigor_expr}
                  AND {exclude_sin_seguro}
                GROUP BY compania
                ORDER BY total DESC
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchall()
            ofertas_estado = conn.execute(
                """
                SELECT estado, COUNT(*) AS total
                FROM seguros_ofertas
                GROUP BY estado
                ORDER BY total DESC
                """
            ).fetchall()
            preferencias = conn.execute(
                """
                SELECT
                  SUM(CASE WHEN prioridad_precio = 1 THEN 1 ELSE 0 END) AS prioriza_precio,
                  SUM(CASE WHEN prioridad_compania = 1 THEN 1 ELSE 0 END) AS prioriza_compania,
                  SUM(CASE WHEN prioridad_coberturas = 1 THEN 1 ELSE 0 END) AS prioriza_coberturas,
                  COUNT(*) AS total
                FROM seguros_preferencias
                """
            ).fetchone()
            json_response(
                self,
                {
                    "por_ramo": por_ramo,
                    "por_compania": [dict(r) for r in por_compania],
                    "ofertas_estado": [dict(r) for r in ofertas_estado],
                    "preferencias": dict(preferencias) if preferencias else {},
                },
            )
            return

        if path == "/api/seguros_eventos":
            seguro_id = params.get("seguro_id", [""])[0]
            cliente_id = params.get("cliente_id", [""])[0]
            if not seguro_id and not cliente_id:
                json_response(self, {"error": "seguro_id o cliente_id requerido"}, status=400)
                return
            where = []
            values = []
            if seguro_id:
                where.append("seguro_id = ?")
                values.append(seguro_id)
            if cliente_id:
                where.append("cliente_id = ?")
                values.append(cliente_id)
            rows = conn.execute(
                f"""
                SELECT id, seguro_id, cliente_id, empresa_id, tipo, fecha, motivo, payload_json, created_at
                FROM seguros_eventos
                WHERE {' AND '.join(where)}
                ORDER BY created_at DESC
                LIMIT 500
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_reclamaciones":
            seguro_id = params.get("seguro_id", [""])[0]
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            if not seguro_id and not cliente_id and not empresa_id:
                json_response(self, {"error": "seguro_id, cliente_id o empresa_id requerido"}, status=400)
                return
            where = []
            values = []
            if seguro_id:
                where.append("seguro_id = ?")
                values.append(seguro_id)
            if cliente_id:
                where.append("cliente_id = ?")
                values.append(cliente_id)
            if empresa_id:
                where.append("empresa_id = ?")
                values.append(empresa_id)
            rows = conn.execute(
                f"""
                SELECT id, seguro_id, cliente_id, empresa_id, estado, canal, fecha_apertura, fecha_cierre,
                       asunto, detalle, resolucion, created_at, updated_at
                FROM seguros_reclamaciones
                WHERE {' AND '.join(where)}
                ORDER BY COALESCE(fecha_apertura, created_at) DESC
                LIMIT 500
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_compliance_kpis":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            uploaded_clause = uploaded_policy_filter()
            total = conn.execute(
                f"SELECT COUNT(*) total FROM seguros WHERE empresa_id = ? AND ({uploaded_clause})",
                (empresa_id,),
            ).fetchone()
            ipid = conn.execute(
                "SELECT COUNT(DISTINCT seguro_id) total FROM seguros_ipid_log WHERE empresa_id = ?",
                (empresa_id,),
            ).fetchone()
            abiertas = conn.execute(
                """
                SELECT COUNT(*) total
                FROM seguros_reclamaciones
                WHERE empresa_id = ? AND LOWER(COALESCE(estado,'')) IN ('abierta','en curso','pendiente')
                """,
                (empresa_id,),
            ).fetchone()
            total_val = int(total["total"] if total and total["total"] is not None else 0)
            ipid_val = int(ipid["total"] if ipid and ipid["total"] is not None else 0)
            abiertas_val = int(abiertas["total"] if abiertas and abiertas["total"] is not None else 0)
            json_response(
                self,
                {
                    "polizas_subidas": total_val,
                    "ipid_registrados": ipid_val,
                    "ipid_pendientes": max(total_val - ipid_val, 0),
                    "reclamaciones_abiertas": abiertas_val,
                },
            )
            return

        if path == "/api/cliente_profesional":
            cliente_id = params.get("cliente_id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT id, cnae, iae, actividad, iban, principal
                FROM cliente_profesional
                WHERE cliente_id = ?
                ORDER BY created_at DESC
                """,
                (cliente_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_modelos":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            scope = params.get("scope", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            if cliente_id:
                rows = conn.execute(
                    """
                    SELECT id, modelo, periodicidad, proxima_fecha, responsable, estado, notas
                    FROM gestoria_modelos
                    WHERE cliente_id = ?
                    ORDER BY proxima_fecha DESC
                    """,
                    (cliente_id,),
                ).fetchall()
                json_response(self, {"rows": [dict(r) for r in rows]})
                return
            today = datetime.now().date()
            next_30 = today + timedelta(days=30)
            where = ["ce.empresa_id = ?"]
            values = [empresa_id]
            if scope == "proximos":
                where.append("m.proxima_fecha IS NOT NULL")
                where.append("date(m.proxima_fecha) BETWEEN date(?) AND date(?)")
                values.extend([today.isoformat(), next_30.isoformat()])
            elif scope == "vencidos":
                where.append("m.proxima_fecha IS NOT NULL")
                where.append("date(m.proxima_fecha) < date(?)")
                where.append("(m.estado IS NULL OR LOWER(m.estado) != 'presentado')")
                values.append(today.isoformat())
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT m.id, m.modelo, m.periodicidad, m.proxima_fecha, m.responsable, m.estado, m.notas,
                       COALESCE(c.nombre, '') AS cliente
                FROM gestoria_modelos m
                JOIN clientes c ON c.id = m.cliente_id
                JOIN clientes_empresas ce ON ce.cliente_id = c.id
                WHERE {where_clause}
                ORDER BY m.proxima_fecha ASC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_trabajos":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            where = []
            values = []
            if cliente_id:
                where.append("gt.cliente_id = ?")
                values.append(cliente_id)
            if empresa_id:
                where.append("gt.empresa_id = ?")
                values.append(empresa_id)
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT gt.id, gt.tipo_trabajo, gt.estado, gt.fecha_inicio, gt.fecha_fin,
                       gt.sla_dias, gt.responsable, gt.importe, gt.notas, gt.cliente_id,
                       COALESCE(c.nombre, '') AS cliente
                FROM gestoria_trabajos gt
                LEFT JOIN clientes c ON c.id = gt.cliente_id
                WHERE {where_clause}
                ORDER BY gt.created_at DESC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_docs":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            service = (params.get("service", [""])[0] or "").strip().lower()
            limit = params.get("limit", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            if cliente_id:
                where = ["cliente_id = ?"]
                values = [cliente_id]
                if service:
                    where.append(
                        "(LOWER(COALESCE(referencia_tipo, '')) = ? OR LOWER(COALESCE(tipo, '')) = ?)"
                    )
                    values.extend([service, service])
                where_clause = " AND ".join(where)
                rows = conn.execute(
                    f"""
                    SELECT id, nombre, tipo, fecha, estado, notas, doc_key, doc_url,
                           referencia_tipo, referencia_id
                    FROM gestoria_docs
                    WHERE {where_clause}
                    ORDER BY created_at DESC
                    """,
                    values,
                ).fetchall()
                json_response(self, {"rows": [dict(r) for r in rows]})
                return
            limit_clause = "LIMIT 50"
            if limit.isdigit():
                limit_clause = f"LIMIT {int(limit)}"
            rows = conn.execute(
                f"""
                SELECT d.id, d.nombre, d.tipo, d.fecha, d.estado, d.notas,
                       COALESCE(c.nombre, '') AS cliente
                FROM gestoria_docs d
                LEFT JOIN clientes c ON c.id = d.cliente_id
                WHERE d.empresa_id = ?
                ORDER BY d.fecha DESC
                {limit_clause}
                """,
                (empresa_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_contabilidad":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            q = params.get("q", [""])[0].strip()
            seguros_only = (params.get("seguros_only", ["0"])[0] or "0").strip() in ("1", "true", "yes")
            where = ["gc.empresa_id = ?"]
            values = [empresa_id]
            if seguros_only:
                where.append(seguros_contabilidad_where_clause("gc"))
            if q:
                where.append("(gc.concepto LIKE ? OR c.nombre LIKE ?)")
                values.extend([f"%{q}%", f"%{q}%"])
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT gc.id, gc.fecha, gc.concepto, gc.gestion, gc.tipo, gc.importe, gc.notas,
                       gc.cliente_id, gc.cliente_ids_json, gc.seguro_id,
                       COALESCE(NULLIF(gc.poliza_numero, ''), s.poliza_numero, '') AS poliza_numero,
                       COALESCE(c.nombre, '') AS cliente
                FROM gestoria_contabilidad gc
                LEFT JOIN clientes c ON c.id = gc.cliente_id
                LEFT JOIN seguros s ON s.id = gc.seguro_id
                WHERE {where_clause}
                ORDER BY gc.fecha DESC
                LIMIT 300
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_facturas":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            where = []
            values = []
            if cliente_id:
                where.append("f.cliente_id = ?")
                values.append(cliente_id)
            if empresa_id:
                where.append("f.empresa_id = ?")
                values.append(empresa_id)
            where_clause = " AND ".join(where) if where else "1=1"
            rows = conn.execute(
                f"""
                SELECT f.id, f.fecha_emision, f.numero, f.tipo, f.total, f.estado_ocr, f.doc_key,
                       COALESCE(t.nombre, '') AS tercero
                FROM gestoria_facturas f
                LEFT JOIN gestoria_terceros t ON t.id = f.tercero_id
                WHERE {where_clause}
                ORDER BY f.created_at DESC
                LIMIT 300
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_asientos":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            where = []
            values = []
            if cliente_id:
                where.append("a.cliente_id = ?")
                values.append(cliente_id)
            if empresa_id:
                where.append("a.empresa_id = ?")
                values.append(empresa_id)
            where_clause = " AND ".join(where) if where else "1=1"
            rows = conn.execute(
                f"""
                SELECT a.id, a.fecha, a.concepto, a.referencia, a.total_debe, a.total_haber,
                       COALESCE(f.numero, '') AS factura_numero, COALESCE(f.doc_key, '') AS factura_doc_key
                FROM gestoria_asientos a
                LEFT JOIN gestoria_facturas f ON f.id = a.factura_id
                WHERE {where_clause}
                ORDER BY a.created_at DESC
                LIMIT 300
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/gestoria_libros":
            empresa_id = params.get("empresa_id", [""])[0]
            cliente_id = (params.get("cliente_id", [""])[0] or "").strip()
            desde = (params.get("desde", [""])[0] or "").strip()
            hasta = (params.get("hasta", [""])[0] or "").strip()
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            date_clause = ""
            values = [empresa_id]
            if cliente_id:
                date_clause += " AND a.cliente_id = ?"
                values.append(cliente_id)
            if desde:
                date_clause += " AND a.fecha >= ?"
                values.append(desde)
            if hasta:
                date_clause += " AND a.fecha <= ?"
                values.append(hasta)
            diario = conn.execute(
                f"""
                SELECT a.id AS asiento_id, a.fecha, a.concepto, a.referencia,
                       l.cuenta, l.descripcion, l.debe, l.haber,
                       l.impuesto_tipo, l.impuesto_pct,
                       COALESCE(t.nombre, '') AS tercero,
                       COALESCE(t.nif, '') AS tercero_nif,
                       COALESCE(f.numero, '') AS factura_numero,
                       COALESCE(f.fecha_emision, '') AS factura_fecha,
                       COALESCE(f.total, 0) AS factura_total,
                       COALESCE(f.tipo, '') AS tipo_factura
                FROM gestoria_asientos a
                JOIN gestoria_asiento_lineas l ON l.asiento_id = a.id
                LEFT JOIN gestoria_terceros t ON t.id = l.tercero_id
                LEFT JOIN gestoria_facturas f ON f.id = a.factura_id
                WHERE a.empresa_id = ? {date_clause}
                ORDER BY a.fecha ASC, a.created_at ASC, l.cuenta ASC
                """,
                values,
            ).fetchall()
            mayor = conn.execute(
                f"""
                SELECT l.cuenta,
                       ROUND(SUM(COALESCE(l.debe, 0)), 2) AS debe,
                       ROUND(SUM(COALESCE(l.haber, 0)), 2) AS haber,
                       ROUND(SUM(COALESCE(l.debe, 0) - COALESCE(l.haber, 0)), 2) AS saldo
                FROM gestoria_asientos a
                JOIN gestoria_asiento_lineas l ON l.asiento_id = a.id
                WHERE a.empresa_id = ? {date_clause}
                GROUP BY l.cuenta
                ORDER BY l.cuenta ASC
                """,
                values,
            ).fetchall()
            iva_values = [empresa_id]
            iva_clause = ""
            if desde:
                iva_clause += " AND f.fecha_emision >= ?"
                iva_values.append(desde)
            if hasta:
                iva_clause += " AND f.fecha_emision <= ?"
                iva_values.append(hasta)
            iva_compras = conn.execute(
                f"""
                SELECT f.id, f.fecha_emision, f.numero, COALESCE(t.nombre, '') AS tercero,
                       f.base_imponible, f.cuota_iva, f.iva_pct, f.total
                FROM gestoria_facturas f
                LEFT JOIN gestoria_terceros t ON t.id = f.tercero_id
                WHERE f.empresa_id = ?
                  AND LOWER(COALESCE(f.tipo, '')) = 'compra'
                  {iva_clause}
                ORDER BY f.fecha_emision ASC, f.created_at ASC
                """,
                iva_values,
            ).fetchall()
            iva_ventas = conn.execute(
                f"""
                SELECT f.id, f.fecha_emision, f.numero, COALESCE(t.nombre, '') AS tercero,
                       f.base_imponible, f.cuota_iva, f.iva_pct, f.total
                FROM gestoria_facturas f
                LEFT JOIN gestoria_terceros t ON t.id = f.tercero_id
                WHERE f.empresa_id = ?
                  AND LOWER(COALESCE(f.tipo, '')) = 'venta'
                  {iva_clause}
                ORDER BY f.fecha_emision ASC, f.created_at ASC
                """,
                iva_values,
            ).fetchall()
            json_response(
                self,
                {
                    "diario": [dict(r) for r in diario],
                    "mayor": [dict(r) for r in mayor],
                    "iva_compras": [dict(r) for r in iva_compras],
                    "iva_ventas": [dict(r) for r in iva_ventas],
                },
            )
            return

        if path == "/api/gestoria_excel_plantilla":
            empresa_id = params.get("empresa_id", [""])[0]
            cliente_id = (params.get("cliente_id", [""])[0] or "").strip()
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            if not OPENPYXL_AVAILABLE:
                json_response(self, {"error": "openpyxl no disponible en servidor"}, status=500)
                return
            diario = conn.execute(
                """
                SELECT a.id AS asiento_id, a.fecha, a.concepto, a.referencia,
                       l.cuenta, l.descripcion, l.debe, l.haber,
                       l.impuesto_tipo, l.impuesto_pct,
                       COALESCE(t.nombre, '') AS tercero,
                       COALESCE(t.nif, '') AS tercero_nif,
                       COALESCE(f.numero, '') AS factura_numero,
                       COALESCE(f.fecha_emision, '') AS factura_fecha,
                       COALESCE(f.total, 0) AS factura_total,
                       COALESCE(f.tipo, '') AS tipo_factura
                FROM gestoria_asientos a
                JOIN gestoria_asiento_lineas l ON l.asiento_id = a.id
                LEFT JOIN gestoria_terceros t ON t.id = l.tercero_id
                LEFT JOIN gestoria_facturas f ON f.id = a.factura_id
                WHERE a.empresa_id = ? AND a.cliente_id = ?
                ORDER BY a.fecha ASC, a.created_at ASC, l.cuenta ASC
                """,
                (empresa_id, cliente_id),
            ).fetchall()
            grouped = {}
            for row in diario:
                key = str(row["asiento_id"] or "").strip() or f"{row['fecha'] or ''}-{row['referencia'] or ''}"
                grouped.setdefault(key, []).append(row)
            output_rows = []
            for _key, lines in grouped.items():
                if not lines:
                    continue
                sample = lines[0]
                base = 0.0
                iva_pct = 0.0
                iva_importe = 0.0
                subcuenta_tercero = ""
                subcuenta_gyi = ""
                tipo_venta = normalize_service_key(sample["tipo_factura"] or "") == "venta"
                for line in lines:
                    cuenta = str(line["cuenta"] or "").strip()
                    debe = float(line["debe"] or 0)
                    haber = float(line["haber"] or 0)
                    imp_tipo = normalize_service_key(line["impuesto_tipo"] or "")
                    if not subcuenta_tercero and cuenta.startswith("4"):
                        subcuenta_tercero = cuenta
                    if not subcuenta_gyi and (cuenta.startswith("6") or cuenta.startswith("7")):
                        subcuenta_gyi = cuenta
                    if imp_tipo == "iva":
                        iva_importe += abs(haber if tipo_venta else debe)
                        if not iva_pct:
                            iva_pct = float(line["impuesto_pct"] or 0)
                    if subcuenta_gyi == cuenta:
                        base += abs(haber if cuenta.startswith("7") else debe)
                total = float(sample["factura_total"] or 0) or (base + iva_importe)
                output_rows.append(
                    [
                        sample["fecha"] or "",
                        sample["factura_fecha"] or "",
                        sample["factura_numero"] or sample["referencia"] or "",
                        sample["concepto"] or "",
                        subcuenta_tercero,
                        sample["tercero_nif"] or "",
                        sample["tercero"] or "",
                        "",
                        "",
                        "",
                        "",
                        round(base, 2) if base else "",
                        round(iva_pct, 2) if iva_pct else "",
                        round(iva_importe, 2) if iva_importe else "",
                        subcuenta_gyi,
                        round(total, 2) if total else "",
                    ]
                )
            if GESTORIA_EXCEL_TEMPLATE.exists():
                wb = load_workbook(GESTORIA_EXCEL_TEMPLATE)
                ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Hoja1"
                headers = [
                    "FECHA ASIENTO",
                    "FECHA FACTURA",
                    "Nº FACTURA",
                    "CONCEPTO",
                    "SUBCUENTA",
                    "NIF",
                    "NOMBRE",
                    "DOMICILIO",
                    "LOCALIDAD",
                    "PROVINCIA",
                    "CODIGO POSTAL",
                    "BASE IMPONIBLE",
                    "% IVA",
                    "IMPORTE IVA",
                    "SUBCUENTA GASTOS/INGRESOS",
                    "IMPORTE (TOTAL)",
                ]
                ws.append(headers)
                ws.append([None] * 16)

            style_row_idx = 2
            max_existing = max(ws.max_row, style_row_idx)
            style_cells = [ws.cell(style_row_idx, col) for col in range(1, 17)]
            # Limpia filas de datos previas (desde la fila 2).
            for row_idx in range(2, max_existing + 1):
                for col in range(1, 17):
                    ws.cell(row_idx, col).value = None

            for offset, row in enumerate(output_rows, start=0):
                row_idx = 2 + offset
                for col in range(1, 17):
                    target = ws.cell(row_idx, col)
                    source = style_cells[col - 1]
                    target._style = shallow_copy(source._style)
                    target.number_format = source.number_format
                    target.protection = shallow_copy(source.protection)
                    target.alignment = shallow_copy(source.alignment)
                    target.font = shallow_copy(source.font)
                    target.fill = shallow_copy(source.fill)
                    target.border = shallow_copy(source.border)
                ws.cell(row_idx, 1).value = row[0]
                ws.cell(row_idx, 2).value = row[1]
                ws.cell(row_idx, 3).value = row[2]
                # Mantiene la lógica de tu plantilla original para el concepto.
                ws.cell(row_idx, 4).value = f'=CONCATENATE(C{row_idx}," ",G{row_idx})'
                ws.cell(row_idx, 5).value = row[4]
                ws.cell(row_idx, 6).value = row[5]
                ws.cell(row_idx, 7).value = row[6]
                ws.cell(row_idx, 8).value = row[7]
                ws.cell(row_idx, 9).value = row[8]
                ws.cell(row_idx, 10).value = row[9]
                ws.cell(row_idx, 11).value = row[10]
                ws.cell(row_idx, 12).value = row[11]
                ws.cell(row_idx, 13).value = row[12]
                ws.cell(row_idx, 14).value = row[13]
                ws.cell(row_idx, 15).value = row[14]
                ws.cell(row_idx, 16).value = row[15]
            bio = BytesIO()
            wb.save(bio)
            payload = bio.getvalue()
            cliente = conn.execute("SELECT nombre FROM clientes WHERE id = ?", (cliente_id,)).fetchone()
            cliente_slug = re.sub(r"[^a-z0-9]+", "_", normalize_lookup_text((cliente["nombre"] if cliente else "cliente"))).strip("_") or "cliente"
            filename = f"plantilla_conversor_asientos_{cliente_slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.send_response(200)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if path == "/api/gestoria_conta_config":
            cliente_id = params.get("cliente_id", [""])[0]
            if not cliente_id:
                json_response(self, {"error": "cliente_id requerido"}, status=400)
                return
            row = conn.execute(
                """
                SELECT id, cliente_id, periodo, fecha_inicio, responsable
                FROM gestoria_conta_config
                WHERE cliente_id = ?
                """,
                (cliente_id,),
            ).fetchone()
            json_response(self, {"row": dict(row) if row else {}})
            return

        if path == "/api/gestoria_conta_tasks":
            cliente_id = params.get("cliente_id", [""])[0]
            empresa_id = params.get("empresa_id", [""])[0]
            estado = params.get("estado", [""])[0]
            periodo = params.get("periodo", [""])[0]
            if not cliente_id and not empresa_id:
                json_response(self, {"error": "cliente_id o empresa_id requerido"}, status=400)
                return
            where = []
            values = []
            if cliente_id:
                where.append("t.cliente_id = ?")
                values.append(cliente_id)
            if periodo:
                where.append("t.periodo = ?")
                values.append(periodo)
            if estado:
                where.append("LOWER(t.estado) = ?")
                values.append(estado.lower())
            if empresa_id:
                where.append("c.empresa_id = ?")
                values.append(empresa_id)
            where_clause = " AND ".join(where) if where else "1=1"
            rows = conn.execute(
                f"""
                SELECT t.id, t.cliente_id, t.periodo, t.tarea, t.estado,
                       t.fecha_limite, t.responsable,
                       COALESCE(c.nombre, '') AS cliente
                FROM gestoria_conta_tasks t
                LEFT JOIN clientes c ON c.id = t.cliente_id
                WHERE {where_clause}
                ORDER BY t.fecha_limite ASC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/usuarios":
            rows = conn.execute(
                "SELECT id, nombre, apellido, usuario, email, servicio, rol, activo FROM usuarios ORDER BY nombre, apellido"
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/auditoria":
            empresa_id = params.get("empresa_id", [""])[0]
            limit = params.get("limit", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            limit_clause = "LIMIT 50"
            if limit.isdigit():
                limit_clause = f"LIMIT {int(limit)}"
            rows = conn.execute(
                f"""
                SELECT a.id, a.entidad, a.entidad_id, a.accion, a.usuario, a.detalles, a.created_at,
                       COALESCE(c.nombre, '') AS cliente
                FROM auditoria a
                LEFT JOIN clientes c ON c.id = a.entidad_id
                WHERE a.empresa_id = ?
                ORDER BY a.created_at DESC
                {limit_clause}
                """,
                (empresa_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/catalogo":
            tipo = params.get("tipo", [""])[0].lower()
            q = params.get("q", [""])[0].strip()
            if tipo not in ("cnae", "iae"):
                json_response(self, {"error": "tipo requerido"}, status=400)
                return
            table = "cnae_catalogo" if tipo == "cnae" else "iae_catalogo"
            if q:
                rows = conn.execute(
                    f"""
                    SELECT codigo, descripcion
                    FROM {table}
                    WHERE codigo LIKE ? OR descripcion LIKE ?
                    ORDER BY codigo
                    LIMIT 30
                    """,
                    (f"%{q}%", f"%{q}%"),
                ).fetchall()
            else:
                rows = conn.execute(
                    f"""
                    SELECT codigo, descripcion
                    FROM {table}
                    ORDER BY codigo
                    LIMIT 30
                    """
                ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/catalogo_match":
            texto = params.get("texto", [""])[0].strip()
            if not texto:
                json_response(self, {"error": "texto requerido"}, status=400)
                return
            def match(table):
                return conn.execute(
                    f"""
                    SELECT codigo, descripcion
                    FROM {table}
                    WHERE LOWER(descripcion) LIKE LOWER(?)
                    ORDER BY CASE
                      WHEN LOWER(descripcion) = LOWER(?) THEN 0
                      ELSE 1
                    END, codigo
                    LIMIT 5
                    """,
                    (f"%{texto}%", texto),
                ).fetchall()
            cnae = match("cnae_catalogo")
            iae = match("iae_catalogo")
            json_response(
                self,
                {
                    "cnae": [dict(r) for r in cnae],
                    "iae": [dict(r) for r in iae],
                },
            )
            return

        if path == "/api/gestoria_dashboard":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            today = datetime.now().date()
            next_30 = today + timedelta(days=30)
            next_14 = today + timedelta(days=14)
            service_filter = (
                "LOWER(ce.servicio) IN ('gestoria', 'gestoría', "
                "'administracion fincas', 'administración fincas')"
            )
            total = conn.execute(
                f"""
                SELECT COUNT(DISTINCT c.id) AS total
                FROM clientes c
                JOIN clientes_empresas ce ON ce.cliente_id = c.id
                WHERE ce.empresa_id = ? AND {service_filter}
                """,
                (empresa_id,),
            ).fetchone()
            activos = conn.execute(
                f"""
                SELECT COUNT(DISTINCT c.id) AS total
                FROM clientes c
                JOIN clientes_empresas ce ON ce.cliente_id = c.id
                WHERE ce.empresa_id = ?
                  AND {service_filter}
                  AND LOWER(COALESCE(ce.estado, '')) = 'alta'
                """,
                (empresa_id,),
            ).fetchone()
            tipos = conn.execute(
                f"""
                SELECT cg.tipo_cliente AS tipo, COUNT(*) AS total
                FROM cliente_gestoria cg
                JOIN clientes_empresas ce ON ce.cliente_id = cg.cliente_id
                WHERE ce.empresa_id = ? AND {service_filter}
                GROUP BY cg.tipo_cliente
                """,
                (empresa_id,),
            ).fetchall()
            tipos_map = {row["tipo"]: row["total"] for row in tipos if row["tipo"]}
            def tipo_count(*labels):
                return sum(tipos_map.get(label, 0) for label in labels)
            modelos_mes = conn.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM gestoria_modelos m
                JOIN clientes_empresas ce ON ce.cliente_id = m.cliente_id
                WHERE ce.empresa_id = ? AND {service_filter}
                  AND m.proxima_fecha IS NOT NULL
                  AND strftime('%Y-%m', m.proxima_fecha) = ?
                """,
                (empresa_id, today.strftime("%Y-%m")),
            ).fetchone()
            modelos = conn.execute(
                f"""
                SELECT c.nombre AS cliente, m.modelo, m.proxima_fecha, m.estado
                FROM gestoria_modelos m
                JOIN clientes c ON c.id = m.cliente_id
                JOIN clientes_empresas ce ON ce.cliente_id = c.id
                WHERE ce.empresa_id = ? AND {service_filter}
                  AND m.proxima_fecha IS NOT NULL
                  AND date(m.proxima_fecha) BETWEEN date(?) AND date(?)
                ORDER BY m.proxima_fecha ASC
                LIMIT 12
                """,
                (empresa_id, today.isoformat(), next_30.isoformat()),
            ).fetchall()
            modelos_vencidos = conn.execute(
                f"""
                SELECT c.nombre AS cliente, m.modelo, m.proxima_fecha, m.estado
                FROM gestoria_modelos m
                JOIN clientes c ON c.id = m.cliente_id
                JOIN clientes_empresas ce ON ce.cliente_id = c.id
                WHERE ce.empresa_id = ? AND {service_filter}
                  AND m.proxima_fecha IS NOT NULL
                  AND date(m.proxima_fecha) < date(?)
                  AND (m.estado IS NULL OR LOWER(m.estado) != 'presentado')
                ORDER BY m.proxima_fecha ASC
                LIMIT 12
                """,
                (empresa_id, today.isoformat()),
            ).fetchall()
            acciones_pendientes = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM acciones a
                WHERE a.empresa_id = ?
                  AND LOWER(a.servicio) = 'gestoria'
                  AND (a.estado IS NULL OR LOWER(a.estado) != 'hecho')
                  AND a.fecha IS NOT NULL
                  AND date(a.fecha) >= date(?)
                """,
                (empresa_id, today.isoformat()),
            ).fetchone()
            acciones = conn.execute(
                """
                SELECT a.fecha, a.hora,
                       COALESCE(c.nombre, a.cliente_nombre) AS cliente,
                       a.tipo, a.estado
                FROM acciones a
                LEFT JOIN clientes c ON c.id = a.cliente_id
                WHERE a.empresa_id = ?
                  AND LOWER(a.servicio) = 'gestoria'
                  AND (a.estado IS NULL OR LOWER(a.estado) != 'hecho')
                  AND a.fecha IS NOT NULL
                  AND date(a.fecha) BETWEEN date(?) AND date(?)
                ORDER BY a.fecha ASC, a.hora ASC
                LIMIT 10
                """,
                (empresa_id, today.isoformat(), next_14.isoformat()),
            ).fetchall()
            acciones_vencidas = conn.execute(
                """
                SELECT a.fecha, a.hora,
                       COALESCE(c.nombre, a.cliente_nombre) AS cliente,
                       a.tipo, a.estado
                FROM acciones a
                LEFT JOIN clientes c ON c.id = a.cliente_id
                WHERE a.empresa_id = ?
                  AND LOWER(a.servicio) = 'gestoria'
                  AND a.fecha IS NOT NULL
                  AND date(a.fecha) < date(?)
                  AND (a.estado IS NULL OR LOWER(a.estado) != 'hecho')
                ORDER BY a.fecha ASC, a.hora ASC
                LIMIT 10
                """,
                (empresa_id, today.isoformat()),
            ).fetchall()
            json_response(
                self,
                {
                    "counts": {
                        "total": total["total"] if total else 0,
                        "activos": activos["total"] if activos else 0,
                        "autonomos": tipo_count("Autónomo", "Autonomo"),
                        "empresas": tipo_count("Empresa", "Empresas"),
                        "puntuales": tipo_count("Puntual", "Puntuales"),
                        "modelos_mes": modelos_mes["total"] if modelos_mes else 0,
                        "acciones_pendientes": acciones_pendientes["total"] if acciones_pendientes else 0,
                    },
                    "modelos": [dict(r) for r in modelos],
                    "modelos_vencidos": [dict(r) for r in modelos_vencidos],
                    "acciones": [dict(r) for r in acciones],
                    "acciones_vencidas": [dict(r) for r in acciones_vencidas],
                },
            )
            return

        if path == "/api/clientes":
            empresa_id = params.get("empresa_id", [""])[0]
            q = params.get("q", [""])[0].strip()
            estado = params.get("estado", [""])[0].strip()
            servicio = params.get("servicio", [""])[0].strip()
            services = parse_services_param(servicio)
            source = (params.get("source", [""])[0] or "").strip().lower()
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            normalized_services = [normalize_service_key(s) for s in services]
            include_id = params.get("include_id", [""])[0] == "1"
            limit_param = params.get("limit", [""])[0].strip()
            select_id = "c.id, " if include_id else ""
            limit_clause = "LIMIT 500"
            if limit_param.isdigit():
                limit_clause = f"LIMIT {int(limit_param)}"
            if source == "seguros" and ("seguros" in normalized_services or not normalized_services):
                where = ["s.cliente_id IS NOT NULL"]
                values = []
                if empresa_id:
                    where.append("s.empresa_id = ?")
                    values.append(empresa_id)
                if q:
                    where.append(
                        "(c.nombre LIKE ? OR c.nif LIKE ? OR c.telefono LIKE ? OR c.email LIKE ?)"
                    )
                    values.extend([f"%{q}%"] * 4)
                if estado:
                    where.append("c.estado = ?")
                    values.append(estado)
                where.append(f"({uploaded_policy_filter('s')} OR ? = 0)")
                values.append(1 if uploaded_only else 0)
                where_clause = f"WHERE {' AND '.join(where)}"
                rows = conn.execute(
                    f"""
                    SELECT
                      {select_id}
                      c.nombre,
                      c.tipo_persona,
                      c.nif,
                      c.telefono,
                      c.email,
                      c.fecha_nacimiento,
                      c.direccion,
                      c.codigo_postal,
                      c.poblacion,
                      c.provincia,
                      GROUP_CONCAT(e.nombre, ' | ') AS empresas,
                      GROUP_CONCAT(COALESCE(NULLIF(ce.servicio, ''), 'seguros'), ' | ') AS servicios
                    FROM clientes c
                    JOIN seguros s ON s.cliente_id = c.id
                    LEFT JOIN clientes_empresas ce ON ce.cliente_id = c.id AND ce.empresa_id = s.empresa_id
                    LEFT JOIN empresas e ON e.id = s.empresa_id
                    {where_clause}
                    GROUP BY c.id
                    ORDER BY c.nombre
                    {limit_clause}
                    """,
                    values,
                ).fetchall()
            else:
                where = []
                values = []
                if services:
                    placeholders = ",".join(["?"] * len(services))
                    where.append(f"LOWER(ce.servicio) IN ({placeholders})")
                    values.extend(services)
                if empresa_id:
                    where.append("ce.empresa_id = ?")
                    values.append(empresa_id)
                if q:
                    where.append(
                        "(c.nombre LIKE ? OR c.nif LIKE ? OR c.telefono LIKE ? OR c.email LIKE ?)"
                    )
                    values.extend([f"%{q}%"] * 4)
                if estado:
                    where.append("c.estado = ?")
                    values.append(estado)
                where_clause = f"WHERE {' AND '.join(where)}" if where else ""
                join_clause = "JOIN clientes_empresas ce ON ce.cliente_id = c.id" if services else "LEFT JOIN clientes_empresas ce ON ce.cliente_id = c.id"
                rows = conn.execute(
                    f"""
                    SELECT
                      {select_id}
                      c.nombre,
                      c.tipo_persona,
                      c.nif,
                      c.telefono,
                      c.email,
                      c.fecha_nacimiento,
                      c.direccion,
                      c.codigo_postal,
                      c.poblacion,
                      c.provincia,
                      GROUP_CONCAT(e.nombre, ' | ') AS empresas,
                      GROUP_CONCAT(ce.servicio, ' | ') AS servicios
                    FROM clientes c
                    {join_clause}
                    LEFT JOIN empresas e ON e.id = ce.empresa_id
                    {where_clause}
                    GROUP BY c.id
                    ORDER BY c.nombre
                    {limit_clause}
                    """,
                    values,
                ).fetchall()
            columns = [
                "nombre",
                "tipo_persona",
                "nif",
                "telefono",
                "email",
                "fecha_nacimiento",
                "direccion",
                "codigo_postal",
                "poblacion",
                "provincia",
                "empresas",
                "servicios",
            ]
            if include_id:
                columns = ["id"] + columns
            json_response(self, {"columns": columns, "rows": [list(r) for r in rows]})
            return

        if path == "/api/inmuebles":
            empresa_id = params.get("empresa_id", [""])[0]
            q = params.get("q", [""])[0].strip()
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            where = ["i.empresa_id = ?"]
            values = [empresa_id]
            if q:
                where.append(
                    "(i.referencia LIKE ? OR i.direccion LIKE ? OR i.zona LIKE ? OR i.estado LIKE ? OR c.nombre LIKE ?)"
                )
                values.extend([f"%{q}%"] * 5)
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT
                  i.id,
                  i.referencia,
                  i.direccion,
                  i.zona,
                  i.estado,
                  GROUP_CONCAT(c.nombre, ' | ') AS propietarios
                FROM inmuebles i
                LEFT JOIN inmueble_propietarios ip ON ip.inmueble_id = i.id
                LEFT JOIN clientes c ON c.id = ip.cliente_id
                WHERE {where_clause}
                GROUP BY i.id
                ORDER BY i.created_at DESC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/inmueble":
            inmueble_id = params.get("id", [""])[0]
            if not inmueble_id:
                json_response(self, {"error": "id requerido"}, status=400)
                return
            inmueble = conn.execute(
                "SELECT * FROM inmuebles WHERE id = ?",
                (inmueble_id,),
            ).fetchone()
            if not inmueble:
                json_response(self, {"error": "Inmueble no encontrado"}, status=404)
                return
            propietarios = conn.execute(
                """
                SELECT c.id, c.nombre, c.nif, c.telefono, c.email
                FROM inmueble_propietarios ip
                JOIN clientes c ON c.id = ip.cliente_id
                WHERE ip.inmueble_id = ?
                """,
                (inmueble_id,),
            ).fetchall()
            docs = conn.execute(
                """
                SELECT nombre, url, tipo
                FROM inmueble_docs
                WHERE inmueble_id = ?
                ORDER BY created_at DESC
                """,
                (inmueble_id,),
            ).fetchall()
            captacion = conn.execute(
                """
                SELECT *
                FROM captaciones
                WHERE inmueble_id = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (inmueble_id,),
            ).fetchone()
            json_response(
                self,
                {
                    "inmueble": dict(inmueble),
                    "propietarios": [dict(r) for r in propietarios],
                    "docs": [dict(r) for r in docs],
                    "captacion": dict(captacion) if captacion else {},
                },
            )
            return

        if path == "/api/inmueble_docs":
            inmueble_id = params.get("inmueble_id", [""])[0]
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT id, nombre, url, tipo, created_at
                FROM inmueble_docs
                WHERE inmueble_id = ?
                ORDER BY created_at DESC
                """,
                (inmueble_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/demandas":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT d.id, d.tipo, d.zona, d.precio_max, d.m2_min,
                       d.habitaciones_min, d.banos_min, d.estado, d.prioridad,
                       c.nombre AS cliente
                FROM demandas d
                LEFT JOIN clientes c ON c.id = d.cliente_id
                WHERE d.empresa_id = ?
                ORDER BY d.created_at DESC
                """,
                (empresa_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/visitas":
            empresa_id = params.get("empresa_id", [""])[0]
            inmueble_id = params.get("inmueble_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            where = ["v.empresa_id = ?"]
            values = [empresa_id]
            if inmueble_id:
                where.append("v.inmueble_id = ?")
                values.append(inmueble_id)
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT v.id, v.fecha, v.hora, v.estado, v.asesor, v.notas,
                       i.direccion AS inmueble,
                       c.nombre AS cliente
                FROM visitas v
                LEFT JOIN inmuebles i ON i.id = v.inmueble_id
                LEFT JOIN demandas d ON d.id = v.demanda_id
                LEFT JOIN clientes c ON c.id = d.cliente_id
                WHERE {where_clause}
                ORDER BY v.fecha DESC, v.hora DESC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/matching":
            empresa_id = params.get("empresa_id", [""])[0]
            demanda_id = params.get("demanda_id", [""])[0]
            if not empresa_id or not demanda_id:
                json_response(self, {"error": "empresa_id y demanda_id requeridos"}, status=400)
                return
            demanda = conn.execute(
                "SELECT * FROM demandas WHERE id = ? AND empresa_id = ?",
                (demanda_id, empresa_id),
            ).fetchone()
            if not demanda:
                json_response(self, {"error": "Demanda no encontrada"}, status=404)
                return
            where = ["empresa_id = ?"]
            values = [empresa_id]
            if demanda["zona"]:
                where.append("LOWER(zona) LIKE ?")
                values.append(f"%{demanda['zona'].lower()}%")
            if demanda["precio_max"]:
                where.append("precio_objetivo <= ?")
                values.append(demanda["precio_max"])
            if demanda["m2_min"]:
                where.append("m2 >= ?")
                values.append(demanda["m2_min"])
            if demanda["habitaciones_min"]:
                where.append("habitaciones >= ?")
                values.append(demanda["habitaciones_min"])
            if demanda["banos_min"]:
                where.append("banos >= ?")
                values.append(demanda["banos_min"])
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT id, referencia, direccion, zona, precio_objetivo, m2, habitaciones, banos, estado
                FROM inmuebles
                WHERE {where_clause}
                ORDER BY created_at DESC
                LIMIT 50
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/inmueble_matching":
            inmueble_id = params.get("inmueble_id", [""])[0]
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            inmueble = conn.execute(
                "SELECT * FROM inmuebles WHERE id = ?",
                (inmueble_id,),
            ).fetchone()
            if not inmueble:
                json_response(self, {"error": "Inmueble no encontrado"}, status=404)
                return
            where = ["d.empresa_id = ?"]
            values = [inmueble["empresa_id"]]
            if inmueble["zona"]:
                where.append("LOWER(d.zona) LIKE ?")
                values.append(f"%{inmueble['zona'].lower()}%")
            if inmueble["precio_objetivo"]:
                where.append("(d.precio_max IS NULL OR d.precio_max >= ?)")
                values.append(inmueble["precio_objetivo"])
            if inmueble["m2"]:
                where.append("(d.m2_min IS NULL OR d.m2_min <= ?)")
                values.append(inmueble["m2"])
            if inmueble["habitaciones"]:
                where.append("(d.habitaciones_min IS NULL OR d.habitaciones_min <= ?)")
                values.append(inmueble["habitaciones"])
            if inmueble["banos"]:
                where.append("(d.banos_min IS NULL OR d.banos_min <= ?)")
                values.append(inmueble["banos"])
            where_clause = " AND ".join(where)
            rows = conn.execute(
                f"""
                SELECT d.id, d.tipo, d.zona, d.precio_max, d.m2_min,
                       d.habitaciones_min, d.banos_min, d.estado,
                       c.nombre AS cliente
                FROM demandas d
                LEFT JOIN clientes c ON c.id = d.cliente_id
                WHERE {where_clause}
                ORDER BY d.created_at DESC
                LIMIT 100
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/inmueble_checklist":
            inmueble_id = params.get("inmueble_id", [""])[0]
            etapa = params.get("etapa", [""])[0]
            if not inmueble_id:
                json_response(self, {"error": "inmueble_id requerido"}, status=400)
                return
            where = ["inmueble_id = ?"]
            values = [inmueble_id]
            if etapa:
                where.append("etapa = ?")
                values.append(etapa)
            rows = conn.execute(
                f"""
                SELECT id, etapa, tarea, estado, responsable, fecha_limite
                FROM inmueble_checklist
                WHERE {' AND '.join(where)}
                ORDER BY created_at ASC
                """,
                values,
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/hipoteca_stats":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return

            total = conn.execute(
                "SELECT COUNT(*) AS total FROM hipotecas WHERE empresa_id = ?",
                (empresa_id,),
            ).fetchone()

            firmadas_mes = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                  AND fecha_firma IS NOT NULL
                  AND strftime('%Y-%m', fecha_firma) = strftime('%Y-%m', 'now', 'localtime')
                """,
                (empresa_id,),
            ).fetchone()

            averages = conn.execute(
                """
                SELECT
                  AVG(
                    CASE
                      WHEN precio IS NOT NULL AND precio > 0 AND importe_hipoteca IS NOT NULL
                        THEN (importe_hipoteca * 100.0 / precio)
                      WHEN porcentaje IS NOT NULL AND porcentaje > 0 AND porcentaje <= 1 THEN porcentaje * 100.0
                      WHEN porcentaje IS NOT NULL AND porcentaje > 1 THEN porcentaje
                      ELSE NULL
                    END
                  ) AS porcentaje_medio,
                  AVG(COALESCE(comision, 0)) AS comision_media
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                """,
                (empresa_id,),
            ).fetchone()

            json_response(
                self,
                {
                    "total": total["total"] if total else 0,
                    "firmadas_mes": firmadas_mes["total"] if firmadas_mes else 0,
                    "porcentaje_medio": averages["porcentaje_medio"] if averages else 0,
                    "comision_media": averages["comision_media"] if averages else 0,
                },
            )
            return

        if path == "/api/hipoteca_dashboard":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return

            current_year = conn.execute("SELECT strftime('%Y','now','localtime') AS y").fetchone()["y"]

            current = conn.execute(
                """
                SELECT
                  COUNT(*) AS total,
                  AVG(
                    CASE
                      WHEN precio IS NOT NULL AND precio > 0 AND importe_hipoteca IS NOT NULL
                        THEN (importe_hipoteca * 100.0 / precio)
                      WHEN porcentaje IS NOT NULL AND porcentaje > 0 AND porcentaje <= 1 THEN porcentaje * 100.0
                      WHEN porcentaje IS NOT NULL AND porcentaje > 1 THEN porcentaje
                      ELSE NULL
                    END
                  ) AS porcentaje_medio,
                  AVG(COALESCE(comision, 0)) AS comision_media
                FROM hipotecas
                WHERE empresa_id = ?
                  AND anio = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                """,
                (empresa_id, current_year),
            ).fetchone()

            firmadas_mes = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                  AND fecha_firma IS NOT NULL
                  AND strftime('%Y-%m', fecha_firma) = strftime('%Y-%m', 'now', 'localtime')
                """,
                (empresa_id,),
            ).fetchone()

            series_totales = conn.execute(
                """
                SELECT anio AS year, COUNT(*) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            series_comision = conn.execute(
                """
                SELECT anio AS year, SUM(COALESCE(comision, 0)) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            series_porcentaje = conn.execute(
                """
                SELECT anio AS year,
                       AVG(
                         CASE
                           WHEN precio IS NOT NULL AND precio > 0 AND importe_hipoteca IS NOT NULL
                             THEN (importe_hipoteca * 100.0 / precio)
                           WHEN porcentaje IS NOT NULL AND porcentaje > 0 AND porcentaje <= 1 THEN porcentaje * 100.0
                           WHEN porcentaje IS NOT NULL AND porcentaje > 1 THEN porcentaje
                           ELSE NULL
                         END
                       ) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            series_entidades = conn.execute(
                """
                SELECT banco AS label, COUNT(*) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND banco IS NOT NULL
                  AND TRIM(banco) != ''
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                GROUP BY banco
                ORDER BY COUNT(*) DESC
                LIMIT 8
                """,
                (empresa_id,),
            ).fetchall()

            series_oficinas = conn.execute(
                """
                SELECT oficina AS label, COUNT(*) AS total
                FROM hipotecas
                WHERE empresa_id = ?
                  AND oficina IS NOT NULL
                  AND TRIM(oficina) != ''
                  AND LOWER(TRIM(estado)) IN ('firmado', 'firmada', 'indemnización', 'indemnizacion')
                GROUP BY oficina
                ORDER BY COUNT(*) DESC
                """,
                (empresa_id,),
            ).fetchall()

            json_response(
                self,
                {
                    "current": {
                        "total": current["total"] if current else 0,
                        "porcentaje_medio": current["porcentaje_medio"] if current else 0,
                        "comision_media": current["comision_media"] if current else 0,
                        "firmadas_mes": firmadas_mes["total"] if firmadas_mes else 0,
                    },
                    "series_totales": [dict(r) for r in series_totales],
                    "series_comision": [dict(r) for r in series_comision],
                    "series_porcentaje": [dict(r) for r in series_porcentaje],
                    "series_entidades": [dict(r) for r in series_entidades],
                    "series_oficinas": [dict(r) for r in series_oficinas],
                },
            )
            return

        if path == "/api/fincas_stats":
            empresa_id = params.get("empresa_id", [""])[0]
            year = params.get("year", [""])[0]
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            if not year:
                year = conn.execute(
                    "SELECT strftime('%Y','now','localtime') AS y"
                ).fetchone()["y"]

            facturado = conn.execute(
                """
                SELECT SUM(COALESCE(comision, 0)) AS total
                FROM movimientos
                WHERE empresa_id = ?
                  AND anio = ?
                  AND LOWER(TRIM(tipo)) = 'ingreso'
                """,
                (empresa_id, year),
            ).fetchone()

            gastos = conn.execute(
                """
                SELECT SUM(COALESCE(comision, 0)) AS total
                FROM movimientos
                WHERE empresa_id = ?
                  AND anio = ?
                  AND LOWER(TRIM(tipo)) = 'gasto'
                """,
                (empresa_id, year),
            ).fetchone()

            empresas_total = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM gestoria
                WHERE empresa_id = ?
                  AND LOWER(TRIM(tipo)) IN ('empresa', 'empresas')
                """,
                (empresa_id,),
            ).fetchone()

            comunidades = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM gestoria
                WHERE empresa_id = ?
                  AND LOWER(TRIM(tipo)) IN ('comunidad', 'comunidad de propietarios', 'comunidades')
                """,
                (empresa_id,),
            ).fetchone()

            autonomos = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM gestoria
                WHERE empresa_id = ?
                  AND LOWER(TRIM(REPLACE(REPLACE(tipo, 'Ó', 'o'), 'ó', 'o'))) IN (
                    'autonomo',
                    'autónomo',
                    'autonomos',
                    'autónomos'
                  )
                """,
                (empresa_id,),
            ).fetchone()

            polizas = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND (COALESCE(poliza_key, '') <> '' OR COALESCE(poliza_url, '') <> '' OR ? = 0)
                  AND LOWER(TRIM(estado)) IN ('en vigor', 'en_vigor', 'vigente', 'poliza', 'póliza', 'poliza en vigor')
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()

            json_response(
                self,
                {
                    "year": year,
                    "facturado": facturado["total"] if facturado else 0,
                    "gastos": gastos["total"] if gastos else 0,
                    "clientes_empresas": empresas_total["total"] if empresas_total else 0,
                    "comunidades": comunidades["total"] if comunidades else 0,
                    "autonomos": autonomos["total"] if autonomos else 0,
                    "polizas_vigor": polizas["total"] if polizas else 0,
                },
            )
            return

        if path == "/api/fincas_seguros_dashboard":
            empresa_id = params.get("empresa_id", [""])[0]
            year = params.get("year", [""])[0]
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            if not year:
                year = conn.execute(
                    "SELECT strftime('%Y','now','localtime') AS y"
                ).fetchone()["y"]

            estado_expr = "LOWER(TRIM(estado))"
            compania_expr = "LOWER(TRIM(compania))"
            year_expr = "STRFTIME('%Y', created_at)"
            exclude_sin_seguro = f"({compania_expr} IS NULL OR {compania_expr} = '' OR {compania_expr} != 'sin seguro')"
            uploaded_clause = uploaded_policy_filter()
            in_vigor_expr = in_vigor_policy_filter()

            current = conn.execute(
                f"""
                SELECT
                  SUM(CASE WHEN {estado_expr} IN ('presupuesto', 'presupuestos') THEN 1 ELSE 0 END) AS presupuesto,
                  SUM(CASE WHEN {estado_expr} IN ('contratada', 'contratado', 'contrato', 'proyecto') THEN 1 ELSE 0 END) AS contratada,
                  SUM(CASE WHEN {in_vigor_expr} THEN 1 ELSE 0 END) AS en_vigor
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {year_expr} = ?
                  AND {exclude_sin_seguro}
                """,
                (empresa_id, 1 if uploaded_only else 0, year),
            ).fetchone()

            totals = conn.execute(
                f"""
                SELECT
                  SUM(CASE WHEN {estado_expr} IN ('presupuesto', 'presupuestos') THEN 1 ELSE 0 END) AS presupuesto,
                  SUM(CASE WHEN {estado_expr} IN ('contratada', 'contratado', 'contrato', 'proyecto') THEN 1 ELSE 0 END) AS contratada,
                  SUM(CASE WHEN {in_vigor_expr} THEN 1 ELSE 0 END) AS en_vigor
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {exclude_sin_seguro}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()

            series = conn.execute(
                f"""
                SELECT
                  {year_expr} AS year,
                  SUM(CASE WHEN {estado_expr} IN ('presupuesto', 'presupuestos') THEN 1 ELSE 0 END) AS presupuesto,
                  SUM(CASE WHEN {estado_expr} IN ('contratada', 'contratado', 'contrato', 'proyecto') THEN 1 ELSE 0 END) AS contratada,
                  SUM(CASE WHEN {in_vigor_expr} THEN 1 ELSE 0 END) AS en_vigor
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {year_expr} IS NOT NULL
                  AND {exclude_sin_seguro}
                GROUP BY {year_expr}
                ORDER BY {year_expr}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchall()

            presupuesto = current["presupuesto"] if current else 0
            contratada = current["contratada"] if current else 0
            en_vigor = current["en_vigor"] if current else 0
            conversion = (en_vigor / presupuesto * 100) if presupuesto else 0
            total_presupuesto = totals["presupuesto"] if totals else 0
            total_contratada = totals["contratada"] if totals else 0
            total_en_vigor = totals["en_vigor"] if totals else 0
            total_global = (total_presupuesto or 0) + (total_contratada or 0) + (total_en_vigor or 0)
            conversion_global = (total_en_vigor / total_presupuesto * 100) if total_presupuesto else 0

            series_payload = []
            for row in series:
                row_dict = dict(row)
                row_dict["conversion"] = (
                    (row_dict.get("en_vigor") or 0) / (row_dict.get("presupuesto") or 0) * 100
                    if (row_dict.get("presupuesto") or 0)
                    else 0
                )
                series_payload.append(row_dict)

            responsables = conn.execute(
                f"""
                SELECT
                  COALESCE(NULLIF(TRIM(colaborador), ''), 'Sin responsable') AS label,
                  COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {in_vigor_expr}
                  AND {exclude_sin_seguro}
                GROUP BY COALESCE(NULLIF(TRIM(colaborador), ''), 'Sin responsable')
                ORDER BY total DESC
                LIMIT 10
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchall()

            facturacion_year = conn.execute(
                f"""
                SELECT COALESCE(
                  SUM(
                    CASE
                      WHEN LOWER(TRIM(COALESCE(gc.tipo, ''))) = 'gasto' THEN 0
                      ELSE COALESCE(gc.importe, 0)
                    END
                  ),
                  0
                ) AS total
                FROM gestoria_contabilidad gc
                WHERE gc.empresa_id = ?
                  AND {seguros_contabilidad_where_clause("gc")}
                  AND STRFTIME('%Y', gc.fecha) = ?
                """,
                (empresa_id, year),
            ).fetchone()
            facturacion_total = conn.execute(
                f"""
                SELECT COALESCE(
                  SUM(
                    CASE
                      WHEN LOWER(TRIM(COALESCE(gc.tipo, ''))) = 'gasto' THEN 0
                      ELSE COALESCE(gc.importe, 0)
                    END
                  ),
                  0
                ) AS total
                FROM gestoria_contabilidad gc
                WHERE gc.empresa_id = ?
                  AND {seguros_contabilidad_where_clause("gc")}
                """,
                (empresa_id,),
            ).fetchone()
            gastos_year = conn.execute(
                """
                SELECT COALESCE(SUM(CASE WHEN LOWER(COALESCE(tipo, '')) = 'gasto' THEN COALESCE(importe, 0) ELSE 0 END), 0) AS total
                FROM gestoria_contabilidad
                WHERE empresa_id = ? AND STRFTIME('%Y', fecha) = ?
                """,
                (empresa_id, year),
            ).fetchone()
            gastos_total = conn.execute(
                """
                SELECT COALESCE(SUM(CASE WHEN LOWER(COALESCE(tipo, '')) = 'gasto' THEN COALESCE(importe, 0) ELSE 0 END), 0) AS total
                FROM gestoria_contabilidad
                WHERE empresa_id = ?
                """,
                (empresa_id,),
            ).fetchone()

            json_response(
                self,
                {
                    "current": {
                        "year": year,
                        "presupuesto": presupuesto or 0,
                        "contratada": contratada or 0,
                        "en_vigor": en_vigor or 0,
                        "conversion": conversion,
                        "presupuesto_total": total_presupuesto or 0,
                        "en_vigor_total": total_en_vigor or 0,
                        "total_global": total_global or 0,
                        "conversion_total": conversion_global,
                        "facturacion_comision": facturacion_year["total"] if facturacion_year else 0,
                        "facturacion_comision_total": facturacion_total["total"] if facturacion_total else 0,
                        "gastos": gastos_year["total"] if gastos_year else 0,
                        "gastos_total": gastos_total["total"] if gastos_total else 0,
                    },
                    "series": series_payload,
                    "responsables": [dict(r) for r in responsables],
                },
            )
            return

        if path == "/api/fincas_alerts":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT
                  tomador,
                  poliza_numero,
                  fecha_efecto,
                  DATE(fecha_efecto, '+1 year') AS fecha_vencimiento
                FROM seguros
                WHERE empresa_id = ?
                  AND fecha_efecto IS NOT NULL
                  AND DATE(fecha_efecto, '+1 year') BETWEEN DATE('now','localtime')
                      AND DATE('now','localtime','+30 days')
                ORDER BY DATE(fecha_efecto, '+1 year') ASC
                LIMIT 50
                """,
                (empresa_id,),
            ).fetchall()
            json_response(
                self,
                {
                    "count": len(rows),
                    "items": [dict(r) for r in rows],
                },
            )
            return

        if path == "/api/seguros_alertas":
            empresa_id = params.get("empresa_id", [""])[0]
            days = params.get("days", ["30"])[0]
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            try:
                days_int = int(days)
            except Exception:
                days_int = 30
            rows = conn.execute(
                f"""
                SELECT
                  id,
                  cliente_id,
                  tomador,
                  poliza_numero,
                  compania,
                  fecha_efecto,
                  COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year')) AS fecha_vencimiento
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_policy_filter()} OR ? = 0)
                  AND COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year')) IS NOT NULL
                  AND DATE(COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year'))) BETWEEN DATE('now','localtime')
                      AND DATE('now','localtime', '+{days_int} days')
                ORDER BY DATE(COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year'))) ASC
                LIMIT 50
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchall()
            json_response(self, {"count": len(rows), "items": [dict(r) for r in rows]})
            return

        if path == "/api/seguros_kpis":
            empresa_id = params.get("empresa_id", [""])[0]
            uploaded_only = (params.get("uploaded_only", ["1"])[0] or "1").strip() in ("1", "true", "yes")
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            estado_expr = "LOWER(TRIM(estado))"
            in_vigor_expr = in_vigor_policy_filter()
            uploaded_clause = uploaded_policy_filter()
            compania_expr = "LOWER(TRIM(compania))"
            exclude_sin_seguro = f"({compania_expr} IS NULL OR {compania_expr} = '' OR {compania_expr} != 'sin seguro')"
            total = conn.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()
            en_vigor = conn.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {in_vigor_expr}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()
            vencen_30 = conn.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year')) IS NOT NULL
                  AND {in_vigor_expr}
                  AND DATE(COALESCE(fecha_vencimiento, DATE(fecha_efecto, '+1 year'))) BETWEEN DATE('now','localtime')
                      AND DATE('now','localtime','+30 days')
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()
            faltantes = conn.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND (
                    tomador IS NULL OR TRIM(tomador) = '' OR
                    poliza_numero IS NULL OR TRIM(poliza_numero) = '' OR
                    compania IS NULL OR TRIM(compania) = '' OR
                    fecha_efecto IS NULL OR TRIM(fecha_efecto) = ''
                  )
                  AND {in_vigor_expr}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()
            quality_rows = conn.execute(
                """
                SELECT calidad_ocr, COUNT(*) AS total
                FROM gestoria_docs
                WHERE empresa_id = ?
                  AND (referencia_tipo = 'seguros' OR tipo = 'Seguros')
                GROUP BY calidad_ocr
                """,
                (empresa_id,),
            ).fetchall()
            primas = conn.execute(
                f"""
                SELECT SUM(prima_total) AS total
                FROM seguros
                WHERE empresa_id = ?
                  AND ({uploaded_clause} OR ? = 0)
                  AND {in_vigor_expr}
                """,
                (empresa_id, 1 if uploaded_only else 0),
            ).fetchone()
            facturacion = conn.execute(
                f"""
                SELECT COALESCE(
                  SUM(
                    CASE
                      WHEN LOWER(TRIM(COALESCE(gc.tipo, ''))) = 'gasto' THEN 0
                      ELSE COALESCE(gc.importe, 0)
                    END
                  ),
                  0
                ) AS total
                FROM gestoria_contabilidad gc
                WHERE gc.empresa_id = ?
                  AND {seguros_contabilidad_where_clause("gc")}
                """,
                (empresa_id,),
            ).fetchone()
            gastos = conn.execute(
                """
                SELECT COALESCE(SUM(CASE WHEN LOWER(COALESCE(tipo, '')) = 'gasto' THEN COALESCE(importe, 0) ELSE 0 END), 0) AS total
                FROM gestoria_contabilidad
                WHERE empresa_id = ?
                """,
                (empresa_id,),
            ).fetchone()
            quality = {"alta": 0, "media": 0, "baja": 0, "desconocida": 0}
            for row in quality_rows:
                key = (row["calidad_ocr"] or "desconocida").lower()
                if key not in quality:
                    key = "desconocida"
                quality[key] += row["total"] or 0
            json_response(
                self,
                {
                    "total": total["total"] if total else 0,
                    "en_vigor": en_vigor["total"] if en_vigor else 0,
                    "vencen_30": vencen_30["total"] if vencen_30 else 0,
                    "faltantes": faltantes["total"] if faltantes else 0,
                    "prima_total": primas["total"] if primas and primas["total"] is not None else 0,
                    "facturacion_comision": facturacion["total"] if facturacion else 0,
                    "gastos": gastos["total"] if gastos else 0,
                    "ocr_quality": quality,
                },
            )
            return

        if path == "/api/seguros_checklist":
            poliza_id = params.get("poliza_id", [""])[0]
            if not poliza_id:
                json_response(self, {"error": "poliza_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT id, poliza_id, tarea, estado, responsable, fecha_limite
                FROM seguros_checklist
                WHERE poliza_id = ?
                ORDER BY created_at
                """,
                (poliza_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/fin_checklist":
            asesoramiento_id = params.get("asesoramiento_id", [""])[0]
            if not asesoramiento_id:
                json_response(self, {"error": "asesoramiento_id requerido"}, status=400)
                return
            rows = conn.execute(
                """
                SELECT id, asesoramiento_id, tarea, estado, responsable, fecha_limite
                FROM fin_checklist
                WHERE asesoramiento_id = ?
                ORDER BY created_at ASC
                """,
                (asesoramiento_id,),
            ).fetchall()
            json_response(self, {"rows": [dict(r) for r in rows]})
            return

        if path == "/api/fin_alertas":
            empresa_id = params.get("empresa_id", [""])[0]
            days = params.get("days", ["30"])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            try:
                days_int = int(days)
            except Exception:
                days_int = 30
            rows = conn.execute(
                """
                SELECT *
                FROM asesoramientos_financiacion
                WHERE empresa_id = ?
                ORDER BY created_at DESC
                LIMIT 200
                """,
                (empresa_id,),
            ).fetchall()
            alerts = []
            cutoff = datetime.now() - timedelta(days=days_int)
            for row in rows:
                row_dict = dict(row)
                missing = fin_missing_fields(row_dict)
                if missing:
                    row_dict["alerta_tipo"] = "faltantes"
                    row_dict["missing_fields"] = missing
                    alerts.append(row_dict)
                    continue
                estado = (row_dict.get("estado") or "").strip().lower()
                if estado in ("en estudio", "pendiente"):
                    fecha_raw = str(row_dict.get("fecha") or "").strip()
                    fecha_dt = None
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            fecha_dt = datetime.strptime(fecha_raw, fmt)
                            break
                        except Exception:
                            continue
                    if fecha_dt and fecha_dt <= cutoff:
                        row_dict["alerta_tipo"] = "seguimiento"
                        alerts.append(row_dict)
            json_response(self, {"count": len(alerts), "items": alerts})
            return

        if path == "/api/fin_kpis":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return
            rows = conn.execute(
                "SELECT estado, calidad_ocr, cliente1_nombre, cliente1_dni, cliente1_telefono, fecha, ingresos_conjuntos FROM asesoramientos_financiacion WHERE empresa_id = ?",
                (empresa_id,),
            ).fetchall()
            total = len(rows)
            estados = {"en_estudio": 0, "aprobado": 0, "descartado": 0, "convertido": 0, "otros": 0}
            faltantes = 0
            quality = {"alta": 0, "media": 0, "baja": 0, "desconocida": 0}
            for row in rows:
                estado = (row["estado"] or "").strip().lower()
                if estado in ("en estudio", "estudio"):
                    estados["en_estudio"] += 1
                elif estado in ("aprobado",):
                    estados["aprobado"] += 1
                elif estado in ("descartado", "rechazado"):
                    estados["descartado"] += 1
                elif estado in ("convertido",):
                    estados["convertido"] += 1
                else:
                    estados["otros"] += 1
                missing = fin_missing_fields(dict(row))
                if missing:
                    faltantes += 1
                key = (row["calidad_ocr"] or "desconocida").lower()
                if key not in quality:
                    key = "desconocida"
                quality[key] += 1
            json_response(
                self,
                {
                    "total": total,
                    "estados": estados,
                    "faltantes": faltantes,
                    "ocr_quality": quality,
                },
            )
            return

        if path == "/api/dashboard":
            empresa_id = params.get("empresa_id", [""])[0]
            if not empresa_id:
                json_response(self, {"error": "empresa_id requerido"}, status=400)
                return

            ventas = conn.execute(
                """
                SELECT anio AS year, COUNT(*) AS total
                FROM movimientos
                WHERE empresa_id = ?
                  AND UPPER(TRIM(concepto)) = 'COMPRAVENTA'
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            ingresos = conn.execute(
                """
                SELECT anio AS year, SUM(COALESCE(comision, 0)) AS total
                FROM movimientos
                WHERE empresa_id = ?
                  AND LOWER(TRIM(tipo)) = 'ingreso'
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            gastos = conn.execute(
                """
                SELECT anio AS year, SUM(COALESCE(comision, 0)) AS total
                FROM movimientos
                WHERE empresa_id = ?
                  AND LOWER(TRIM(tipo)) = 'gasto'
                GROUP BY anio
                ORDER BY anio
                """,
                (empresa_id,),
            ).fetchall()

            alquileres = conn.execute(
                """
                SELECT STRFTIME('%Y', fecha) AS year,
                       COUNT(*) AS total,
                       SUM(COALESCE(precio, 0)) AS facturado
                FROM alquileres
                WHERE empresa_id = ?
                  AND fecha IS NOT NULL
                GROUP BY STRFTIME('%Y', fecha)
                ORDER BY STRFTIME('%Y', fecha)
                """,
                (empresa_id,),
            ).fetchall()

            json_response(
                self,
                {
                    "ventas": [dict(r) for r in ventas],
                    "ingresos": [dict(r) for r in ingresos],
                    "gastos": [dict(r) for r in gastos],
                    "alquileres": [dict(r) for r in alquileres],
                },
            )
            return

        if path == "/api/tabla":
            tabla = params.get("tabla", ["movimientos"])[0]
            if tabla not in TABLES:
                json_response(self, {"error": "Tabla no valida"}, status=400)
                return

            empresa_id = params.get("empresa_id", [""])[0]
            q = params.get("q", [""])[0].strip()
            year_filter = params.get("year", [""])[0].strip()
            field_filter = params.get("field", [""])[0].strip()
            estado_filter = params.get("estado", [""])[0].strip()
            tipo_filter = params.get("tipo", [""])[0].strip()
            perfil_filter = params.get("perfil", [""])[0].strip()
            limit_param = params.get("limit", [""])[0].strip()
            include_id = params.get("include_id", ["0"])[0] == "1"

            columns = [
                r["name"]
                for r in conn.execute(f"PRAGMA table_info({tabla})").fetchall()
            ]
            hidden = {"empresa_id", "created_at", "updated_at"}
            visible_columns = [col for col in columns if col not in hidden]
            if not include_id:
                visible_columns = [col for col in visible_columns if col != "id"]
            text_columns = [col for col in visible_columns]

            where = []
            values = []
            if empresa_id:
                where.append("t.empresa_id = ?")
                values.append(empresa_id)

            if year_filter and tabla == "movimientos":
                where.append("t.anio = ?")
                values.append(year_filter)

            if q:
                if field_filter and field_filter in visible_columns:
                    where.append(f"t.{field_filter} LIKE ?")
                    values.append(f"%{q}%")
                else:
                    likes = " OR ".join([f"t.{col} LIKE ?" for col in text_columns])
                    where.append(f"({likes})")
                    values.extend([f"%{q}%"] * len(text_columns))

            if estado_filter and "estado" in visible_columns:
                where.append("t.estado = ?")
                values.append(estado_filter)
            if tipo_filter and "tipo" in visible_columns:
                where.append("t.tipo = ?")
                values.append(tipo_filter)
            if perfil_filter and "perfil" in visible_columns:
                where.append("t.perfil LIKE ?")
                values.append(f"%{perfil_filter}%")

            where_clause = f"WHERE {' AND '.join(where)}" if where else ""
            select_cols = ", ".join([f"t.{col}" for col in visible_columns])
            limit_clause = "LIMIT 200"
            if limit_param.isdigit():
                limit_clause = f"LIMIT {int(limit_param)}"
            query = (
                f"SELECT e.nombre AS empresa, {select_cols} "
                f"FROM {tabla} t "
                "LEFT JOIN empresas e ON e.id = t.empresa_id "
                f"{where_clause} "
                f"{limit_clause}"
            )
            rows = conn.execute(query, values).fetchall()
            json_response(
                self,
                {"columns": ["empresa"] + visible_columns, "rows": [list(r) for r in rows]},
            )
            return

        json_response(self, {"error": "Endpoint no valido"}, status=404)


def main():
    parser = argparse.ArgumentParser(description="ERP Modernia local server.")
    parser.add_argument("--db", default=str(DB_CONFIGURED), help="SQLite path.")
    parser.add_argument("--ocr-db", default=str(OCR_DB_CONFIGURED), help="SQLite OCR jobs path.")
    parser.add_argument("--ocr-workers", type=int, default=OCR_WORKERS, help="Numero de workers OCR en paralelo.")
    parser.add_argument("--host", default="127.0.0.1", help="Host.")
    env_port = os.environ.get("PORT")
    try:
        env_port = int(env_port) if env_port else None
    except ValueError:
        env_port = None
    parser.add_argument("--port", type=int, default=env_port or 8000, help="Port.")
    args = parser.parse_args()

    ensure_tables(args.db)
    ensure_ocr_tables(args.ocr_db)
    Handler.db_path = args.db
    Handler.ocr_db_path = args.ocr_db
    ocr_workers = max(1, min(8, int(args.ocr_workers or 1)))
    workers = []
    for idx in range(ocr_workers):
        worker = threading.Thread(
            target=ocr_worker_loop,
            args=(args.ocr_db, args.db),
            name=f"ocr-worker-{idx+1}",
            daemon=True,
        )
        worker.start()
        workers.append(worker)
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(
        f"Servidor activo en http://{args.host}:{args.port} · db={Path(args.db).resolve()} · "
        f"ocr_db={Path(args.ocr_db).resolve()} · ocr_workers={ocr_workers}"
    )
    server.serve_forever()


if __name__ == "__main__":
    main()
