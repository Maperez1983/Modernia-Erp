#!/usr/bin/env python3
"""
Importa una hoja "BDT Hipotecas" (xls/xlsx) a la tabla hipotecas.

Uso:
  python3 scripts/import_bdt_hipotecas_excel.py \
    --excel "/ruta/archivo.xlsx" \
    --db "data/erp_import2.sqlite"
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sqlite3
import sys
import unicodedata
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


FIN_COMPANY = "Financiaciones Modernia"

TARGET_FIELDS = [
    "cliente",
    "banco",
    "precio",
    "importe_hipoteca",
    "porcentaje",
    "entrada",
    "comision",
    "oficina",
    "fecha_encargo",
    "encargo",
    "tipo_hipoteca",
    "fecha_firma",
    "cesion",
    "comision_juan",
    "comision_modernia",
    "inmobiliaria_compra",
    "asesor",
    "estado",
    "anio",
]


HEADER_ALIASES: Dict[str, Sequence[str]] = {
    "cliente": ("cliente", "nombre", "nombre_cliente", "titular", "nombre_razon_social"),
    "banco": ("banco", "entidad", "entidad_financiera"),
    "precio": ("precio", "precio_compra", "valor_compra"),
    "importe_hipoteca": ("importe_hipoteca", "importe", "capital"),
    "porcentaje": ("porcentaje", "ltv", "porcentaje_hipoteca"),
    "entrada": ("entrada", "ahorro", "aportacion"),
    "comision": ("comision",),
    "oficina": ("oficina", "sucursal"),
    "fecha_encargo": ("fecha_encargo", "fecha_encargo_firma", "fecha"),
    "encargo": ("encargo",),
    "tipo_hipoteca": ("tipo_hipoteca", "tipo"),
    "fecha_firma": ("fecha_firma", "firma"),
    "cesion": ("cesion",),
    "comision_juan": ("comision_juan", "comisionjuan"),
    "comision_modernia": ("comision_modernia", "comisionmodernia"),
    "inmobiliaria_compra": ("inmobiliaria_compra", "inmobiliaria", "agencia"),
    "asesor": ("asesor", "responsable"),
    "estado": ("estado",),
    "anio": ("anio", "año"),
    "id": ("id",),
}


def norm_text(value: object) -> str:
    txt = "" if value is None else str(value)
    txt = txt.strip()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = txt.lower()
    txt = re.sub(r"[^a-z0-9]+", "_", txt).strip("_")
    return txt


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace("€", "").replace("%", "").replace(" ", "")
    if "," in txt and "." in txt:
        txt = txt.replace(".", "").replace(",", ".")
    elif "," in txt:
        txt = txt.replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


def parse_date(value: object, datemode: Optional[int] = None) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and datemode is not None:
        try:
            # Para xls via xlrd.
            import xlrd  # type: ignore

            y, m, d, *_ = xlrd.xldate_as_tuple(float(value), datemode)
            return dt.date(y, m, d).isoformat()
        except Exception:
            return ""
    txt = str(value).strip()
    if not txt:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def excel_cell_to_value(cell: object) -> object:
    return cell


@dataclass
class SheetData:
    headers: List[str]
    rows: List[List[object]]
    datemode: Optional[int] = None


def read_sheet_xls(path: Path, forced_sheet: Optional[str] = None) -> SheetData:
    import xlrd  # type: ignore

    book = xlrd.open_workbook(str(path))
    sheet_name = forced_sheet or pick_sheet_name(book.sheet_names())
    if not sheet_name:
        raise RuntimeError("No se encontró hoja tipo 'BDT Hipotecas' en el Excel.")
    sh = book.sheet_by_name(sheet_name)
    headers, start_row = extract_headers_xls(sh)
    rows: List[List[object]] = []
    for r in range(start_row, sh.nrows):
        row = [sh.cell_value(r, c) for c in range(sh.ncols)]
        if any(str(v).strip() for v in row):
            rows.append(row)
    return SheetData(headers=headers, rows=rows, datemode=book.datemode)


def read_sheet_xlsx(path: Path, forced_sheet: Optional[str] = None) -> SheetData:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheet_name = forced_sheet or pick_sheet_name(wb.sheetnames)
    if not sheet_name:
        raise RuntimeError("No se encontró hoja tipo 'BDT Hipotecas' en el Excel.")
    ws = wb[sheet_name]
    raw_rows = [list(row) for row in ws.iter_rows(values_only=True)]
    headers, start_row = extract_headers_rows(raw_rows)
    rows: List[List[object]] = []
    for row in raw_rows[start_row:]:
        if any(str(v).strip() for v in row if v is not None):
            rows.append(row)
    return SheetData(headers=headers, rows=rows, datemode=None)


def pick_sheet_name(sheet_names: Iterable[str]) -> Optional[str]:
    names = list(sheet_names)
    exact = [n for n in names if norm_text(n) in {"bdt_hipotecas", "bdt_hipoteca"}]
    if exact:
        return exact[0]
    contains = [n for n in names if "bdt" in norm_text(n) and "hipotec" in norm_text(n)]
    if contains:
        return contains[0]
    return None


def extract_headers_xls(sheet) -> Tuple[List[str], int]:
    raw_rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    return extract_headers_rows(raw_rows)


def extract_headers_rows(raw_rows: Sequence[Sequence[object]]) -> Tuple[List[str], int]:
    for i, row in enumerate(raw_rows):
        texts = [norm_text(v) for v in row if str(v).strip()]
        if not texts:
            continue
        score = sum(1 for t in texts if t in _all_aliases_flat())
        if score >= 2 or ("cliente" in texts and ("banco" in texts or "entidad" in texts)):
            headers = [norm_text(v) for v in row]
            return headers, i + 1
    # Fallback: primera fila no vacía
    for i, row in enumerate(raw_rows):
        if any(str(v).strip() for v in row):
            return [norm_text(v) for v in row], i + 1
    raise RuntimeError("No se pudieron detectar cabeceras en la hoja.")


def _all_aliases_flat() -> set:
    s = set()
    for aliases in HEADER_ALIASES.values():
        s.update(aliases)
    return {norm_text(x) for x in s}


def build_index_map(headers: Sequence[str]) -> Dict[str, int]:
    header_to_idx = {norm_text(h): i for i, h in enumerate(headers)}
    field_idx: Dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            key = norm_text(alias)
            if key in header_to_idx:
                field_idx[field] = header_to_idx[key]
                break
    return field_idx


def get_row_value(row: Sequence[object], idx_map: Dict[str, int], field: str) -> object:
    idx = idx_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def to_record(row: Sequence[object], idx_map: Dict[str, int], datemode: Optional[int]) -> Dict[str, object]:
    rec: Dict[str, object] = {}
    for field in TARGET_FIELDS:
        raw = get_row_value(row, idx_map, field)
        if field in {"precio", "importe_hipoteca", "porcentaje", "entrada", "comision", "cesion", "comision_juan", "comision_modernia"}:
            rec[field] = parse_number(raw)
        elif field in {"fecha_encargo", "fecha_firma"}:
            rec[field] = parse_date(raw, datemode=datemode)
        elif field == "anio":
            num = parse_number(raw)
            if num is not None:
                rec[field] = int(num)
            else:
                rec[field] = None
        else:
            rec[field] = str(raw).strip() if raw is not None else ""
    if not rec.get("anio"):
        base = str(rec.get("fecha_firma") or rec.get("fecha_encargo") or "")
        rec["anio"] = int(base[:4]) if re.match(r"^\d{4}-\d{2}-\d{2}$", base) else None
    if not rec.get("estado"):
        rec["estado"] = "Pendiente"
    return rec


def fingerprint(rec: Dict[str, object]) -> str:
    def clean(v: object) -> str:
        return re.sub(r"\s+", " ", str(v or "").strip().lower())

    parts = [
        clean(rec.get("cliente")),
        clean(rec.get("banco")),
        clean(rec.get("oficina")),
        clean(rec.get("fecha_encargo")),
        clean(rec.get("fecha_firma")),
        clean(rec.get("importe_hipoteca")),
    ]
    return "|".join(parts)


def load_existing_map(conn: sqlite3.Connection, empresa_id: str) -> Dict[str, str]:
    rows = conn.execute(
        """
        SELECT id, cliente, banco, oficina, fecha_encargo, fecha_firma, importe_hipoteca
        FROM hipotecas
        WHERE empresa_id = ?
        """,
        (empresa_id,),
    ).fetchall()
    data: Dict[str, str] = {}
    for row in rows:
        rec = {
            "cliente": row[1],
            "banco": row[2],
            "oficina": row[3],
            "fecha_encargo": row[4],
            "fecha_firma": row[5],
            "importe_hipoteca": row[6],
        }
        data[fingerprint(rec)] = row[0]
    return data


def import_rows(db_path: Path, sheet: SheetData) -> Tuple[int, int, int]:
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    now = dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")
    try:
        empresa = conn.execute(
            "SELECT id FROM empresas WHERE nombre = ? LIMIT 1",
            (FIN_COMPANY,),
        ).fetchone()
        if not empresa:
            raise RuntimeError(f"No existe empresa '{FIN_COMPANY}' en la BD.")
        empresa_id = empresa[0]
        idx_map = build_index_map(sheet.headers)
        if "cliente" not in idx_map:
            raise RuntimeError("No se encontró columna de cliente en la hoja.")

        existing = load_existing_map(conn, empresa_id)
        inserted = updated = skipped = 0

        for row in sheet.rows:
            rec = to_record(row, idx_map, sheet.datemode)
            if not str(rec.get("cliente") or "").strip():
                skipped += 1
                continue

            rec_id = ""
            if "id" in idx_map:
                rec_id_raw = get_row_value(row, idx_map, "id")
                rec_id = str(rec_id_raw).strip()

            fp = fingerprint(rec)
            if not rec_id:
                rec_id = existing.get(fp, "")
            if not rec_id:
                rec_id = str(uuid.uuid4())
                conn.execute(
                    """
                    INSERT INTO hipotecas (
                      id, empresa_id, cliente, banco, precio, importe_hipoteca, porcentaje, entrada,
                      comision, oficina, fecha_encargo, encargo, tipo_hipoteca, fecha_firma, cesion,
                      comision_juan, comision_modernia, inmobiliaria_compra, asesor, estado, anio,
                      created_at, updated_at
                    ) VALUES (
                      ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                    )
                    """,
                    (
                        rec_id,
                        empresa_id,
                        rec.get("cliente"),
                        rec.get("banco"),
                        rec.get("precio"),
                        rec.get("importe_hipoteca"),
                        rec.get("porcentaje"),
                        rec.get("entrada"),
                        rec.get("comision"),
                        rec.get("oficina"),
                        rec.get("fecha_encargo"),
                        rec.get("encargo"),
                        rec.get("tipo_hipoteca"),
                        rec.get("fecha_firma"),
                        rec.get("cesion"),
                        rec.get("comision_juan"),
                        rec.get("comision_modernia"),
                        rec.get("inmobiliaria_compra"),
                        rec.get("asesor"),
                        rec.get("estado"),
                        rec.get("anio"),
                        now,
                        now,
                    ),
                )
                existing[fp] = rec_id
                inserted += 1
            else:
                conn.execute(
                    """
                    UPDATE hipotecas SET
                      cliente = ?, banco = ?, precio = ?, importe_hipoteca = ?, porcentaje = ?, entrada = ?,
                      comision = ?, oficina = ?, fecha_encargo = ?, encargo = ?, tipo_hipoteca = ?,
                      fecha_firma = ?, cesion = ?, comision_juan = ?, comision_modernia = ?,
                      inmobiliaria_compra = ?, asesor = ?, estado = ?, anio = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        rec.get("cliente"),
                        rec.get("banco"),
                        rec.get("precio"),
                        rec.get("importe_hipoteca"),
                        rec.get("porcentaje"),
                        rec.get("entrada"),
                        rec.get("comision"),
                        rec.get("oficina"),
                        rec.get("fecha_encargo"),
                        rec.get("encargo"),
                        rec.get("tipo_hipoteca"),
                        rec.get("fecha_firma"),
                        rec.get("cesion"),
                        rec.get("comision_juan"),
                        rec.get("comision_modernia"),
                        rec.get("inmobiliaria_compra"),
                        rec.get("asesor"),
                        rec.get("estado"),
                        rec.get("anio"),
                        now,
                        rec_id,
                    ),
                )
                updated += 1

        conn.commit()
        return inserted, updated, skipped
    finally:
        conn.close()


def read_sheet(path: Path, forced_sheet: Optional[str] = None) -> SheetData:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return read_sheet_xls(path, forced_sheet=forced_sheet)
    if suffix in {".xlsx", ".xlsm"}:
        return read_sheet_xlsx(path, forced_sheet=forced_sheet)
    raise RuntimeError(f"Formato no soportado: {path.suffix}")


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Importador BDT Hipotecas desde Excel")
    parser.add_argument("--excel", required=True, help="Ruta al Excel fuente (.xls/.xlsx)")
    parser.add_argument("--db", default="data/erp_import2.sqlite", help="Ruta a la BD sqlite")
    parser.add_argument("--sheet", default="", help="Nombre exacto de la hoja a importar")
    return parser.parse_args(argv)


def main(argv: Sequence[str]) -> int:
    args = parse_args(argv)
    excel_path = Path(args.excel).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()
    if not excel_path.exists():
        print(f"ERROR: Excel no encontrado: {excel_path}")
        return 1
    if not db_path.exists():
        print(f"ERROR: Base de datos no encontrada: {db_path}")
        return 1
    try:
        sheet = read_sheet(excel_path, forced_sheet=(args.sheet.strip() or None))
        inserted, updated, skipped = import_rows(db_path, sheet)
        print(f"OK: importación completada. Insertadas={inserted}, Actualizadas={updated}, Omitidas={skipped}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
