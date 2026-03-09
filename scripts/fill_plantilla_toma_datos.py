#!/usr/bin/env python3
import argparse
import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def norm(text):
    raw = str(text or "").strip().lower()
    raw = raw.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return re.sub(r"\s+", " ", raw).strip()


HEADER_ALIASES = {
    "fecha": ("fecha", "fecha factura", "fecha emision"),
    "numero": ("numero", "num factura", "factura", "n factura"),
    "tipo": ("tipo", "tipo factura"),
    "tercero": ("proveedor", "acreedor", "cliente", "tercero", "razon social", "nombre"),
    "nif": ("nif", "cif", "dni"),
    "base_imponible": ("base imponible", "base", "bi"),
    "cuota_iva": ("iva", "cuota iva"),
    "cuota_irpf": ("irpf", "retencion"),
    "total": ("total", "importe total"),
    "descripcion": ("concepto", "descripcion"),
    "cuenta": ("cuenta", "cuenta contable"),
    "asiento_id": ("asiento", "id asiento"),
}


def find_header_map(sheet):
    best = {}
    best_row = 1
    for row_idx in range(1, 40):
        row_values = [norm(sheet.cell(row=row_idx, column=col).value) for col in range(1, min(80, sheet.max_column + 1))]
        current = {}
        for col_idx, value in enumerate(row_values, start=1):
            if not value:
                continue
            for canonical, aliases in HEADER_ALIASES.items():
                if canonical in current:
                    continue
                if any(alias in value for alias in aliases):
                    current[canonical] = col_idx
        if len(current) > len(best):
            best = current
            best_row = row_idx
    return best_row, best


def read_rows(conn, empresa_nombre):
    row = conn.execute("SELECT id FROM empresas WHERE nombre = ?", (empresa_nombre,)).fetchone()
    if not row:
        raise SystemExit(f"Empresa no encontrada: {empresa_nombre}")
    empresa_id = row["id"]
    return conn.execute(
        """
        SELECT f.fecha_emision, f.numero, f.tipo, COALESCE(t.nombre, '') AS tercero, COALESCE(t.nif, '') AS nif,
               f.base_imponible, f.cuota_iva, f.cuota_irpf, f.total, f.descripcion,
               COALESCE(t.cuenta_contable, '') AS cuenta, COALESCE(a.id, '') AS asiento_id
        FROM gestoria_facturas f
        LEFT JOIN gestoria_terceros t ON t.id = f.tercero_id
        LEFT JOIN gestoria_asientos a ON a.factura_id = f.id
        WHERE f.empresa_id = ?
        ORDER BY f.fecha_emision ASC, f.created_at ASC
        """,
        (empresa_id,),
    ).fetchall()


def main():
    parser = argparse.ArgumentParser(description="Rellena la plantilla de toma de datos con facturas OCR.")
    parser.add_argument("--db", default="data/erp_import2.sqlite", help="Ruta al sqlite")
    parser.add_argument("--empresa", required=True, help="Nombre empresa exacto")
    parser.add_argument("--template", required=True, help="Ruta plantilla xlsx")
    parser.add_argument("--output", required=True, help="Ruta salida xlsx")
    args = parser.parse_args()

    db_path = Path(args.db).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not db_path.exists():
        raise SystemExit(f"DB no encontrada: {db_path}")
    if not template_path.exists():
        raise SystemExit(f"Plantilla no encontrada: {template_path}")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    rows = read_rows(conn, args.empresa)
    conn.close()

    wb = load_workbook(template_path)
    sheet = wb.active
    header_row, col_map = find_header_map(sheet)
    if not col_map:
        raise SystemExit("No se detectaron columnas compatibles en la plantilla.")

    write_row = header_row + 1
    for item in rows:
        for key, col in col_map.items():
            sheet.cell(row=write_row, column=col, value=item[key] if key in item.keys() else "")
        write_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"OK: {len(rows)} filas exportadas en {output_path}")


if __name__ == "__main__":
    main()

