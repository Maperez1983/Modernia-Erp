import argparse
import json
import sys
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from web.server import OPENPYXL_AVAILABLE, _iivtnu_norm_text, _iivtnu_postal_variants


def build_index_spain(municipios_csv: Path):
    import csv
    import re

    idx = {}  # {(norm_name, norm_prov): ine}
    name_to = {}  # {norm_name: {ine,...}}
    collisions = {}
    with municipios_csv.open("r", encoding="utf-8", newline="") as handle:
        r = csv.DictReader(handle)
        for row in r:
            ine = str(row.get("municipio_id") or row.get("municipio") or "").strip().zfill(5)
            nombre = str(row.get("nombre") or "").strip()
            prov = str(row.get("provincia_nombre") or row.get("provincia") or "").strip()
            if not ine or not prov or not nombre:
                continue
            variants = _iivtnu_postal_variants(nombre)
            if nombre not in variants:
                variants.append(nombre)

            # Variantes adicionales para nombres oficiales INE tipo:
            # - "Alacant/Alicante" <-> "Alicante/Alacant"
            # - "Castelló de la Plana/Castellón de la Plana" -> "Castelló de la Plana"
            if "/" in nombre:
                parts = [p.strip() for p in str(nombre).split("/") if str(p).strip()]
                if len(parts) == 2:
                    a, b = parts[0], parts[1]
                    variants.extend([a, b, f"{b}/{a}"])

            # - "Coruña, A" -> "A Coruña" / "Coruña (A)"
            m = re.match(r"^(.+),\s*(EL|LA|LOS|LAS|A|O|OS|AS)$", str(nombre).strip(), flags=re.IGNORECASE)
            if m:
                base = str(m.group(1) or "").strip()
                art = str(m.group(2) or "").strip()
                if base and art:
                    variants.extend([f"{art} {base}", f"{base} ({art})"])

            for v in variants:
                nn = _iivtnu_norm_text(v)
                if nn:
                    name_to.setdefault(nn, set()).add(ine)
            for v in variants:
                key = (_iivtnu_norm_text(v), _iivtnu_norm_text(prov))
                if not key[0] or not key[1]:
                    continue
                if key in idx and idx[key] != ine:
                    collisions.setdefault(key, set()).update({idx[key], ine})
                    continue
                idx[key] = ine
    return idx, name_to, collisions


def parse_xlsx(path: Path):
    if not OPENPYXL_AVAILABLE:
        raise SystemExit("openpyxl no está disponible")
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True)
    ws = wb["Orden ALFABETICO"] if "Orden ALFABETICO" in wb.sheetnames else wb[wb.sheetnames[0]]
    max_col = int(getattr(ws, "max_column", 0) or 0)
    rows = []  # [(muni, prov_or_empty, tipo_float)]
    # Cabecera real: fila 8; datos desde 9.
    for r in range(9, ws.max_row + 1):
        muni = ws.cell(r, 1).value
        prov = ws.cell(r, 2).value if max_col >= 3 else ""
        tipo = ws.cell(r, 3).value if max_col >= 3 else ws.cell(r, 2).value
        if muni is None and prov is None and tipo is None:
            continue
        muni = str(muni or "").strip()
        prov = str(prov or "").strip()
        if not muni:
            continue
        if muni.strip().upper() in {"MEDIA CAPITALES", "MEDIA DE CAPITALES"}:
            continue
        if tipo is None or str(tipo).strip() == "":
            continue
        try:
            tipo_f = float(tipo)
        except Exception:
            continue
        if tipo_f < 0 or tipo_f > 60:
            continue
        rows.append((muni, prov, round(tipo_f, 4)))
    return rows


def main():
    ap = argparse.ArgumentParser(description="Genera catálogo de tipos IIVTNU (España: capitales) por INE desde XLSX OTA Málaga.")
    ap.add_argument("--xlsx-2024", required=True, help="Ruta al XLSX Tipo IIVTNU 2024 - Capitales")
    ap.add_argument("--xlsx-2025", required=True, help="Ruta al XLSX Tipo IIVTNU 2025 - Capitales")
    ap.add_argument(
        "--ine-municipios-csv",
        default=str(ROOT / "data" / "catalogos" / "ine_municipios_ine_2026.csv"),
        help="CSV con municipios (municipio_id, nombre, provincia_nombre).",
    )
    ap.add_argument("--out", default=str(ROOT / "data" / "catalogos" / "iivtnu_tipo_gravamen_capitales.min.json"))
    ap.add_argument(
        "--source-url",
        default="https://ota.malaga.eu/Estadisticas-por-impuestos/iivtnu-plusvalia/tipos-de-gravamen/",
        help="URL de referencia (fuente)",
    )
    args = ap.parse_args()

    idx, name_to, collisions = build_index_spain(Path(args.ine_municipios_csv))
    years = {}
    missing = {}
    missing_tipo = {}
    for year, xlsx in (("2024", args.xlsx_2024), ("2025", args.xlsx_2025)):
        rows = parse_xlsx(Path(xlsx))
        ymap = {}
        miss = []
        miss_tipo = []
        for muni, prov, tipo in rows:
            ine = ""
            if prov:
                key = (_iivtnu_norm_text(muni), _iivtnu_norm_text(prov))
                ine = idx.get(key, "")
                if not ine:
                    found = ""
                    for v in _iivtnu_postal_variants(muni):
                        ine2 = idx.get((_iivtnu_norm_text(v), _iivtnu_norm_text(prov)), "")
                        if ine2:
                            found = ine2
                            break
                    ine = found
            if not ine:
                # Capitales OTA: el XLSX suele venir sin provincia; intentamos match por nombre único.
                nn = _iivtnu_norm_text(muni)
                candidates = sorted(list(name_to.get(nn) or []))
                if len(candidates) == 1:
                    ine = candidates[0]
                elif not candidates:
                    # intenta variantes
                    for v in _iivtnu_postal_variants(muni):
                        nn2 = _iivtnu_norm_text(v)
                        candidates2 = sorted(list(name_to.get(nn2) or []))
                        if len(candidates2) == 1:
                            ine = candidates2[0]
                            break
            if not ine:
                miss.append(f"{prov}|{muni}".strip("|"))
                continue
            ine = str(ine).zfill(5)
            if float(tipo or 0.0) <= 0:
                miss_tipo.append(ine)
                continue
            ymap[ine] = float(tipo)
        years[year] = ymap
        missing[year] = sorted(set(miss))
        missing_tipo[year] = sorted(set(miss_tipo))

    out = {
        "scope": "España (capitales de provincia)",
        "source": {
            "2024": {"label": "OTA Málaga · Tipo IIVTNU 2024 - Capitales (XLSX)", "url": str(args.source_url)},
            "2025": {"label": "OTA Málaga · Tipo IIVTNU 2025 - Capitales (XLSX)", "url": str(args.source_url)},
        },
        "years": years,
        "missing": {k: v for k, v in missing.items() if v},
        "missing_tipo": {k: v for k, v in missing_tipo.items() if v},
        "collisions": {f"{k[0]}|{k[1]}": sorted(list(v)) for k, v in collisions.items()} if collisions else {},
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"OK: {out_path} (2024={len(years.get('2024') or {})} 2025={len(years.get('2025') or {})})")
    if out.get("missing"):
        for year, miss in out.get("missing", {}).items():
            print(f"WARNING: missing {year}: {len(miss)}")


if __name__ == "__main__":
    main()
