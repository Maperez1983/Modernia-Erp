#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from web.server import (
    DB_DEFAULT,
    apply_gestoria_import_lote,
    ensure_tables,
    normalize_lookup_text,
    open_sqlite_conn,
    parse_invoice_text,
    refresh_gestoria_import_lote_totals,
    upsert_gestoria_import_document,
)


TEMPLATE_SHEET = "Hoja1"
TEMPLATE_CATEGORY_ROWS = {
    "SEGURO LOCAL": "SEGURO LOCAL",
    "SEGURO RC": "SEGURO RC",
    "ALQUILER LOCAL": "ALQUILER LOCAL",
    "SALARIOS": "SALARIOS",
    "SUMINISTROS": "SUMINISTROS",
    "LOPD": "LOPD",
    "PRL": "PRL",
    "CANAL DE DENUNCIAS": "CANAL DE DENUNCIAS",
    "SEGURO CONVENIO": "SEGURO CONVENIO",
    "GESTORIA": "GESTORIA",
    "SALUD": "SALUD",
    "ILT": "ILT",
}

CATEGORY_RULES = [
    ("SEGURO LOCAL", ("SEGURO LOCAL", "POLIZA LOCAL", "SEGURO OFICINA", "SEGURO NEGOCIO")),
    ("SEGURO RC", ("RESPONSABILIDAD CIVIL", "SEGURO RC", "POLIZA RC")),
    ("ALQUILER LOCAL", ("ALQUILER LOCAL", "ARRENDAMIENTO", "RENTA MENSUAL", "ALQUILER OFICINA")),
    ("SALARIOS", ("NOMINA", "SALARIO", "SEGURO SOCIAL", "TC1", "TC2")),
    ("SUMINISTROS", ("SUMINISTRO", "LUZ", "ELECTRICIDAD", "INTERNET", "TELEFONIA", "GAS")),
    ("LOPD", ("LOPD", "RGPD", "PROTECCION DE DATOS")),
    ("PRL", ("PRL", "PREVENCION DE RIESGOS", "VIGILANCIA DE LA SALUD")),
    ("CANAL DE DENUNCIAS", ("CANAL DE DENUNCIAS", "WHISTLEBLOW", "INFORMANTE")),
    ("SEGURO CONVENIO", ("SEGURO CONVENIO", "CONVENIO COLECTIVO")),
    ("GESTORIA", ("GESTORIA", "ASESORIA", "ASESORIA FISCAL", "ASESORIA LABORAL")),
    ("SALUD", ("SEGURO SALUD", "SALUD COLECTIVO", "SALUD INDIVIDUAL", "MEDICO", "ADESLAS", "SANITAS", "DKV")),
    ("ILT", ("ILT", "INCAPACIDAD LABORAL TEMPORAL", "INCAPACIDAD TEMPORAL")),
]

OUT_OF_SCOPE_RULES = [
    ("COMIDAS", ("COMIDA", "DESAYUNO", "CENA", "CAFES", "CAFÉ", "BEBIDA", "RESTAURANTE", "SUPERMERCADO", "CERVEZA", "CHUCHES", "SAN MIGUEL")),
    ("VIAJE", ("HOTEL", "PARKING", "PEAJE", "REPOSTAJE", "COMBUSTIBLE")),
]

DOC_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg"}
FILENAME_PREFILTER_RULES = [
    ("INGRESO", ("FACTURA_GAPP", "GAPP MONTAJES", "FRA MARZO GAPP", "GAPP ELEVADORES")),
    ("COMIDAS", ("COMIDA", "DESAYUNO", "CENA", "BEBIDA", "CAFES", "CAFE", "INVITACION")),
    ("VIAJE", ("PARKING", "PEAJE", "REPOSTAJE", "COMBUSTIBLE", "GASOLINA", "HOTEL", "RECEIPT_BI", "TICKETS.PDF", "BIL_")),
    ("SUMINISTROS", ("MATERIAL", "MATERIALES", "OBRAMAT", "OPTIMUS", "LEROY", "FERRETERIA", "LIMPIEZA", "HTM", "MANDOS", "CAJA FUERTE", "CERRAJER", "ROPA TRABAJO")),
]

SUPPLIER_CATEGORY_RULES = [
    ("ALQUILER LOCAL", ("ACTIVOS INMOBILIARIOS GILDUSA", "ACTIVOS INMOVILIARIOS GILDUSA")),
    ("SUMINISTROS", ("OPTIMUS", "OBRAMAT", "AUTOPUERTAS FUENGIROLA", "HTM", "LEROY", "INTHER", "HONGSHUK", "COMASUR", "ADAIRA", "AUTOMATISMO MARBELLA")),
]

TRUSTED_OCR_SUPPLIERS = (
    "OBRAMAT",
    "LEROY",
    "OPTIMUS",
    "HTM",
    "COMASUR",
    "ADAIRA",
    "ACTIVOS INMOVILIARIOS GILDUSA",
    "ACTIVOS INMOBILIARIOS GILDUSA",
    "AUTOMATISMO MARBELLA",
)
EXPENSE_TEMPLATE_CATEGORIES = set(TEMPLATE_CATEGORY_ROWS)
INVALID_NUMBER_TOKENS = {
    "PAGADO",
    "EMITIDA",
    "CALLE",
    "SIMPLIFICADA",
    "FACTURA",
    "GAPP",
    "ORIGINAL",
}
INVALID_VENDOR_TOKENS = {
    "Q",
    "QO",
    "Q°",
    "ORIGINAL",
    "PAGADO",
    "EMITIDA",
    "MES FEBRERO",
}


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_decimal(value: Any) -> float:
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    raw = raw.replace("EUR", "").replace("€", "").replace(" ", "")
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", ".", "-", "-."}:
        return 0.0
    try:
        return round(float(raw), 2)
    except ValueError:
        return 0.0


def extract_filename_date(path: Path) -> str:
    match = re.search(r"(\d{1,2})[.\-_/](\d{1,2})[.\-_/](\d{2,4})", path.stem)
    if not match:
        return ""
    day, month, year = match.groups()
    if len(year) == 2:
        year = f"20{year}"
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def extract_amount_by_labels(text: str, labels: tuple[str, ...]) -> float:
    for label in labels:
        pattern = rf"{label}\s*[:\-]?\s*(-?[0-9][0-9\., ]{{0,20}}(?:€|EUR)?)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount = parse_decimal(match.group(1))
            if amount != 0:
                return amount
    return 0.0


def extract_fallback_total(text: str) -> float:
    values: list[float] = []
    for match in re.finditer(r"(?<![\d.,-])(-?\d{1,4}[.,]\d{2})(?![\d%]|[.,]\d)", text):
        amount = parse_decimal(match.group(1))
        if amount != 0 and abs(amount) < 100000:
            values.append(amount)
    if not values:
        return 0.0
    negatives = [value for value in values if value < 0]
    if negatives and "ABONO" in norm(text):
        return round(min(negatives), 2)
    positives = [value for value in values if value > 0]
    if not positives:
        return round(max(values, key=abs), 2)
    return round(max(positives), 2)


def detect_document_type(path: Path, text: str, parsed: dict[str, Any]) -> str:
    filename = norm(path.name)
    haystack = norm(text)
    if "FACTURA_GAPP" in filename or "GESTIONES COMERCIALES" in haystack:
        return "venta"
    if any(token in filename for token in ("COMBUSTIBLE", "PARKING", "PEAJE", "REPOSTAJE", "HOTEL", "OPTIMUS", "OBRAMAT", "HTM", "LEROY", "MANDOS", "ARRENDAMIENTO", "ALQUILER", "MATERIAL", "MATERIALES", "LIMPIEZA", "CERRAJER", "ROPA TRABAJO")):
        return "compra"
    if any(token in haystack for token in ("OPTIMUS", "OBRAMAT", "HTM", "LEROY", "GILDUSA", "COMASUR", "ADAIRA", "AUTOMATISMO MARBELLA", "INTHER")):
        return "compra"
    if "TOTAL A COBRAR" in haystack or "FACTURA EMITIDA" in haystack:
        return "venta"
    if "TOTAL FACTURA" in haystack:
        return "compra"
    return parsed.get("tipo") or "compra"


def infer_vendor(text: str, source_path: Path) -> str:
    sold_by = re.search(r"Vendido por\s+([^\n\r]{3,120})", text, re.IGNORECASE)
    if sold_by:
        return re.sub(r"\s+", " ", sold_by.group(1)).strip(" .:-")[:120]
    lines = [line.strip(" .:-") for line in str(text or "").splitlines()]
    for line in lines[:12]:
        if not line:
            continue
        if len(line) < 3:
            continue
        line = re.sub(r"\s+", " ", line).strip()
        upper = norm(line)
        if " CLIENTE " in f" {upper} ":
            left = re.split(r"\bCLIENTE\b", line, flags=re.IGNORECASE)[0].strip(" .:-")
            if len(left) >= 3:
                return left[:120]
        if " FACTURA " in f" {upper} ":
            left = re.split(r"\bFACTURA\b", line, flags=re.IGNORECASE)[0].strip(" .:-")
            if len(left) >= 3:
                return left[:120]
        if any(token in upper for token in ("FACTURA", "FECHA", "TOTAL", "CLIENTE", "BASE", "IVA")):
            continue
        if re.search(r"\b\d{5}\b", line):
            continue
        if sum(ch.isdigit() for ch in line) >= 6:
            continue
        return line[:120]
    return source_path.stem


def looks_like_own_company(name: str) -> bool:
    return "GAPP" in norm(name)


def looks_like_suspicious_invoice_number(value: str) -> bool:
    token = norm(value)
    if not token:
        return False
    if token in INVALID_NUMBER_TOKENS:
        return True
    if re.fullmatch(r"\d{1,2}", token):
        return True
    if re.fullmatch(r"\d{4}", token):
        return True
    return False


def looks_like_suspicious_vendor(value: str) -> bool:
    token = norm(value)
    if not token:
        return True
    if token in INVALID_VENDOR_TOKENS:
        return True
    if len(token) <= 3:
        return True
    letters = re.sub(r"[^A-Z]", "", token)
    if len(letters) <= 2:
        return True
    return False


def is_implausible_vat(base: float, cuota_iva: float, total: float) -> bool:
    base = float(base or 0.0)
    cuota_iva = float(cuota_iva or 0.0)
    total = float(total or 0.0)
    if cuota_iva < 0:
        return True
    if cuota_iva <= 0:
        return False
    if total > 0 and cuota_iva > total + 0.01:
        return True
    if base > 0:
        iva_pct = cuota_iva / max(base, 0.01)
        if iva_pct > 0.30:
            return True
        if total > 0 and abs((base + cuota_iva) - total) > max(1.0, total * 0.15):
            return True
    return False


def extract_text(path: Path, pdf_pages: int = 2) -> tuple[str, str, str]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text, err = pdftotext_extract_fast(path, pages=pdf_pages)
        if norm(text):
            return text, "pdftotext", err
        text, err = ocr_pdf_first_page_fast(path)
        return text, "ocr_pdf_first_page", err
    if suffix in {".png", ".jpg", ".jpeg"}:
        text, err = ocr_image_file_fast(path)
        return text, "ocr_image_file", err
    return "", "unsupported", "extension no soportada"


def pdftotext_extract_fast(path: Path, pages: int = 2, timeout_seconds: int = 12) -> tuple[str, str]:
    cmd = shutil.which("pdftotext") or "/opt/homebrew/bin/pdftotext" or "/usr/local/bin/pdftotext"
    if not cmd or not os.path.exists(cmd):
        return "", "pdftotext no encontrado"
    with tempfile.TemporaryDirectory() as tmpdir:
        out_txt = Path(tmpdir) / "out.txt"
        try:
            subprocess.run(
                [cmd, "-layout", "-nopgbrk", "-f", "1", "-l", str(max(1, pages)), str(path), str(out_txt)],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired:
            return "", "pdftotext timeout"
        except subprocess.CalledProcessError as exc:
            return "", (exc.stderr or "").strip()
        if not out_txt.exists():
            return "", "pdftotext sin salida"
        return out_txt.read_text(encoding="utf-8", errors="ignore"), ""


def tesseract_stdout(image_path: Path, timeout_seconds: int = 20, psm: int = 6) -> tuple[str, str]:
    cmd = shutil.which("tesseract") or "/opt/homebrew/bin/tesseract" or "/usr/local/bin/tesseract"
    if not cmd or not os.path.exists(cmd):
        return "", "tesseract no encontrado"
    try:
        result = subprocess.run(
            [cmd, str(image_path), "stdout", "-l", "spa+eng", "--oem", "1", "--psm", str(psm)],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=timeout_seconds,
        )
        return result.stdout or "", ""
    except subprocess.TimeoutExpired:
        return "", "tesseract timeout"
    except subprocess.CalledProcessError as exc:
        return "", (exc.stderr or "").strip()


def score_ocr_text(text: str) -> int:
    normalized = norm(text)
    if not normalized:
        return 0
    score = min(len(normalized), 800) // 8
    for token, weight in (
        ("FACTURA", 30),
        ("FECHA", 10),
        ("TOTAL", 20),
        ("IVA", 10),
        ("CLIENTE", 5),
        ("TARJ", 10),
        ("EUR", 10),
    ):
        if token in normalized:
            score += weight
    if re.search(r"(?<![\d.,-])\d{1,4}[.,]\d{2}(?![\d%]|[.,]\d)", text):
        score += 25
    if extract_amount_by_labels(text, ("total factura", "total", "total tti", "total tu", "importe total")) > 0:
        score += 30
    return score


def needs_ocr_rescue(text: str) -> bool:
    normalized = norm(text)
    if not normalized:
        return True
    if len(normalized) < 120:
        return True
    if not re.search(r"(?<![\d.,-])\d{1,4}[.,]\d{2}(?![\d%]|[.,]\d)", text):
        return True
    if "FACTURA" not in normalized and "TOTAL" not in normalized:
        return True
    return False


def sips_transform(src: Path, out: Path, rotate: int | None = None, max_size: int | None = None) -> bool:
    cmd = ["/usr/bin/sips"]
    if rotate is not None:
        cmd.extend(["-r", str(rotate)])
    if max_size is not None:
        cmd.extend(["-Z", str(max_size)])
    cmd.extend([str(src), "--out", str(out)])
    try:
        subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=20)
        return out.exists()
    except Exception:
        return False


def ocr_image_file_fast(path: Path) -> tuple[str, str]:
    best_text, best_err = tesseract_stdout(path, psm=6)
    best_score = score_ocr_text(best_text)
    if not needs_ocr_rescue(best_text):
        return best_text, best_err
    with tempfile.TemporaryDirectory() as tmpdir:
        candidates: list[tuple[str, str]] = [(best_text, best_err)]
        rescue_psm11, rescue_err = tesseract_stdout(path, psm=11)
        candidates.append((rescue_psm11, rescue_err))
        big_path = Path(tmpdir) / "big.jpg"
        if sips_transform(path, big_path, max_size=2600):
            candidates.append(tesseract_stdout(big_path, psm=11))
        for angle in (90, 270):
            rotated = Path(tmpdir) / f"rot_{angle}.jpg"
            if sips_transform(path, rotated, rotate=angle, max_size=2600):
                candidates.append(tesseract_stdout(rotated, psm=11))
        for text, err in candidates:
            score = score_ocr_text(text)
            if score > best_score:
                best_text, best_err, best_score = text, err, score
    return best_text, best_err


def ocr_pdf_first_page_fast(path: Path, timeout_seconds: int = 20) -> tuple[str, str]:
    pdftoppm = shutil.which("pdftoppm") or "/opt/homebrew/bin/pdftoppm" or "/usr/local/bin/pdftoppm"
    if not pdftoppm or not os.path.exists(pdftoppm):
        return "", "pdftoppm no encontrado"
    with tempfile.TemporaryDirectory() as tmpdir:
        out_base = Path(tmpdir) / "page"
        try:
            subprocess.run(
                [pdftoppm, "-png", "-f", "1", "-singlefile", str(path), str(out_base)],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.PIPE,
                text=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired:
            return "", "pdftoppm timeout"
        except subprocess.CalledProcessError as exc:
            return "", (exc.stderr or "").strip()
        image_path = out_base.with_suffix(".png")
        if not image_path.exists():
            return "", "pdftoppm sin salida"
        return tesseract_stdout(image_path, timeout_seconds=timeout_seconds)


def enrich_parsed(path: Path, text: str, parsed: dict[str, Any]) -> dict[str, Any]:
    result = dict(parsed or {})
    result.setdefault("raw_text", text.strip())
    result["tipo"] = detect_document_type(path, text, result)
    if not result.get("fecha"):
        result["fecha"] = extract_filename_date(path)
    guessed_vendor = infer_vendor(text, path)
    if not result.get("tercero") or looks_like_own_company(result.get("tercero") or ""):
        result["tercero"] = guessed_vendor
    if not result.get("numero"):
        number = re.search(r"\b([A-Z]{0,4}\d{2,}[A-Z0-9/\-]*)\b", norm(path.stem))
        result["numero"] = number.group(1) if number else path.stem[:60]
    if looks_like_suspicious_invoice_number(result.get("numero") or ""):
        alt_number_match = re.search(
            r"(?:FACTURA|FRA|N[OU]?M(?:ERO)?)[^A-Z0-9]{0,8}([A-Z]{1,4}[\d/\-]{2,}[A-Z0-9/\-]*)",
            norm(text),
        )
        alt_number = alt_number_match.group(1) if alt_number_match else ""
        result["numero"] = alt_number if alt_number and not looks_like_suspicious_invoice_number(alt_number) else ""
    total = float(result.get("total") or 0.0)
    if total <= 0:
        total = extract_amount_by_labels(
            text,
            (
                "total factura",
                "importe total",
                "total a pagar",
                "total pagar",
                "total iva incluido",
                "importe pagado",
                "total",
                "renta mensual",
                "datafono",
                "datofono",
                "importe",
            ),
        )
        if total <= 0:
            total = extract_fallback_total(text)
        result["total"] = total
    if float(result.get("base_imponible") or 0.0) <= 0 and total > 0:
        base = extract_amount_by_labels(text, ("base imponible", "subtotal", "base"))
        if base <= 0:
            base = total
        result["base_imponible"] = round(base, 2)
    if float(result.get("base_imponible") or 0.0) == 0.0:
        signed_base = extract_amount_by_labels(text, ("base imponible", "bases iva", "total bruto"))
        if signed_base != 0:
            result["base_imponible"] = round(signed_base, 2)
    if float(result.get("total") or 0.0) == 0.0:
        signed_total = extract_amount_by_labels(text, ("total factura", "total"))
        if signed_total != 0:
            result["total"] = round(signed_total, 2)
    if float(result.get("cuota_iva") or 0.0) <= 0 and float(result.get("total") or 0.0) > float(result.get("base_imponible") or 0.0):
        result["cuota_iva"] = round(float(result["total"]) - float(result["base_imponible"]), 2)
    result["descripcion"] = result.get("descripcion") or f"{result.get('numero') or 'Documento'} · {result.get('tercero') or path.stem}"
    return result


def classify_record(path: Path, parsed: dict[str, Any]) -> tuple[str, float, str]:
    corpus = " ".join(
        [
            path.name,
            parsed.get("tercero") or "",
            parsed.get("descripcion") or "",
            parsed.get("raw_text") or "",
        ]
    )
    haystack = norm(corpus)
    for category, tokens in SUPPLIER_CATEGORY_RULES:
        for token in tokens:
            if norm(token) in haystack:
                return category, 0.96, f"proveedor:{token}"
    for bucket, tokens in OUT_OF_SCOPE_RULES:
        for token in tokens:
            if norm(token) in haystack:
                return bucket, 0.95, f"fuera_plantilla:{token}"
    for category, tokens in CATEGORY_RULES:
        for token in tokens:
            if norm(token) in haystack:
                return category, 0.92, f"regla:{token}"
    return "SIN_CATEGORIA", 0.0, "sin_regla"


def preclassify_from_filename(path: Path) -> tuple[str, float, str]:
    haystack = norm(path.name)
    for category, tokens in FILENAME_PREFILTER_RULES:
        for token in tokens:
            if norm(token) in haystack:
                return category, 0.99, f"prefiltro_nombre:{token}"
    return "", 0.0, ""


def infer_target_year(records: list[dict[str, Any]]) -> int | None:
    years: dict[int, int] = defaultdict(int)
    for row in records:
        fecha = str(row.get("fecha") or "")
        if re.match(r"^\d{4}-\d{2}-\d{2}$", fecha):
            years[int(fecha[:4])] += 1
    if not years:
        return None
    return max(years, key=years.get)


def review_record(record: dict[str, Any], target_year: int | None = None) -> tuple[str, str]:
    flags: list[str] = []
    category = record.get("categoria_excel") or ""
    if category not in TEMPLATE_CATEGORY_ROWS:
        return "OK", ""
    amount = float(record.get("importe_agregado") or 0.0)
    confidence = float(record.get("confianza_categoria") or 0.0)
    method = record.get("ocr_metodo") or ""
    vendor = norm(record.get("tercero") or "")
    fecha = str(record.get("fecha") or "")
    motivo = str(record.get("motivo_categoria") or "")
    tipo = str(record.get("tipo") or "").strip().lower()
    numero = str(record.get("numero") or "").strip()
    base = float(record.get("base_imponible") or 0.0)
    cuota_iva = float(record.get("cuota_iva") or 0.0)
    total = float(record.get("total") or 0.0)

    if category in {"ALQUILER LOCAL", "SUMINISTROS"} and amount <= 0:
        flags.append("importe_cero")
    trusted_ocr = any(token in vendor for token in TRUSTED_OCR_SUPPLIERS)

    if confidence < 0.95 and not (
        category == "ALQUILER LOCAL" and "GILDUSA" in vendor
    ):
        flags.append("clasificacion_debil")
    if method == "ocr_image_file" and amount >= 200 and not trusted_ocr:
        flags.append("importe_alto_por_ocr")
    if target_year and re.match(r"^\d{4}-\d{2}-\d{2}$", fecha) and int(fecha[:4]) != target_year:
        flags.append("ano_distinto")
    if "pagado en dic 2025" in norm(record.get("archivo") or ""):
        flags.append("posible_otro_periodo")
    if vendor in {"", "PAGADO", "MES FEBRERO"}:
        flags.append("proveedor_dudoso")
    if category == "SUMINISTROS" and any(token in motivo for token in ("regla:AGUA", "regla:TELEFONO")):
        flags.append("regla_generica")
    if category in EXPENSE_TEMPLATE_CATEGORIES and tipo == "venta":
        flags.append("tipo_venta_en_gasto")
    if looks_like_suspicious_invoice_number(numero):
        flags.append("numero_dudoso")
    if looks_like_suspicious_vendor(record.get("tercero") or "") and not trusted_ocr:
        flags.append("proveedor_dudoso")
    if is_implausible_vat(base, cuota_iva, total):
        flags.append("iva_inverosimil")

    state = "OK" if not flags else "REVISAR"
    deduped = []
    for flag in flags:
        if flag not in deduped:
            deduped.append(flag)
    return state, ",".join(deduped)


def canonical_filename_stem(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    return norm(stem)


def mark_filename_duplicates(records: list[dict[str, Any]]) -> None:
    preferred: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in records:
        key = (
            canonical_filename_stem(record.get("archivo") or ""),
            str(record.get("fecha") or ""),
            str(record.get("categoria_excel") or ""),
        )
        current = preferred.get(key)
        score = (
            1 if record.get("estado_revision") == "OK" else 0,
            float(record.get("importe_agregado") or 0.0),
            float(record.get("confianza_categoria") or 0.0),
        )
        if current is None:
            preferred[key] = record
            record["_dup_score"] = score
            continue
        current_score = current.get("_dup_score") or (
            1 if current.get("estado_revision") == "OK" else 0,
            float(current.get("importe_agregado") or 0.0),
            float(current.get("confianza_categoria") or 0.0),
        )
        if score > current_score:
            preferred[key] = record
            record["_dup_score"] = score
    for record in records:
        key = (
            canonical_filename_stem(record.get("archivo") or ""),
            str(record.get("fecha") or ""),
            str(record.get("categoria_excel") or ""),
        )
        chosen = preferred.get(key)
        filename = str(record.get("archivo") or "")
        if chosen is not record and re.search(r"\(\d+\)", filename):
            record["estado_revision"] = "DUPLICADO"
            record["motivos_revision"] = "duplicado_nombre"
    for record in records:
        record.pop("_dup_score", None)


def row_map_from_sheet(sheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for row_idx in range(1, sheet.max_row + 1):
        label = norm(sheet.cell(row=row_idx, column=1).value)
        if label:
            mapping[label] = row_idx
    return mapping


def write_template_output(template_path: Path, output_path: Path, category_totals: dict[str, float]) -> None:
    wb = load_workbook(template_path)
    sheet = wb[TEMPLATE_SHEET] if TEMPLATE_SHEET in wb.sheetnames else wb.active
    row_map = row_map_from_sheet(sheet)
    for category, label in TEMPLATE_CATEGORY_ROWS.items():
        row_idx = row_map.get(norm(label))
        if not row_idx:
            continue
        sheet.cell(row=row_idx, column=2, value=round(category_totals.get(category, 0.0), 2))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def scan_documents(input_dir: Path, pdf_pages: int = 2, limit: int = 0) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    candidates = sorted(p for p in input_dir.iterdir() if p.is_file() and p.suffix.lower() in DOC_EXTENSIONS)
    if limit > 0:
        candidates = candidates[:limit]
    for path in candidates:
        pre_category, pre_confidence, pre_reason = preclassify_from_filename(path)
        if pre_category in {"COMIDAS", "VIAJE", "INGRESO"}:
            text = ""
            method = "filename_prefilter"
            err = ""
            parsed = enrich_parsed(path, text, {"tipo": "venta" if pre_category == "INGRESO" else "compra"})
            category, confidence, reason = pre_category, pre_confidence, pre_reason
        else:
            text, method, err = extract_text(path, pdf_pages=pdf_pages)
            parsed = enrich_parsed(path, text, parse_invoice_text(text))
            category, confidence, reason = classify_record(path, parsed)
            if pre_category and category == "SIN_CATEGORIA":
                category, confidence, reason = pre_category, pre_confidence, pre_reason
        if category in EXPENSE_TEMPLATE_CATEGORIES and parsed.get("tipo") == "venta":
            parsed["tipo"] = "compra"
        if category == "INGRESO":
            parsed["tipo"] = "venta"
        amount = round(float(parsed.get("total") or parsed.get("base_imponible") or 0.0), 2)
        records.append(
            {
                "archivo": path.name,
                "ruta": str(path),
                "fecha": parsed.get("fecha") or "",
                "numero": parsed.get("numero") or "",
                "tercero": parsed.get("tercero") or "",
                "nif": parsed.get("nif") or "",
                "tipo": parsed.get("tipo") or "",
                "base_imponible": round(float(parsed.get("base_imponible") or 0.0), 2),
                "cuota_iva": round(float(parsed.get("cuota_iva") or 0.0), 2),
                "total": round(float(parsed.get("total") or 0.0), 2),
                "importe_agregado": amount,
                "categoria_excel": category,
                "confianza_categoria": confidence,
                "motivo_categoria": reason,
                "ocr_metodo": method,
                "ocr_error": err or "",
                "descripcion": parsed.get("descripcion") or "",
            }
        )
    target_year = infer_target_year(records)
    for record in records:
        estado, motivos = review_record(record, target_year=target_year)
        record["estado_revision"] = estado
        record["motivos_revision"] = motivos
    mark_filename_duplicates(records)
    return records


def build_category_totals(records: list[dict[str, Any]], include_needs_review: bool = False) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for row in records:
        category = row["categoria_excel"]
        if category not in TEMPLATE_CATEGORY_ROWS:
            continue
        if not include_needs_review and row.get("estado_revision") == "REVISAR":
            continue
        totals[category] += float(row.get("importe_agregado") or 0.0)
    return {key: round(value, 2) for key, value in totals.items()}


def write_csv(output_path: Path, records: list[dict[str, Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "archivo",
        "fecha",
        "numero",
        "tercero",
        "nif",
        "tipo",
        "base_imponible",
        "cuota_iva",
        "total",
        "importe_agregado",
        "categoria_excel",
        "confianza_categoria",
        "motivo_categoria",
        "ocr_metodo",
        "ocr_error",
        "estado_revision",
        "motivos_revision",
        "descripcion",
        "ruta",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def write_json(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_empresa_id(conn, empresa_ref: str) -> str:
    ref = str(empresa_ref or "").strip()
    if not ref:
        raise ValueError("empresa_id o empresa_nombre requerido")
    row = conn.execute("SELECT id FROM empresas WHERE id = ? LIMIT 1", (ref,)).fetchone()
    if row:
        return str(row["id"])
    normalized = normalize_lookup_text(ref)
    row = conn.execute("SELECT id, nombre FROM empresas").fetchall()
    for item in row:
        if normalize_lookup_text(item["nombre"] or "") == normalized:
            return str(item["id"])
    raise ValueError(f"Empresa no encontrada: {ref}")


def create_local_import_lote(
    conn,
    empresa_id: str,
    cliente_id: str | None,
    periodo: str | None,
    origen: str | None,
    carpeta_origen: str | None,
    template_path: str | None,
    notas: str | None,
) -> str:
    lote_id = os.urandom(16).hex()
    conn.execute(
        """
        INSERT INTO gestoria_import_lotes (
          id, empresa_id, cliente_id, origen, estado, periodo, carpeta_origen,
          template_path, notas, created_at, updated_at
        ) VALUES (
          ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now')
        )
        """,
        (
            lote_id,
            empresa_id,
            cliente_id,
            origen or "script_local",
            "nuevo",
            periodo,
            carpeta_origen,
            template_path,
            notas,
        ),
    )
    return lote_id


def import_records_to_local_db(
    db_path: Path,
    empresa_ref: str,
    cliente_id: str | None,
    periodo: str | None,
    records: list[dict[str, Any]],
    apply_ok: bool = False,
    origen: str = "script_local",
    carpeta_origen: str | None = None,
    template_path: str | None = None,
    notas: str | None = None,
) -> dict[str, Any]:
    ensure_tables(db_path)
    conn = open_sqlite_conn(db_path, with_row_factory=True)
    try:
        empresa_id = resolve_empresa_id(conn, empresa_ref)
        lote_id = create_local_import_lote(
            conn,
            empresa_id=empresa_id,
            cliente_id=cliente_id,
            periodo=periodo,
            origen=origen,
            carpeta_origen=carpeta_origen,
            template_path=template_path,
            notas=notas,
        )
        inserted = 0
        for record in records:
            upsert_gestoria_import_document(conn, lote_id, empresa_id, cliente_id, record, "now")
            inserted += 1
        lote = refresh_gestoria_import_lote_totals(conn, lote_id, "now")
        applied_result = {"applied": [], "errors": [], "lote": lote}
        if apply_ok:
            applied_result = apply_gestoria_import_lote(conn, lote_id, empresa_id, "now")
        conn.commit()
        return {
            "db_path": str(db_path),
            "empresa_id": empresa_id,
            "cliente_id": cliente_id,
            "lote_id": lote_id,
            "inserted": inserted,
            "applied": len(applied_result.get("applied") or []),
            "apply_errors": len(applied_result.get("errors") or []),
            "lote": applied_result.get("lote") or lote,
        }
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Lee facturas/tickets y rellena el Excel mensual de GAPP.")
    parser.add_argument("--input-dir", required=True, help="Carpeta con PDFs e imágenes")
    parser.add_argument("--template", required=True, help="Excel plantilla")
    parser.add_argument("--output-excel", required=True, help="Excel de salida")
    parser.add_argument("--output-csv", help="CSV detalle de revisión")
    parser.add_argument("--output-json", help="JSON detalle de revisión")
    parser.add_argument("--pdf-pages", type=int, default=2, help="Páginas máximas a leer con pdftotext")
    parser.add_argument("--limit", type=int, default=0, help="Limita el número de documentos procesados")
    parser.add_argument("--include-needs-review", action="store_true", help="Incluye documentos marcados para revisión en los totales del Excel")
    parser.add_argument("--db-path", help=f"Base SQLite local para importar el lote (por defecto {DB_DEFAULT})")
    parser.add_argument("--empresa-ref", help="ID o nombre de la empresa destino en la base local")
    parser.add_argument("--cliente-id", help="Cliente destino en gestoria")
    parser.add_argument("--periodo", help="Periodo del lote, por ejemplo 2026-03")
    parser.add_argument("--import-to-db", action="store_true", help="Carga el resultado en la base local del CRM")
    parser.add_argument("--apply-ok-to-db", action="store_true", help="Tras cargar, aplica automaticamente los documentos OK a facturas/asientos")
    parser.add_argument("--origen-lote", default="script_local", help="Origen a guardar en gestoria_import_lotes")
    parser.add_argument("--notas-lote", help="Notas del lote en la base local")
    args = parser.parse_args()

    input_dir = Path(args.input_dir).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    output_excel = Path(args.output_excel).expanduser().resolve()
    output_csv = Path(args.output_csv).expanduser().resolve() if args.output_csv else None
    output_json = Path(args.output_json).expanduser().resolve() if args.output_json else None

    if not input_dir.exists():
        raise SystemExit(f"Carpeta no encontrada: {input_dir}")
    if not template_path.exists():
        raise SystemExit(f"Plantilla no encontrada: {template_path}")

    records = scan_documents(input_dir, pdf_pages=max(args.pdf_pages, 1), limit=max(args.limit, 0))
    category_totals = build_category_totals(records, include_needs_review=args.include_needs_review)
    write_template_output(template_path, output_excel, category_totals)

    if output_csv:
        write_csv(output_csv, records)
    if output_json:
        payload = {
            "input_dir": str(input_dir),
            "template": str(template_path),
            "output_excel": str(output_excel),
            "category_totals": category_totals,
            "records": records,
        }
    else:
        payload = None

    db_result = None
    if args.import_to_db or args.apply_ok_to_db:
        empresa_ref = str(args.empresa_ref or "").strip()
        if not empresa_ref:
            raise SystemExit("--empresa-ref es obligatorio cuando usas --import-to-db o --apply-ok-to-db")
        db_path = Path(args.db_path).expanduser().resolve() if args.db_path else DB_DEFAULT
        db_result = import_records_to_local_db(
            db_path=db_path,
            empresa_ref=empresa_ref,
            cliente_id=str(args.cliente_id or "").strip() or None,
            periodo=str(args.periodo or "").strip() or None,
            records=records,
            apply_ok=bool(args.apply_ok_to_db),
            origen=str(args.origen_lote or "").strip() or "script_local",
            carpeta_origen=str(input_dir),
            template_path=str(template_path),
            notas=str(args.notas_lote or "").strip() or None,
        )
        if payload is not None:
            payload["db_import"] = db_result

    if output_json and payload is not None:
        write_json(output_json, payload)

    print(f"OK: {len(records)} documentos analizados")
    review_count = sum(1 for row in records if row.get("estado_revision") == "REVISAR")
    duplicate_count = sum(1 for row in records if row.get("estado_revision") == "DUPLICADO")
    print(f"REVISAR: {review_count}")
    print(f"DUPLICADO: {duplicate_count}")
    for category in TEMPLATE_CATEGORY_ROWS:
        print(f"{category}: {category_totals.get(category, 0.0):.2f}")
    if db_result:
        print(f"DB LOTE: {db_result['lote_id']}")
        print(f"DB INSERTADOS: {db_result['inserted']}")
        print(f"DB APLICADOS: {db_result['applied']}")
        print(f"DB ERRORES_APLICACION: {db_result['apply_errors']}")


if __name__ == "__main__":
    main()
