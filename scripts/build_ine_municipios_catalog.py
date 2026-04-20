import argparse
import csv
import sys
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from web.server import OPENPYXL_AVAILABLE, POSTAL_PROVINCES


def main():
    ap = argparse.ArgumentParser(description="Convierte el XLSX del INE (diccionario municipios) a CSV para uso en la app.")
    ap.add_argument("--ine-xlsx", required=True, help="Ruta al diccionario INE (p.ej. diccionario26.xlsx).")
    ap.add_argument("--out", default=str(ROOT / "data" / "catalogos" / "ine_municipios_ine_2026.csv"))
    args = ap.parse_args()

    if not OPENPYXL_AVAILABLE:
        raise SystemExit("openpyxl no está disponible")
    from openpyxl import load_workbook

    wb = load_workbook(str(Path(args.ine_xlsx)), data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Header row is expected at row 2.
    headers = [str(ws.cell(2, c).value or "").strip().upper() for c in range(1, 10)]
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}
    required = ("CPRO", "CMUN", "DC", "NOMBRE")
    if not all(k in col for k in required):
        raise SystemExit(f"XLSX inesperado, faltan columnas: {required} (headers={headers})")

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as handle:
        w = csv.writer(handle)
        w.writerow(["municipio_id", "provincia_id", "cmun", "dc", "nombre", "provincia_nombre"])
        for r in range(3, ws.max_row + 1):
            prov_id = str(ws.cell(r, col["CPRO"]).value or "").strip().zfill(2)
            cmun = str(ws.cell(r, col["CMUN"]).value or "").strip().zfill(3)
            dc = str(ws.cell(r, col["DC"]).value or "").strip()
            nombre = str(ws.cell(r, col["NOMBRE"]).value or "").strip()
            if not prov_id or not cmun or not nombre:
                continue
            municipio_id = f"{prov_id}{cmun}"
            provincia_nombre = POSTAL_PROVINCES.get(prov_id, "")
            w.writerow([municipio_id, prov_id, cmun, dc, nombre, provincia_nombre])

    print(f"OK: {out_path}")


if __name__ == "__main__":
    main()

