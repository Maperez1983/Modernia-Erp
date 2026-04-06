#!/usr/bin/env python3
"""
Convierte exports "DIARIO" (Monitor AON) a la plantilla "Plantilla MiConversor (empresas).xlsx".

Entrada:
  - Una carpeta raíz con subcarpetas que contienen ficheros `DIARIO*.xlsx`.
  - Una plantilla XLSX con una hoja `Hoja1` y las cabeceras:
    FECHA ASIENTO, FECHA FACTURA, Nº FACTURA, CONCEPTO, SUBCUENTA, NIF, NOMBRE,
    DOMICILIO, LOCALIDAD, PROVINCIA, CODIGO POSTAL, BASE IMPONIBLE, % IVA,
    IMPORTE IVA, SUBCUENTA GASTOS/INGRESOS, IMPORTE (TOTAL)

Salida:
  - Un XLSX por carpeta (autónomo) encontrado, con 2 hojas extra:
    - CONTROL_LISTADO: facturas numeradas y ordenadas por fecha.
    - CONTROL_TOTALIZADOR: totales (base, IVA, retenciones, número de facturas).
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


_DATE_RE = re.compile(r"Fecha:\s*(\d{2}/\d{2}/\d{4})", re.IGNORECASE)
_INVOICE_RE = re.compile(r"\bN\s*/\s*Fra\s*:\s*([^\s]+)\b", re.IGNORECASE)
_INVOICE_RE_ALT = re.compile(r"\bN[ºo]?\s*Fra\s*:\s*([^\s]+)\b", re.IGNORECASE)


def _parse_spanish_date(value: str) -> Optional[_dt.date]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return _dt.datetime.strptime(raw, "%d/%m/%Y").date()
    except Exception:
        return None


def _safe_str(value) -> str:
    return str(value or "").strip()


def _net_amount(debe, haber) -> float:
    try:
        d = float(debe or 0.0)
    except Exception:
        d = 0.0
    try:
        h = float(haber or 0.0)
    except Exception:
        h = 0.0
    return d - h


def _is_third_party_account(account: str) -> bool:
    a = (account or "").strip()
    return a.startswith(("43", "40", "41", "44"))


def _is_vat_account(account: str) -> bool:
    a = (account or "").strip()
    return a.startswith(("472", "477"))


def _vat_bucket(account: str) -> str:
    a = (account or "").strip()
    if a.startswith("477"):
        return "repercutido"
    if a.startswith("472"):
        return "soportado"
    return "desconocido"


def _is_withholding_account(account: str) -> bool:
    a = (account or "").strip()
    return a.startswith(("473", "4751", "4750"))


def _guess_invoice_number(concept: str, document: str) -> str:
    c = str(concept or "").strip()
    for rx in (_INVOICE_RE, _INVOICE_RE_ALT):
        m = rx.search(c)
        if m:
            return m.group(1).strip()
    # fallback: el documento (E-000001) es mejor que vacío
    return str(document or "").strip()


@dataclass(frozen=True)
class DiarioLine:
    account: str
    description: str
    concept: str
    debe: float
    haber: float
    contra: str
    document: str

    @property
    def net(self) -> float:
        return float(self.debe or 0.0) - float(self.haber or 0.0)


@dataclass
class InvoiceRecord:
    fecha: _dt.date
    documento: str
    numero_factura: str
    tercero_subcuenta: str
    tercero_nombre: str
    base_subcuenta: str
    base: float
    iva: float
    iva_pct: float
    retencion: float
    total: float
    iva_bucket: str
    source_file: str


def _parse_diario_xlsx(path: Path) -> List[InvoiceRecord]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        current_date: Optional[_dt.date] = None
        docs: Dict[str, List[DiarioLine]] = {}
        doc_date: Dict[str, _dt.date] = {}

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cell0 = row[0]
            if isinstance(cell0, str):
                m = _DATE_RE.search(cell0)
                if m:
                    parsed = _parse_spanish_date(m.group(1))
                    if parsed:
                        current_date = parsed
                    continue

            account = _safe_str(row[0] if len(row) > 0 else "")
            if not account:
                continue
            # Saltar cabecera
            if account.upper() == "CUENTA":
                continue
            if current_date is None:
                continue

            description = _safe_str(row[1] if len(row) > 1 else "")
            concept = _safe_str(row[2] if len(row) > 2 else "")
            debe = float(row[3] or 0.0) if len(row) > 3 else 0.0
            haber = float(row[4] or 0.0) if len(row) > 4 else 0.0
            contra = _safe_str(row[5] if len(row) > 5 else "")
            document = _safe_str(row[6] if len(row) > 6 else "")
            if not document:
                # sin documento no podemos agrupar de forma fiable
                continue

            line = DiarioLine(
                account=account,
                description=description,
                concept=concept,
                debe=float(debe or 0.0),
                haber=float(haber or 0.0),
                contra=contra,
                document=document,
            )
            docs.setdefault(document, []).append(line)
            doc_date.setdefault(document, current_date)

        records: List[InvoiceRecord] = []
        for document, lines in docs.items():
            fecha = doc_date.get(document)
            if not fecha:
                continue

            # tercero
            third_party_lines = [ln for ln in lines if _is_third_party_account(ln.account)]
            third_party = None
            if third_party_lines:
                third_party = max(third_party_lines, key=lambda ln: abs(ln.net))
            tercero_subcuenta = third_party.account if third_party else ""
            tercero_nombre = third_party.description if third_party else ""

            # num factura (mejor primer concepto útil)
            numero = ""
            for ln in lines:
                numero = _guess_invoice_number(ln.concept, document)
                if numero:
                    break

            # IVA / retenciones / base
            vat_lines = [ln for ln in lines if _is_vat_account(ln.account)]
            ret_lines = [ln for ln in lines if _is_withholding_account(ln.account)]
            base_lines = [ln for ln in lines if (not _is_third_party_account(ln.account) and not _is_vat_account(ln.account) and not _is_withholding_account(ln.account))]

            base = sum(abs(ln.net) for ln in base_lines)
            iva = sum(abs(ln.net) for ln in vat_lines)
            ret = sum(abs(ln.net) for ln in ret_lines)

            iva_pct = 0.0
            if base and iva:
                try:
                    iva_pct = round((iva / base) * 100.0, 2)
                except Exception:
                    iva_pct = 0.0

            base_subcuenta = ""
            if base_lines:
                base_subcuenta = max(base_lines, key=lambda ln: abs(ln.net)).account

            iva_bucket = "desconocido"
            if vat_lines:
                buckets = {_vat_bucket(ln.account) for ln in vat_lines}
                if len(buckets) == 1:
                    iva_bucket = next(iter(buckets))

            total = 0.0
            if third_party_lines:
                total = sum(abs(ln.net) for ln in third_party_lines)
            else:
                # fallback
                total = max(0.0, base + iva - ret)

            records.append(
                InvoiceRecord(
                    fecha=fecha,
                    documento=document,
                    numero_factura=numero,
                    tercero_subcuenta=tercero_subcuenta,
                    tercero_nombre=tercero_nombre,
                    base_subcuenta=base_subcuenta,
                    base=float(base or 0.0),
                    iva=float(iva or 0.0),
                    iva_pct=float(iva_pct or 0.0),
                    retencion=float(ret or 0.0),
                    total=float(total or 0.0),
                    iva_bucket=iva_bucket,
                    source_file=str(path),
                )
            )

        records.sort(key=lambda r: (r.fecha, r.numero_factura, r.documento))
        return records
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _iter_diario_files(root: Path) -> Iterable[Path]:
    for p in root.rglob("*.xlsx"):
        name = p.name.lower()
        if "diario" in name:
            yield p


def _write_output_workbook(template_path: Path, out_path: Path, records: List[InvoiceRecord], *, title: str):
    wb = load_workbook(template_path)
    if "Hoja1" not in wb.sheetnames:
        raise SystemExit("La plantilla debe tener una hoja llamada 'Hoja1'.")
    ws = wb["Hoja1"]

    # Escribir desde la fila 2 (la fila 1 son cabeceras)
    start_row = 2
    # Limpiar filas existentes salvo la 2 (que trae la fórmula de CONCEPTO). Mantener fórmula al copiarla.
    if ws.max_row > start_row:
        ws.delete_rows(start_row + 1, ws.max_row - start_row)

    # Copiar estilo/formula base de la fila 2 si existe
    base_formula = ws.cell(start_row, 4).value
    if not base_formula:
        base_formula = "=+CONCATENATE(C{row},\" \",G{row})"

    for idx, rec in enumerate(records):
        r = start_row + idx
        ws.cell(r, 1).value = rec.fecha
        ws.cell(r, 2).value = rec.fecha
        ws.cell(r, 3).value = rec.numero_factura
        ws.cell(r, 4).value = str(base_formula).replace("C2", f"C{r}").replace("G2", f"G{r}").replace("{row}", str(r))
        ws.cell(r, 5).value = rec.tercero_subcuenta
        ws.cell(r, 6).value = ""  # NIF (no aparece en DIARIO)
        ws.cell(r, 7).value = rec.tercero_nombre
        ws.cell(r, 8).value = ""
        ws.cell(r, 9).value = ""
        ws.cell(r, 10).value = ""
        ws.cell(r, 11).value = ""
        ws.cell(r, 12).value = float(rec.base or 0.0)
        ws.cell(r, 13).value = float(rec.iva_pct or 0.0)
        ws.cell(r, 14).value = float(rec.iva or 0.0)
        ws.cell(r, 15).value = rec.base_subcuenta
        ws.cell(r, 16).value = float(rec.total or 0.0)

        # formats
        for c in (1, 2):
            ws.cell(r, c).number_format = "dd/mm/yyyy"
        for c in (12, 14, 16):
            ws.cell(r, c).number_format = "#,##0.00"
        ws.cell(r, 13).number_format = "0.00"

    # Hoja de control: listado
    if "CONTROL_LISTADO" in wb.sheetnames:
        del wb["CONTROL_LISTADO"]
    wsl = wb.create_sheet("CONTROL_LISTADO")
    wsl.append(["#", "FECHA", "Nº FACTURA", "DOCUMENTO", "SUBCUENTA", "NOMBRE", "BASE", "IVA", "RETENCION", "TOTAL", "IVA_TIPO", "ORIGEN"])
    for i, rec in enumerate(sorted(records, key=lambda r: (r.fecha, r.numero_factura, r.documento)), start=1):
        wsl.append(
            [
                i,
                rec.fecha,
                rec.numero_factura,
                rec.documento,
                rec.tercero_subcuenta,
                rec.tercero_nombre,
                rec.base,
                rec.iva,
                rec.retencion,
                rec.total,
                rec.iva_bucket,
                os.path.basename(rec.source_file),
            ]
        )
    for r in range(2, wsl.max_row + 1):
        wsl.cell(r, 2).number_format = "dd/mm/yyyy"
        for c in (7, 8, 9, 10):
            wsl.cell(r, c).number_format = "#,##0.00"

    # Hoja de control: totalizador
    if "CONTROL_TOTALIZADOR" in wb.sheetnames:
        del wb["CONTROL_TOTALIZADOR"]
    wst = wb.create_sheet("CONTROL_TOTALIZADOR")
    wst["A1"].value = "RESUMEN"
    wst["B1"].value = title
    wst["A3"].value = "Nº FACTURAS"
    wst["B3"].value = len(records)
    base_sum = sum(float(r.base or 0.0) for r in records)
    iva_sum = sum(float(r.iva or 0.0) for r in records)
    ret_sum = sum(float(r.retencion or 0.0) for r in records)
    total_sum = sum(float(r.total or 0.0) for r in records)
    wst["A4"].value = "BASE IMPONIBLE (SUMA)"
    wst["B4"].value = base_sum
    wst["A5"].value = "IVA (SUMA)"
    wst["B5"].value = iva_sum
    wst["A6"].value = "RETENCIONES (SUMA)"
    wst["B6"].value = ret_sum
    wst["A7"].value = "TOTAL (SUMA)"
    wst["B7"].value = total_sum

    iva_rep = sum(float(r.iva or 0.0) for r in records if r.iva_bucket == "repercutido")
    iva_sop = sum(float(r.iva or 0.0) for r in records if r.iva_bucket == "soportado")
    wst["A9"].value = "IVA REPERCUTIDO (SUMA)"
    wst["B9"].value = iva_rep
    wst["A10"].value = "IVA SOPORTADO (SUMA)"
    wst["B10"].value = iva_sop

    for cell in ("B4", "B5", "B6", "B7", "B9", "B10"):
        wst[cell].number_format = "#,##0.00"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Carpeta raíz con exports AON (contiene DIARIO*.xlsx).")
    ap.add_argument("--template", required=True, help="Plantilla XLSX 'MiConversor (empresas)'.")
    ap.add_argument("--out-dir", default="reports/aon_miconversor", help="Carpeta de salida (por defecto en el repo).")
    ap.add_argument("--one-file", action="store_true", help="Genera un único XLSX combinando todas las DIARIO encontradas.")
    args = ap.parse_args(argv)

    root = Path(args.root).expanduser().resolve()
    template = Path(args.template).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()

    if not root.exists():
        raise SystemExit(f"No existe root: {root}")
    if not template.exists():
        raise SystemExit(f"No existe template: {template}")

    diario_files = sorted(set(_iter_diario_files(root)))
    if not diario_files:
        raise SystemExit("No he encontrado ningún fichero DIARIO*.xlsx en la carpeta indicada.")

    if args.one_file:
        all_records: List[InvoiceRecord] = []
        for f in diario_files:
            all_records.extend(_parse_diario_xlsx(f))
        title = f"AON DIARIO ({len(diario_files)} ficheros)"
        out_path = out_dir / "miconversor_autonomos.xlsx"
        _write_output_workbook(template, out_path, all_records, title=title)
        print(str(out_path))
        return 0

    # Un XLSX por carpeta contenedora del DIARIO
    by_folder: Dict[Path, List[Path]] = {}
    for f in diario_files:
        by_folder.setdefault(f.parent, []).append(f)

    for folder, files in sorted(by_folder.items(), key=lambda kv: str(kv[0])):
        records: List[InvoiceRecord] = []
        for f in sorted(files):
            records.extend(_parse_diario_xlsx(f))
        if not records:
            continue
        # Nombre de salida estable sin exponer datos: usar el último segmento de la carpeta
        safe_folder = folder.name.replace(" ", "_")
        out_path = out_dir / f"{safe_folder}_miconversor.xlsx"
        _write_output_workbook(template, out_path, records, title=safe_folder)
        print(str(out_path))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

