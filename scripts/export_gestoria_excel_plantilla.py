#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from copy import copy as shallow_copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from web.server import (  # noqa: E402
    GESTORIA_EXCEL_TEMPLATE,
    ensure_tables,
    normalize_lookup_text,
    open_sqlite_conn,
    normalize_service_key,
)


def build_rows(conn, empresa_id: str, cliente_id: str) -> list[list[object]]:
    diario = conn.execute(
        """
        SELECT a.id AS asiento_id, a.fecha, a.concepto, a.referencia,
               l.cuenta, l.descripcion, l.debe, l.haber,
               l.impuesto_tipo, l.impuesto_pct,
               COALESCE(t.nombre, '') AS tercero,
               COALESCE(t.nif, '') AS tercero_nif,
               COALESCE(f.numero, '') AS factura_numero,
               COALESCE(f.fecha_emision, '') AS factura_fecha,
               COALESCE(f.total, 0) AS factura_total,
               COALESCE(f.tipo, '') AS tipo_factura
        FROM gestoria_asientos a
        JOIN gestoria_asiento_lineas l ON l.asiento_id = a.id
        LEFT JOIN gestoria_terceros t ON t.id = l.tercero_id
        LEFT JOIN gestoria_facturas f ON f.id = a.factura_id
        WHERE a.empresa_id = ? AND a.cliente_id = ?
        ORDER BY a.fecha ASC, a.created_at ASC, l.cuenta ASC
        """,
        (empresa_id, cliente_id),
    ).fetchall()
    grouped: dict[str, list[dict]] = {}
    for row in diario:
        key = str(row["asiento_id"] or "").strip() or f"{row['fecha'] or ''}-{row['referencia'] or ''}"
        grouped.setdefault(key, []).append(row)
    output_rows: list[list[object]] = []
    for _key, lines in grouped.items():
        if not lines:
            continue
        sample = lines[0]
        base = 0.0
        iva_pct = 0.0
        iva_importe = 0.0
        subcuenta_tercero = ""
        subcuenta_gyi = ""
        tipo_venta = normalize_service_key(sample["tipo_factura"] or "") == "venta"
        for line in lines:
            cuenta = str(line["cuenta"] or "").strip()
            debe = float(line["debe"] or 0)
            haber = float(line["haber"] or 0)
            imp_tipo = normalize_service_key(line["impuesto_tipo"] or "")
            if not subcuenta_tercero and cuenta.startswith("4"):
                subcuenta_tercero = cuenta
            if not subcuenta_gyi and (cuenta.startswith("6") or cuenta.startswith("7")):
                subcuenta_gyi = cuenta
            if imp_tipo == "iva":
                iva_importe += abs(haber if tipo_venta else debe)
                if not iva_pct:
                    iva_pct = float(line["impuesto_pct"] or 0)
            if subcuenta_gyi == cuenta:
                base += abs(haber if cuenta.startswith("7") else debe)
        total = float(sample["factura_total"] or 0) or (base + iva_importe)
        output_rows.append(
            [
                sample["fecha"] or "",
                sample["factura_fecha"] or "",
                sample["factura_numero"] or sample["referencia"] or "",
                sample["concepto"] or "",
                subcuenta_tercero,
                sample["tercero_nif"] or "",
                sample["tercero"] or "",
                "",
                "",
                "",
                "",
                round(base, 2) if base else "",
                round(iva_pct, 2) if iva_pct else "",
                round(iva_importe, 2) if iva_importe else "",
                subcuenta_gyi,
                round(total, 2) if total else "",
            ]
        )
    return output_rows


def build_workbook(rows: list[list[object]], template_path: Path | None) -> Workbook:
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
    elif GESTORIA_EXCEL_TEMPLATE.exists():
        wb = load_workbook(GESTORIA_EXCEL_TEMPLATE)
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        headers = [
            "FECHA ASIENTO",
            "FECHA FACTURA",
            "Nº FACTURA",
            "CONCEPTO",
            "SUBCUENTA",
            "NIF",
            "NOMBRE",
            "DOMICILIO",
            "LOCALIDAD",
            "PROVINCIA",
            "CODIGO POSTAL",
            "BASE IMPONIBLE",
            "% IVA",
            "IMPORTE IVA",
            "SUBCUENTA GASTOS/INGRESOS",
            "IMPORTE (TOTAL)",
        ]
        ws.append(headers)
        ws.append([None] * 16)

    style_row_idx = 2
    max_existing = max(ws.max_row, style_row_idx)
    style_cells = [ws.cell(style_row_idx, col) for col in range(1, 17)]
    for row_idx in range(2, max_existing + 1):
        for col in range(1, 17):
            ws.cell(row_idx, col).value = None

    for offset, row in enumerate(rows, start=0):
        row_idx = 2 + offset
        for col in range(1, 17):
            target = ws.cell(row_idx, col)
            source = style_cells[col - 1]
            target._style = shallow_copy(source._style)
            target.number_format = source.number_format
            target.protection = shallow_copy(source.protection)
            target.alignment = shallow_copy(source.alignment)
            target.font = shallow_copy(source.font)
            target.fill = shallow_copy(source.fill)
            target.border = shallow_copy(source.border)
        ws.cell(row_idx, 1).value = row[0]
        ws.cell(row_idx, 2).value = row[1]
        ws.cell(row_idx, 3).value = row[2]
        ws.cell(row_idx, 4).value = f'=CONCATENATE(C{row_idx}," ",G{row_idx})'
        ws.cell(row_idx, 5).value = row[4]
        ws.cell(row_idx, 6).value = row[5]
        ws.cell(row_idx, 7).value = row[6]
        ws.cell(row_idx, 8).value = row[7]
        ws.cell(row_idx, 9).value = row[8]
        ws.cell(row_idx, 10).value = row[9]
        ws.cell(row_idx, 11).value = row[10]
        ws.cell(row_idx, 12).value = row[11]
        ws.cell(row_idx, 13).value = row[12]
        ws.cell(row_idx, 14).value = row[13]
        ws.cell(row_idx, 15).value = row[14]
        ws.cell(row_idx, 16).value = row[15]

    return wb


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", normalize_lookup_text(value).lower()).strip("_")
    return slug or "cliente"


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta la Plantilla conversor asientos facturas desde la BD local.")
    parser.add_argument("--db", required=True, help="Ruta sqlite")
    parser.add_argument("--empresa-id", required=True, help="empresa_id (UUID) en la BD")
    parser.add_argument("--cliente-id", required=True, help="cliente_id en la BD")
    parser.add_argument("--output", required=True, help="Ruta salida xlsx")
    parser.add_argument("--template", help="Ruta plantilla xlsx (opcional)")
    args = parser.parse_args()

    db_path = Path(args.db).expanduser().resolve()
    out_path = Path(args.output).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve() if args.template else None
    ensure_tables(db_path)
    conn = open_sqlite_conn(db_path, with_row_factory=True)
    try:
        rows = build_rows(conn, str(args.empresa_id).strip(), str(args.cliente_id).strip())
        wb = build_workbook(rows, template_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        cliente = conn.execute("SELECT nombre FROM clientes WHERE id = ?", (str(args.cliente_id).strip(),)).fetchone()
        print(f"OK: {len(rows)} filas · cliente={slugify((cliente['nombre'] if cliente else '') or '')} · output={out_path}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
