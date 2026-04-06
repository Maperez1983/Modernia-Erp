#!/usr/bin/env python3
"""
Importa PDFs/JPG/PNG de facturas/tickets y genera un Excel basado en
"Plantilla MiConversor (empresas).xlsx" + 2 hojas de control.

Fuente:
  - Documentos (pdf/jpg/jpeg/png) en una carpeta (recursivo).

Salida:
  - Un XLSX con:
    - Hoja1 (la plantilla MiConversor)
    - CONTROL_LISTADO (facturas numeradas y ordenadas por fecha)
    - CONTROL_TOTALIZADOR (totales: base, IVA soportado/repercutido, retenciones, nº facturas)

Notas:
  - Si faltan NIF/domicilio, se dejan en blanco.
  - Se generan subcuentas de tercero determinísticas (410/430) por NIF o nombre.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import re
import sys
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.build_gapp_facturas_excel import extract_text, enrich_parsed, infer_vendor, norm  # noqa: E402
from web.server import infer_expense_account, infer_revenue_account, parse_invoice_text


DOC_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg"}


def _parse_iso_date(value: str) -> Optional[_dt.date]:
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return _dt.datetime.strptime(raw, fmt).date()
        except Exception:
            pass
    return None


def _extract_filename_date(path: Path) -> Optional[_dt.date]:
    m = re.search(r"(\d{1,2})[.\-_/](\d{1,2})[.\-_/](\d{2,4})", path.stem)
    if not m:
        return None
    day, month, year = m.groups()
    if len(year) == 2:
        year = f"20{year}"
    try:
        return _dt.date(int(year), int(month), int(day))
    except Exception:
        return None


def _safe_float(value) -> float:
    try:
        return float(value or 0.0)
    except Exception:
        return 0.0


def _infer_iva_pct(parsed: dict) -> float:
    pct = _safe_float(parsed.get("iva_pct"))
    if pct > 0:
        return round(pct, 2)
    base = _safe_float(parsed.get("base_imponible"))
    iva = _safe_float(parsed.get("cuota_iva"))
    if base > 0 and iva > 0:
        return round((iva / base) * 100.0, 2)
    return 0.0


def _pad_account(prefix: str, seq: int) -> str:
    prefix = str(prefix or "").strip()
    return f"{prefix}{seq:06d}"


def _guess_gasto_ingreso_account(parsed: dict) -> str:
    concepto = parsed.get("descripcion") or parsed.get("numero") or ""
    kind = (parsed.get("tipo") or "").strip().lower() or "compra"
    if kind == "venta":
        base = infer_revenue_account(concepto)
        return f"{base}000000"
    base = infer_expense_account(concepto)
    return f"{base}000000"


def _iter_docs(root: Path) -> Iterable[Path]:
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() in DOC_EXTENSIONS:
            yield p


def _sanitize_filename(value: str) -> str:
    raw = str(value or "").strip()
    raw = re.sub(r"\s+", " ", raw)
    raw = re.sub(r"[^A-Za-z0-9._ -]+", "", raw)
    raw = raw.strip(" ._-")
    raw = raw.replace(" ", "_")
    return raw[:120] or "autonomo"


@dataclass
class DocRecord:
    fecha: _dt.date
    numero: str
    tercero: str
    nif: str
    tipo: str
    base: float
    iva: float
    irpf: float
    total: float
    iva_pct: float
    tercero_subcuenta: str
    gasto_ingreso_subcuenta: str
    iva_bucket: str
    source_file: str
    ocr_method: str
    ocr_error: str


def _parse_doc(path: Path, *, pdf_pages: int = 2) -> DocRecord:
    text, method, err = extract_text(path, pdf_pages=pdf_pages)
    parsed = enrich_parsed(path, text, parse_invoice_text(text))
    tercero = (parsed.get("tercero") or "").strip()
    if not tercero:
        tercero = infer_vendor(text, path)
        parsed["tercero"] = tercero
    nif = (parsed.get("nif") or "").strip()
    tipo = (parsed.get("tipo") or "").strip().lower() or "compra"

    fecha = _parse_iso_date(parsed.get("fecha") or "") or _extract_filename_date(path) or _dt.date.today()
    numero = (parsed.get("numero") or "").strip() or path.stem

    base = round(_safe_float(parsed.get("base_imponible")), 2)
    iva = round(_safe_float(parsed.get("cuota_iva")), 2)
    irpf = round(_safe_float(parsed.get("cuota_irpf")), 2)
    total = round(_safe_float(parsed.get("total")) or round(base + iva - irpf, 2), 2)
    iva_pct = _infer_iva_pct(parsed)

    gasto_ingreso = _guess_gasto_ingreso_account(parsed)
    iva_bucket = "repercutido" if tipo == "venta" else "soportado"

    return DocRecord(
        fecha=fecha,
        numero=numero,
        tercero=tercero,
        nif=nif,
        tipo=tipo,
        base=base,
        iva=iva,
        irpf=irpf,
        total=total,
        iva_pct=iva_pct,
        tercero_subcuenta="",
        gasto_ingreso_subcuenta=gasto_ingreso,
        iva_bucket=iva_bucket,
        source_file=str(path),
        ocr_method=method,
        ocr_error=err or "",
    )

def _assign_tercero_subcuentas(records: List[DocRecord]) -> None:
    """
    Asigna `tercero_subcuenta` de forma estable/determinística para evitar dependencia del orden/threads.
    - Compras -> prefijo 410
    - Ventas  -> prefijo 430
    Key preferida: NIF, si no, nombre normalizado.
    """
    buckets: Dict[str, Dict[str, str]] = {"410": {}, "430": {}}
    for prefix in ("410", "430"):
        keys = set()
        for r in records:
            p = "430" if (r.tipo or "").strip().lower() == "venta" else "410"
            if p != prefix:
                continue
            key = (r.nif or "").strip() or norm(r.tercero or "")
            key = key.strip() or os.path.basename(r.source_file)
            keys.add(key)
        for idx, key in enumerate(sorted(keys), start=1):
            buckets[prefix][key] = _pad_account(prefix, idx)
    for r in records:
        prefix = "430" if (r.tipo or "").strip().lower() == "venta" else "410"
        key = (r.nif or "").strip() or norm(r.tercero or "")
        key = key.strip() or os.path.basename(r.source_file)
        r.tercero_subcuenta = buckets[prefix].get(key) or _pad_account(prefix, 1)


def _scan_folder_docs(folder: Path, *, pdf_pages: int, workers: int = 1, verbose: bool = False) -> List[DocRecord]:
    docs = sorted(set(_iter_docs(folder)))
    if not docs:
        return []
    workers = int(workers or 1)
    workers = max(1, min(8, workers))
    records: List[DocRecord] = []
    if workers == 1:
        for idx, p in enumerate(docs, start=1):
            records.append(_parse_doc(p, pdf_pages=pdf_pages))
            if verbose and (idx % 50 == 0):
                print(f"[{folder.name}] {idx}/{len(docs)}")
        _assign_tercero_subcuentas(records)
        return records
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(_parse_doc, p, pdf_pages=pdf_pages): p for p in docs}
        done = 0
        for fut in as_completed(futs):
            records.append(fut.result())
            done += 1
            if verbose and (done % 50 == 0):
                print(f"[{folder.name}] {done}/{len(docs)}")
    _assign_tercero_subcuentas(records)
    return records

def _write_output(template_path: Path, out_path: Path, records: List[DocRecord], *, title: str):
    wb = load_workbook(template_path)
    if "Hoja1" not in wb.sheetnames:
        raise SystemExit("La plantilla debe tener una hoja llamada 'Hoja1'.")
    ws = wb["Hoja1"]

    start_row = 2
    if ws.max_row > start_row:
        ws.delete_rows(start_row + 1, ws.max_row - start_row)

    base_formula = ws.cell(start_row, 4).value
    if not base_formula:
        base_formula = "=+CONCATENATE(C{row},\" \",G{row})"

    records_sorted = sorted(records, key=lambda r: (r.fecha, r.numero, r.tercero))
    for idx, rec in enumerate(records_sorted):
        r = start_row + idx
        ws.cell(r, 1).value = rec.fecha
        ws.cell(r, 2).value = rec.fecha
        ws.cell(r, 3).value = rec.numero
        ws.cell(r, 4).value = str(base_formula).replace("C2", f"C{r}").replace("G2", f"G{r}").replace("{row}", str(r))
        ws.cell(r, 5).value = rec.tercero_subcuenta
        ws.cell(r, 6).value = rec.nif
        ws.cell(r, 7).value = rec.tercero
        ws.cell(r, 8).value = ""
        ws.cell(r, 9).value = ""
        ws.cell(r, 10).value = ""
        ws.cell(r, 11).value = ""
        ws.cell(r, 12).value = float(rec.base or 0.0)
        ws.cell(r, 13).value = float(rec.iva_pct or 0.0)
        ws.cell(r, 14).value = float(rec.iva or 0.0)
        ws.cell(r, 15).value = rec.gasto_ingreso_subcuenta
        ws.cell(r, 16).value = float(rec.total or 0.0)

        for c in (1, 2):
            ws.cell(r, c).number_format = "dd/mm/yyyy"
        for c in (12, 14, 16):
            ws.cell(r, c).number_format = "#,##0.00"
        ws.cell(r, 13).number_format = "0.00"

    if "CONTROL_LISTADO" in wb.sheetnames:
        del wb["CONTROL_LISTADO"]
    wsl = wb.create_sheet("CONTROL_LISTADO")
    wsl.append(
        [
            "#",
            "FECHA",
            "Nº FACTURA",
            "TERCERO",
            "NIF",
            "TIPO",
            "BASE",
            "IVA",
            "IRPF/RET",
            "TOTAL",
            "IVA_TIPO",
            "OCR_METODO",
            "OCR_ERROR",
            "ORIGEN",
        ]
    )
    for i, rec in enumerate(records_sorted, start=1):
        wsl.append(
            [
                i,
                rec.fecha,
                rec.numero,
                rec.tercero,
                rec.nif,
                rec.tipo,
                rec.base,
                rec.iva,
                rec.irpf,
                rec.total,
                rec.iva_bucket,
                rec.ocr_method,
                rec.ocr_error,
                os.path.basename(rec.source_file),
            ]
        )
    for r in range(2, wsl.max_row + 1):
        wsl.cell(r, 2).number_format = "dd/mm/yyyy"
        for c in (7, 8, 9, 10):
            wsl.cell(r, c).number_format = "#,##0.00"

    if "CONTROL_TOTALIZADOR" in wb.sheetnames:
        del wb["CONTROL_TOTALIZADOR"]
    wst = wb.create_sheet("CONTROL_TOTALIZADOR")
    wst["A1"].value = "RESUMEN"
    wst["B1"].value = title
    wst["A3"].value = "Nº FACTURAS"
    wst["B3"].value = len(records_sorted)
    base_sum = sum(float(r.base or 0.0) for r in records_sorted)
    iva_sum = sum(float(r.iva or 0.0) for r in records_sorted)
    ret_sum = sum(float(r.irpf or 0.0) for r in records_sorted)
    total_sum = sum(float(r.total or 0.0) for r in records_sorted)
    wst["A4"].value = "BASE IMPONIBLE (SUMA)"
    wst["B4"].value = base_sum
    wst["A5"].value = "IVA (SUMA)"
    wst["B5"].value = iva_sum
    wst["A6"].value = "RETENCIONES (SUMA)"
    wst["B6"].value = ret_sum
    wst["A7"].value = "TOTAL (SUMA)"
    wst["B7"].value = total_sum
    iva_rep = sum(float(r.iva or 0.0) for r in records_sorted if r.iva_bucket == "repercutido")
    iva_sop = sum(float(r.iva or 0.0) for r in records_sorted if r.iva_bucket == "soportado")
    wst["A9"].value = "IVA REPERCUTIDO (SUMA)"
    wst["B9"].value = iva_rep
    wst["A10"].value = "IVA SOPORTADO (SUMA)"
    wst["B10"].value = iva_sop
    for cell in ("B4", "B5", "B6", "B7", "B9", "B10"):
        wst[cell].number_format = "#,##0.00"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", required=True, help="Carpeta con facturas/tickets (pdf/jpg/png).")
    ap.add_argument("--template", required=True, help="Plantilla MiConversor (empresas).xlsx")
    ap.add_argument("--out", default="reports/miconversor_autonomos_ocr.xlsx", help="Ruta de salida (XLSX) para modo single.")
    ap.add_argument("--out-dir", default="", help="Carpeta de salida para modo batch (un XLSX por subcarpeta).")
    ap.add_argument("--pdf-pages", type=int, default=2, help="Máx páginas PDF a leer/OCR.")
    ap.add_argument("--workers", type=int, default=1, help="Paralelismo (1-8) al procesar documentos.")
    ap.add_argument("--batch-top-level", action="store_true", help="Genera un XLSX por cada subcarpeta inmediata dentro de --root.")
    ap.add_argument("--summary", default="", help="Ruta JSON de resumen (opcional).")
    ap.add_argument("--max-folders", type=int, default=0, help="Límite de carpetas en modo batch (0 = sin límite).")
    ap.add_argument("--verbose", action="store_true", help="Imprime progreso (útil para lotes grandes).")
    args = ap.parse_args(argv)

    root = Path(args.root).expanduser().resolve()
    template = Path(args.template).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve() if str(args.out_dir or "").strip() else None
    summary_path = Path(args.summary).expanduser().resolve() if str(args.summary or "").strip() else None

    if not root.exists():
        raise SystemExit(f"No existe root: {root}")
    if not template.exists():
        raise SystemExit(f"No existe template: {template}")

    pdf_pages = max(1, int(args.pdf_pages))
    workers = max(1, min(8, int(args.workers or 1)))
    verbose = bool(args.verbose)

    if not args.batch_top_level:
        records = _scan_folder_docs(root, pdf_pages=pdf_pages, workers=workers, verbose=verbose)
        if not records:
            raise SystemExit("No he encontrado PDFs/JPG/PNG en la ruta indicada.")
        title = root.name
        _write_output(template, out_path, records, title=title)
        if summary_path:
            payload = {
                "mode": "single",
                "root": str(root),
                "docs": len(records),
                "with_total": sum(1 for r in records if (r.total or 0.0) > 0),
                "zero_total": sum(1 for r in records if (r.total or 0.0) <= 0),
            }
            summary_path.parent.mkdir(parents=True, exist_ok=True)
            summary_path.write_text(__import__("json").dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(str(out_path))
        return 0

    if out_dir is None:
        out_dir = ROOT / "reports" / "miconversor_autonomos_batch"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Batch por subcarpeta inmediata
    subfolders = [p for p in sorted(root.iterdir()) if p.is_dir() and not p.name.startswith(".")]
    if args.max_folders and args.max_folders > 0:
        subfolders = subfolders[: int(args.max_folders)]

    summary = {
        "mode": "batch_top_level",
        "root": str(root),
        "out_dir": str(out_dir),
        "folders": [],
        "totals": {"folders": 0, "docs": 0, "with_total": 0, "zero_total": 0},
    }

    for folder in subfolders:
        records = _scan_folder_docs(folder, pdf_pages=pdf_pages, workers=workers, verbose=verbose)
        if not records:
            continue
        if verbose:
            print(f"[DONE] {folder.name} docs={len(records)}")
        safe = _sanitize_filename(folder.name)
        out_file = out_dir / f"{safe}_miconversor.xlsx"
        _write_output(template, out_file, records, title=folder.name)
        folder_summary = {
            "folder": folder.name,
            "docs": len(records),
            "with_total": sum(1 for r in records if (r.total or 0.0) > 0),
            "zero_total": sum(1 for r in records if (r.total or 0.0) <= 0),
            "out": str(out_file),
        }
        summary["folders"].append(folder_summary)
        summary["totals"]["folders"] += 1
        summary["totals"]["docs"] += int(folder_summary["docs"])
        summary["totals"]["with_total"] += int(folder_summary["with_total"])
        summary["totals"]["zero_total"] += int(folder_summary["zero_total"])
        print(str(out_file))

    if summary_path:
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.write_text(__import__("json").dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
