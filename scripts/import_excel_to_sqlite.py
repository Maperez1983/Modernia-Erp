#!/usr/bin/env python3
import argparse
import datetime as dt
import re
import sqlite3
import unicodedata
import uuid
from pathlib import Path

from openpyxl import load_workbook

COMPANIES = [
    "Inversure",
    "Financiaciones Modernia",
    "Estudio Velazquez 2012 SL",
    "Fincas Velazquez",
    "Grupo Modernia",
]

SL_COMPANY_MAP = {
    "estudio velazquez": "Estudio Velazquez 2012 SL",
    "financiaciones": "Financiaciones Modernia",
    "fincas velazquez": "Fincas Velazquez",
    "grupo modernia": "Grupo Modernia",
    "inversure": "Inversure",
}

SHEETS = {
    "BDT": {
        "table": "movimientos",
        "map": {
            "CONCEPTO": "concepto",
            "PISOS VENDIDOS": "pisos_vendidos",
            "COMISION": "comision",
            "ASESOR": "asesor",
            "ANO": "anio",
            "SL": "sl",
            "MES": "mes",
            "TIPO": "tipo",
        },
        "columns": [
            "empresa_id",
            "concepto",
            "pisos_vendidos",
            "comision",
            "asesor",
            "anio",
            "mes",
            "sl",
            "tipo",
            "created_at",
            "updated_at",
        ],
        "required": ["concepto"],
        "empresa_resolver": "sl",
    },
    "BDT SEGUROS": {
        "table": "seguros",
        "map": {
            "MES DE CREACION": "mes_creacion",
            "FECHA EFECTO": "fecha_efecto",
            "NOMBRE DEL TOMADOR": "tomador",
            "COMPANIA": "compania",
            "RAMO": "ramo",
            "NUMERO DE POLIZA": "poliza_numero",
            "PRIMA NETA": "prima_neta",
            "PRIMA TOTAL": "prima_total",
            "COMISION": "comision",
            "PRODUCCION": "produccion",
            "COLABORADOR": "colaborador",
            "ESTADO": "estado",
        },
        "columns": [
            "empresa_id",
            "mes_creacion",
            "fecha_efecto",
            "tomador",
            "compania",
            "ramo",
            "poliza_numero",
            "prima_neta",
            "prima_total",
            "comision",
            "produccion",
            "colaborador",
            "estado",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Fincas Velazquez",
    },
    "BDT CLIENTE GESTORÍA": {
        "table": "gestoria",
        "map": {
            "CLIENTE": "cliente",
            "FECHA": "fecha",
            "CUOTA": "cuota",
            "PRECIO": "precio",
            "TIPO": "tipo",
            "PERFIL": "perfil",
            "ESTADO": "estado",
        },
        "columns": [
            "empresa_id",
            "cliente",
            "fecha",
            "cuota",
            "precio",
            "tipo",
            "perfil",
            "estado",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Fincas Velazquez",
    },
    "BDT HIPOTECA": {
        "table": "hipotecas",
        "map": {
            "CLIENTE": "cliente",
            "BANCO": "banco",
            "PRECIO": "precio",
            "IMPORTE HIPOTECA": "importe_hipoteca",
            "PORCENTAJE": "porcentaje",
            "ENTRADA": "entrada",
            "COMISION": "comision",
            "OFICINA": "oficina",
            "FECHA ENCARGO": "fecha_encargo",
            "ENCARGO": "encargo",
            "TIPO HIPOTECA": "tipo_hipoteca",
            "FECHA FIRMA": "fecha_firma",
            "CESION": "cesion",
            "COMISION JUAN": "comision_juan",
            "COMISION MODERNIA": "comision_modernia",
            "INMOBILIARIA COMPRA": "inmobiliaria_compra",
            "ASESOR": "asesor",
            "ESTADO": "estado",
            "ANO": "anio",
        },
        "columns": [
            "empresa_id",
            "cliente",
            "banco",
            "precio",
            "importe_hipoteca",
            "porcentaje",
            "entrada",
            "comision",
            "oficina",
            "fecha_encargo",
            "encargo",
            "tipo_hipoteca",
            "fecha_firma",
            "cesion",
            "comision_juan",
            "comision_modernia",
            "inmobiliaria_compra",
            "asesor",
            "estado",
            "anio",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Financiaciones Modernia",
    },
    "ALQUILERES": {
        "table": "alquileres",
        "map": {
            "FECHA": "fecha",
            "DIRECCION": "direccion",
            "PROPIETARIO": "propietario",
            "TELEFONO": "telefono",
            "PRECIO": "precio",
            "SEGURO": "seguro",
            "HACIENDA": "hacienda",
            "COMISION": "comision",
            "IMPORTE COMISION": "importe_comision",
            "TOTAL": "total",
            "INQUILINO": "inquilino",
            "TELEFONO2": "telefono2",
            "AGENTE": "agente",
            "N ALQUILERES": "numero_alquileres",
            "TIPO": "tipo",
            "OFICINA": "oficina",
        },
        "columns": [
            "empresa_id",
            "fecha",
            "direccion",
            "propietario",
            "telefono",
            "precio",
            "seguro",
            "hacienda",
            "comision",
            "importe_comision",
            "total",
            "inquilino",
            "telefono2",
            "agente",
            "numero_alquileres",
            "tipo",
            "oficina",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Estudio Velazquez 2012 SL",
    },
    "INVERSORES": {
        "table": "inversores",
        "map": {
            "NOMBRE Y APELLIDOS": "nombre",
            "APORTACION": "aportacion",
            "FECHA": "fecha",
            "PROYECTO": "proyecto",
        },
        "columns": [
            "empresa_id",
            "nombre",
            "aportacion",
            "fecha",
            "proyecto",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Inversure",
    },
    "BDT INVERSURE": {
        "table": "inversure_operaciones",
        "map": {
            "PROYECTO": "proyecto",
            "PRECIO": "precio",
            "CONCEPTO": "concepto",
            "TIPO": "tipo",
            "SUJETO": "sujeto",
            "FECHA": "fecha",
        },
        "columns": [
            "empresa_id",
            "proyecto",
            "precio",
            "concepto",
            "tipo",
            "sujeto",
            "fecha",
            "created_at",
            "updated_at",
        ],
        "empresa_fija": "Inversure",
    },
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_header_row(ws, max_rows=20, max_cols=80):
    for r in range(1, max_rows + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)]
        if sum(1 for v in values if v not in (None, "")) >= 3:
            return r, values
    return None, []


def to_iso_date(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return value


def to_scalar(value):
    if value is None or value == "":
        return None
    value = to_iso_date(value)
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def now_iso():
    return dt.datetime.utcnow().replace(microsecond=0).isoformat()


def load_schema(conn, schema_path):
    sql = Path(schema_path).read_text(encoding="utf-8")
    conn.executescript(sql)


def ensure_companies(conn):
    now = now_iso()
    ids = {}
    for name in COMPANIES:
        company_id = str(uuid.uuid4())
        conn.execute(
            "INSERT OR IGNORE INTO empresas (id, nombre, activo, created_at, updated_at) "
            "VALUES (?, ?, 1, ?, ?)",
            (company_id, name, now, now),
        )
        row = conn.execute(
            "SELECT id FROM empresas WHERE nombre = ?",
            (name,),
        ).fetchone()
        ids[name] = row[0]
    return ids


def normalize_company_key(value):
    if not value:
        return ""
    return str(value).strip().lower()


def import_sheet(conn, ws, config, empresa_id, companies, batch_size=500):
    header_row, header_values = detect_header_row(ws)
    if not header_row:
        return 0

    index_by_header = {}
    for idx, value in enumerate(header_values, start=1):
        key = normalize_header(value)
        if key and key not in index_by_header:
            index_by_header[key] = idx

    mapping = config["map"]
    columns = config["columns"]
    table = config["table"]
    now = now_iso()
    inserted = 0

    placeholders = ", ".join(["?"] * (len(columns) + 1))
    column_sql = ", ".join(["id"] + columns)
    insert_sql = f"INSERT INTO {table} ({column_sql}) VALUES ({placeholders})"

    resolver = config.get("empresa_resolver")
    empresa_fija = config.get("empresa_fija")
    required = config.get("required", [])
    rows = []
    max_col = max(index_by_header.values()) if index_by_header else 0
    iter_rows = ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        max_col=max_col,
        values_only=True,
    )
    for row in iter_rows:
        row_values = []
        empty = True
        for header_key in mapping.keys():
            idx = index_by_header.get(header_key)
            val = row[idx - 1] if idx else None
            val = to_scalar(val)
            if val not in (None, ""):
                empty = False
            row_values.append(val)
        if empty:
            continue

        data = {mapping[k]: v for k, v in zip(mapping.keys(), row_values)}
        if required and any(not data.get(field) for field in required):
            continue
        resolved_empresa_id = empresa_id
        if not resolved_empresa_id and empresa_fija:
            resolved_empresa_id = companies.get(empresa_fija)
        if not resolved_empresa_id and resolver == "sl":
            sl_value = data.get("sl")
            company_name = SL_COMPANY_MAP.get(normalize_company_key(sl_value))
            if company_name:
                resolved_empresa_id = companies.get(company_name)

        row = []
        for col in columns:
            if col == "empresa_id":
                row.append(resolved_empresa_id)
            elif col in ("created_at", "updated_at"):
                row.append(now)
            else:
                row.append(data.get(col))

        rows.append((str(uuid.uuid4()), *row))
        if len(rows) >= batch_size:
            conn.executemany(insert_sql, rows)
            inserted += len(rows)
            rows.clear()

    if rows:
        conn.executemany(insert_sql, rows)
        inserted += len(rows)
        rows.clear()

    return inserted


def main():
    parser = argparse.ArgumentParser(
        description="Import EMPRESAS MIGUEL.xlsm into a SQLite database."
    )
    parser.add_argument("--excel", required=True, help="Path to the Excel file.")
    parser.add_argument("--db", default="data/erp.sqlite", help="SQLite output path.")
    parser.add_argument(
        "--empresa",
        help="Nombre de empresa para asignar todos los registros (opcional).",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    db_path = Path(args.db)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_path.as_posix())
    conn.execute("PRAGMA journal_mode=OFF;")
    conn.execute("PRAGMA synchronous=OFF;")
    load_schema(conn, Path(__file__).resolve().parent.parent / "schema.sql")
    companies = ensure_companies(conn)

    empresa_id = None
    if args.empresa:
        empresa_id = companies.get(args.empresa)
        if not empresa_id:
            raise SystemExit(
                f"Empresa desconocida: {args.empresa}. Opciones: {', '.join(COMPANIES)}"
            )

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    total = 0
    for sheet_name, config in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        inserted = import_sheet(conn, ws, config, empresa_id, companies)
        total += inserted
        print(f"{sheet_name}: {inserted} filas")

    conn.commit()
    conn.close()
    print(f"Importacion completada. Total filas: {total}. DB: {db_path}")


if __name__ == "__main__":
    main()
