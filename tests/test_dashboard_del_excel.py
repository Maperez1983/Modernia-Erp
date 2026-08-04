"""El Excel del listado sale con una hoja de dashboard.

El fichero traía el detalle y unas cifras sueltas. Para ver con qué banco se firma
más, qué inmobiliaria trae más operaciones o cómo se reparte la comisión había que
montarse una tabla dinámica a mano cada vez.

Dos decisiones que no son de adorno:

- **Se agrupa por el nombre normalizado, no por el literal.** En producción
  convivían "Malaga Norte", "MALAGA NORTE" y "MÁLAGA NORTE": la misma oficina
  partida en tres barras, cada una con un trozo del volumen. Agrupadas son 4
  operaciones y 533.600 €. Se enseña la grafía más usada, que es la que el equipo
  reconoce.
- **Los gráficos son nativos de Excel**, no imágenes: se recalculan si alguien
  edita las tablas de apoyo, que quedan a la vista debajo de cada uno.

Y una salvaguarda: si el dashboard falla por lo que sea, el Excel se entrega igual
con su detalle. Quien pidió un listado no se queda sin listado porque un gráfico
no cuadre.
"""

import unittest
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
SERVER = (RAIZ / "web" / "server.py").read_text(encoding="utf-8")

try:
    import sys

    sys.path.insert(0, str(RAIZ))
    from web import server
    from openpyxl import Workbook

    LISTO = True
except Exception:  # pragma: no cover
    LISTO = False


def bloque():
    i = SERVER.index("def add_hipotecas_dashboard_sheet")
    return SERVER[i: SERVER.index("\ndef ", i + 10)]


class LaHojaSeAnadeSinRomperElExcelTests(unittest.TestCase):
    def test_existe(self):
        self.assertIn("def add_hipotecas_dashboard_sheet(", SERVER)

    def test_se_coloca_la_primera(self):
        """Al abrir el fichero se ve el resumen, no una tabla de 24 filas."""
        self.assertIn('wb.create_sheet("Dashboard", 0)', bloque())

    def test_si_falla_el_dashboard_el_listado_se_entrega_igual(self):
        i = SERVER.index("def build_hipotecas_listado_excel_workbook")
        cuerpo = SERVER[i: SERVER.index("\ndef ", i + 10)]
        self.assertIn("add_hipotecas_dashboard_sheet(wb, normalized_items", cuerpo)
        self.assertIn("except Exception:", cuerpo)


@unittest.skipUnless(LISTO, "hace falta openpyxl y poder importar web.server")
class LoQuePinta(unittest.TestCase):
    def setUp(self):
        self.items = [
            {"banco": "Banco Santander", "inmobiliaria": "MALAGA NORTE", "importe_hipoteca": 100000,
             "honorarios": 3000, "comision_juan": 1000, "comision_modernia": 2000, "fecha_firma": "2026-01-15"},
            {"banco": "Banco Santander", "inmobiliaria": "Malaga Norte", "importe_hipoteca": 200000,
             "honorarios": 4000, "comision_juan": 1500, "comision_modernia": 2500, "fecha_firma": "2026-02-20"},
            {"banco": "BBVA", "inmobiliaria": "MÁLAGA NORTE", "importe_hipoteca": 50000,
             "honorarios": 1000, "comision_juan": 0, "comision_modernia": 1000, "fecha_firma": "2026-02-25"},
        ]
        wb = Workbook()
        wb.active.title = "otra"
        self.hoja = server.add_hipotecas_dashboard_sheet(wb, self.items, "2026", "Financiaciones Modernia")
        self.valores = [
            [c for c in fila] for fila in self.hoja.iter_rows(max_col=4, values_only=True)
        ]

    def test_hay_los_cuatro_graficos(self):
        """Bancos, inmobiliarias, comisión por mes y reparto."""
        self.assertEqual(len(self.hoja._charts), 4)

    def test_las_tres_grafias_de_la_misma_oficina_son_una_sola_fila(self):
        # Se compara sin tildes: al empatar en frecuencia gana "MÁLAGA NORTE", que es
        # la grafía correcta en castellano. Lo que importa es que salga una sola fila.
        import unicodedata

        def sin_tildes(valor):
            texto = unicodedata.normalize("NFKD", str(valor or ""))
            return "".join(c for c in texto if not unicodedata.combining(c)).upper()

        filas = [f for f in self.valores if f[0] and "MALAGA NORTE" in sin_tildes(f[0])]
        self.assertEqual(len(filas), 1, f"la oficina sale partida: {filas}")
        self.assertEqual(filas[0][1], 3)
        self.assertEqual(filas[0][2], 350000)

    def test_los_kpis_cuadran(self):
        fila = self.valores[3]  # la de las cifras
        self.assertEqual(fila[0], 3)          # operaciones
        self.assertEqual(fila[2], 350000)     # volumen

    def test_la_comision_por_mes_suma_el_total(self):
        i = next(n for n, f in enumerate(self.valores) if f[0] == "Comisión cobrada por mes")
        meses = []
        for fila in self.valores[i + 2:]:
            if not fila[0] or not str(fila[0]).startswith("2026-"):
                break
            meses.append(fila[1])
        self.assertEqual(sum(meses), 8000)

    def test_el_reparto_no_pierde_dinero(self):
        i = next(n for n, f in enumerate(self.valores) if f[0] == "Reparto de la comisión")
        reparto = [f[1] for f in self.valores[i + 2:] if f[0] and f[1] is not None]
        self.assertEqual(sum(reparto), 8000)

    def test_un_importe_decimal_no_lo_revienta(self):
        """Las columnas de dinero son NUMERIC: Postgres devuelve Decimal."""
        from decimal import Decimal

        items = [dict(self.items[0], importe_hipoteca=Decimal("100000.55"), honorarios=Decimal("3000.45"))]
        wb = Workbook()
        wb.active.title = "otra"
        hoja = server.add_hipotecas_dashboard_sheet(wb, items, "2026", "X")
        self.assertIsNotNone(hoja)


if __name__ == "__main__":
    unittest.main()
