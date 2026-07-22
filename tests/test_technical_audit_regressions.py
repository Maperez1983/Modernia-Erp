import os
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from PIL import Image, ImageDraw

from web import pdf_utils
from web import ocr_service
from web import public_links
from web import security_utils
from web import server


class TechnicalAuditRegressionTests(unittest.TestCase):
    def _make_workspace_scope_conn(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_companies (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              legacy_empresa_id TEXT,
              nombre TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE workspace_empresas (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL
            );
            CREATE TABLE workspace_miembros (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              usuario_id TEXT NOT NULL,
              rol TEXT NOT NULL DEFAULT 'Miembro'
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              rol TEXT,
              servicio TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            """
        )
        conn.execute(
            """
            INSERT INTO workspace_companies (id, workspace_id, legacy_empresa_id, nombre, activo)
            VALUES ('wc-1', 'ws-1', 'emp-1', 'Empresa Uno', 1)
            """
        )
        conn.execute(
            """
            INSERT INTO workspace_miembros (id, workspace_id, usuario_id, rol)
            VALUES ('wm-1', 'ws-1', 'u-1', 'Miembro')
            """
        )
        conn.execute(
            """
            INSERT INTO usuarios (id, rol, servicio, activo)
            VALUES ('u-1', 'Miembro', 'Gestoría', 1)
            """
        )
        conn.commit()
        return conn

    def _make_workspace_rrhh_conn(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        server.ensure_usuarios_schema(conn)
        server.ensure_workspace_core_tables(conn)
        server.ensure_workspace_product_tables(conn)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS empresas (
              id TEXT PRIMARY KEY,
              nombre TEXT NOT NULL,
              activo INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        conn.commit()
        return conn

    def _make_auth_user_conn(self, *, user_id="u-auth-1", usuario="Mperez", email="mperez@example.com", password_hash=None):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        server.ensure_usuarios_schema(conn)
        server.ensure_auth_invites_table(conn)
        conn.execute(
            """
            INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, password_hash, activo, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            (
                user_id,
                "Miguel",
                "Perez",
                usuario,
                email,
                "Gestoría",
                "Miembro",
                password_hash,
                1,
            ),
        )
        conn.commit()
        return conn

    def test_resolve_external_ocr_config_prefers_explicit_env_path(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            explicit_path = tmp_path / "explicit-creds.json"
            standard_path = tmp_path / "standard-creds.json"
            explicit_path.write_text("{}", encoding="utf-8")
            standard_path.write_text("{}", encoding="utf-8")

            with mock.patch.dict(
                os.environ,
                {
                    "OCR_GOOGLE_APPLICATION_CREDENTIALS": str(explicit_path),
                    "GOOGLE_APPLICATION_CREDENTIALS": str(standard_path),
                },
                clear=True,
            ):
                expected = (str(explicit_path), "")
                self.assertEqual(server._resolve_external_ocr_config(), expected)
                self.assertEqual(ocr_service._resolve_external_ocr_config(), expected)

            with mock.patch.dict(
                os.environ,
                {
                    "GOOGLE_APPLICATION_CREDENTIALS": str(standard_path),
                },
                clear=True,
            ):
                expected = (str(standard_path), "")
                self.assertEqual(server._resolve_external_ocr_config(), expected)
                self.assertEqual(ocr_service._resolve_external_ocr_config(), expected)

    def test_resolve_external_ocr_config_rejects_missing_directory_and_non_json_paths(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            missing_path = tmp_path / "missing-creds.json"
            directory_path = tmp_path / "creds-dir"
            non_json_path = tmp_path / "creds.txt"
            directory_path.mkdir()
            non_json_path.write_text("{}", encoding="utf-8")

            with mock.patch.dict(
                os.environ,
                {"OCR_GOOGLE_APPLICATION_CREDENTIALS": str(missing_path)},
                clear=True,
            ):
                self.assertEqual(server._resolve_external_ocr_config(), ("", ""))
                self.assertEqual(ocr_service._resolve_external_ocr_config(), ("", ""))

            with mock.patch.dict(
                os.environ,
                {"OCR_GOOGLE_APPLICATION_CREDENTIALS": str(directory_path)},
                clear=True,
            ):
                self.assertEqual(server._resolve_external_ocr_config(), ("", ""))
                self.assertEqual(ocr_service._resolve_external_ocr_config(), ("", ""))

            with mock.patch.dict(
                os.environ,
                {"OCR_GOOGLE_APPLICATION_CREDENTIALS": str(non_json_path)},
                clear=True,
            ):
                self.assertEqual(server._resolve_external_ocr_config(), ("", ""))
                self.assertEqual(ocr_service._resolve_external_ocr_config(), ("", ""))

    def test_resolve_external_ocr_config_returns_empty_without_variables(self):
        with mock.patch.dict(os.environ, {}, clear=True):
            self.assertEqual(server._resolve_external_ocr_config(), ("", ""))
            self.assertEqual(ocr_service._resolve_external_ocr_config(), ("", ""))

    def test_resolve_external_ocr_config_does_not_autodiscover_vision_sa_json(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            web_root = tmp_path / "web"
            web_root.mkdir()
            (tmp_path / "vision-sa.json").write_text("{}", encoding="utf-8")

            with mock.patch.object(server, "ROOT", web_root):
                cwd_error = AssertionError("cwd lookup is not allowed")
                with mock.patch.object(server.Path, "cwd", side_effect=cwd_error):
                    with mock.patch.object(ocr_service.Path, "cwd", side_effect=cwd_error):
                        with mock.patch.dict(os.environ, {}, clear=True):
                            self.assertEqual(server._resolve_external_ocr_config(), ("", ""))
                            self.assertEqual(ocr_service._resolve_external_ocr_config(), ("", ""))

    def test_login_recovery_payload_includes_recovery_fields(self):
        payload = server.build_login_recovery_payload("  Mperez  ")
        self.assertEqual(
            payload,
            {
                "recovery_available": True,
                "recovery_login": "Mperez",
                "recovery_message": "Si la cuenta existe, te enviaremos un enlace para recuperar el acceso.",
            },
        )
        self.assertEqual(server.build_login_recovery_payload(""), {})

    def test_recovery_invite_roundtrip_allows_password_reset_with_existing_hash(self):
        with mock.patch.object(server.os, "urandom", return_value=b"\x01" * 16):
            initial_hash = server.hash_password("InitialPwd123")
        conn = self._make_auth_user_conn(password_hash=initial_hash)
        try:
            result = server.issue_access_recovery_invite(conn, "Mperez")
            self.assertEqual(result["user_id"], "u-auth-1")
            self.assertEqual(result["usuario"], "Mperez")
            self.assertEqual(result["email"], "mperez@example.com")
            self.assertTrue(result["token"])

            stored = conn.execute(
                """
                SELECT invite_token, invite_expires_at, invite_sent_at, password_hash
                FROM usuarios
                WHERE id = ?
                """,
                ("u-auth-1",),
            ).fetchone()
            self.assertEqual(stored["invite_token"], result["token"])
            self.assertTrue(str(stored["invite_expires_at"] or "").strip())
            self.assertTrue(str(stored["invite_sent_at"] or "").strip())
            self.assertEqual(stored["password_hash"], initial_hash)

            status_payload, status_code = server.build_auth_invite_status_response(conn, result["token"])
            self.assertEqual(status_code, 200)
            self.assertTrue(status_payload["ok"])
            self.assertTrue(status_payload["valid"])
            self.assertEqual(status_payload["mode"], "recovery")
            self.assertFalse(status_payload["activated"])
            self.assertEqual(status_payload["user"]["usuario"], "Mperez")
            self.assertEqual(status_payload["user"]["email"], "mperez@example.com")

            apply_payload, apply_status = server.apply_auth_invite_password(conn, result["token"], "NuevaClave123")
            self.assertEqual(apply_status, 200)
            self.assertEqual(apply_payload, {"ok": True})

            updated = conn.execute(
                """
                SELECT password_hash, invite_token, invite_expires_at, invite_sent_at
                FROM usuarios
                WHERE id = ?
                """,
                ("u-auth-1",),
            ).fetchone()
            self.assertTrue(str(updated["password_hash"] or "").strip())
            self.assertTrue(server.verify_password("NuevaClave123", updated["password_hash"]))
            self.assertFalse(str(updated["invite_token"] or "").strip())
            self.assertFalse(str(updated["invite_expires_at"] or "").strip())
            self.assertFalse(str(updated["invite_sent_at"] or "").strip())

            invite_row = conn.execute(
                "SELECT used_at, notes FROM auth_invites WHERE token = ?",
                (result["token"],),
            ).fetchone()
            self.assertTrue(str(invite_row["used_at"] or "").strip())
            self.assertEqual(invite_row["notes"], "access_recovery")
        finally:
            conn.close()

    def test_activation_invite_with_existing_password_remains_rejected(self):
        with mock.patch.object(server.os, "urandom", return_value=b"\x02" * 16):
            initial_hash = server.hash_password("InitialPwd123")
        conn = self._make_auth_user_conn(user_id="u-auth-2", usuario="Activacion", email="activacion@example.com", password_hash=initial_hash)
        try:
            invite = server._issue_auth_invite(
                conn,
                "Activacion",
                notes="usuarios_invitar_v2",
                clear_password=False,
                activate_user=False,
            )
            status_payload, status_code = server.build_auth_invite_status_response(conn, invite["token"])
            self.assertEqual(status_code, 200)
            self.assertFalse(status_payload["valid"])
            self.assertEqual(status_payload["mode"], "activation")
            self.assertTrue(status_payload["activated"])

            apply_payload, apply_status = server.apply_auth_invite_password(conn, invite["token"], "OtraClave123")
            self.assertEqual(apply_status, 409)
            self.assertEqual(apply_payload["error"], "La cuenta ya está activada")
        finally:
            conn.close()

    def test_external_ocr_functions_match_new_module(self):
        fake_response = mock.MagicMock()
        fake_response.read.return_value = (
            b'{"responses":[{"fullTextAnnotation":{"text":"OCR OK"}}]}'
        )
        fake_response.__enter__.return_value = fake_response
        fake_response.__exit__.return_value = None

        helper = mock.Mock(return_value=("", "vision-key"))

        with mock.patch.object(server, "_resolve_external_ocr_config", helper):
            with mock.patch.object(ocr_service.urllib.request, "urlopen", return_value=fake_response) as urlopen_mock:
                self.assertEqual(
                    server.external_ocr_available(),
                    ocr_service.external_ocr_available(resolver=server._resolve_external_ocr_config),
                )
                self.assertEqual(
                    server.ocr_image_external(b"image-bytes"),
                    ocr_service.ocr_image_external(b"image-bytes", resolver=server._resolve_external_ocr_config),
                )

        self.assertEqual(helper.call_count, 4)
        self.assertEqual(urlopen_mock.call_count, 2)

    def test_docai_ocr_matches_new_module_without_credentials(self):
        with mock.patch.dict(os.environ, {}, clear=True):
            self.assertEqual(
                server.ocr_image_docai(b"image-bytes", "image/png"),
                ocr_service.ocr_image_docai(b"image-bytes", "image/png"),
            )
            self.assertEqual(server.docai_available(), ocr_service.docai_available())

    def test_docai_mapping_helpers_match_new_module(self):
        doc_fields = {
            "nombre y apellidos 1": "Ana López",
            "nombre y apellidos 2": "Luis Pérez",
            "dni 1": "12345678A",
            "dni 2": "87654321B",
            "telefono 1": "600111222",
            "telefono 2": "600333444",
            "correo electronico 1": "ana@example.com",
            "correo electronico 2": "luis@example.com",
            "fecha nacimiento 1": "1990-01-02",
            "fecha nacimiento 2": "1992-03-04",
            "estado civil 1": "Casada",
            "estado civil 2": "Soltero",
            "hijos 1": "1",
            "hijos 2": "0",
            "profesion 1": "Asesora",
            "profesion 2": "Consultor",
            "tipo contrato 1": "Indefinido",
            "tipo contrato 2": "Temporal",
            "ingresos nomina 1": "1200",
            "ingresos nomina 2": "1500",
            "patrimonio alquiler 1": "0",
            "patrimonio alquiler 2": "1",
            "prestamos 1": "0",
            "prestamos 2": "1",
        }
        poliza_fields = {
            "tomador": "Ana López",
            "dni": "12345678A",
            "telefono": "600111222",
            "correo electronico": "ana@example.com",
            "direccion": "Calle Falsa 123",
            "compania": "Aseguradora Demo",
            "ramo": "Hogar",
            "poliza": "POL123456",
            "fecha efecto": "2026-01-01",
            "fecha vencimiento": "2027-01-01",
            "prima neta": "120,00",
            "prima total": "145,20",
        }

        self.assertEqual(server.map_docai_fields(doc_fields), ocr_service.map_docai_fields(doc_fields))
        self.assertEqual(
            server.map_docai_poliza_fields(poliza_fields),
            ocr_service.map_docai_poliza_fields(poliza_fields),
        )

    def test_resolve_gestoria_factura_cliente_id_scopes_nif_and_name_by_empresa(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Factura", "99999999Z", "2026-03-01", "2026-03-01"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-right", "e1", "Cliente Factura", "12345678A", "2026-03-02", "2026-03-02"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, estado, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("link-right", "c-right", "e1", "Gestoria", "Activo", "2026-03-02", "2026-03-02"),
        )

        try:
            by_nif = server.resolve_gestoria_factura_cliente_id(conn, "e1", "12345678A", "")
            by_name = server.resolve_gestoria_factura_cliente_id(conn, "e1", "", "Cliente Factura")
        finally:
            conn.close()

        self.assertEqual(by_nif, "c-right")
        self.assertEqual(by_name, "c-right")

    def test_resolve_gestoria_factura_cliente_id_requires_empresa_scope(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-global", "", "Cliente Global", "12345678A", "2026-03-01", "2026-03-01"),
        )

        try:
            result = server.resolve_gestoria_factura_cliente_id(conn, "", "12345678A", "Cliente Global")
        finally:
            conn.close()

        self.assertIsNone(result)

    def test_compute_workspace_rrhh_productividad_facturacion_anual_handles_empty_year_filter(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-gestoria", "Cliente Gestoria", "11111111A", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-fincas", "Cliente Fincas", "22222222B", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-gestoria", "c-gestoria", "emp-1", "gestoria", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-fincas", "c-fincas", "emp-1", "administración fincas", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-1", "ws-1", "emp-1", "c-gestoria", "gestoria", "2026-01-15", "", 100.0, 1),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-2", "ws-1", "emp-1", "c-fincas", "administración fincas", "2026-02-20", "", 200.0, 0),
        )

        try:
            result = server.compute_workspace_rrhh_productividad_facturacion_anual(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                servicio_keys={"gestoria", "administración fincas"},
                ejercicio="",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 2)
        self.assertEqual(result["kpis"]["facturado_total"], 300.0)
        self.assertEqual(result["kpis"]["comision_total"], 30.0)
        self.assertEqual(len(result["items"]), 2)

    def test_compute_workspace_rrhh_economicos_dashboard_aggregates_gestoria_and_fincas(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            CREATE TABLE cliente_gestoria (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              mod_renta INTEGER,
              renta_detalles TEXT
            );
            CREATE TABLE workspace_rrhh_documentos (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              persona_id TEXT NOT NULL,
              tipo TEXT,
              nombre TEXT,
              fecha_emision TEXT,
              created_at TEXT,
              doc_key TEXT,
              doc_url TEXT,
              nomina_ocr_status TEXT,
              nomina_ocr_confidence REAL,
              nomina_ocr_error TEXT,
              nomina_ocr_json TEXT,
              nomina_ocr_updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-gestoria", "Cliente Gestoria", "11111111A", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-fincas", "Cliente Fincas", "22222222B", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-gestoria", "c-gestoria", "emp-1", "gestoria", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-fincas", "c-fincas", "emp-1", "administración fincas", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-1", "ws-1", "emp-1", "c-gestoria", "gestoria", "2026-01-15", "", 100.0, 1),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-2", "ws-1", "emp-1", "c-fincas", "administración fincas", "2026-02-20", "", 200.0, 0),
        )
        conn.execute(
            "INSERT INTO cliente_gestoria (id, cliente_id, mod_renta, renta_detalles) VALUES (?, ?, ?, ?)",
            ("cg-1", "c-gestoria", 0, "[]"),
        )

        try:
            result = server.compute_workspace_rrhh_economicos_dashboard(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["facturado_total"], 300.0)
        self.assertEqual(result["kpis"]["comision_total"], 30.0)
        self.assertEqual(result["services"]["gestoria"]["kpis"]["facturado_total"], 100.0)
        self.assertEqual(result["services"]["fincas"]["kpis"]["facturado_total"], 200.0)
        self.assertEqual(result["pending"]["total"], 1)

    def test_compute_workspace_rrhh_productividad_routes_service_aliases(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        facturacion_calls = []

        def fake_facturacion_anual(conn_arg, workspace_id, empresa_id, persona_id, *, servicio_keys, ejercicio=""):
            facturacion_calls.append(
                {
                    "workspace_id": workspace_id,
                    "empresa_id": empresa_id,
                    "persona_id": persona_id,
                    "servicio_keys": {str(item) for item in servicio_keys},
                    "ejercicio": ejercicio,
                }
            )
            return {"kpis": {"origin": "facturacion", "ejercicio": ejercicio}, "items": []}

        with mock.patch.object(server, "compute_workspace_rrhh_productividad_renta", return_value={"kpis": {"origin": "renta"}, "items": []}) as renta_mock:
            with mock.patch.object(server, "compute_workspace_rrhh_productividad_seguros", return_value={"kpis": {"origin": "seguros"}, "items": []}) as seguros_mock:
                with mock.patch.object(server, "compute_workspace_rrhh_productividad_hipotecas", return_value={"kpis": {"origin": "hipotecas"}, "items": []}) as hipotecas_mock:
                    with mock.patch.object(server, "compute_workspace_rrhh_productividad_facturacion_anual", side_effect=fake_facturacion_anual) as facturacion_mock:
                        renta = server.compute_workspace_rrhh_productividad(conn, "ws-1", "emp-1", "p-1", "renta", ejercicio="2026")
                        seguros = server.compute_workspace_rrhh_productividad(conn, "ws-1", "emp-1", "p-1", "seguros", ejercicio="2026")
                        hipotecas = server.compute_workspace_rrhh_productividad(conn, "ws-1", "emp-1", "p-1", "hipotecas", ejercicio="2026")
                        gestoria = server.compute_workspace_rrhh_productividad(conn, "ws-1", "emp-1", "p-1", "gestoría", ejercicio="2026")
                        fincas = server.compute_workspace_rrhh_productividad(conn, "ws-1", "emp-1", "p-1", "administración fincas", ejercicio="2026")

        conn.close()

        self.assertEqual(renta, {"kpis": {"origin": "renta"}, "items": []})
        self.assertEqual(seguros, {"kpis": {"origin": "seguros"}, "items": []})
        self.assertEqual(hipotecas, {"kpis": {"origin": "hipotecas"}, "items": []})
        self.assertEqual(gestoria, {"kpis": {"origin": "facturacion", "ejercicio": "2026"}, "items": []})
        self.assertEqual(fincas, {"kpis": {"origin": "facturacion", "ejercicio": "2026"}, "items": []})

        self.assertEqual(renta_mock.call_count, 1)
        self.assertEqual(seguros_mock.call_count, 1)
        self.assertEqual(hipotecas_mock.call_count, 1)
        self.assertEqual(facturacion_mock.call_count, 2)
        self.assertEqual(facturacion_calls[0]["servicio_keys"], {"gestoria", "gestoría"})
        self.assertEqual(facturacion_calls[0]["ejercicio"], "2026")
        self.assertEqual(
            facturacion_calls[1]["servicio_keys"],
            {
                "fincas",
                "administracion fincas",
                "administración fincas",
                "administracion de fincas",
                "administración de fincas",
                "admin fincas",
                "admin de fincas",
            },
        )

    def test_compute_workspace_rrhh_productividad_facturacion_anual_matches_real_fincas_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              usuario TEXT,
              email TEXT,
              nombre TEXT,
              apellido TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO usuarios (id, usuario, email, nombre, apellido) VALUES (?, ?, ?, ?, ?)",
            ("u-1", "mper", "mper@example.com", "Miguel", "Perez"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-fincas", "Cliente Fincas", "22222222B", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-fincas", "c-fincas", "emp-1", "Administración de fincas", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-fincas", "ws-1", "emp-1", "c-fincas", "Administración de fincas", "2026-02-20", "", 200.0, 0),
        )

        try:
            result = server.compute_workspace_rrhh_productividad(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                "administración fincas",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 1)
        self.assertEqual(result["kpis"]["facturado_total"], 200.0)
        self.assertEqual(result["kpis"]["comision_total"], 20.0)
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["cliente_nombre"], "Cliente Fincas")

    def test_compute_workspace_rrhh_productividad_facturacion_anual_matches_commercial_fincas_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              usuario TEXT,
              email TEXT,
              nombre TEXT,
              apellido TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO usuarios (id, usuario, email, nombre, apellido) VALUES (?, ?, ?, ?, ?)",
            ("u-1", "mper", "mper@example.com", "Miguel", "Perez"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-admin-fincas", "Cliente Alias", "33333333C", "u-1", "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-admin-fincas", "c-admin-fincas", "emp-1", "Administracion Fincas Velazquez", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-admin-fincas", "ws-1", "emp-1", "c-admin-fincas", "Administracion Fincas Velazquez", "2026-03-20", "", 150.0, 1),
        )

        try:
            result = server.compute_workspace_rrhh_productividad(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                "admin fincas",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 1)
        self.assertEqual(result["kpis"]["facturado_total"], 150.0)
        self.assertEqual(result["kpis"]["comision_total"], 15.0)
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["cliente_nombre"], "Cliente Alias")

    def test_compute_workspace_rrhh_productividad_facturacion_anual_uses_clientes_empresas_capture_for_commercial_fincas_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              usuario TEXT,
              email TEXT,
              nombre TEXT,
              apellido TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO usuarios (id, usuario, email, nombre, apellido) VALUES (?, ?, ?, ?, ?)",
            ("u-1", "mper", "mper@example.com", "Miguel", "Perez"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-admin-fincas", "Cliente Alias", "44444444D", None, "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-admin-fincas", "c-admin-fincas", "emp-1", "Administracion Fincas Velazquez", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-admin-fincas", "ws-1", "emp-1", "c-admin-fincas", "Administracion Fincas Velazquez", "2026-03-20", "", 150.0, 1),
        )

        try:
            result = server.compute_workspace_rrhh_productividad(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                "admin fincas",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 1)
        self.assertEqual(result["kpis"]["facturado_total"], 150.0)
        self.assertEqual(result["kpis"]["comision_total"], 15.0)
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["cliente_nombre"], "Cliente Alias")

    def test_compute_workspace_rrhh_productividad_facturacion_anual_matches_community_alias_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              usuario TEXT,
              email TEXT,
              nombre TEXT,
              apellido TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO usuarios (id, usuario, email, nombre, apellido) VALUES (?, ?, ?, ?, ?)",
            ("u-1", "mper", "mper@example.com", "Miguel", "Perez"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-comunidades", "Cliente Comunidad", "55555555E", None, "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-comunidades", "c-comunidades", "emp-1", "Comunidad de propietarios Velazquez", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-comunidades", "ws-1", "emp-1", "c-comunidades", "Comunidad de propietarios Velazquez", "2026-04-10", "", 180.0, 1),
        )

        try:
            result = server.compute_workspace_rrhh_productividad(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                "fincas",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 1)
        self.assertEqual(result["kpis"]["facturado_total"], 180.0)
        self.assertEqual(result["kpis"]["comision_total"], 18.0)
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["cliente_nombre"], "Cliente Comunidad")

    def test_compute_workspace_rrhh_productividad_facturacion_anual_uses_clientes_empresas_capture_for_community_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              nombre TEXT,
              usuario_id TEXT
            );
            CREATE TABLE usuarios (
              id TEXT PRIMARY KEY,
              usuario TEXT,
              email TEXT,
              nombre TEXT,
              apellido TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              captado_por_user_id TEXT,
              procedencia_canal TEXT
            );
            CREATE TABLE workspace_facturacion (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              cliente_id TEXT,
              servicio TEXT,
              fecha_emision TEXT,
              responsable TEXT,
              subtotal REAL,
              cobrada INTEGER
            );
            """
        )
        conn.execute(
            "INSERT INTO workspace_registro_personal (id, workspace_id, nombre, usuario_id) VALUES (?, ?, ?, ?)",
            ("p-1", "ws-1", "Persona Demo", "u-1"),
        )
        conn.execute(
            "INSERT INTO usuarios (id, usuario, email, nombre, apellido) VALUES (?, ?, ?, ?, ?)",
            ("u-1", "mper", "mper@example.com", "Miguel", "Perez"),
        )
        conn.execute(
            "INSERT INTO clientes (id, nombre, nif, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?)",
            ("c-comunidades", "Cliente Comunidad", "55555555E", None, "web"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, captado_por_user_id, procedencia_canal) VALUES (?, ?, ?, ?, ?, ?)",
            ("ce-comunidades", "c-comunidades", "emp-1", "Comunidades Velazquez", "u-1", "web"),
        )
        conn.execute(
            """
            INSERT INTO workspace_facturacion (
              id, workspace_id, empresa_id, cliente_id, servicio, fecha_emision, responsable, subtotal, cobrada
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("f-comunidades", "ws-1", "emp-1", "c-comunidades", "Comunidades Velazquez", "2026-04-10", "", 180.0, 1),
        )

        try:
            result = server.compute_workspace_rrhh_productividad(
                conn,
                "ws-1",
                "emp-1",
                "p-1",
                "fincas",
                ejercicio="2026",
            )
        finally:
            conn.close()

        self.assertEqual(result["kpis"]["clientes"], 1)
        self.assertEqual(result["kpis"]["facturado_total"], 180.0)
        self.assertEqual(result["kpis"]["comision_total"], 18.0)
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["cliente_nombre"], "Cliente Comunidad")

    def test_gestoria_service_sql_condition_matches_full_fincas_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              servicio TEXT
            );
            """
        )
        conn.executemany(
            "INSERT INTO clientes_empresas (id, servicio) VALUES (?, ?)",
            [
                ("ce-1", "gestoria"),
                ("ce-2", "administración de fincas"),
                ("ce-3", "Administracion Fincas Velazquez"),
            ],
        )

        clause = server.gestoria_service_sql_condition("ce")
        row = conn.execute(
            f"SELECT COUNT(*) AS total FROM clientes_empresas ce WHERE {clause}"
        ).fetchone()
        conn.close()

        self.assertEqual(row["total"], 3)

    def test_gestoria_service_sql_condition_matches_community_label(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              servicio TEXT
            );
            """
        )
        conn.execute("INSERT INTO clientes_empresas (id, servicio) VALUES (?, ?)", ("ce-1", "Comunidades Velazquez"))

        clause = server.gestoria_service_sql_condition("ce")
        row = conn.execute(
            f"SELECT COUNT(*) AS total FROM clientes_empresas ce WHERE {clause}"
        ).fetchone()
        conn.close()

        self.assertEqual(row["total"], 1)

    def test_resolve_seguros_ocr_cliente_id_scopes_nif_and_name_by_empresa(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Seguro", "12345678A", "2026-03-03", "2026-03-03"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-right", "e1", "Cliente Seguro", "12345678A", "2026-03-01", "2026-03-01"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, estado, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("link-right", "c-right", "e1", "Seguros", "Activo", "2026-03-01", "2026-03-01"),
        )

        try:
            by_nif = server.resolve_seguros_ocr_cliente_id(conn, "e1", "12345678A", "")
            by_name = server.resolve_seguros_ocr_cliente_id(conn, "e1", "", "Cliente Seguro")
        finally:
            conn.close()

        self.assertEqual(by_nif, "c-right")
        self.assertEqual(by_name, "c-right")

    def test_resolve_seguros_ocr_cliente_id_requires_empresa_scope(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
            ("c-global", "", "Cliente Seguro Global", "12345678A", "2026-03-01", "2026-03-01"),
        )

        try:
            result = server.resolve_seguros_ocr_cliente_id(conn, "", "12345678A", "Cliente Seguro Global")
        finally:
            conn.close()

        self.assertIsNone(result)

    def test_ensure_cliente_for_financiacion_requires_empresa_scope(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              fecha_nacimiento TEXT,
              direccion TEXT,
              procedencia_canal TEXT,
              procedencia_detalle TEXT,
              procedencia_user_id TEXT,
              estado TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )

        try:
            result = server.ensure_cliente_for_financiacion(
                conn,
                "",
                "Cliente Financiacion",
                "12345678A",
                "2026-03-01",
                {"telefono": "600000000", "email": "cliente@example.com"},
            )
            count = conn.execute("SELECT COUNT(*) AS total FROM clientes").fetchone()["total"]
        finally:
            conn.close()

        self.assertIsNone(result)
        self.assertEqual(count, 0)

    def test_resolve_cliente_lookup_row_scopes_workspace_company_ids_by_workspace(self):
        conn = self._make_workspace_scope_conn()
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Fuera", "12345678A", "", "", "2026-03-04", "2026-03-04"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-right", "emp-1", "Cliente Dentro", "12345678A", "", "", "2026-03-01", "2026-03-01"),
        )
        conn.execute(
            "INSERT INTO workspace_empresas (id, workspace_id, empresa_id) VALUES (?, ?, ?)",
            ("we-1", "ws-1", "e1"),
        )

        try:
            row = server.resolve_cliente_lookup_row(conn, "12345678A", workspace_id="ws-1")
        finally:
            conn.close()

        self.assertIsNotNone(row)
        self.assertEqual(row["id"], "c-right")

    def test_resolve_cliente_duplicate_id_scopes_workspace_company_ids_by_workspace(self):
        conn = self._make_workspace_scope_conn()
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Duplicado", "12345678A", "", "", "2026-03-04", "2026-03-04"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-right", "emp-1", "Cliente Duplicado", "12345678A", "", "", "2026-03-01", "2026-03-01"),
        )

        try:
            by_nif = server.resolve_cliente_duplicate_id(
                conn,
                "",
                "12345678A",
                workspace_id="ws-1",
            )
            by_name = server.resolve_cliente_duplicate_id(
                conn,
                "Cliente Duplicado",
                "",
                workspace_id="ws-1",
            )
        finally:
            conn.close()

        self.assertEqual(by_nif, "c-right")
        self.assertEqual(by_name, "c-right")

    def test_resolve_clientes_by_nif_rows_scopes_workspace_and_legacy_company_ids(self):
        conn = self._make_workspace_scope_conn()
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Fuera", "12345678A", "", "", "2026-03-04", "2026-03-04"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-right", "emp-1", "Cliente Dentro", "12345678A", "", "", "2026-03-01", "2026-03-01"),
        )

        try:
            by_workspace = server.resolve_clientes_by_nif_rows(
                conn,
                "12345678A",
                workspace_id="ws-1",
                limit=6,
            )
            by_empresa = server.resolve_clientes_by_nif_rows(
                conn,
                "12345678A",
                empresa_id="emp-1",
                limit=6,
            )
        finally:
            conn.close()

        self.assertEqual([row["id"] for row in by_workspace], ["c-right"])
        self.assertEqual([row["id"] for row in by_empresa], ["c-right"])

    def test_resolve_clientes_by_nif_rows_matches_commercial_community_label_for_fincas(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              workspace_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-community", "emp-1", "Cliente Comunidad", "12345678A", "", "", "2026-03-01", "2026-03-01"),
        )
        conn.execute(
            """
            INSERT INTO clientes_empresas (
              id, cliente_id, empresa_id, workspace_id, servicio, estado, fecha_inicio, fecha_fin, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "ce-community",
                "c-community",
                "emp-1",
                "ws-1",
                "Comunidades Velazquez",
                "Activo",
                "2026-03-01",
                "",
                "2026-03-01",
                "2026-03-01",
            ),
        )

        try:
            rows = server.resolve_clientes_by_nif_rows(
                conn,
                "12345678A",
                services=["fincas"],
                workspace_id="ws-1",
            )
        finally:
            conn.close()

        self.assertEqual([row["id"] for row in rows], ["c-community"])

    def test_ensure_workspace_budget_client_scopes_workspace_company_ids_by_workspace(self):
        conn = self._make_workspace_scope_conn()
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              telefono TEXT,
              email TEXT,
              procedencia_canal TEXT,
              procedencia_detalle TEXT,
              procedencia_user_id TEXT,
              estado TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              fecha_inicio TEXT,
              fecha_fin TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, procedencia_canal, procedencia_detalle, procedencia_user_id, estado, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-other", "e2", "Cliente Presupuesto", "12345678A", "", "", "", "", None, "Lead", "2026-03-04", "2026-03-04"),
        )
        conn.execute(
            "INSERT INTO clientes (id, empresa_id, nombre, nif, telefono, email, procedencia_canal, procedencia_detalle, procedencia_user_id, estado, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("c-right", "emp-1", "Cliente Presupuesto", "12345678A", "", "", "", "", None, "Lead", "2026-03-01", "2026-03-01"),
        )
        conn.execute(
            "INSERT INTO clientes_empresas (id, cliente_id, empresa_id, servicio, estado, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("link-right", "c-right", "emp-1", "Gestoría", "Presupuesto", "2026-03-01", "2026-03-01"),
        )

        try:
            cliente_id, cliente_nombre = server.ensure_workspace_budget_client(
                conn,
                workspace_id="ws-1",
                empresa_id="emp-1",
                servicio="gestoria",
                cliente_nif="12345678A",
                cliente_lookup="Cliente Presupuesto",
                now="2026-03-04",
            )
        finally:
            conn.close()

        self.assertEqual(cliente_id, "c-right")
        self.assertEqual(cliente_nombre, "Cliente Presupuesto")

    def test_workspace_budget_action_service_keeps_fincas_separate_from_gestoria(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE acciones (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              servicio TEXT,
              cliente_id TEXT,
              cliente_nombre TEXT,
              fecha TEXT,
              tipo TEXT,
              responsable TEXT,
              estado TEXT,
              notas TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )

        gestoria_service = server.workspace_budget_action_service("gestoria")
        fincas_service = server.workspace_budget_action_service("administración fincas")
        self.assertEqual(gestoria_service, "gestoria")
        self.assertEqual(fincas_service, "fincas")

        action_id = server.upsert_workspace_budget_action(
            conn,
            empresa_id="emp-1",
            servicio=fincas_service,
            cliente_id="c-1",
            cliente_nombre="Cliente Fincas",
            fecha="2026-03-05",
            tipo="Seguimiento",
            responsable="Persona Demo",
            estado="Pendiente",
            notas="Seguimiento de fincas",
            now="2026-03-05T10:00:00+00:00",
        )
        row = conn.execute("SELECT servicio, cliente_nombre, tipo FROM acciones WHERE id = ?", (action_id,)).fetchone()
        conn.close()

        self.assertEqual(row["servicio"], "fincas")
        self.assertEqual(row["cliente_nombre"], "Cliente Fincas")
        self.assertEqual(row["tipo"], "Seguimiento")

    def test_fetch_empresa_presupuestos_includes_legacy_fincas_labels(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE empresas (
              id TEXT PRIMARY KEY,
              nombre TEXT
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              nombre TEXT
            );
            CREATE TABLE workspace_presupuestos (
              id TEXT PRIMARY KEY,
              workspace_id TEXT,
              empresa_id TEXT,
              cliente_id TEXT,
              servicio TEXT,
              titulo TEXT,
              estado TEXT,
              fecha TEXT,
              fecha_seguimiento TEXT,
              encargo_estado TEXT,
              fecha_encargo TEXT,
              total REAL,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE workspace_presupuesto_lineas (
              id TEXT PRIMARY KEY,
              presupuesto_id TEXT,
              orden INTEGER,
              categoria TEXT,
              concepto TEXT,
              cantidad REAL,
              unidad TEXT,
              precio_unitario REAL,
              descuento_pct REAL,
              total_linea REAL,
              created_at TEXT,
              updated_at TEXT
            );
            """
        )
        conn.executemany(
            "INSERT INTO empresas (id, nombre) VALUES (?, ?)",
            [("emp-1", "Empresa Demo")],
        )
        conn.executemany(
            "INSERT INTO clientes (id, nombre) VALUES (?, ?)",
            [("c-1", "Cliente Uno"), ("c-2", "Cliente Dos"), ("c-3", "Cliente Tres"), ("c-4", "Cliente Cuatro")],
        )
        conn.executemany(
            """
            INSERT INTO workspace_presupuestos (
              id, workspace_id, empresa_id, cliente_id, servicio, titulo, estado, fecha, fecha_seguimiento,
              encargo_estado, fecha_encargo, total, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("p-1", "ws-1", "emp-1", "c-1", "administración de fincas", "Propuesta Acústica", "Estudio", "2026-01-10", "2026-01-11", "", "", 100.0, "2026-01-10", "2026-01-10"),
                ("p-2", "ws-1", "emp-1", "c-2", "admin de fincas", "Propuesta Alias", "Estudio", "2026-01-12", "2026-01-13", "", "", 200.0, "2026-01-12", "2026-01-12"),
                ("p-3", "ws-1", "emp-1", "c-3", "Fincas Velazquez", "Propuesta Comercial", "Estudio", "2026-01-14", "2026-01-15", "", "", 300.0, "2026-01-14", "2026-01-14"),
                ("p-4", "ws-1", "emp-1", "c-4", "Gestoría", "Propuesta Gestoria", "Estudio", "2026-01-16", "2026-01-17", "", "", 400.0, "2026-01-16", "2026-01-16"),
            ],
        )

        try:
            result = server.fetch_empresa_presupuestos(conn, "emp-1", servicio="fincas")
        finally:
            conn.close()

        titles = {row["titulo"] for row in result["rows"]}
        self.assertEqual(titles, {"Propuesta Acústica", "Propuesta Alias", "Propuesta Comercial"})
        self.assertEqual(len(result["rows"]), 3)

    def test_fetch_workspace_company_ids_does_not_backfill_every_company(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE workspaces (
              id TEXT PRIMARY KEY,
              slug TEXT,
              nombre TEXT
            );
            CREATE TABLE empresas (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE workspace_empresas (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT NOT NULL,
              rol TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE workspace_companies (
              workspace_id TEXT NOT NULL,
              legacy_empresa_id TEXT,
              activo INTEGER NOT NULL DEFAULT 1,
              nombre TEXT
            );
            CREATE TABLE workspace_registro_personal (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              empresa_id TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            """
        )
        conn.executemany(
            "INSERT INTO empresas (id, nombre, activo) VALUES (?, ?, 1)",
            [("emp-1", "Empresa Uno"), ("emp-2", "Empresa Dos")],
        )
        conn.execute("INSERT INTO workspaces (id, slug, nombre) VALUES (?, ?, ?)", ("ws-1", "ws-1", "Workspace 1"))
        conn.commit()

        ids = server.fetch_workspace_company_ids(conn, "ws-1")
        self.assertEqual(ids, [])
        count = conn.execute("SELECT COUNT(*) AS total FROM workspace_empresas").fetchone()["total"]
        self.assertEqual(count, 0)
        conn.close()

    def test_resolve_workspace_scope_empresa_ids_does_not_broaden_scope_with_mismatched_empresa_id(self):
        conn = self._make_workspace_scope_conn()
        try:
            ids = server.resolve_workspace_scope_empresa_ids(conn, "ws-1", empresa_id="emp-2")
            self.assertEqual(ids, ["emp-1"])
        finally:
            conn.close()

    def test_resolve_workspace_time_toggle_persona_id_allows_manager_targets_and_keeps_self_service(self):
        conn = self._make_workspace_rrhh_conn()
        try:
            conn.execute(
                """
                INSERT INTO workspaces (id, nombre, slug, estado, plan, kind, created_at, updated_at)
                VALUES ('ws-1', 'Workspace 1', 'ws-1', 'Activo', 'Enterprise', 'Directo', datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_companies (id, workspace_id, legacy_empresa_id, nombre, activo, created_at, updated_at)
                VALUES ('wc-1', 'ws-1', 'emp-1', 'Empresa Uno', 1, datetime('now'), datetime('now'))
                """
            )
            conn.execute("INSERT INTO empresas (id, nombre, activo) VALUES ('emp-1', 'Empresa Uno', 1)")
            conn.execute(
                """
                INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, activo, registro_horario_activo, created_at, updated_at)
                VALUES ('u-manager', 'Manager', 'User', 'manager', 'manager@example.com', 'Gestoría', 'Administrador', 1, 1, datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, activo, registro_horario_activo, created_at, updated_at)
                VALUES ('u-worker', 'Worker', 'User', 'worker', 'worker@example.com', 'Gestoría', 'Miembro', 1, 1, datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_miembros (id, workspace_id, usuario_id, rol, created_at, updated_at)
                VALUES ('wm-worker', 'ws-1', 'u-worker', 'Miembro', datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_registro_personal (
                  id, workspace_id, empresa_id, empresa_manual, usuario_id, usuario_manual, source,
                  nombre, activo, created_at, updated_at
                ) VALUES (
                  'persona-self', 'ws-1', 'emp-1', 1, 'u-worker', 1, 'manual',
                  'Persona Worker', 1, datetime('now'), datetime('now')
                )
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_registro_personal (
                  id, workspace_id, empresa_id, empresa_manual, usuario_id, usuario_manual, source,
                  nombre, activo, created_at, updated_at
                ) VALUES (
                  'persona-target', 'ws-1', 'emp-1', 1, 'u-target', 1, 'manual',
                  'Persona Target', 1, datetime('now'), datetime('now')
                )
                """
            )
            conn.commit()

            manager_session = {"user_id": "u-manager", "rol": "Admin", "usuario": "Manager", "email": "manager@example.com"}
            worker_session = {"user_id": "u-worker", "rol": "Miembro", "usuario": "Worker", "email": "worker@example.com"}

            self.assertEqual(
                server.resolve_workspace_time_toggle_persona_id(conn, manager_session, "ws-1", "persona-target"),
                ("persona-target", ""),
            )
            self.assertEqual(
                server.resolve_workspace_time_toggle_persona_id(conn, worker_session, "ws-1", ""),
                ("persona-self", ""),
            )
            self.assertEqual(
                server.resolve_workspace_time_toggle_persona_id(conn, worker_session, "ws-1", "persona-target"),
                ("", "No autorizado"),
            )
        finally:
            conn.close()

    def test_fetch_workspace_personal_and_time_users_include_linked_rows_without_memberships(self):
        conn = self._make_workspace_rrhh_conn()
        try:
            conn.execute(
                """
                INSERT INTO workspaces (id, nombre, slug, estado, plan, kind, created_at, updated_at)
                VALUES ('ws-1', 'Workspace 1', 'ws-1', 'Activo', 'Enterprise', 'Directo', datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_companies (id, workspace_id, legacy_empresa_id, nombre, activo, created_at, updated_at)
                VALUES ('wc-1', 'ws-1', 'emp-1', 'Empresa Uno', 1, datetime('now'), datetime('now'))
                """
            )
            conn.execute("INSERT INTO empresas (id, nombre, activo) VALUES ('emp-1', 'Empresa Uno', 1)")
            conn.execute(
                """
                INSERT INTO usuarios (id, nombre, apellido, usuario, email, servicio, rol, activo, registro_horario_activo, created_at, updated_at)
                VALUES ('u-1', 'Persona', 'RRHH', 'u1', 'u1@example.com', 'Gestoría', 'Miembro', 1, 1, datetime('now'), datetime('now'))
                """
            )
            conn.execute(
                """
                INSERT INTO workspace_registro_personal (
                  id, workspace_id, empresa_id, empresa_manual, usuario_id, usuario_manual, source,
                  nombre, activo, created_at, updated_at
                ) VALUES (
                  'persona-1', 'ws-1', 'emp-1', 1, 'u-1', 1, 'manual',
                  'Persona RRHH', 1, datetime('now'), datetime('now')
                )
                """
            )
            conn.commit()

            personal = server.fetch_workspace_personal(conn, "ws-1", only_active=False, limit=10)
            time_users = server.fetch_workspace_time_users(conn, "ws-1", only_enabled=False, limit=10)
        finally:
            conn.close()

        self.assertEqual(len(personal["rows"]), 1)
        self.assertEqual(personal["rows"][0]["usuario_id"], "u-1")
        self.assertEqual(personal["rows"][0]["empresa_id"], "emp-1")
        self.assertEqual(personal["rows"][0]["empresa_nombre"], "Empresa Uno")

        self.assertEqual(len(time_users["rows"]), 1)
        self.assertEqual(time_users["rows"][0]["id"], "u-1")
        self.assertEqual(time_users["rows"][0]["empresa_id"], "emp-1")
        self.assertEqual(time_users["rows"][0]["empresa_nombre"], "Empresa Uno")

    def test_resolve_empresa_id_for_request_rejects_workspace_company_mismatch(self):
        conn = self._make_workspace_scope_conn()
        try:
            with mock.patch.object(server, "WORKSPACE_MEMBERSHIP_ENFORCE", True):
                eid, wc_id, err = server.resolve_empresa_id_for_request(
                    conn,
                    {"user_id": "u-1", "rol": "Miembro", "servicio": "Gestoría"},
                    workspace_id="ws-1",
                    empresa_id="emp-2",
                    workspace_company_id="wc-1",
                    write=False,
                )
            self.assertEqual((eid, wc_id, err), ("", "wc-1", "workspace_company_id no coincide con empresa_id"))
        finally:
            conn.close()

    def test_resolve_empresa_id_for_request_resolves_matching_workspace_company(self):
        conn = self._make_workspace_scope_conn()
        try:
            with mock.patch.object(server, "WORKSPACE_MEMBERSHIP_ENFORCE", True):
                eid, wc_id, err = server.resolve_empresa_id_for_request(
                    conn,
                    {"user_id": "u-1", "rol": "Miembro", "servicio": "Gestoría"},
                    workspace_id="ws-1",
                    empresa_id="",
                    workspace_company_id="wc-1",
                    write=False,
                )
            self.assertEqual((eid, wc_id, err), ("emp-1", "wc-1", ""))
        finally:
            conn.close()

    def test_resolve_scoped_record_access_distinguishes_missing_and_foreign_records(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE inmuebles (
              id TEXT PRIMARY KEY,
              empresa_id TEXT
            );
            CREATE TABLE demandas (
              id TEXT PRIMARY KEY,
              empresa_id TEXT
            );
            """
        )
        conn.execute("INSERT INTO inmuebles (id, empresa_id) VALUES (?, ?)", ("inm-foreign", "emp-2"))
        conn.execute("INSERT INTO demandas (id, empresa_id) VALUES (?, ?)", ("dem-foreign", "emp-2"))
        try:
            inm_status = server.resolve_scoped_record_access(
                conn,
                "inm-foreign",
                "emp-1",
                table="inmuebles",
                fetch_fn=server.fetch_inmueble_for_empresa,
            )
            dem_status = server.resolve_scoped_record_access(
                conn,
                "dem-foreign",
                "emp-1",
                table="demandas",
                fetch_fn=server.fetch_demanda_for_empresa,
            )
            missing_status = server.resolve_scoped_record_access(
                conn,
                "inm-missing",
                "emp-1",
                table="inmuebles",
                fetch_fn=server.fetch_inmueble_for_empresa,
            )
        finally:
            conn.close()

        self.assertEqual(inm_status, "forbidden")
        self.assertEqual(dem_status, "forbidden")
        self.assertEqual(missing_status, "missing")

    def test_resolve_cliente_scope_access_distinguishes_allowed_forbidden_and_missing(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT
            );
            """
        )
        conn.execute("INSERT INTO clientes (id, empresa_id) VALUES (?, ?)", ("cli-ok", "emp-1"))
        conn.execute("INSERT INTO clientes (id, empresa_id) VALUES (?, ?)", ("cli-foreign", "emp-2"))
        conn.execute("INSERT INTO clientes_empresas (id, cliente_id, empresa_id) VALUES (?, ?, ?)", ("link-ok", "cli-ok", "emp-1"))
        try:
            allowed = server.resolve_cliente_scope_access(conn, "cli-ok", empresa_id="emp-1")
            forbidden = server.resolve_cliente_scope_access(conn, "cli-foreign", empresa_id="emp-1")
            missing = server.resolve_cliente_scope_access(conn, "cli-missing", empresa_id="emp-1")
        finally:
            conn.close()

        self.assertEqual(allowed, "ok")
        self.assertEqual(forbidden, "forbidden")
        self.assertEqual(missing, "missing")

    def test_resolve_request_legacy_empresa_id_normalizes_workspace_company_sources(self):
        conn = self._make_workspace_scope_conn()
        try:
            self.assertEqual(
                server.resolve_request_legacy_empresa_id(
                    conn,
                    workspace_id="ws-1",
                    empresa_id="",
                    workspace_company_id="wc-1",
                ),
                "emp-1",
            )
            self.assertEqual(
                server.resolve_request_legacy_empresa_id(
                    conn,
                    workspace_id="",
                    empresa_id="wc-1",
                    workspace_company_id="",
                ),
                "emp-1",
            )
            self.assertEqual(
                server.resolve_request_legacy_empresa_id(
                    conn,
                    workspace_id="ws-1",
                    empresa_id="emp-2",
                    workspace_company_id="wc-1",
                ),
                "emp-2",
            )
            self.assertEqual(
                server.resolve_request_legacy_empresa_id(
                    conn,
                    workspace_id="",
                    empresa_id="emp-2",
                    workspace_company_id="",
                ),
                "emp-2",
            )
        finally:
            conn.close()

    def test_resolve_payload_legacy_empresa_id_normalizes_workspace_payload(self):
        conn = self._make_workspace_scope_conn()
        try:
            session = {"user_id": "u-1", "rol": "Miembro", "servicio": "Gestoría"}
            self.assertEqual(
                server.resolve_payload_legacy_empresa_id(
                    conn,
                    session,
                    {"workspace_id": "ws-1", "workspace_company_id": "wc-1"},
                    write=True,
                ),
                ("emp-1", ""),
            )
            self.assertEqual(
                server.resolve_payload_legacy_empresa_id(
                    conn,
                    session,
                    {"workspace_id": "ws-1", "workspace_company_id": "wc-1", "empresa_id": "emp-2"},
                    write=True,
                ),
                ("", ""),
            )
        finally:
            conn.close()

    def test_external_base_url_ignores_host_headers(self):
        handler = SimpleNamespace(
            headers={
                "Host": "attacker.example",
                "X-Forwarded-Host": "attacker.example",
                "X-Forwarded-Proto": "https",
            }
        )
        with mock.patch.dict(os.environ, {}, clear=True):
            expected = "http://localhost:8000"
            self.assertEqual(server.Handler._external_base_url(handler), expected)
            self.assertEqual(public_links.external_base_url(), expected)

    def test_public_link_helpers_match_new_module(self):
        payload = {"id": "req-1", "doc_nombre": "Contrato", "otp_required": 1}
        with mock.patch.dict(
            os.environ,
            {
                "APP_BASE_URL": "https://crm.example.com",
                "PUBLIC_URL": "https://public.example.com",
            },
            clear=True,
        ):
            self.assertEqual(server.configured_app_base_url(), public_links.configured_app_base_url())
            self.assertEqual(
                server.resolve_public_link_base_url(""),
                public_links.resolve_public_link_base_url(""),
            )
            self.assertEqual(
                server.build_public_fragment_url("activar_token", "abc"),
                public_links.build_public_fragment_url("activar_token", "abc"),
            )
            self.assertEqual(
                server.build_public_fragment_url("token", "abc", base_url="https://crm.example.com", path="/kiosk"),
                public_links.build_public_fragment_url("token", "abc", base_url="https://crm.example.com", path="/kiosk"),
            )
        with mock.patch.object(public_links.secrets, "token_urlsafe", return_value="fixed-token"):
            self.assertEqual(server.make_signature_token(), public_links.make_signature_token())
            self.assertEqual(
                server.hash_signature_token("abc123"),
                public_links.hash_signature_token("abc123"),
            )
            self.assertEqual(
                server.signature_request_public_payload(payload, token="ignored"),
                public_links.signature_request_public_payload(payload, token="ignored"),
            )

    def test_safe_json_object_matches_new_module(self):
        payload = {"foo": 1, "bar": ["a", "b"]}
        invalid = "not json"
        self.assertEqual(server._safe_json_object(payload), security_utils._safe_json_object(payload))
        self.assertEqual(server._safe_json_object(invalid), security_utils._safe_json_object(invalid))

    def test_security_helpers_match_new_module(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            base_path = Path(tmpdir)
            target_path = base_path / "safe" / "file.txt"
            target_path.parent.mkdir(parents=True)

            sample_html = "<p>Hello</p><script>ignored()</script>"
            sample_strict_html = "<p>Hello</p>\n\n<p>World</p>"
            sample_json = '{"ok": true}'
            placeholder_key = "0123456789abcdef0123456789abcdef"
            s3_key = r"folder\\nested\\file.pdf"

            with mock.patch.object(server, "COPILOT_WEB_ALLOWED_DOMAINS", {"example.com"}):
                self.assertEqual(server._ct_eq("abc", "abc"), security_utils._ct_eq("abc", "abc"))
                self.assertEqual(server._ct_eq("abc", "xyz"), security_utils._ct_eq("abc", "xyz"))
                self.assertEqual(server._normalize_s3_key(s3_key), security_utils._normalize_s3_key(s3_key))
                self.assertEqual(
                    server._iter_s3_legacy_key_candidates(placeholder_key),
                    security_utils._iter_s3_legacy_key_candidates(placeholder_key),
                )
                self.assertEqual(server._is_public_doc_url("https://example.com/doc.pdf"), security_utils._is_public_doc_url("https://example.com/doc.pdf"))
                self.assertEqual(server._looks_like_placeholder_doc_key(placeholder_key), security_utils._looks_like_placeholder_doc_key(placeholder_key))
                self.assertEqual(server._normalize_doc_key_for_ui("s3://bucket/doc.pdf"), security_utils._normalize_doc_key_for_ui("s3://bucket/doc.pdf"))
                self.assertEqual(server._safe_json_object(sample_json), security_utils._safe_json_object(sample_json))
                self.assertEqual(server.html_to_text(sample_html), security_utils.html_to_text(sample_html))
                self.assertEqual(server._html_to_text(sample_strict_html), security_utils._html_to_text(sample_strict_html))
                self.assertEqual(server._extract_title("<html><head><title>  Demo  </title></head></html>"), security_utils._extract_title("<html><head><title>  Demo  </title></head></html>"))
                self.assertEqual(server._pdf_escape("a(b)\\c"), security_utils._pdf_escape("a(b)\\c"))
                self.assertEqual(server.safe_resolve_under(base_path, "safe/file.txt"), security_utils.safe_resolve_under(base_path, "safe/file.txt"))
                self.assertIsNone(server.safe_resolve_under(base_path, "../escape.txt"))
                self.assertEqual(server._domain_is_allowed("example.com"), security_utils._domain_is_allowed("example.com", {"example.com"}))
                self.assertFalse(server._domain_is_allowed("localhost"))

            with mock.patch.object(security_utils.socket, "getaddrinfo", return_value=[(None, None, None, None, ("127.0.0.1", 0))]):
                self.assertEqual(
                    server._hostname_resolves_to_disallowed_ip("example.com"),
                    security_utils._hostname_resolves_to_disallowed_ip("example.com"),
                )

            with mock.patch.object(security_utils.socket, "getaddrinfo", return_value=[(None, None, None, None, ("8.8.8.8", 0))]):
                self.assertEqual(
                    server._hostname_resolves_to_disallowed_ip("example.com"),
                    security_utils._hostname_resolves_to_disallowed_ip("example.com"),
                )

    def test_workspace_service_alone_does_not_grant_privileged_session(self):
        self.assertFalse(server.workspace_session_is_privileged({"rol": "", "servicio": "Administración"}))
        self.assertFalse(server.workspace_session_is_privileged({"rol": "", "servicio": "Control"}))

    def test_login_rate_limit_persists_across_memory_reset(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = os.path.join(tmpdir, "auth.sqlite")

            def open_auth_store_conn(with_row_factory=True):
                conn = sqlite3.connect(db_path)
                if with_row_factory:
                    conn.row_factory = sqlite3.Row
                return conn

            with mock.patch.object(server, "open_auth_store_conn", side_effect=open_auth_store_conn):
                for _ in range(server.LOGIN_RATE_MAX_ATTEMPTS):
                    server.register_login_attempt("1.2.3.4", "alice", ok=False)

                allowed, retry_after = server.check_login_rate_limit("1.2.3.4", "alice")
                self.assertFalse(allowed)
                self.assertGreaterEqual(retry_after, 1)

                allowed_after_reset, retry_after_after_reset = server.check_login_rate_limit("1.2.3.4", "alice")
                self.assertFalse(allowed_after_reset)
                self.assertGreaterEqual(retry_after_after_reset, 1)

                server.register_login_attempt("1.2.3.4", "alice", ok=True)
                allowed_after_success, retry_after_after_success = server.check_login_rate_limit("1.2.3.4", "alice")
                self.assertTrue(allowed_after_success)
                self.assertEqual(retry_after_after_success, 0)

    def test_signature_public_payload_uses_fragment_links(self):
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(
            """
            CREATE TABLE inmuebles (
              id TEXT PRIMARY KEY,
              empresa_id TEXT NOT NULL
            );
            CREATE TABLE inmueble_docs (
              id TEXT PRIMARY KEY,
              inmueble_id TEXT NOT NULL,
              nombre TEXT,
              url TEXT
            );
            """
        )
        conn.execute("INSERT INTO inmuebles (id, empresa_id) VALUES (?, ?)", ("inm-1", "emp-1"))
        conn.execute(
            "INSERT INTO inmueble_docs (id, inmueble_id, nombre, url) VALUES (?, ?, ?, ?)",
            ("doc-1", "inm-1", "Contrato", "/uploads/contrato.pdf"),
        )
        conn.commit()

        result = server.create_inmueble_signature_request(
            conn,
            empresa_id="emp-1",
            inmueble_id="inm-1",
            doc_id="doc-1",
            doc_url="/uploads/contrato.pdf",
            doc_nombre="Contrato",
            signer_nombre="Persona",
            signer_nif="12345678A",
            signer_email="firma@example.com",
            signer_telefono="600000000",
            purpose="Firma",
            otp_required=False,
            expires_days=15,
            created_by="tester",
            now="2026-07-13T12:00:00+00:00",
        )
        self.assertTrue(result["public_url"].startswith("/#firma_inmo="))
        row = server._signature_request_row_by_token(conn, result["token"])
        public = server.signature_request_public_payload(row, token=result["token"])
        self.assertEqual(public, public_links.signature_request_public_payload(row, token=result["token"]))
        self.assertEqual(public["doc_public_url"], "/api/inmueble_signature_document")
        self.assertNotIn("?token=", public["doc_public_url"])
        conn.close()

    def test_pdf_helpers_match_new_module(self):
        canvas_a = Image.new("RGB", (800, 400), "white")
        canvas_b = Image.new("RGB", (800, 400), "white")
        draw_a = ImageDraw.Draw(canvas_a)
        draw_b = ImageDraw.Draw(canvas_b)

        font_server = server._document_font(18, bold=True)
        font_module = pdf_utils._document_font(18, bold=True)

        self.assertEqual(server._pdf_escape("a(b)\\c"), pdf_utils._pdf_escape("a(b)\\c"))
        self.assertEqual(server._pdf_wrap_lines("Uno dos tres", width=4), pdf_utils._pdf_wrap_lines("Uno dos tres", width=4))
        self.assertEqual(
            server._pdf_wrap_lines("Linea 1\n\nLinea 2", width=20),
            pdf_utils._pdf_wrap_lines("Linea 1\n\nLinea 2", width=20),
        )
        self.assertEqual(
            server._pdf_wrap_lines_px(draw_a, "Banco Santander Modernia", font_server, 120),
            pdf_utils._pdf_wrap_lines_px(draw_b, "Banco Santander Modernia", font_module, 120),
        )
        self.assertEqual(
            server._pdf_draw_justified_paragraph(draw_a, 10, 10, 180, "uno dos tres cuatro", font_server, (0, 0, 0)),
            pdf_utils._pdf_draw_justified_paragraph(draw_b, 10, 10, 180, "uno dos tres cuatro", font_module, (0, 0, 0)),
        )
        self.assertEqual(server._pdf_format_number("1234.5", 1), pdf_utils._pdf_format_number("1234.5", 1))
        self.assertEqual(font_server.getbbox("Ag"), font_module.getbbox("Ag"))
        self.assertEqual(font_server.getbbox("Modernia"), font_module.getbbox("Modernia"))
        self.assertEqual(server._parse_pdf_color("#abc"), pdf_utils._parse_pdf_color("#abc"))
        self.assertEqual(server._parse_pdf_color("#abcdef"), pdf_utils._parse_pdf_color("#abcdef"))
        self.assertEqual(
            server._pil_multiline(draw_a, "Texto largo para probar el salto de línea", font_server, width=10),
            pdf_utils._pil_multiline(draw_b, "Texto largo para probar el salto de línea", font_module, width=10),
        )
        self.assertEqual(
            server.resolve_hipoteca_bank_brand("Banco Santander S.A."),
            pdf_utils.resolve_hipoteca_bank_brand("Banco Santander S.A."),
        )
        self.assertEqual(
            server.build_hipoteca_bank_logo_meta("Banco Santander S.A."),
            pdf_utils.build_hipoteca_bank_logo_meta("Banco Santander S.A."),
        )
        self.assertEqual(
            server.normalize_hipoteca_pdf_sort_order("ASCENDENTE"),
            pdf_utils.normalize_hipoteca_pdf_sort_order("ASCENDENTE"),
        )
        self.assertEqual(
            server._logo_badge_info_from_path("/assets/logos/santander.svg"),
            pdf_utils._logo_badge_info_from_path("/assets/logos/santander.svg"),
        )

        badge_server = server._build_logo_badge_image(
            "Banco Santander",
            color="#e30613",
            short="BS",
            logo_on_dark=True,
            max_width=280,
        )
        badge_module = pdf_utils._build_logo_badge_image(
            "Banco Santander",
            color="#e30613",
            short="BS",
            logo_on_dark=True,
            max_width=280,
        )
        self.assertEqual(badge_server.size, badge_module.size)
        self.assertEqual(badge_server.mode, badge_module.mode)
        self.assertEqual(badge_server.tobytes(), badge_module.tobytes())


class GestoriaServerRouteRegressionTests(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.tmpdir.name) / "gestoria.sqlite"
        os.environ["APP_DB_BACKEND"] = "sqlite"
        self._old_db_ready = getattr(server.Handler, "_db_ready", False)
        self._old_db_ready_last_error = getattr(server.Handler, "_db_ready_last_error", "")
        self._old_allow_reuse_port = getattr(server.ThreadingHTTPServer, "allow_reuse_port", False)
        self._old_allow_reuse_address = getattr(server.ThreadingHTTPServer, "allow_reuse_address", True)
        server.ThreadingHTTPServer.allow_reuse_port = False
        server.ThreadingHTTPServer.allow_reuse_address = True
        bootstrap_conn = sqlite3.connect(self.db_path)
        bootstrap_conn.executescript(
            """
            CREATE TABLE empresas (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE workspace_companies (
              id TEXT PRIMARY KEY,
              workspace_id TEXT NOT NULL,
              legacy_empresa_id TEXT,
              nombre TEXT,
              activo INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE clientes (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              nombre TEXT,
              nif TEXT,
              estado TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            );
            CREATE TABLE clientes_empresas (
              id TEXT PRIMARY KEY,
              cliente_id TEXT,
              empresa_id TEXT,
              servicio TEXT,
              estado TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE cliente_gestoria (
              id TEXT PRIMARY KEY,
              cliente_id TEXT UNIQUE,
              mod_renta INTEGER,
              renta_detalles TEXT,
              created_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            );
            CREATE TABLE gestoria_terceros (
              id TEXT PRIMARY KEY,
              nombre TEXT,
              nif TEXT
            );
            CREATE TABLE gestoria_facturas (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              cliente_id TEXT,
              tercero_id TEXT,
              tipo TEXT,
              numero TEXT,
              fecha_emision TEXT,
              descripcion TEXT,
              base_imponible REAL,
              cuota_iva REAL,
              cuota_irpf REAL,
              total REAL,
              iva_pct REAL,
              doc_key TEXT,
              estado_ocr TEXT,
              archivo_hash TEXT,
              dedupe_key TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE gestoria_asientos (
              id TEXT PRIMARY KEY,
              empresa_id TEXT,
              cliente_id TEXT,
              factura_id TEXT,
              fecha TEXT,
              concepto TEXT,
              referencia TEXT,
              created_at TEXT,
              updated_at TEXT
            );
            CREATE TABLE gestoria_asiento_lineas (
              id TEXT PRIMARY KEY,
              asiento_id TEXT,
              tercero_id TEXT,
              cuenta TEXT,
              descripcion TEXT,
              debe REAL,
              haber REAL,
              impuesto_tipo TEXT,
              impuesto_pct REAL
            );
            """
        )
        bootstrap_conn.commit()
        bootstrap_conn.close()
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        server.Handler._db_ready = True
        server.Handler._db_ready_last_error = ""
        self._seed_gestoria_workspace()

        server.Handler.db_path = str(self.db_path)
        server.Handler.ocr_db_path = str(Path(self.tmpdir.name) / "ocr.sqlite")
        server.Handler._gestoria_dashboard_cache.clear()

        self._session_data = {"user_id": "u-admin", "rol": "ADMINISTRADOR", "servicio": "Gestoría"}
        self._session_patch = mock.patch.object(server.Handler, "_current_session", lambda _handler: self._session_data)
        self._session_patch.start()
        self.base_url = "http://127.0.0.1"

    def _call_gestoria_route(self, path):
        from io import BytesIO

        class DirectHandler(server.Handler):
            def __init__(self, route_path, session_data):
                self.path = route_path
                self.command = "GET"
                self.request_version = "HTTP/1.1"
                self.headers = {}
                self.rfile = None
                self.wfile = BytesIO()
                self.server = SimpleNamespace(server_name="127.0.0.1", server_port=0)
                self.client_address = ("127.0.0.1", 0)
                self._session_data = session_data
                self._status = 0
                self._headers = []
                self.close_connection = True

            def _current_session(self):
                return self._session_data

            def send_response(self, code, message=None):
                self._status = int(code or 0)

            def send_header(self, key, value):
                self._headers.append((str(key), str(value)))

            def end_headers(self):
                return None

            def log_message(self, *_args, **_kwargs):
                return None

        handler = DirectHandler(path, self._session_data)
        server.Handler.do_GET(handler)
        headers = {key: value for key, value in handler._headers}
        return handler._status, headers, handler.wfile.getvalue()

    def tearDown(self):
        try:
            self._session_patch.stop()
        except Exception:
            pass
        try:
            self.conn.close()
        except Exception:
            pass
        try:
            server.Handler._db_ready = self._old_db_ready
            server.Handler._db_ready_last_error = self._old_db_ready_last_error
        except Exception:
            pass
        try:
            self.tmpdir.cleanup()
        except Exception:
            pass
        try:
            server.ThreadingHTTPServer.allow_reuse_port = self._old_allow_reuse_port
            server.ThreadingHTTPServer.allow_reuse_address = self._old_allow_reuse_address
        except Exception:
            pass

    def _seed_gestoria_workspace(self):
        now = "2026-07-22T10:00:00+00:00"
        self.conn.execute(
            """
            INSERT INTO workspace_companies (id, workspace_id, legacy_empresa_id, nombre, activo)
            VALUES (?, ?, ?, ?, 1)
            """,
            ("wc-1", "ws-1", "emp-1", "Empresa Uno"),
        )
        self.conn.execute(
            """
            INSERT INTO clientes (id, empresa_id, nombre, nif, estado, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'Activo', datetime(?), datetime(?))
            """,
            ("cli-1", "emp-1", "Cliente Uno", "12345678Z", now, now),
        )
        self.conn.execute(
            """
            INSERT INTO cliente_gestoria (id, cliente_id, mod_renta, created_at, updated_at)
            VALUES (?, ?, 1, datetime(?), datetime(?))
            """,
            ("cg-1", "cli-1", now, now),
        )
        self.conn.execute(
            """
            INSERT INTO gestoria_terceros (id, nombre, nif)
            VALUES (?, ?, ?)
            """,
            ("t-1", "Proveedor Uno", "B12345678"),
        )
        self.conn.execute(
            """
            INSERT INTO gestoria_facturas (
              id, empresa_id, cliente_id, tercero_id, tipo, numero, fecha_emision, descripcion,
              base_imponible, cuota_iva, cuota_irpf, total, iva_pct, doc_key,
              estado_ocr, archivo_hash, dedupe_key, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "f-1",
                "emp-1",
                "cli-1",
                "t-1",
                "compra",
                "F-2026-001",
                "2026-07-11",
                "Factura de prueba",
                100.0,
                21.0,
                0.0,
                121.0,
                21.0,
                "doc-1",
                "pendiente",
                "hash-1",
                "",
                now,
                now,
            ),
        )
        for idx in range(2, 5):
            self.conn.execute(
                """
                INSERT INTO cliente_gestoria (id, cliente_id, mod_renta, created_at, updated_at)
                VALUES (?, ?, 1, datetime(?), datetime(?))
                """,
                (f"cg-extra-{idx}", f"renta-global-{idx}", now, now),
            )
        self.conn.commit()

    def test_gestoria_excel_plantilla_route_returns_xlsx(self):
        status, headers, body = self._call_gestoria_route("/api/gestoria_excel_plantilla?workspace_id=ws-1&cliente_id=cli-1")
        self.assertEqual(status, 200)
        self.assertEqual(
            headers.get("Content-Type"),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        disposition = headers.get("Content-Disposition") or ""
        self.assertIn("plantilla_conversor_asientos", disposition)
        self.assertGreater(len(body), 0)
        self.assertEqual(body[:2], b"PK")
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(body), data_only=False)
        self.assertIn("Listado Facturas", wb.sheetnames)
        self.assertIn("Control IVA", wb.sheetnames)
        self.assertEqual(wb["Listado Facturas"]["A2"].value, 1)
        self.assertEqual(wb["Listado Facturas"]["D2"].value, "F-2026-001")
        self.assertEqual(wb["Control IVA"]["B2"].value, '=COUNTIF(\'Listado Facturas\'!C2:C2,"compra")')

    def test_gestoria_dashboard_cache_is_scoped_by_access_mode(self):
        self._session_data = {"user_id": "u-admin", "rol": "ADMINISTRADOR", "servicio": "Gestoría"}
        admin_status, admin_headers, admin_body = self._call_gestoria_route("/api/gestoria_dashboard?workspace_id=ws-1")
        self.assertEqual(admin_status, 200)
        self.assertEqual(admin_headers.get("Content-Type"), "application/json; charset=utf-8")
        admin_payload = json.loads(admin_body.decode("utf-8"))

        self._session_data = {"user_id": "u-member", "rol": "Miembro", "servicio": "Gestoría"}
        member_status, member_headers, member_body = self._call_gestoria_route("/api/gestoria_dashboard?workspace_id=ws-1")
        self.assertEqual(member_status, 200)
        self.assertEqual(member_headers.get("Content-Type"), "application/json; charset=utf-8")
        member_payload = json.loads(member_body.decode("utf-8"))

        self.assertEqual(admin_payload["counts"]["total"], 1)
        self.assertEqual(member_payload["counts"]["total"], 1)
        self.assertEqual(admin_payload["counts"]["clientes_renta_global"], 4)
        self.assertEqual(member_payload["counts"]["clientes_renta_global"], 4)
        self.assertEqual(
            set(server.Handler._gestoria_dashboard_cache.keys()),
            {"emp-1::full", "emp-1::limited"},
        )
