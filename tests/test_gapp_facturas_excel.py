import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.build_gapp_facturas_excel import (
    build_category_totals,
    canonical_supplier_from_reason,
    canonical_supplier_name,
    canonical_filename_stem,
    classify_record,
    detect_document_type,
    enrich_parsed,
    extract_totals_triplet,
    extract_fallback_total,
    import_records_to_local_db,
    looks_like_own_company,
    mark_filename_duplicates,
    preclassify_from_filename,
    review_record,
    write_template_output,
)
from web.server import ensure_tables, open_sqlite_conn


class GappFacturasExcelTests(unittest.TestCase):
    def test_classify_alquiler_local(self):
        path = Path("/tmp/Factura Alquiler Local Enero 2026.pdf")
        parsed = {
            "tipo": "compra",
            "tercero": "Activos Inmobiliarios Gildusa SL",
            "descripcion": "Arrendamiento garaje Enero 2026",
            "raw_text": "Arrendamiento garaje situado... RENTA MENSUAL 180,00 EUR",
        }
        category, confidence, reason = classify_record(path, parsed)
        self.assertEqual(category, "ALQUILER LOCAL")
        self.assertGreaterEqual(confidence, 0.9)
        self.assertTrue("regla" in reason or "proveedor" in reason)

    def test_classify_restaurante_as_out_of_scope(self):
        path = Path("/tmp/Comida 05.02.2026.jpeg")
        parsed = {
            "tipo": "compra",
            "tercero": "Restaurante Hecho y Miel",
            "descripcion": "Ticket comida",
            "raw_text": "RESTAURANTE TOTAL 26,50",
        }
        category, _confidence, _reason = classify_record(path, parsed)
        self.assertEqual(category, "COMIDAS")

    def test_enrich_uses_filename_date_and_total(self):
        path = Path("/tmp/parking 30.01.2026.jpeg")
        parsed = enrich_parsed(
            path,
            "Factura Simplificada 104750\nImporte Pagado: 3,80 EUR",
            {"tipo": "compra", "base_imponible": 0.0, "cuota_iva": 0.0, "total": 0.0},
        )
        self.assertEqual(parsed["fecha"], "2026-01-30")
        self.assertAlmostEqual(parsed["total"], 3.8)
        self.assertTrue(parsed["tercero"])

    def test_enrich_detects_total_iva_incluido(self):
        path = Path("/tmp/Factura Alquiler Local Enero 2026.pdf")
        parsed = enrich_parsed(
            path,
            "Arrendamiento garaje Enero 2026\nTOTAL IVA INCLUIDO: 180,00 EUR",
            {"tipo": "compra", "base_imponible": 0.0, "cuota_iva": 0.0, "total": 0.0},
        )
        self.assertAlmostEqual(parsed["total"], 180.0)

    def test_client_word_does_not_force_ingreso(self):
        path = Path("/tmp/factura combustible 23-02-2026.pdf")
        parsed = {
            "tipo": "venta",
            "tercero": "PETROPRIX ENERGIA S.L.",
            "descripcion": "Factura combustible",
            "raw_text": "CLIENTE: GAPP MONTAJES\nDiesel\nTotal factura 50,00",
        }
        category, _confidence, _reason = classify_record(path, parsed)
        self.assertNotEqual(category, "INGRESO")

    def test_detect_document_type_marks_supplier_invoice_as_compra(self):
        path = Path("/tmp/Factura_2026-01-21_FM26_571.pdf")
        kind = detect_document_type(path, "Cliente GAPP\nTOTAL FACTURA 45,79", {"tipo": "venta"})
        self.assertEqual(kind, "compra")

    def test_detect_document_type_marks_materials_image_as_compra(self):
        path = Path("/tmp/Materiales 26.02.2026.jpeg")
        kind = detect_document_type(path, "Cliente GAPP\nimporte 12,49", {"tipo": "venta"})
        self.assertEqual(kind, "compra")

    def test_prefilter_marks_materiales_as_suministros(self):
        path = Path("/tmp/Materiales 26.02.2026.jpeg")
        category, confidence, _reason = preclassify_from_filename(path)
        self.assertEqual(category, "SUMINISTROS")
        self.assertGreaterEqual(confidence, 0.9)

    def test_classify_supplier_rule_for_suministros(self):
        path = Path("/tmp/Factura HTM 13.01.2026.pdf")
        parsed = {
            "tipo": "compra",
            "tercero": "HTM",
            "descripcion": "Material ferreteria",
            "raw_text": "HTM tornillos brocas TOTAL FACTURA 59,47",
        }
        category, confidence, reason = classify_record(path, parsed)
        self.assertEqual(category, "SUMINISTROS")
        self.assertGreaterEqual(confidence, 0.9)
        self.assertIn("proveedor", reason)

    def test_own_company_name_is_detected(self):
        self.assertTrue(looks_like_own_company("GAPP ELEVADORES S.L"))
        self.assertFalse(looks_like_own_company("OPTIMUS TINEO S. PEDRO S.L."))

    def test_canonical_supplier_name_normalizes_known_alias(self):
        self.assertEqual(canonical_supplier_name("OPTIMUS TINEO s. pEpp,"), "OPTIMUS TINEO S. PEDRO S.L")
        self.assertEqual(canonical_supplier_name("ACTIVOS INMOBILIARIOS GILDUSA"), "ACTIVOS INMOVILIARIOS GILDUSA S.L")

    def test_canonical_supplier_from_reason_uses_supplier_rule(self):
        self.assertEqual(canonical_supplier_from_reason("proveedor:LEROY"), "LEROY MERLIN MARBELLA")

    def test_extract_fallback_total_picks_largest_positive_amount(self):
        text = "Subtotal 53,33 IVA 11,20 Total 64,53"
        self.assertAlmostEqual(extract_fallback_total(text), 64.53)

    def test_extract_totals_triplet_reads_summary_line(self):
        text = """
        DESGLOSE TOTALES
        FECHA METODO DE PAGO CANT. TASA TOTALBI TOTALIVA TOTAL
        26/01/2026 Tarjeta Bancaria SA 62.07 1 IVA 21% 51.30 10.77 62.07
        """
        self.assertEqual(extract_totals_triplet(text), (51.3, 10.77, 62.07))

    def test_review_record_marks_suspicious_ocr_amount(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 296.61,
                "confianza_categoria": 0.92,
                "ocr_metodo": "ocr_image_file",
                "tercero": "PROVEEDOR DESCONOCIDO",
                "fecha": "2026-02-19",
                "motivo_categoria": "regla:SUMINISTRO",
                "archivo": "Materiales 19.02.2026.jpeg",
            },
            target_year=2026,
        )
        self.assertEqual(state, "REVISAR")
        self.assertIn("importe_alto_por_ocr", reasons)

    def test_review_record_trusted_supplier_can_skip_high_ocr_flag(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 296.61,
                "confianza_categoria": 0.96,
                "ocr_metodo": "ocr_image_file",
                "tercero": "LEROY MERLIN MARBELLA",
                "fecha": "2026-02-19",
                "motivo_categoria": "proveedor:LEROY",
                "archivo": "Materiales 19.02.2026.jpeg",
            },
            target_year=2026,
        )
        self.assertEqual(state, "OK")
        self.assertEqual(reasons, "")

    def test_review_record_marks_implausible_vat(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 9.0,
                "confianza_categoria": 0.96,
                "ocr_metodo": "ocr_image_file",
                "tercero": "OBRAMAT",
                "fecha": "2026-01-09",
                "motivo_categoria": "proveedor:OBRAMAT",
                "archivo": "Factura Obramat 09.01.2026.jpg.jpeg",
                "numero": "09",
                "tipo": "compra",
                "base_imponible": 9.0,
                "cuota_iva": 21.0,
                "total": 9.0,
            },
            target_year=2026,
        )
        self.assertEqual(state, "REVISAR")
        self.assertIn("iva_inverosimil", reasons)

    def test_review_record_marks_expense_detected_as_sale(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 40.39,
                "confianza_categoria": 0.96,
                "ocr_metodo": "ocr_image_file",
                "tercero": "COMASUR",
                "fecha": "2026-02-09",
                "motivo_categoria": "proveedor:COMASUR",
                "archivo": "Material limpieza 9.02.2026.jpeg",
                "numero": "GAPP",
                "tipo": "venta",
                "base_imponible": 40.39,
                "cuota_iva": 0.0,
                "total": 40.39,
            },
            target_year=2026,
        )
        self.assertEqual(state, "REVISAR")
        self.assertIn("tipo_venta_en_gasto", reasons)

    def test_review_record_marks_suspicious_invoice_number(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 99.55,
                "confianza_categoria": 0.96,
                "ocr_metodo": "ocr_image_file",
                "tercero": "OBRAMAT",
                "fecha": "2026-03-02",
                "motivo_categoria": "proveedor:OBRAMAT",
                "archivo": "Obramat 2.03.2026.jpeg",
                "numero": "emitida",
                "tipo": "compra",
                "base_imponible": 99.55,
                "cuota_iva": 0.0,
                "total": 99.55,
            },
            target_year=2026,
        )
        self.assertEqual(state, "REVISAR")
        self.assertIn("numero_dudoso", reasons)

    def test_enrich_uses_triplet_to_fix_implausible_vat(self):
        path = Path("/tmp/Factura Obramat 26.01.2026.jpg")
        parsed = enrich_parsed(
            path,
            """
            OBRAMAT
            DESGLOSE TOTALES
            FECHA METODO DE PAGO CANT. TASA TOTALBI TOTALIVA TOTAL
            26/01/2026 Tarjeta Bancaria SA 62.07 1 IVA 21% 51.30 10.77 62.07
            """,
            {"tipo": "compra", "base_imponible": 62.07, "cuota_iva": 21.0, "total": 62.07, "tercero": "Original ,"},
        )
        self.assertEqual(parsed["tercero"], "OBRAMAT")
        self.assertAlmostEqual(parsed["base_imponible"], 51.3, places=2)
        self.assertAlmostEqual(parsed["cuota_iva"], 10.77, places=2)
        self.assertAlmostEqual(parsed["total"], 62.07, places=2)

    def test_enrich_detects_supplier_company_suffix(self):
        path = Path("/tmp/Materiales 2 - 02.03.2026.jpeg")
        parsed = enrich_parsed(
            path,
            """
            Factura Simplificada
            Jiménez Maña Recambios, S.L.U. Store
            TOTAL DOCUMENTO
            4,22 21,0 0,89 0,00 5,11€
            """,
            {"tipo": "compra"},
        )
        self.assertIn("Jim", parsed["tercero"])

    def test_review_record_does_not_flag_bad_vendor_if_supplier_rule_is_trusted(self):
        state, reasons = review_record(
            {
                "categoria_excel": "SUMINISTROS",
                "importe_agregado": 25.69,
                "confianza_categoria": 0.96,
                "ocr_metodo": "ocr_image_file",
                "tercero": "a7? @ &",
                "fecha": "2026-02-05",
                "motivo_categoria": "proveedor:LEROY",
                "archivo": "Factura Leroy 5-02-2026.jpeg",
                "numero": "036-000s",
                "tipo": "compra",
                "base_imponible": 25.69,
                "cuota_iva": 0.0,
                "total": 25.69,
            },
            target_year=2026,
        )
        self.assertEqual(state, "OK")
        self.assertNotIn("proveedor_dudoso", reasons)

    def test_review_record_ignores_non_template_categories(self):
        state, reasons = review_record(
            {
                "categoria_excel": "VIAJE",
                "importe_agregado": 0.0,
                "confianza_categoria": 0.5,
                "ocr_metodo": "ocr_image_file",
                "tercero": "",
                "fecha": "2025-02-24",
                "motivo_categoria": "prefiltro_nombre:HOTEL",
                "archivo": "hotel 24-25.02.2026.jpeg",
            },
            target_year=2026,
        )
        self.assertEqual(state, "OK")
        self.assertEqual(reasons, "")

    def test_build_category_totals_excludes_review_by_default(self):
        totals = build_category_totals(
            [
                {"categoria_excel": "SUMINISTROS", "importe_agregado": 100.0, "estado_revision": "OK"},
                {"categoria_excel": "SUMINISTROS", "importe_agregado": 50.0, "estado_revision": "REVISAR"},
            ]
        )
        self.assertEqual(totals, {"SUMINISTROS": 100.0})

    def test_build_category_totals_can_include_review(self):
        totals = build_category_totals(
            [
                {"categoria_excel": "SUMINISTROS", "importe_agregado": 100.0, "estado_revision": "OK"},
                {"categoria_excel": "SUMINISTROS", "importe_agregado": 50.0, "estado_revision": "REVISAR"},
            ],
            include_needs_review=True,
        )
        self.assertEqual(totals, {"SUMINISTROS": 150.0})

    def test_canonical_filename_stem_removes_copy_suffix(self):
        self.assertEqual(
            canonical_filename_stem("Materiales 19.02.2026 (2).jpeg"),
            canonical_filename_stem("Materiales 19.02.2026.jpeg"),
        )

    def test_mark_filename_duplicates_marks_copy_record(self):
        rows = [
            {
                "archivo": "Materiales 19.02.2026.jpeg",
                "fecha": "2026-02-19",
                "categoria_excel": "SUMINISTROS",
                "estado_revision": "OK",
                "motivos_revision": "",
                "importe_agregado": 71.99,
                "confianza_categoria": 0.96,
            },
            {
                "archivo": "Materiales 19.02.2026 (2).jpeg",
                "fecha": "2026-02-19",
                "categoria_excel": "SUMINISTROS",
                "estado_revision": "REVISAR",
                "motivos_revision": "importe_alto_por_ocr",
                "importe_agregado": 296.61,
                "confianza_categoria": 0.99,
            },
        ]
        mark_filename_duplicates(rows)
        self.assertEqual(rows[0]["estado_revision"], "OK")
        self.assertEqual(rows[1]["estado_revision"], "DUPLICADO")

    def test_write_template_output_replaces_category_values(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            template = Path(tmpdir) / "template.xlsx"
            output = Path(tmpdir) / "output.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Hoja1"
            ws["A1"] = "GAPP"
            ws["A2"] = "SEGURO LOCAL"
            ws["B2"] = 0
            ws["A3"] = "ALQUILER LOCAL"
            ws["B3"] = 0
            wb.save(template)

            write_template_output(template, output, {"SEGURO LOCAL": 120.5, "ALQUILER LOCAL": 300.0})

            result = load_workbook(output)
            sheet = result["Hoja1"]
            self.assertAlmostEqual(sheet["B2"].value, 120.5)
            self.assertAlmostEqual(sheet["B3"].value, 300.0)

    def test_build_category_totals_ignores_non_template_categories(self):
        totals = build_category_totals(
            [
                {"categoria_excel": "ALQUILER LOCAL", "importe_agregado": 180.0},
                {"categoria_excel": "ALQUILER LOCAL", "importe_agregado": 120.0},
                {"categoria_excel": "COMIDAS", "importe_agregado": 50.0},
            ]
        )
        self.assertEqual(totals, {"ALQUILER LOCAL": 300.0})

    def test_import_records_to_local_db_creates_lote(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "crm.sqlite"
            ensure_tables(db_path)
            conn = open_sqlite_conn(db_path, with_row_factory=True)
            try:
                conn.execute(
                    "INSERT INTO empresas (id, nombre, created_at, updated_at) VALUES (?, ?, datetime('now'), datetime('now'))",
                    ("e1", "Empresa Demo"),
                )
                conn.execute(
                    "INSERT INTO clientes (id, nombre, created_at, updated_at) VALUES (?, ?, datetime('now'), datetime('now'))",
                    ("c1", "Cliente Demo"),
                )
                conn.commit()
            finally:
                conn.close()

            result = import_records_to_local_db(
                db_path=db_path,
                empresa_ref="Empresa Demo",
                cliente_id="c1",
                periodo="2026-03",
                records=[
                    {
                        "archivo": "factura_luz.pdf",
                        "fecha": "2026-03-08",
                        "numero": "F-2026-88",
                        "tercero": "Comercial Electrica",
                        "nif": "B76543210",
                        "tipo": "compra",
                        "base_imponible": 200.0,
                        "cuota_iva": 42.0,
                        "total": 242.0,
                        "categoria_excel": "SUMINISTROS",
                        "descripcion": "Factura luz marzo",
                        "estado_revision": "OK",
                    }
                ],
                apply_ok=False,
            )
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["applied"], 0)

            conn = open_sqlite_conn(db_path, with_row_factory=True)
            try:
                lote = conn.execute(
                    "SELECT total_documentos, total_ok, total_revisar FROM gestoria_import_lotes WHERE id = ?",
                    (result["lote_id"],),
                ).fetchone()
                self.assertEqual(lote["total_documentos"], 1)
                self.assertEqual(lote["total_ok"], 1)
                self.assertEqual(lote["total_revisar"], 0)
            finally:
                conn.close()

    def test_import_records_to_local_db_can_apply_ok(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "crm.sqlite"
            ensure_tables(db_path)
            conn = open_sqlite_conn(db_path, with_row_factory=True)
            try:
                conn.execute(
                    "INSERT INTO empresas (id, nombre, created_at, updated_at) VALUES (?, ?, datetime('now'), datetime('now'))",
                    ("e1", "Empresa Demo"),
                )
                conn.execute(
                    "INSERT INTO clientes (id, nombre, created_at, updated_at) VALUES (?, ?, datetime('now'), datetime('now'))",
                    ("c1", "Cliente Demo"),
                )
                conn.commit()
            finally:
                conn.close()

            result = import_records_to_local_db(
                db_path=db_path,
                empresa_ref="e1",
                cliente_id="c1",
                periodo="2026-03",
                records=[
                    {
                        "archivo": "factura_alquiler.pdf",
                        "fecha": "2026-03-01",
                        "numero": "ALQ-03",
                        "tercero": "Activos Inmobiliarios Gildusa",
                        "nif": "B12345678",
                        "tipo": "compra",
                        "base_imponible": 180.0,
                        "cuota_iva": 0.0,
                        "total": 180.0,
                        "categoria_excel": "ALQUILER LOCAL",
                        "descripcion": "Alquiler marzo",
                        "estado_revision": "OK",
                    }
                ],
                apply_ok=True,
            )
            self.assertEqual(result["inserted"], 1)
            self.assertEqual(result["applied"], 1)
            self.assertEqual(result["apply_errors"], 0)

            conn = open_sqlite_conn(db_path, with_row_factory=True)
            try:
                factura = conn.execute(
                    "SELECT numero, origen_importacion FROM gestoria_facturas LIMIT 1"
                ).fetchone()
                asiento = conn.execute(
                    "SELECT total_debe, total_haber FROM gestoria_asientos LIMIT 1"
                ).fetchone()
                self.assertEqual(factura["numero"], "ALQ-03")
                self.assertEqual(factura["origen_importacion"], "gestoria_import")
                self.assertAlmostEqual(float(asiento["total_debe"] or 0), 180.0, places=2)
                self.assertAlmostEqual(float(asiento["total_haber"] or 0), 180.0, places=2)
            finally:
                conn.close()


if __name__ == "__main__":
    unittest.main()
