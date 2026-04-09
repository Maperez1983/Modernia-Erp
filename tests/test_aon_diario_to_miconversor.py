import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.import_aon_diario_to_miconversor import main


class AonDiarioToMiConversorTests(unittest.TestCase):
    def _build_template(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        ws.append(
            [
                "FECHA ASIENTO",
                "FECHA FACTURA",
                "Nº FACTURA",
                "CONCEPTO",
                "SUBCUENTA",
                "NIF",
                "NOMBRE",
                "DOMICILIO",
                "LOCALIDAD",
                "PROVINCIA",
                "CODIGO POSTAL",
                "BASE IMPONIBLE",
                "% IVA",
                "IMPORTE IVA",
                "SUBCUENTA GASTOS/INGRESOS",
                "IMPORTE (TOTAL)",
            ]
        )
        ws.append([None, None, None, '=+CONCATENATE(C2," ",G2)', None, None, None, None, None, None, None, None, None, None, None, None])
        wb.save(path)

    def _build_diario(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Diario"
        ws.append(["LISTADO DIARIO DE MOVIMIENTOS"])
        ws.append(["CUENTA", "DESCRIPCIÓN", "CONCEPTO", "DEBE", "HABER", "CONTRAP.", "DOCUMENTO"])
        ws.append(["Fecha: 21/02/2025 Nº diario: 1"])
        # Venta: base 2000, IVA 420 (21%), retención 380, total 2040
        ws.append(["430000001", "CLIENTE EJEMPLO SL", "N/Fra: 000001", 2040.0, 0.0, "700000000", "E-000001"])
        ws.append(["477000000", "Hacienda Pública, IVA repercutido.", "N/Fra: 000001", 0.0, 420.0, "430000001", "E-000001"])
        ws.append(["473000000", "Hacienda Pública, retenciones y pagos a cuenta.", "N/Fra: 000001", 380.0, 0.0, "430000001", "E-000001"])
        ws.append(["700000000", "Ventas de mercaderías.", "N/Fra: 000001", 0.0, 2000.0, "430000001", "E-000001"])
        wb.save(path)

    def test_builds_output_with_control_sheets(self):
        with tempfile.TemporaryDirectory() as td:
            td = Path(td)
            root = td / "root"
            folder = root / "A001"
            folder.mkdir(parents=True)
            diario = folder / "DIARIO.xlsx"
            template = td / "template.xlsx"
            out_dir = td / "out"
            self._build_diario(diario)
            self._build_template(template)

            rc = main(["--root", str(root), "--template", str(template), "--out-dir", str(out_dir)])
            self.assertEqual(rc, 0)

            outs = list(out_dir.glob("*.xlsx"))
            self.assertEqual(len(outs), 1)
            out_path = outs[0]

            wb = load_workbook(out_path, data_only=True)
            self.assertIn("Hoja1", wb.sheetnames)
            self.assertIn("CONTROL_LISTADO", wb.sheetnames)
            self.assertIn("CONTROL_TOTALIZADOR", wb.sheetnames)

            ws = wb["Hoja1"]
            # header + 1 data row
            self.assertGreaterEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(2, 3).value, "000001")
            self.assertAlmostEqual(float(ws.cell(2, 12).value), 2000.0, places=2)
            self.assertAlmostEqual(float(ws.cell(2, 14).value), 420.0, places=2)
            self.assertAlmostEqual(float(ws.cell(2, 16).value), 2040.0, places=2)

            wst = wb["CONTROL_TOTALIZADOR"]
            self.assertEqual(int(wst["B3"].value), 1)
            self.assertAlmostEqual(float(wst["B4"].value), 2000.0, places=2)
            self.assertAlmostEqual(float(wst["B5"].value), 420.0, places=2)
            self.assertAlmostEqual(float(wst["B6"].value), 380.0, places=2)
            self.assertAlmostEqual(float(wst["B7"].value), 2040.0, places=2)

